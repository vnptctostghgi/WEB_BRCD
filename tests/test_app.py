import os
import json
import base64
import importlib.util
import sqlite3
import threading
import time
import uuid
from io import BytesIO
from datetime import UTC, datetime, timedelta
from pathlib import Path
from zipfile import ZipFile

import openpyxl
import pytest

os.environ["DB_MOCK_MODE"] = "true"
os.environ["INTERNAL_API_MOCK_MODE"] = "true"
os.environ["INTERNAL_API_URL"] = "http://10.92.17.88:8000/api/du-lieu-web"
os.environ["INTERNAL_API_TOKEN"] = "test-worker-token"
os.environ["APP_DATABASE_BACKEND"] = "sqlite"
os.environ["APP_DATABASE_PATH"] = "data/test_app.db"
os.environ["INITIAL_ADMIN_USERNAME"] = "admin"
os.environ["INITIAL_ADMIN_PASSWORD"] = "Admin@Brcd2026!"

test_database = Path("data/test_app.db")
test_database.unlink(missing_ok=True)

from fastapi.testclient import TestClient

from app.application.telegram_notifier import sanitize_alert_details, sanitize_alert_text
from app.application.database_service import DatabaseService
from app.data_access.supabase_repository import SupabaseRepository
from app.main import app
from app.presentation import routes
from app.settings import Settings, get_settings


def load_api_middleware_module():
    module_path = Path("docs/api_trung_gian_drive_export.py")
    spec = importlib.util.spec_from_file_location("api_trung_gian_drive_export_test", module_path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def login(client: TestClient, username: str = "admin", password: str = "Admin@Brcd2026!") -> None:
    response = client.post("/api/auth/login", json={"username": username, "password": password})
    assert response.status_code == 200


def test_unauthenticated_user_is_redirected_to_login() -> None:
    with TestClient(app) as client:
        response = client.get("/", follow_redirects=False)
        assert response.status_code == 303
        assert response.headers["location"] == "/login"


def test_login_uses_optimized_static_assets() -> None:
    with TestClient(app) as client:
        response = client.get("/login")
        assert response.status_code == 200
        assert "/static/tailwind-built.css?v=1" in response.text
        assert "cdn.tailwindcss.com" not in response.text
        assert "/static/login-hero-900.webp" in response.text
        assert "/static/images/system-logo-96.webp" in response.text
        assert "/static/login.js?v=10" in response.text
        login_js = client.get("/static/login.js?v=10")
        assert login_js.status_code == 200
        assert "function safeNextPath" in login_js.text
        assert 'return "/";' in login_js.text


def test_authenticated_login_page_returns_empty_shell() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.get("/login?next=/publicmessages", follow_redirects=False)
        assert response.status_code == 303
        assert response.headers["location"] == "/"


def test_admin_can_login_and_open_dashboard() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.get("/")
        assert response.status_code == 200
        assert 'rel="icon" type="image/png" href="/static/images/system-logo.png"' in response.text
        assert "/static/tailwind-built.css?v=1" in response.text
        assert "cdn.tailwindcss.com" not in response.text
        assert "dashboard-tab-fiber" not in response.text
        assert 'data-feature-code="internalemail"' in response.text
        assert 'data-feature-code="publicmessages"' in response.text
        assert "Đào dữ liệu SQL" in response.text
        assert "Báo cáo mới" in response.text
        assert "Quản trị người dùng" in response.text


def test_feature_path_opens_current_app_shell() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.get("/quantrimenu")
        assert response.status_code == 200
        assert 'data-feature-code="quantrimenu"' in response.text
        assert 'class="app-view menu-admin-view"' in response.text
        assert 'class="menu-admin-toolbar"' in response.text
        assert 'id="menu-layout-search"' in response.text
        assert 'id="create-menu"' in response.text
        assert "Menu / chức năng" in response.text
        assert "Lưu menu" in response.text

        email_response = client.get("/internalemail")
        assert email_response.status_code == 200
        assert 'id="view-internal-email"' in email_response.text
        assert 'data-internal-email-tab="messages"' in email_response.text
        assert 'data-internal-email-tab="email"' in email_response.text
        assert 'data-internal-email-panel="messages"' in email_response.text
        assert 'data-internal-email-panel="email"' in email_response.text
        assert 'id="internal-email-otp-rule-form"' in email_response.text
        assert 'id="internal-email-otp-rules-table"' in email_response.text
        assert 'id="internal-email-refresh-existing"' in email_response.text
        assert 'data-mobile-tab="email"' not in email_response.text

        public_response = client.get("/publicmessages")
        assert public_response.status_code == 200
        assert 'id="view-public-messages"' in public_response.text
        assert "/static/app.js?v=227" in public_response.text
        assert "/static/styles.css?v=138" in public_response.text
        assert "fonts.googleapis.com" not in public_response.text
        assert 'href="/api/navigation"' not in public_response.text
        public_js = client.get("/static/app.js?v=227")
        assert public_js.status_code == 200
        assert "function bindPublicMessagesEvents" in public_js.text
        assert "function renderPublicMessages" in public_js.text
        assert "function startPublicMessagesAutoRefresh" in public_js.text
        assert "function generateUserPasswordFromDialog" in public_js.text
        assert "/generate-password" in public_js.text
        assert "Tạo mã dùng một lần" in public_js.text
        assert "Mật khẩu chính của tài khoản vẫn giữ nguyên" in public_js.text
        assert "function collapseNavigationTree" in public_js.text
        assert "function dedupeFeaturesForDisplay" in public_js.text
        assert "/static/internal-email.js?v=8" in public_js.text
        assert "function readCachedNavigation" in public_js.text
        assert "function fetchNavigation" in public_js.text
        assert "function warmSystemSecondarySections" in public_js.text
        assert 'event.target.closest("#create-menu")' in public_js.text
        assert "await createMenu(button)" in public_js.text
        assert "await loadMenuLayout({ focusCode:" in public_js.text
        assert 'menuLayoutSearchQuery = event.currentTarget.value || ""' in public_js.text
        assert "function filterMenuLayoutRows" in public_js.text
        assert "function normalizeMenuSearchText" in public_js.text
        assert "function fillDynamicReportSelect()" in public_js.text
        assert "function fillOneBssRunSelect()" in public_js.text
        assert "function oneBssIsRegionParameterKey(key)" in public_js.text
        assert '|| /^P_TINH$/i.test(text)' in public_js.text
        assert "Object.keys(output).find(oneBssIsRegionParameterKey)" in public_js.text
        assert '$("#new-dashboard-page")?.addEventListener("click", createDashboardPage)' in public_js.text
        assert 'event.target.closest("#save-dashboard-layout")' in public_js.text
        assert "await saveDashboardLayout(button)" in public_js.text
        assert '$("#add-dashboard-row")?.addEventListener("click", () => addDashboardRow($("#dashboard-row-type")?.value || "2_columns"))' in public_js.text
        assert "total_data_card" in public_js.text
        assert "function renderRuntimeTotalDataCardWidget" in public_js.text
        assert "data-save-google-drive-folder" in public_js.text
        assert "/api/google-drive/oauth/folder" in public_js.text
        assert "/api/google-drive/oauth/status" in public_js.text
        assert "Link lưu báo cáo" not in public_js.text
        assert 'data-inline-onebss-field="storage_link"' not in public_js.text
        assert "/static/workstation.js?v=6" in public_js.text
        assert "window.VNPTReportsRuntime?.fillOneBssRunSelect?.()" in public_js.text
        assert "/static/reports-runtime.js?v=15" in public_js.text
        reports_runtime_js = client.get("/static/reports-runtime.js?v=15")
        assert reports_runtime_js.status_code == 200
        assert "fillDynamicReportSelect, fillOneBssRunSelect }" in reports_runtime_js.text
        assert "fillOneBssRunSelect }" in reports_runtime_js.text
        assert "function dynamicReportProgressHtml" in reports_runtime_js.text
        assert "function dynamicReportProgressHint" in reports_runtime_js.text
        assert "function dynamicReportDateMs" in reports_runtime_js.text
        assert "worker_state" in reports_runtime_js.text
        assert "no_sql_worker" in reports_runtime_js.text
        assert "Chưa có máy trạm nhận lệnh" in reports_runtime_js.text
        assert "Máy trạm${workerId} đã nhận lệnh" in reports_runtime_js.text
        assert "progress_steps" in reports_runtime_js.text
        assert "client_request_id" in reports_runtime_js.text
        assert "dataset.onebssSubmitting" in reports_runtime_js.text
        assert "started_at_label" in reports_runtime_js.text
        workstation_js = client.get("/static/workstation.js?v=6")
        assert workstation_js.status_code == 200
        assert "worker.version" in workstation_js.text
        assert "worker.roles" in workstation_js.text
        assert 'data-workstation-action="test"' in workstation_js.text
        assert "/api/admin/workstation/profile" in workstation_js.text
        assert "oracle_config_ready" in workstation_js.text
        assert "Oracle dong bo" in workstation_js.text
        assert 'Promise.allSettled([' in public_js.text
        assert '$("#connection-picker")?.addEventListener("change", renderConnectionsTable)' in public_js.text
        assert "function permissionDisplayFeatures" in public_js.text
        assert "function isTechnicalPermissionFeature" in public_js.text
        assert "function renderPermissionTree" in public_js.text
        assert "function syncPermissionTreeForSelectedUsers" in public_js.text
        assert "function setAllPermissionChecks" in public_js.text
        assert "function setAllUserSelection" in public_js.text
        assert "filter((feature) => !isTechnicalPermissionFeature(feature))" in public_js.text
        assert "async function logoutFromClient" in public_js.text
        assert 'window.location.replace("/login")' in public_js.text
        assert "/api/admin/public-messages/feed?limit=100" not in public_js.text
        assert "const PUBLIC_MESSAGES_LIMIT = 10" in public_js.text
        assert 'params.set("after", publicMessagesCursor)' in public_js.text
        public_css = client.get("/static/styles.css?v=138")
        assert public_css.status_code == 200
        assert ".sql-progress-hint" in public_css.text
        assert "Compact desktop rail" in public_css.text
        assert ".sidebar:not(.menu-open) #main-navigation" in public_css.text
        assert ".permission-node" in public_css.text
        assert ".permission-children" in public_css.text
        assert ".dynamic-report-export-table a.table-action" in public_css.text
        assert ".sql-progress-step" in public_css.text
        assert ".runtime-total-data-card" in public_css.text
        assert ".password-admin-tools" in public_css.text
        assert "text-decoration-color: rgba(11, 99, 182, .42)" in public_css.text
        assert ".menu-admin-toolbar" in public_css.text
        assert ".ftp-run-form" in public_css.text
        assert ".ftp-source-item" in public_css.text
        assert ".menu-layout-table" in public_css.text
        assert ".menu-code-badge" in public_css.text
        assert "grid-template-columns: minmax(210px, .85fr) minmax(260px, 1fr) auto auto" in public_css.text

        permissions_response = client.get("/phanquyennguoidung")
        assert permissions_response.status_code == 200
        assert 'id="permission-selection-status"' in permissions_response.text
        assert 'data-permission-select="all"' in permissions_response.text
        assert 'data-permission-select="none"' in permissions_response.text
        assert 'data-user-select="all"' in permissions_response.text
        assert 'data-user-select="none"' in permissions_response.text


def test_admin_can_open_workstation_overview_and_download_setup_package() -> None:
    with TestClient(app) as client:
        login(client)
        overview = client.get("/api/admin/workstation/overview")
        assert overview.status_code == 200
        payload = overview.json()
        assert payload["hardware_profile"]["cpu"] == "Core i3"
        assert any(role["code"] == "onebss_worker" for role in payload["roles"])
        assert payload["setup"]["package_url"].startswith("/api/admin/workstation/setup-package?v=")
        assert payload["setup"]["package_version"] == routes.WORKSTATION_SETUP_PACKAGE_VERSION
        saved_oracle = client.put(
            "/api/admin/connections/oracle_agency_db",
            json={
                "name": "Oracle DB noi bo",
                "connection_type": "oracle",
                "description": "Oracle cho may tram",
                "config": {
                    "host": "10.10.10.20",
                    "port": 1521,
                    "service": "ONEBSS",
                    "username": "REPORT_USER",
                    "password": "oracle-secret",
                },
                "is_active": True,
            },
        )
        assert saved_oracle.status_code == 200
        configured_overview = client.get("/api/admin/workstation/overview")
        assert configured_overview.status_code == 200
        assert configured_overview.json()["config"]["oracle_config_ready"] is True
        assert configured_overview.json()["config"]["oracle_password_configured"] is True

        package = client.get("/api/admin/workstation/setup-package")
        assert package.status_code == 200
        assert package.headers["content-type"] == "application/zip"
        assert package.headers["cache-control"] == "no-store, max-age=0"
        assert routes.WORKSTATION_SETUP_PACKAGE_VERSION in package.headers["content-disposition"]
        workstation_page = client.get("/maytram")
        assert workstation_page.status_code == 200
        assert 'id="view-workstation"' in workstation_page.text
        assert f"/api/admin/workstation/setup-package?v={routes.WORKSTATION_SETUP_PACKAGE_VERSION}" in workstation_page.text
        assert "setup-package?v=20260730-synced-oracle-v5" not in workstation_page.text
        assert "setup-package?v=20260730-synced-oracle-v6" not in workstation_page.text
        assert "setup-package?v=20260730-synced-oracle-v7" not in workstation_page.text
        with ZipFile(BytesIO(package.content)) as archive:
            names = set(archive.namelist())
            config_text = archive.read("VNPTCTO_WORKSTATION_SETUP/workstation-install-config.ps1").decode("utf-8")
            readme_text = archive.read("VNPTCTO_WORKSTATION_SETUP/README_SETUP.txt").decode("utf-8")
            setup_bat = archive.read("VNPTCTO_WORKSTATION_SETUP/SETUP_VNPTCTO_WORKSTATION.bat").decode("utf-8")
            background_bat = archive.read("VNPTCTO_WORKSTATION_SETUP/START_ONEBSS_WORKER_BACKGROUND.bat").decode("utf-8")
            install_autostart_bat = archive.read("VNPTCTO_WORKSTATION_SETUP/INSTALL_ONEBSS_WORKER_AUTOSTART.bat").decode("utf-8")
            uninstall_autostart_bat = archive.read("VNPTCTO_WORKSTATION_SETUP/UNINSTALL_ONEBSS_WORKER_AUTOSTART.bat").decode("utf-8")
            setup_script = archive.read("VNPTCTO_WORKSTATION_SETUP/scripts/setup_vnptcto_workstation.ps1").decode("utf-8")
            install_task_script = archive.read("VNPTCTO_WORKSTATION_SETUP/scripts/install_onebss_worker_task.ps1").decode("utf-8")
            background_worker_script = archive.read("VNPTCTO_WORKSTATION_SETUP/scripts/run_onebss_worker_background.ps1").decode("utf-8")
            start_worker_script = archive.read("VNPTCTO_WORKSTATION_SETUP/scripts/start_onebss_worker.ps1").decode("utf-8")
            worker_script = archive.read("VNPTCTO_WORKSTATION_SETUP/scripts/onebss_workstation_worker.py").decode("utf-8")
            uninstall_task_script = archive.read("VNPTCTO_WORKSTATION_SETUP/scripts/uninstall_onebss_worker_task.ps1").decode("utf-8")
            health_script = archive.read("VNPTCTO_WORKSTATION_SETUP/scripts/test_vnptcto_workstation.ps1").decode("utf-8")
            api_task_script = archive.read("VNPTCTO_WORKSTATION_SETUP/docs/install_api_trung_gian_task.ps1").decode("utf-8")
        assert "VNPTCTO_WORKSTATION_SETUP/SETUP_VNPTCTO_WORKSTATION.bat" in names
        assert "VNPTCTO_WORKSTATION_SETUP/workstation-install-config.ps1" in names
        assert "VNPTCTO_WORKSTATION_SETUP/scripts/setup_vnptcto_workstation.ps1" in names
        assert "VNPTCTO_WORKSTATION_SETUP/scripts/test_vnptcto_workstation.ps1" in names
        assert "VNPTCTO_WORKSTATION_SETUP/scripts/run_onebss_worker_background.ps1" in names
        assert "InternalApiToken = 'test-worker-token'" in config_text
        assert "OracleDbDsn = '10.10.10.20:1521/ONEBSS'" in config_text
        assert "OracleDbHost = '10.10.10.20'" in config_text
        assert "OracleDbPort = '1521'" in config_text
        assert "OracleDbService = 'ONEBSS'" in config_text
        assert "OracleDbSid" in config_text
        assert "OracleDbUser = 'REPORT_USER'" in config_text
        assert "OracleDbPass = 'oracle-secret'" in config_text
        assert "OneBssTaskTimeoutSeconds = '1200'" in config_text
        assert "OneBssOtpWaitSeconds = '180'" in config_text
        assert "OneBssGridTimeoutSeconds = '90'" in config_text
        assert "OneBssProcessingTimeoutRetryAttempts = '3'" in config_text
        assert "OneBssProcessingTimeoutRetryDelaySeconds = '8'" in config_text
        assert "SqlWorkerTimeoutSeconds = '1800'" in config_text
        assert "WorkerMaxConcurrentTasks = '4'" in config_text
        assert "OneBssWorkerMaxTasks = '2'" in config_text
        assert "SqlWorkerMaxTasks = '2'" in config_text
        assert "FtpWorkerMaxTasks = '2'" in config_text
        assert "ExportPageSize = '20000'" in config_text
        assert "ExportMaxRows = '1000000'" in config_text
        assert "Khong can go token" in readme_text
        assert "Oracle" in readme_text
        assert "Read-Host $Prompt" not in setup_script
        assert "Nhap INTERNAL_API_TOKEN" not in setup_script
        assert "Test-PythonLauncher" in setup_script
        assert "Test-PythonLauncher" in start_worker_script
        assert "Python.Python.3.12" in start_worker_script
        assert "WorkerDriveUploadApiUrl = 'http://127.0.0.1:8000/api/du-lieu-web'" in config_text
        assert "GoogleDriveOauthRefreshToken" in config_text
        assert 'Set-UserEnvironment "ONEBSS_DRIVE_UPLOAD_API_URL" $WorkerDriveUploadApiUrl' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_TASK_TIMEOUT_SECONDS" $onebssTaskTimeoutSeconds' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_GRID_TIMEOUT_SECONDS" $onebssGridTimeoutSeconds' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_PROCESSING_TIMEOUT_RETRY_ATTEMPTS" $onebssProcessingTimeoutRetryAttempts' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_PROCESSING_TIMEOUT_RETRY_DELAY_SECONDS" $onebssProcessingTimeoutRetryDelaySeconds' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_WORKER_OTP_WAIT_SECONDS" $onebssOtpWaitSeconds' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_WORKER_DISABLE_TASK_GUARD" "1"' in setup_script
        assert 'Set-UserEnvironment "SQL_WORKER_POLL_SECONDS" "10"' in setup_script
        assert 'Set-UserEnvironment "FTP_WORKER_POLL_SECONDS" "30"' in setup_script
        assert 'Set-UserEnvironment "VNPTCTO_WORKER_MAX_CONCURRENT_TASKS" $workerMaxConcurrentTasks' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_WORKER_MAX_CONCURRENT_TASKS" $workerMaxConcurrentTasks' in setup_script
        assert 'Set-UserEnvironment "ONEBSS_WORKER_MAX_ONEBSS_TASKS" $onebssWorkerMaxTasks' in setup_script
        assert 'Set-UserEnvironment "SQL_WORKER_MAX_CONCURRENT_TASKS" $sqlWorkerMaxTasks' in setup_script
        assert 'Set-UserEnvironment "FTP_WORKER_MAX_CONCURRENT_TASKS" $ftpWorkerMaxTasks' in setup_script
        assert "ONEBSS_TASK_TIMEOUT_SECONDS" in start_worker_script
        assert routes.WORKSTATION_SETUP_PACKAGE_VERSION.endswith("v42")
        assert 'WORKER_VERSION = "2026.08.20-sql-cancel-v42"' in worker_script
        assert "WorkerConcurrencyTracker" in worker_script
        assert "WorkerTaskDispatcher" in worker_script
        assert "normalize_ftp_variable_value" in worker_script
        assert "ONEBSS_GRID_TIMEOUT_SECONDS" in start_worker_script
        assert "ONEBSS_PROCESSING_TIMEOUT_RETRY_ATTEMPTS" in start_worker_script
        assert "ONEBSS_PROCESSING_TIMEOUT_RETRY_DELAY_SECONDS" in start_worker_script
        assert "ONEBSS_WORKER_OTP_WAIT_SECONDS" in start_worker_script
        assert "ONEBSS_WORKER_DISABLE_TASK_GUARD" in start_worker_script
        assert '$env:ONEBSS_WORKER_DISABLE_TASK_GUARD = "1"' in start_worker_script
        assert "SQL_WORKER_POLL_SECONDS" in start_worker_script
        assert "FTP_WORKER_POLL_SECONDS" in start_worker_script
        assert "VNPTCTO_WORKER_MAX_CONCURRENT_TASKS" in start_worker_script
        assert "ONEBSS_WORKER_MAX_CONCURRENT_TASKS" in start_worker_script
        assert "ONEBSS_WORKER_MAX_ONEBSS_TASKS" in start_worker_script
        assert "SQL_WORKER_MAX_CONCURRENT_TASKS" in start_worker_script
        assert "FTP_WORKER_MAX_CONCURRENT_TASKS" in start_worker_script
        assert "GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON_BASE64" in setup_script
        assert "OracleDbDsn" in setup_script
        assert "OracleDbHost" in setup_script
        assert "Stop-ApiMiddlewareProcesses" in setup_script
        assert "Stop-WorkstationWorkerProcesses" in setup_script
        assert "Start-WorkstationWorkerNow" in setup_script
        assert "Get-WorkstationWorkerProcesses" in setup_script
        assert "Wait-WorkstationWorkerStable" in setup_script
        assert "Get-WorkstationWorkerPythonProcesses" in setup_script
        assert "Worker da chay qua Scheduled Task, Python worker PID" in setup_script
        assert "onebss_workstation_worker.py" in setup_script
        assert "Stop-ScheduledTask -TaskName \"VNPTCTO OneBSS Worker\"" in setup_script
        assert "Start-Process -FilePath \"powershell.exe\"" in setup_script
        assert "Goi cai dat thieu cau hinh Oracle" in setup_script
        assert "DB_DSN=$(DotEnvValue $oracleDbDsn)" in setup_script
        assert "DB_HOST=$(DotEnvValue $oracleDbHost)" in setup_script
        assert "DB_SID=$(DotEnvValue $oracleDbSid)" in setup_script
        assert "EXPORT_PAGE_SIZE=$(DotEnvValue $exportPageSize)" in setup_script
        assert 'Set-DotEnvValue $apiEnv "DB_DSN" $oracleDbDsn' in setup_script
        assert 'Set-DotEnvValue $apiEnv "DB_PASS" $oracleDbPass' in setup_script
        assert 'Set-DotEnvValue $apiEnv "EXPORT_PAGE_SIZE" $exportPageSize' in setup_script
        assert 'Set-UserEnvironment "SQL_WORKER_TIMEOUT_SECONDS" $sqlWorkerTimeoutSeconds' in setup_script
        assert "$canStartWorkerNow = -not [string]::IsNullOrWhiteSpace($InternalApiToken)" in setup_script
        assert "drive-oauth-token.json" in setup_script
        assert "Ensure-ApiDriveOauthFiles" in setup_script
        assert '$driveAuthMode = "oauth"' in setup_script
        assert '$driveAuthMode = "service_account"' not in setup_script
        assert "LeastPrivilege" not in setup_script
        assert "LeastPrivilege" not in install_task_script
        assert "-RunLevel Limited" in setup_script
        assert "-RunLevel Limited" in install_task_script
        assert "Get-InteractiveTaskUserCandidates" in setup_script
        assert "Get-InteractiveTaskUserCandidates" in install_task_script
        assert "run_onebss_worker_background.ps1" in install_task_script
        assert "-WindowStyle Hidden" in install_task_script
        assert "-WindowStyle Hidden" in setup_script
        assert "onebss-worker.log" in background_worker_script
        assert "onebss-worker-error.log" in background_worker_script
        assert "start_onebss_worker.ps1" in background_worker_script
        assert "Out-File -FilePath '$safeLogFile' -Append -Encoding UTF8" in background_worker_script
        assert '$ErrorActionPreference = "Continue"' in start_worker_script
        assert '$ErrorActionPreference = "Continue"' in setup_script
        assert "System.Management.Automation.ErrorRecord" in start_worker_script
        assert "System.Management.Automation.ErrorRecord" in setup_script
        assert "Tu khoi dong lai worker sau $RestartDelaySeconds giay." in background_worker_script
        assert "New-WorkerMutexName" in background_worker_script
        assert "Da co wrapper worker dang chay" in background_worker_script
        assert "RepetitionInterval (New-TimeSpan -Minutes 2)" in setup_script
        assert "[Security.Principal.WindowsIdentity]::GetCurrent().Name" in install_task_script
        assert "whoami.exe" in install_task_script
        assert "-UserId $env:USERNAME" not in setup_script
        assert "-UserId $env:USERNAME" not in install_task_script
        assert 'Invoke-External "python" "-m" "venv"' not in start_worker_script
        assert "Hay cai Cloudflare Tunnel service rieng" not in api_task_script
        assert "bo qua Cloudflare Tunnel" in api_task_script
        assert "Stop-LocalApiProcesses" in api_task_script
        assert "config-status" in api_task_script
        assert "Bo qua vi chua cai cloudflared" in health_script
        assert "health-check-2026.08.03-failover" in health_script
        assert "Khong tim thay Scheduled Task; khong bat buoc neu Local API dang OK." in health_script
        assert "Wait-WorkerProcessesStable" in health_script
        assert "Get-WorkerPythonProcesses" in health_script
        assert "Get-WorkerPythonInstances" in health_script
        assert "Get-WorkerWrapperProcesses" in health_script
        assert "Dang co nhieu worker chay trung" in health_script
        assert "Test-OneBssWorkerTaskUsesBackgroundWorker" in health_script
        assert "run_onebss_worker_background.ps1" in health_script
        assert "Da start Scheduled Task, Python worker PID" in health_script
        assert "Da start worker fallback" in health_script
        assert "Da start launcher" not in health_script
        assert "Start-WorkerIfMissing" in health_script
        assert "Worker process" in health_script
        assert '$heartbeatRoles = @("health_check")' in health_script
        assert 'if ($workerStart.Ok)' in health_script
        assert '"sql_report_worker"' in health_script
        assert 'roles = $heartbeatRoles' in health_script
        assert "Local API config" in health_script
        assert "config-status" in health_script
        assert "workstation-setup-error.log" in setup_script
        assert "ONEBSS_WORKER_DISABLE_TASK_GUARD" in uninstall_task_script
        assert "ONEBSS_GRID_TIMEOUT_SECONDS" in uninstall_task_script
        assert "ONEBSS_WORKER_OTP_WAIT_SECONDS" in uninstall_task_script
        assert "-NoPause" in setup_bat
        assert "-NoPause" in background_bat
        assert "-NoPause" in install_autostart_bat
        assert "-NoPause" in uninstall_autostart_bat
        assert "Verb RunAs" in uninstall_autostart_bat
        assert 'cd /d "%TEMP%"' in uninstall_autostart_bat
        assert "[switch]$NoPause" in uninstall_task_script
        assert "[switch]$KeepInstallDir" in uninstall_task_script
        assert "VNPTCTO Workstation Health Check" in uninstall_task_script
        assert "VNPTCTO API Trung Gian" in uninstall_task_script
        assert "VNPTCTO API Watchdog" in uninstall_task_script
        assert "Stop-VnptctoProcesses" in uninstall_task_script
        assert "Remove-UserEnvironment" in uninstall_task_script
        assert "Remove-SafeDirectory" in uninstall_task_script
        assert "VNPTCTO_WORKSTATION_ROOT" in uninstall_task_script
        assert "C:\\VNPTCTO" in uninstall_task_script
        assert "ONEBSS_DRIVE_UPLOAD_API_URL" in uninstall_task_script
        assert "SQL_WORKER_POLL_SECONDS" in uninstall_task_script
        assert "FTP_WORKER_POLL_SECONDS" in uninstall_task_script
        assert "VNPTCTO_WORKER_MAX_CONCURRENT_TASKS" in uninstall_task_script
        assert "ONEBSS_WORKER_MAX_CONCURRENT_TASKS" in uninstall_task_script
        assert "ONEBSS_WORKER_MAX_ONEBSS_TASKS" in uninstall_task_script
        assert "SQL_WORKER_MAX_CONCURRENT_TASKS" in uninstall_task_script
        assert "FTP_WORKER_MAX_CONCURRENT_TASKS" in uninstall_task_script
        assert "timeout /t" not in setup_bat.lower()
        assert "\npause" not in setup_bat.lower()


def test_workstation_oracle_password_can_fallback_to_environment(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("DB_PASS", "env-oracle-secret")
    get_settings.cache_clear()
    try:
        with TestClient(app) as client:
            login(client)
            saved_oracle = client.put(
                "/api/admin/connections/oracle_agency_db",
                json={
                    "name": "Oracle DB noi bo",
                    "connection_type": "oracle",
                    "description": "Oracle cho may tram",
                    "config": {
                        "host": "10.10.10.20",
                        "port": 1521,
                        "service": "ONEBSS",
                        "username": "REPORT_USER",
                        "password": "",
                    },
                    "is_active": True,
                },
            )
            assert saved_oracle.status_code == 200
            overview = client.get("/api/admin/workstation/overview")
            assert overview.status_code == 200
            assert overview.json()["config"]["oracle_config_ready"] is True
            package = client.get("/api/admin/workstation/setup-package")
            assert package.status_code == 200
            with ZipFile(BytesIO(package.content)) as archive:
                config_text = archive.read("VNPTCTO_WORKSTATION_SETUP/workstation-install-config.ps1").decode("utf-8")
            assert "OracleDbPass = 'env-oracle-secret'" in config_text
    finally:
        get_settings.cache_clear()


def test_workstation_overview_reports_missing_oracle_password() -> None:
    with TestClient(app) as client:
        login(client)
        saved_oracle = client.put(
            "/api/admin/connections/oracle_agency_db",
            json={
                "name": "Oracle DB noi bo",
                "connection_type": "oracle",
                "description": "Oracle cho may tram",
                "config": {
                    "host": "10.10.10.20",
                    "port": 1521,
                    "service": "ONEBSS",
                    "username": "REPORT_USER",
                    "password": "",
                },
                "is_active": True,
            },
        )
        assert saved_oracle.status_code == 200
        overview = client.get("/api/admin/workstation/overview")
        assert overview.status_code == 200
        config = overview.json()["config"]
        assert config["oracle_config_ready"] is False
        assert "DB_PASS" in config["oracle_missing_items"]


def test_api_middleware_uses_oracle_dsn_and_rejects_bequeath(monkeypatch: pytest.MonkeyPatch) -> None:
    module = load_api_middleware_module()
    for key in (
        "DB_DSN",
        "ORACLE_DSN",
        "TNS_DSN",
        "DB_CONNECT_STRING",
        "ORACLE_CONNECT_STRING",
        "DB_HOST",
        "DB_SERVICE",
        "DB_SID",
        "DB_USER",
        "DB_PASS",
    ):
        monkeypatch.delenv(key, raising=False)

    monkeypatch.setenv("DB_USER", "REPORT_USER")
    monkeypatch.setenv("DB_PASS", "oracle-secret")
    monkeypatch.setenv("DB_DSN", "10.92.53.53:1521/DBCTO")
    config = module.oracle_connection_config()
    assert config["source"] == "DB_DSN"
    assert config["dsn"] == "10.92.53.53:1521/DBCTO"

    monkeypatch.setenv("DB_DSN", "/")
    with pytest.raises(RuntimeError, match="bequeath"):
        module.oracle_connection_config()

    monkeypatch.setenv("DB_HOST", "10.92.53.53")
    monkeypatch.setenv("DB_SERVICE", "DBCTO")
    fallback = module.oracle_connection_config()
    assert fallback["source"] == "DB_HOST/DB_SERVICE"
    assert "10.92.53.53" in fallback["dsn"]
    assert "DBCTO" in fallback["dsn"]


def test_api_middleware_config_status_reports_version_without_password(monkeypatch: pytest.MonkeyPatch) -> None:
    module = load_api_middleware_module()
    monkeypatch.setenv("DB_USER", "REPORT_USER")
    monkeypatch.setenv("DB_PASS", "oracle-secret")
    monkeypatch.setenv("DB_DSN", "10.92.53.53:1521/DBCTO")
    monkeypatch.setenv("GOOGLE_DRIVE_FOLDER_ID", "drive-folder")
    payload = module.config_status()

    assert payload["version"] == module.API_MIDDLEWARE_VERSION
    assert payload["oracle_config_ok"] is True
    assert payload["dsn_source"] == "DB_DSN"
    assert payload["db_pass_configured"] is True
    assert "oracle-secret" not in json.dumps(payload)


def test_api_middleware_formats_oracle_date_binds() -> None:
    module = load_api_middleware_module()
    sql = "SELECT * FROM dual WHERE ngay >= TO_DATE(:P_TUNGAY, 'YYYYMMDD') AND ngay < TO_DATE(:P_DENNGAY, 'YYYY-MM-DD')"
    binds = module.normalize_binds_for_sql(sql, {"P_TUNGAY": "01/07/2026", "P_DENNGAY": "30/07/2026"})

    assert binds["P_TUNGAY"] == "20260701"
    assert binds["P_DENNGAY"] == "2026-07-30"


def test_api_middleware_reads_json_files_with_utf8_bom(tmp_path: Path) -> None:
    module = load_api_middleware_module()
    token_file = tmp_path / "drive-oauth-token.json"
    token_file.write_text('\ufeff{"refresh_token":"r","client_id":"c","client_secret":"s","token_uri":"https://oauth2.googleapis.com/token"}', encoding="utf-8")

    assert module.read_json_file(token_file)["refresh_token"] == "r"


def test_api_middleware_reads_service_account_base64_with_utf8_bom(monkeypatch: pytest.MonkeyPatch) -> None:
    module = load_api_middleware_module()
    info = {
        "type": "service_account",
        "client_email": "svc@example.iam.gserviceaccount.com",
        "private_key": "-----BEGIN PRIVATE KEY-----\\nabc\\n-----END PRIVATE KEY-----\\n",
        "token_uri": "https://oauth2.googleapis.com/token",
    }
    raw = "\ufeff" + json.dumps(info)
    monkeypatch.setenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON_BASE64", base64.b64encode(raw.encode("utf-8")).decode("ascii"))
    monkeypatch.delenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_JSON", raising=False)
    monkeypatch.delenv("GOOGLE_DRIVE_SERVICE_ACCOUNT_FILE", raising=False)

    assert module.load_service_account_info()["client_email"] == info["client_email"]


def test_api_middleware_run_sql_falls_back_when_paged_wrapper_has_duplicate_columns() -> None:
    module = load_api_middleware_module()

    class FakeCursor:
        def __init__(self) -> None:
            self.description: list[tuple[str]] = []
            self.executed: list[tuple[str, dict]] = []
            self.direct_rows = [("skip-left", "skip-right", "skip-name"), ("left", "right", "name")]

        def execute(self, sql: str, binds: dict) -> None:
            self.executed.append((sql, binds))
            if "OFFSET" in sql:
                raise Exception("ORA-00918: column ambiguously defined")
            self.description = [("ID",), ("ID",), ("TEN_TB",)]

        def fetchmany(self, size: int) -> list[tuple]:
            rows = self.direct_rows[:size]
            self.direct_rows = self.direct_rows[size:]
            return rows

    cursor = FakeCursor()
    columns, rows = module.fetch_page(cursor, "SELECT a.id, b.id, b.ten_tb FROM rpt", {"P": "1"}, page=2, page_size=1)

    assert columns == ["ID", "ID_2", "TEN_TB"]
    assert rows == [{"ID": "left", "ID_2": "right", "TEN_TB": "name"}]
    assert len(cursor.executed) == 2
    assert "OFFSET" in cursor.executed[0][0]
    assert cursor.executed[1] == ("SELECT a.id, b.id, b.ten_tb FROM rpt", {"P": "1"})


def test_api_middleware_streams_excel_export_without_count_or_offset(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    module = load_api_middleware_module()

    def fail_count_or_page(*args, **kwargs):
        raise AssertionError("Export must stream the SELECT once, not count/page it first.")

    class FakeCursor:
        description = [("MA_TB",), ("DOANH_THU",)]

        def __init__(self) -> None:
            self.arraysize = 0
            self.prefetchrows = 0
            self.executed: list[tuple[str, dict]] = []
            self.batches = [[("TB001", 1000), ("TB002", 2000)], []]

        def execute(self, sql: str, binds: dict) -> None:
            self.executed.append((sql, binds))

        def fetchmany(self, size: int) -> list[tuple]:
            return self.batches.pop(0)

    monkeypatch.setattr(module, "count_rows", fail_count_or_page)
    monkeypatch.setattr(module, "fetch_page", fail_count_or_page)

    cursor = FakeCursor()
    target = tmp_path / "export.xlsx"
    result = module.write_export_to_excel(cursor, "SELECT ma_tb, doanh_thu FROM rpt", {"P": "X"}, target, 20000, 100)

    assert result["streaming"] is True
    assert result["fetch_size"] == 20000
    assert result["rows"] == 2
    assert cursor.executed == [("SELECT ma_tb, doanh_thu FROM rpt", {"P": "X"})]
    workbook = openpyxl.load_workbook(target, read_only=True)
    assert len(list(workbook.active.iter_rows(values_only=True))) == 3


def test_api_middleware_export_file_names_include_job_id() -> None:
    module = load_api_middleware_module()

    assert module.file_name_with_job_id("crs_20260811_120000.xlsx", "SQL-RUN-001") == "crs_20260811_120000_SQL-RUN-001.xlsx"
    assert module.file_name_with_job_id("crs_SQL-RUN-001.xlsx", "SQL-RUN-001") == "crs_SQL-RUN-001.xlsx"


def test_workstation_heartbeat_uses_worker_token() -> None:
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/workstation/heartbeat",
                json={"worker_id": "ws-test", "status": "idle", "roles": ["onebss_worker"]},
                headers={"Authorization": "Bearer test-worker-token"},
            )
            assert response.status_code == 200
            assert response.json()["worker_id"] == "ws-test"

            login(client)
            overview = client.get("/api/admin/workstation/overview")
            workers = overview.json()["workers"]
            assert any(worker["worker_id"] == "ws-test" and worker["status"] == "online" for worker in workers)

            sql_worker_state = routes._dynamic_report_sql_worker_state()
            assert sql_worker_state["status"] == "no_sql_worker"
            assert sql_worker_state["online_count"] == 1
            assert sql_worker_state["sql_count"] == 0
            assert "worker sql" in sql_worker_state["message"].lower()

            sql_response = client.post(
                "/api/workstation/heartbeat",
                json={"worker_id": "ws-sql", "status": "idle", "roles": ["onebss_worker", "sql_report_worker"]},
                headers={"Authorization": "Bearer test-worker-token"},
            )
            assert sql_response.status_code == 200
            sql_worker_state = routes._dynamic_report_sql_worker_state()
            assert sql_worker_state["status"] == "ready"
            assert sql_worker_state["online_count"] == 2
            assert sql_worker_state["sql_count"] == 1
    finally:
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()


def test_admin_can_test_save_and_delete_workstation_profile() -> None:
    worker_id = f"ws-admin-{uuid.uuid4().hex[:8]}"
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    try:
        with TestClient(app) as client:
            login(client)
            saved = client.post(
                "/api/admin/workstation/profile",
                json={"worker_id": worker_id, "display_name": "May tram so 1", "priority": 1, "enabled": True},
            )
            assert saved.status_code == 200
            assert saved.json()["profile"]["display_name"] == "May tram so 1"
            assert saved.json()["profile"]["priority"] == 1

            heartbeat = client.post(
                "/api/workstation/heartbeat",
                json={
                    "worker_id": worker_id,
                    "status": "idle",
                    "roles": ["onebss_worker", "sql_report_worker"],
                    "version": "worker-test",
                    "message": "Worker dang cho lenh.",
                    "details": {"pid": 1234, "worker_process": "Dang chay PID: 1234"},
                },
                headers={"Authorization": "Bearer test-worker-token"},
            )
            assert heartbeat.status_code == 200

            overview = client.get("/api/admin/workstation/overview")
            workers = overview.json()["workers"]
            assert any(worker["worker_id"] == worker_id and worker["display_name"] == "May tram so 1" and worker["priority"] == 1 for worker in workers)

            tested = client.post(f"/api/admin/workstation/{worker_id}/test")
            assert tested.status_code == 200
            body = tested.json()
            assert body["ok"] is True
            assert {check["code"] for check in body["checks"]} >= {"connection", "worker", "background", "priority"}
            background_check = next(check for check in body["checks"] if check["code"] == "background")
            assert background_check["status"] == "ok"

            deleted = client.delete(f"/api/admin/workstation/{worker_id}")
            assert deleted.status_code == 200
            assert deleted.json()["profile"]["deleted"] is True
            overview_after_delete = client.get("/api/admin/workstation/overview")
            assert all(worker["worker_id"] != worker_id for worker in overview_after_delete.json()["workers"])
    finally:
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()


def test_sql_claim_preserves_background_process_details() -> None:
    worker_id = f"ws-bg-{uuid.uuid4().hex[:8]}"
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    try:
        with TestClient(app) as client:
            heartbeat = client.post(
                "/api/workstation/heartbeat",
                json={
                    "worker_id": worker_id,
                    "status": "idle",
                    "roles": ["onebss_worker", "sql_report_worker"],
                    "version": "worker-test",
                    "message": "Worker dang cho lenh.",
                    "details": {
                        "pid": 4321,
                        "worker_process": "Dang chay PID: 4321",
                        "python": "3.12",
                    },
                },
                headers={"Authorization": "Bearer test-worker-token"},
            )
            assert heartbeat.status_code == 200

            claim = client.post(
                "/api/sql-worker/tasks/claim",
                json={"worker_id": worker_id},
                headers={"Authorization": "Bearer test-worker-token"},
            )
            assert claim.status_code == 200

            login(client)
            tested = client.post(f"/api/admin/workstation/{worker_id}/test")
            assert tested.status_code == 200
            body = tested.json()
            assert body["ok"] is True
            background_check = next(check for check in body["checks"] if check["code"] == "background")
            assert background_check["status"] == "ok"
            assert "4321" in background_check["message"]
    finally:
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()


def test_empty_onebss_worker_claim_cache_skips_repeated_repository_lookup(monkeypatch: pytest.MonkeyPatch) -> None:
    worker_id = f"ws-empty-{uuid.uuid4().hex[:8]}"
    calls = {"claim": 0}

    class EmptyRepository:
        def list_system_connections(self) -> list[dict]:
            return []

        def claim_next_onebss_report_run(self, worker_id: str) -> None:
            calls["claim"] += 1
            return None

    routes.invalidate_worker_claim_empty_cache()
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    monkeypatch.setattr(routes, "build_app_repository", lambda: EmptyRepository())
    monkeypatch.setattr(routes, "_expire_stale_onebss_worker_runs", lambda *args, **kwargs: None)
    try:
        with TestClient(app) as client:
            headers = {"Authorization": "Bearer test-worker-token"}
            first = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": worker_id}, headers=headers)
            second = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": worker_id}, headers=headers)

        assert first.status_code == 200
        assert first.json()["task"] is None
        assert second.status_code == 200
        assert second.json()["cached_empty"] is True
        assert calls["claim"] == 1
    finally:
        routes.invalidate_worker_claim_empty_cache()
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()


def test_sql_worker_queue_invalidates_empty_claim_cache(monkeypatch: pytest.MonkeyPatch) -> None:
    class PersistOnlyRepository:
        def save_report_run(self, payload: dict) -> None:
            return None

    job_id = f"sql-cache-{uuid.uuid4().hex}"
    routes.invalidate_worker_claim_empty_cache()
    routes.mark_worker_claim_empty("sql")
    assert routes.worker_claim_empty_cached("sql") is True
    monkeypatch.setattr(routes, "build_app_repository", lambda: PersistOnlyRepository())
    try:
        routes._set_dynamic_report_run_job(
            job_id,
            status="queued_worker",
            message="Da gui lenh SQL cho may tram.",
            payload={"ma_bao_cao": "CACHE_TEST"},
            report_code="CACHE_TEST",
            report_name="Cache test",
        )
        assert routes.worker_claim_empty_cached("sql") is False
    finally:
        with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
            routes.DYNAMIC_REPORT_RUN_JOBS.pop(job_id, None)
        routes.invalidate_worker_claim_empty_cache()


def test_task_report_auto_sql_worker_claim_collects_all_pages(monkeypatch: pytest.MonkeyPatch) -> None:
    report_code = f"TASK_AUTO_SQL_{uuid.uuid4().hex[:8].upper()}"
    job_id = f"task-auto-sql-{uuid.uuid4().hex}"
    headers = {"Authorization": "Bearer test-worker-token"}
    prepared_args = {}

    class FakeDatabaseService:
        def prepare_dynamic_report_query(self, **kwargs):
            prepared_args.update(kwargs)
            return {
                "ok": True,
                "ten_bao_cao": "Task auto all pages",
                "ma_bao_cao": kwargs["ma_bao_cao"],
                "cau_lenh_sql": "SELECT ma_tb FROM css_cto.rpt_task_auto",
                "tham_so": kwargs.get("filters") or {},
                "page": kwargs.get("page"),
                "page_size": kwargs.get("page_size"),
                "report": {"ma_bao_cao": kwargs["ma_bao_cao"], "ten_bao_cao": "Task auto all pages"},
            }

    monkeypatch.setattr(routes, "build_database_service", lambda: FakeDatabaseService())
    monkeypatch.setattr(routes, "_next_dynamic_report_export_worker_job", lambda *args, **kwargs: None)
    monkeypatch.setattr(routes, "_next_dashboard_refresh_worker_job", lambda: None)
    routes.invalidate_worker_claim_empty_cache("sql")
    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
    with routes.DASHBOARD_REFRESH_JOBS_LOCK:
        routes.DASHBOARD_REFRESH_JOBS.clear()
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    try:
        with TestClient(app) as client:
            queued_at = time.time() + 100000
            with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
                routes.DYNAMIC_REPORT_RUN_JOBS[job_id] = {
                    "job_id": job_id,
                    "status": "queued_worker",
                    "created_at": queued_at,
                    "updated_at": queued_at,
                    "created_by": "pytest",
                    "report_code": report_code,
                    "report_name": "Task auto all pages",
                    "payload": {
                        "ma_bao_cao": report_code,
                        "filters": {},
                        "page": 1,
                        "page_size": 20000,
                        "collect_all_pages": True,
                        "max_rows": 50000,
                    },
                }

            claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-task-auto-sql"}, headers=headers)
            assert claim.status_code == 200
            task = claim.json()["task"]
            assert task["run_id"] == job_id
            assert task["task_type"] == "dynamic_report_load"
            assert task["query"]["pagination"] == {"page": 1, "page_size": 20000}
            assert task["query"]["collect_all_pages"] is True
            assert task["query"]["max_rows"] == 50000
            assert prepared_args["page_size"] == 20000
    finally:
        with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
            routes.DYNAMIC_REPORT_RUN_JOBS.pop(job_id, None)
        with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
            routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
        with routes.DASHBOARD_REFRESH_JOBS_LOCK:
            routes.DASHBOARD_REFRESH_JOBS.clear()
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()
        routes.invalidate_worker_claim_empty_cache()


def test_workstation_priority_blocks_lower_priority_onebss_claim() -> None:
    primary_id = f"ws-primary-{uuid.uuid4().hex[:8]}"
    secondary_id = f"ws-secondary-{uuid.uuid4().hex[:8]}"
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    try:
        with TestClient(app) as client:
            login(client)
            client.post("/api/admin/workstation/profile", json={"worker_id": primary_id, "display_name": "May 1", "priority": 1, "enabled": True})
            client.post("/api/admin/workstation/profile", json={"worker_id": secondary_id, "display_name": "May 2", "priority": 2, "enabled": True})
            client.post(
                "/api/workstation/heartbeat",
                json={"worker_id": primary_id, "status": "idle", "roles": ["onebss_worker"], "version": "worker-primary"},
                headers={"Authorization": "Bearer test-worker-token"},
            )
            created = client.post(
                "/api/admin/onebss-reports",
                json={
                    "ten_bao_cao": "OneBSS priority",
                    "danh_sach_bien": ["P_TUNGAY"],
                    "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_PRIORITY&name=Test",
                    "storage_link": "",
                },
            )
            assert created.status_code == 200
            code = created.json()["ma_bao_cao"]
            queued = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
            assert queued.status_code == 200

            headers = {"Authorization": "Bearer test-worker-token"}
            lower_claim = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": secondary_id}, headers=headers)
            assert lower_claim.status_code == 200
            assert lower_claim.json()["task"] is None
            assert lower_claim.json()["status"] == "waiting_priority"

            with routes.WORKSTATION_HEARTBEATS_LOCK:
                routes.WORKSTATION_HEARTBEATS[primary_id]["received_at_ts"] = (
                    time.time() - routes.WORKSTATION_READY_HEARTBEAT_SECONDS - 5
                )
            fallback_claim = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": secondary_id}, headers=headers)
            assert fallback_claim.status_code == 200
            assert fallback_claim.json()["task"]["ma_bao_cao"] == code
    finally:
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()


def test_sql_worker_state_accepts_legacy_role_and_busy_task_heartbeat() -> None:
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    try:
        routes._record_workstation_heartbeat(
            "ws-legacy",
            "idle",
            "May tram cu van gui role SQL cu.",
            roles=["onebss_worker", "sql_export_worker"],
        )
        sql_worker_state = routes._dynamic_report_sql_worker_state()
        assert sql_worker_state["status"] == "ready"
        assert sql_worker_state["sql_count"] == 1

        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()
        routes._record_workstation_heartbeat(
            "ws-busy",
            "busy",
            "May tram dang xu ly OneBSS.",
            details={"run_id": "RUN-BUSY", "task_type": "onebss"},
        )
        busy_state = routes._dynamic_report_sql_worker_state()
        assert busy_state["status"] == "busy"
        assert busy_state["sql_count"] == 1
    finally:
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()


def test_favicon_redirects_to_system_logo() -> None:
    with TestClient(app) as client:
        response = client.get("/favicon.ico", follow_redirects=False)
        assert response.status_code == 307
        assert response.headers["location"] == "/static/images/system-logo.png"


def test_static_assets_are_cacheable() -> None:
    with TestClient(app) as client:
        response = client.get("/static/app.js?v=53")
        assert response.status_code == 200
        assert response.headers["cache-control"] == "public, max-age=31536000, immutable"
        assert response.headers["x-content-type-options"] == "nosniff"
        assert response.headers["x-frame-options"] == "SAMEORIGIN"
        assert response.headers["referrer-policy"] == "strict-origin-when-cross-origin"


def test_mobile_gateway_admin_lists_are_clamped_to_twenty_rows() -> None:
    with TestClient(app) as client:
        login(client)
        sms = client.get("/api/admin/mobile-gateway/sms?page=1&page_size=100")
        notifications = client.get("/api/admin/mobile-gateway/notifications?page=1&page_size=100")
        commands = client.get("/api/admin/mobile-gateway/commands?limit=100")
        assert sms.status_code == 200
        assert notifications.status_code == 200
        assert commands.status_code == 200
        assert sms.json()["page_size"] == 20
        assert notifications.json()["page_size"] == 20
        assert len(commands.json()["commands"]) <= 20


def test_production_startup_validation_rejects_unsafe_defaults() -> None:
    settings = Settings(
        app_env="production",
        session_secret="change-this-session-secret",
        initial_admin_password="ChangeMe123!",
        internal_api_mock_mode=True,
    )
    with pytest.raises(RuntimeError) as error:
        settings.validate_for_startup()
    message = str(error.value)
    assert "SESSION_SECRET" in message
    assert "INITIAL_ADMIN_PASSWORD" in message
    assert "INTERNAL_API_MOCK_MODE" in message
    assert "ChangeMe123" not in message


def test_telegram_alert_sanitizer_redacts_secrets() -> None:
    text = sanitize_alert_text("token=abc123 password:super-secret cookie=session-id")
    assert "abc123" not in text
    assert "super-secret" not in text
    assert "session-id" not in text
    assert "[redacted]" in text

    details = sanitize_alert_details({"telegram_token": "abc123", "url": "/x?password=super-secret"})
    assert details["telegram_token"] == "[redacted]"
    assert "super-secret" not in details["url"]


def test_admin_navigation_payload_combines_features_and_layouts() -> None:
    with TestClient(app) as client:
        assert client.get("/api/navigation").status_code == 401
        login(client)
        response = client.get("/api/navigation")
        assert response.status_code == 200
        payload = response.json()
        assert any(feature["code"] == "quantrimenu" for feature in payload["features"])
        assert any(feature["code"] == "internalemail" for feature in payload["features"])
        assert any(feature["code"] == "internal_email.view" for feature in payload["features"])
        assert any(feature["code"] == "publicmessages" for feature in payload["features"])
        assert any(feature["code"] == "public_messages.view" for feature in payload["features"])
        assert any(layout["page_id"] == "DASHBOARD_KINH_DOANH" for layout in payload["dashboard_layouts"])


def test_viewer_navigation_includes_parent_for_granted_child_dashboard() -> None:
    with TestClient(app) as client:
        login(client)
        saved = client.post(
            "/api/admin/dashboard-layouts",
            json={
                "page_id": "DASHBOARD_VIEWER_CHILD",
                "page_name": "Dashboard Viewer Child",
                "parent_code": "baocaomoi",
                "layout": {
                    "tabs": [
                        {
                            "tab_id": "tab_a",
                            "tab_name": "Tab A",
                            "order": 1,
                            "grid_layout": [
                                {
                                    "row_id": 1,
                                    "layout_type": "1_column",
                                    "widgets": [
                                        {
                                            "position": 1,
                                            "type": "text_title",
                                            "title": "Viewer dashboard",
                                            "text_content": "Only granted viewer can see this item.",
                                        }
                                    ],
                                }
                            ],
                        }
                    ],
                },
            },
        )
        assert saved.status_code == 200
        feature_code = saved.json()["feature_code"]
        created = client.post(
            "/api/admin/users",
            json={
                "username": "viewer_navigation",
                "full_name": "Viewer Navigation",
                "password": "Viewer@Navigation123",
                "role": "viewer",
            },
        )
        assert created.status_code == 200
        viewer_id = created.json()["user"]["id"]
        assert client.put(
            f"/api/admin/users/{viewer_id}/permissions",
            json={"feature_codes": [feature_code]},
        ).status_code == 200

        client.post("/api/auth/logout")
        login(client, "viewer_navigation", "Viewer@Navigation123")
        response = client.get("/api/navigation")
        assert response.status_code == 200
        payload = response.json()
        feature_codes = {feature["code"] for feature in payload["features"]}
        assert "baocaomoi" in feature_codes
        assert feature_code in feature_codes
        assert "quantriweb" not in feature_codes
        assert [layout["page_id"] for layout in payload["dashboard_layouts"]] == ["DASHBOARD_VIEWER_CHILD"]

        page = client.get(f"/{feature_code}")
        assert page.status_code == 200
        assert "/static/app.js?v=227" in page.text
        assert "dashboard-designed-section" in page.text

        detail = client.get("/api/dashboard-layouts/DASHBOARD_VIEWER_CHILD")
        assert detail.status_code == 200
        assert detail.json()["feature_code"] == feature_code
        tab_data = client.get("/api/dashboard-layouts/DASHBOARD_VIEWER_CHILD/tabs/tab_a/data")
        assert tab_data.status_code == 200
        assert client.get("/api/dashboard-layouts/DASHBOARD_KINH_DOANH").status_code == 403


def test_admin_can_toggle_and_demo_renew_user_billing() -> None:
    with TestClient(app) as client:
        login(client)
        plans_response = client.get("/api/admin/billing/plans")
        assert plans_response.status_code == 200
        plan_codes = {plan["code"] for plan in plans_response.json()["plans"]}
        assert {"monthly", "quarterly", "six_months", "yearly"} <= plan_codes

        created = client.post(
            "/api/admin/users",
            json={
                "username": "viewer_billing",
                "full_name": "Viewer Billing",
                "password": "Viewer@Billing123",
                "role": "viewer",
            },
        )
        assert created.status_code == 200
        viewer_id = created.json()["user"]["id"]

        expired = client.put(
            f"/api/admin/users/{viewer_id}/billing",
            json={"billing_enabled": True, "plan_code": "monthly", "expires_at": "2020-01-01"},
        )
        assert expired.status_code == 200
        assert expired.json()["billing"]["billing_status"] == "expired"

        client.post("/api/auth/logout")
        login(client, "viewer_billing", "Viewer@Billing123")
        assert client.get("/").status_code == 200
        blocked = client.get("/api/navigation")
        assert blocked.status_code == 402

        client.post("/api/auth/logout")
        login(client)
        renewed = client.post(
            f"/api/admin/users/{viewer_id}/billing/renew",
            json={"plan_code": "six_months"},
        )
        assert renewed.status_code == 200
        payload = renewed.json()
        assert payload["invoice"]["status"] == "paid"
        assert payload["subscription"]["billing_status"] == "active"
        assert payload["subscription"]["billing_total_months"] == 7

        client.post("/api/auth/logout")
        login(client, "viewer_billing", "Viewer@Billing123")
        assert client.get("/api/navigation").status_code == 200


def test_user_login_is_limited_to_two_active_devices() -> None:
    username = f"viewer_device_{uuid.uuid4().hex[:8]}"
    password = "Viewer@Device123"
    with TestClient(app) as admin_client:
        login(admin_client)
        created = admin_client.post(
            "/api/admin/users",
            json={
                "username": username,
                "full_name": "Viewer Device Limit",
                "password": password,
                "role": "viewer",
            },
        )
        assert created.status_code == 200

    with TestClient(app) as first_client, TestClient(app) as second_client, TestClient(app) as third_client:
        login(first_client, username, password)
        login(second_client, username, password)
        assert first_client.get("/api/auth/me").status_code == 200
        assert second_client.get("/api/auth/me").status_code == 200

        login(third_client, username, password)
        assert third_client.get("/api/auth/me").status_code == 200
        assert first_client.get("/api/auth/me").status_code == 401
        assert second_client.get("/api/auth/me").status_code == 200


def test_five_failed_logins_send_telegram_alert(monkeypatch) -> None:
    sent_messages = []
    routes.FAILED_LOGIN_COUNTS.clear()

    def fake_send_message(self, title, message, details=None):
        sent_messages.append((title, message, details))
        return True

    monkeypatch.setattr("app.presentation.routes.TelegramNotifier.send_message", fake_send_message)
    with TestClient(app) as client:
        for _ in range(4):
            response = client.post("/api/auth/login", json={"username": "bad_admin", "password": "wrong"})
            assert response.status_code == 401
        assert sent_messages == []
        response = client.post("/api/auth/login", json={"username": "bad_admin", "password": "wrong"})
        assert response.status_code == 401
        assert len(sent_messages) == 1
        assert sent_messages[0][0] == "Canh bao dang nhap sai"
        response = client.post("/api/auth/login", json={"username": "bad_admin", "password": "wrong"})
        assert response.status_code == 401
        assert len(sent_messages) == 2


def test_telegram_alert_test_route_requires_admin(monkeypatch) -> None:
    sent_messages = []

    def fake_send_message(self, title, message, details=None):
        sent_messages.append((title, message, details))
        return True

    monkeypatch.setattr("app.presentation.routes.TelegramNotifier.send_message", fake_send_message)
    with TestClient(app) as client:
        response = client.get("/api/test/telegram-alert")
        assert response.status_code == 401
        assert sent_messages == []

        login(client)
        response = client.get("/api/test/telegram-alert")
        assert response.status_code == 200
        assert response.json()["ok"] is True
        assert sent_messages[0][0] == "TEST Telegram"
        assert "[TEST]" in sent_messages[0][1]
        assert sent_messages[0][2]["actor"] == "admin"


def test_user_import_rejects_large_file() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.post(
            "/api/admin/users/import",
            files={
                "file": (
                    "users.xlsx",
                    b"x" * (routes.MAX_USER_IMPORT_BYTES + 1),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 413


def test_zalo_webhook_rejects_invalid_secret(monkeypatch) -> None:
    monkeypatch.setenv("ZALO_WEBHOOK_SECRET", "zalo-secret-123")
    get_settings.cache_clear()
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/zalo/webhook",
                headers={"X-Bot-Api-Secret-Token": "wrong-secret"},
                json={"ok": True, "result": {"event_name": "message.text.received"}},
            )
            assert response.status_code == 403
    finally:
        get_settings.cache_clear()


def test_zalo_webhook_accepts_text_and_auto_replies(monkeypatch) -> None:
    sent_messages = []

    def fake_send_message(self, chat_id, text, parse_mode=None):
        sent_messages.append((chat_id, text, parse_mode))
        return True

    monkeypatch.setenv("ZALO_WEBHOOK_SECRET", "zalo-secret-123")
    monkeypatch.setenv("ZALO_BOT_TOKEN", "123456:test-token")
    get_settings.cache_clear()
    monkeypatch.setattr("app.presentation.routes.ZaloBotClient.send_message", fake_send_message)
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/zalo/webhook",
                headers={"X-Bot-Api-Secret-Token": "zalo-secret-123"},
                json={
                    "ok": True,
                    "result": {
                        "event_name": "message.text.received",
                        "message": {
                            "chat": {"id": "chat-001", "chat_type": "PRIVATE"},
                            "text": "ping",
                        },
                    },
                },
            )
            assert response.status_code == 200
            assert response.json()["auto_replied"] is True
            assert sent_messages == [("chat-001", "pong", None)]
    finally:
        get_settings.cache_clear()


def test_zalo_webhook_understands_mentioned_ping(monkeypatch) -> None:
    sent_messages = []

    def fake_send_message(self, chat_id, text, parse_mode=None):
        sent_messages.append((chat_id, text, parse_mode))
        return True

    monkeypatch.setenv("ZALO_WEBHOOK_SECRET", "zalo-secret-123")
    monkeypatch.setenv("ZALO_BOT_TOKEN", "123456:test-token")
    get_settings.cache_clear()
    monkeypatch.setattr("app.presentation.routes.ZaloBotClient.send_message", fake_send_message)
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/zalo/webhook",
                headers={"X-Bot-Api-Secret-Token": " zalo-secret-123 "},
                json={
                    "ok": True,
                    "result": {
                        "event_name": "message.text.received",
                        "message": {
                            "chat": {"id": "group-001", "chat_type": "GROUP"},
                            "text": "@Bot VNPT Can Tho ping",
                        },
                    },
                },
            )
            assert response.status_code == 200
            assert response.json()["auto_replied"] is True
            assert sent_messages == [("group-001", "pong", None)]
    finally:
        get_settings.cache_clear()


def test_zalo_webhook_accepts_result_json_string(monkeypatch) -> None:
    sent_messages = []

    def fake_send_message(self, chat_id, text, parse_mode=None):
        sent_messages.append((chat_id, text, parse_mode))
        return True

    monkeypatch.setenv("ZALO_WEBHOOK_SECRET", "zalo-secret-123")
    monkeypatch.setenv("ZALO_BOT_TOKEN", "123456:test-token")
    get_settings.cache_clear()
    monkeypatch.setattr("app.presentation.routes.ZaloBotClient.send_message", fake_send_message)
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/zalo/webhook",
                headers={"X-Bot-Api-Secret-Token": "zalo-secret-123"},
                json={
                    "ok": True,
                    "result": json.dumps({
                        "event_name": "message.text.received",
                        "message": {
                            "from": {"id": "user-json-001", "display_name": "Json User"},
                            "chat": {"id": "chat-json-001", "chat_type": "PRIVATE"},
                            "text": "ping",
                            "message_id": "msg-json-001",
                        },
                    }),
                },
            )
            assert response.status_code == 200
            payload = response.json()
            assert payload["chat_id"] == "chat-json-001"
            assert payload["text"] == "ping"
            assert payload["message_id"] == "msg-json-001"
            assert sent_messages == [("chat-json-001", "pong", None)]
    finally:
        get_settings.cache_clear()


def test_admin_can_view_zalo_message_logs(monkeypatch) -> None:
    def fake_send_message(self, chat_id, text, parse_mode=None):
        return True

    monkeypatch.setenv("ZALO_WEBHOOK_SECRET", "zalo-secret-123")
    monkeypatch.setenv("ZALO_BOT_TOKEN", "123456:test-token")
    get_settings.cache_clear()
    monkeypatch.setattr("app.presentation.routes.ZaloBotClient.send_message", fake_send_message)
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/zalo/webhook",
                headers={"X-Bot-Api-Secret-Token": "zalo-secret-123"},
                json={
                    "ok": True,
                    "result": {
                        "event_name": "message.text.received",
                        "message": {
                            "from": {"id": "user-log-001", "display_name": "Nguoi test log"},
                            "chat": {"id": "group-log-001", "chat_type": "GROUP"},
                            "message_id": "msg-log-001",
                            "text": "@Bot VNPT Can Tho ghi log",
                        },
                    },
                },
            )
            assert response.status_code == 200
            login(client)
            logs_response = client.get("/api/admin/zalo/message-logs?limit=20")
            assert logs_response.status_code == 200
            logs = logs_response.json()["logs"]
            assert any(log["direction"] == "in" and log["chat_id"] == "group-log-001" and "ghi log" in log["text"] for log in logs)
            assert any(log["direction"] == "out" and log["chat_id"] == "group-log-001" and log["ok"] is True for log in logs)
    finally:
        get_settings.cache_clear()


def test_admin_can_send_zalo_test_message_to_latest_chat(monkeypatch) -> None:
    sent_messages = []

    def fake_send_message(self, chat_id, text, parse_mode=None):
        sent_messages.append((chat_id, text, parse_mode))
        return True

    monkeypatch.setenv("ZALO_WEBHOOK_SECRET", "zalo-secret-123")
    monkeypatch.setenv("ZALO_BOT_TOKEN", "123456:test-token")
    get_settings.cache_clear()
    monkeypatch.setattr("app.presentation.routes.ZaloBotClient.send_message", fake_send_message)
    try:
        with TestClient(app) as client:
            response = client.post(
                "/api/zalo/webhook",
                headers={"X-Bot-Api-Secret-Token": "zalo-secret-123"},
                json={
                    "ok": True,
                    "result": {
                        "event_name": "message.text.received",
                        "message": {
                            "from": {"id": "user-manual-001", "display_name": "Manual User"},
                            "chat": {"id": "chat-manual-001", "chat_type": "PRIVATE"},
                            "message_id": "msg-manual-001",
                            "text": "hello",
                        },
                    },
                },
            )
            assert response.status_code == 200
            login(client)
            response = client.post("/api/admin/zalo/send-test-message", json={})
            assert response.status_code == 200
            assert response.json()["chat_id"] == "chat-manual-001"
            assert sent_messages[-1] == ("chat-manual-001", "Tin nhan test tu Bot VNPT Can Tho.", None)
    finally:
        get_settings.cache_clear()


def test_admin_can_setup_zalo_webhook(monkeypatch) -> None:
    def fake_configure_webhook(self):
        return {"ok": True, "message": "Da cai dat webhook Zalo Bot.", "details": {"webhook_url": "https://vnptcto.com/api/zalo/webhook"}}

    monkeypatch.setattr("app.presentation.routes.ZaloBotClient.configure_webhook", fake_configure_webhook)
    with TestClient(app) as client:
        login(client)
        response = client.post("/api/admin/zalo/webhook/setup")
        assert response.status_code == 200
        assert response.json()["details"]["webhook_url"] == "https://vnptcto.com/api/zalo/webhook"


def test_system_connections_include_zalo_bot() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.get("/api/admin/connections")
        assert response.status_code == 200
        connections = response.json()["connections"]
        codes = {connection["code"] for connection in connections}
        assert "zalo_bot" in codes
        assert "oracle_agency_db" in codes
        ftp = next(connection for connection in connections if connection["code"] == "ftp_storage")
        assert ftp["connection_type"] == "ftp"
        assert ftp["config"]["host"] == "10.159.23.100"
        assert ftp["config"]["username"] == "thangph.cto"
        assert "password" not in ftp["config"]
        assert "password" in ftp["protected_config_keys"]
        oracle = next(connection for connection in connections if connection["code"] == "oracle_agency_db")
        assert oracle["connection_type"] == "oracle"
        assert "password" not in oracle["config"]
        assert "password" in oracle["protected_config_keys"]


def test_admin_can_manage_report_links_and_active_links_are_public() -> None:
    with TestClient(app) as client:
        login(client)
        active_payload = {
            "ten_bao_cao": "Bao cao link active",
            "link": "https://docs.google.com/spreadsheets/d/sheet-link-active/edit#gid=0",
            "link_type": "google_sheet",
            "is_active": True,
        }
        created = client.post("/api/admin/report-links", json=active_payload)
        assert created.status_code == 200
        active_code = created.json()["ma_bao_cao"]
        assert active_code.startswith("LINK")

        inactive = client.post(
            "/api/admin/report-links",
            json={
                "ten_bao_cao": "Bao cao link inactive",
                "link": "https://drive.google.com/file/d/pdf-link-inactive/view",
                "link_type": "pdf",
                "is_active": False,
            },
        )
        assert inactive.status_code == 200

        form = client.post(
            "/api/admin/report-links",
            json={
                "ten_bao_cao": "Bieu mau Google",
                "link": "https://docs.google.com/forms/d/e/form-link/viewform",
                "link_type": "google_form",
                "is_active": True,
            },
        )
        assert form.status_code == 200

        duplicate_name = client.post(
            "/api/admin/report-links",
            json={**active_payload, "link": "https://docs.google.com/spreadsheets/d/another-sheet/edit"},
        )
        assert duplicate_name.status_code == 400
        assert "Ten bao cao" in duplicate_name.json()["detail"]

        duplicate_link = client.post(
            "/api/admin/report-links",
            json={**active_payload, "ten_bao_cao": "Bao cao link khac ten"},
        )
        assert duplicate_link.status_code == 400
        assert "Link nay" in duplicate_link.json()["detail"]

        admin_links = client.get("/api/report-links")
        assert admin_links.status_code == 200
        admin_payload = admin_links.json()["links"]
        active = next(item for item in admin_payload if item["ma_bao_cao"] == active_code)
        inactive_item = next(item for item in admin_payload if item["ten_bao_cao"] == "Bao cao link inactive")
        form_item = next(item for item in admin_payload if item["ten_bao_cao"] == "Bieu mau Google")
        assert active["download_url"].endswith(f"/api/report-links/{created.json()['id']}/download")
        assert inactive_item["is_active"] is False
        assert inactive_item["can_download"] is True
        assert form_item["can_download"] is False
        assert form_item["download_url"] == ""

        created_user = client.post(
            "/api/admin/users",
            json={
                "username": "viewer_report_links",
                "full_name": "Viewer Report Links",
                "role": "viewer",
                "password": "Viewer@Links2026!",
            },
        )
        assert created_user.status_code == 200
        client.post("/api/auth/logout")
        login(client, "viewer_report_links", "Viewer@Links2026!")

        navigation = client.get("/api/navigation")
        assert navigation.status_code == 200
        feature_codes = {feature["code"] for feature in navigation.json()["features"]}
        assert "baocaomoi" in feature_codes
        assert "linkbaocao" in feature_codes

        viewer_links = client.get("/api/report-links")
        assert viewer_links.status_code == 200
        viewer_names = {item["ten_bao_cao"] for item in viewer_links.json()["links"]}
        assert "Bao cao link active" in viewer_names
        assert "Bieu mau Google" in viewer_names
        assert "Bao cao link inactive" not in viewer_names


def test_internal_api_client_uses_active_connection_config() -> None:
    class FakeRepository:
        def get_system_connection_by_code(self, code: str) -> dict:
            assert code == "internal_fastapi_api"
            return {
                "connection_type": "internal_api",
                "is_active": True,
                "config": {
                    "url": "https://current-internal-api.example/api/du-lieu-web",
                    "mock_mode": "true",
                    "token": "connection-token",
                },
            }

    settings = Settings(
        internal_api_url="https://old-env-url.example/api/du-lieu-web",
        internal_api_mock_mode=False,
        internal_api_token="env-token",
    )

    client = routes.InternalApiClient.from_repository(settings, FakeRepository())

    assert client.api_url == "https://current-internal-api.example/api/du-lieu-web"
    assert client.mock_mode is True
    assert client.token == "connection-token"
    assert client.health_check()["api_url"] == "https://current-internal-api.example/api/du-lieu-web"


def test_internal_api_dns_error_message_points_to_tunnel_config() -> None:
    message = DatabaseService._internal_api_connection_message(Exception("[Errno 11001] getaddrinfo failed"))

    assert "Không phân giải được tên miền API dữ liệu nội bộ" in message
    assert "URL tunnel" in message


def test_seed_current_connections_preserves_internal_api_admin_config(tmp_path) -> None:
    repository = routes.AppRepository(str(tmp_path / "app.db"))
    repository.initialize("admin", "Admin@Brcd2026!")
    repository.upsert_system_connection(
        "internal_fastapi_api",
        "API du lieu noi bo",
        "internal_api",
        "Custom internal API",
        {
            "url": "https://current-internal-api.example/api/du-lieu-web",
            "mock_mode": False,
            "secret_ref": "INTERNAL_API_TOKEN",
        },
        True,
    )

    routes.ConnectionService(
        repository,
        Settings(
            internal_api_url="https://old-env-url.example/api/du-lieu-web",
            internal_api_mock_mode=True,
        ),
    ).seed_current_connections()

    stored = repository.get_system_connection_by_code("internal_fastapi_api")
    assert stored["config"]["url"] == "https://current-internal-api.example/api/du-lieu-web"
    assert stored["config"]["mock_mode"] is False


def test_seed_current_connections_preserves_ftp_admin_config(tmp_path) -> None:
    repository = routes.AppRepository(str(tmp_path / "app.db"))
    repository.initialize("admin", "Admin@Brcd2026!")
    repository.upsert_system_connection(
        "ftp_storage",
        "FTP rieng",
        "ftp",
        "Custom FTP",
        {
            "host": "10.1.1.10",
            "port": 2121,
            "username": "custom-user",
            "password": "custom-password",
            "passive": False,
            "timeout_seconds": 90,
            "secret_ref": "FTP_PASSWORD",
        },
        True,
    )

    routes.ConnectionService(repository, Settings()).seed_current_connections()

    stored = repository.get_system_connection_by_code("ftp_storage")
    assert stored["name"] == "FTP rieng"
    assert stored["config"]["host"] == "10.1.1.10"
    assert stored["config"]["port"] == 2121
    assert stored["config"]["username"] == "custom-user"
    assert stored["config"]["password"] == "custom-password"
    assert stored["config"]["passive"] is False
    assert stored["config"]["timeout_seconds"] == 90


def test_admin_can_manage_zalo_auto_messages_and_captures(monkeypatch) -> None:
    monkeypatch.setenv("APP_PUBLIC_URL", "https://vnptcto.com")
    get_settings.cache_clear()
    tiny_png = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO+/p9sAAAAASUVORK5CYII="
    try:
        with TestClient(app) as client:
            login(client)
            created = client.post(
                "/api/admin/zalo/auto-messages",
                json={
                    "name": "Dashboard 7h",
                    "page_url": "/dashboard",
                    "page_label": "Dashboard kinh doanh",
                    "schedule_type": "Daily",
                    "run_time": "07:00",
                    "target_type": "group",
                    "chat_id": "group-auto-001",
                    "caption": "Anh chup dashboard",
                    "is_active": True,
                },
            )
            assert created.status_code == 200
            schedule = created.json()["schedule"]
            assert schedule["schedule_id"].startswith("ZALO")

            listed = client.get("/api/admin/zalo/auto-messages")
            assert listed.status_code == 200
            assert any(item["schedule_id"] == schedule["schedule_id"] for item in listed.json()["schedules"])

            uploaded = client.post(
                f"/api/admin/zalo/auto-messages/{schedule['schedule_id']}/captures",
                json={"image_base64": f"data:image/png;base64,{tiny_png}", "mime_type": "image/png", "page_url": "/dashboard"},
            )
            assert uploaded.status_code == 200
            capture = uploaded.json()["capture"]
            assert uploaded.json()["capture_url"].startswith("https://vnptcto.com/api/zalo/auto-message-captures/")

            denied = client.get(f"/api/zalo/auto-message-captures/{capture['capture_id']}?token=wrong")
            assert denied.status_code == 404
            image = client.get(f"/api/zalo/auto-message-captures/{capture['capture_id']}?token={capture['public_token']}")
            assert image.status_code == 200
            assert image.headers["content-type"].startswith("image/png")
            assert image.content.startswith(b"\x89PNG")
    finally:
        get_settings.cache_clear()


def test_active_zalo_auto_message_requires_one_target() -> None:
    with TestClient(app) as client:
        login(client)
        no_target = client.post(
            "/api/admin/zalo/auto-messages",
            json={
                "name": "Bao cao chua chon dich",
                "page_url": "/dashboard",
                "schedule_type": "Daily",
                "run_time": "07:00",
                "target_type": "group",
                "chat_id": "",
                "is_active": True,
            },
        )
        assert no_target.status_code == 400
        assert "Chua chon" in no_target.json()["detail"]

        many_targets = client.post(
            "/api/admin/zalo/auto-messages",
            json={
                "name": "Bao cao nhieu dich",
                "page_url": "/dashboard",
                "schedule_type": "Daily",
                "run_time": "07:00",
                "target_type": "person",
                "chat_id": "chat-a,chat-b",
                "is_active": True,
            },
        )
        assert many_targets.status_code == 400
        assert "1 chat_id" in many_targets.json()["detail"]

        disabled_draft = client.post(
            "/api/admin/zalo/auto-messages",
            json={
                "name": "Ban nhap Zalo",
                "page_url": "/dashboard",
                "schedule_type": "Daily",
                "run_time": "07:00",
                "target_type": "group",
                "chat_id": "",
                "is_active": False,
            },
        )
        assert disabled_draft.status_code == 200
        assert disabled_draft.json()["schedule"]["is_active"] is False


def test_zalo_auto_capture_session_cookie_opens_authenticated_page() -> None:
    from app.application.zalo_auto_message_service import signed_capture_session_cookie

    with TestClient(app) as client:
        repository = routes.build_app_repository()
        user = repository.get_user_by_username(get_settings().initial_admin_username)
        assert user is not None
        client.cookies.set("brcd_session", signed_capture_session_cookie(get_settings(), user))

        response = client.get("/api/auth/me")
        assert response.status_code == 200
        assert response.json()["user"]["username"] == user["username"]


def test_zalo_auto_capture_playwright_cookie_uses_url_only(monkeypatch) -> None:
    from app.application.zalo_auto_message_service import playwright_session_cookie

    monkeypatch.setenv("APP_PUBLIC_URL", "https://vnptcto.com")
    get_settings.cache_clear()
    try:
        with TestClient(app):
            repository = routes.build_app_repository()
            user = repository.get_user_by_username(get_settings().initial_admin_username)
            assert user is not None

            cookie = playwright_session_cookie(get_settings(), user)
            assert cookie["url"] == "https://vnptcto.com"
            assert "path" not in cookie
            assert cookie["secure"] is True
    finally:
        get_settings.cache_clear()


def test_zalo_auto_capture_uses_dashboard_area(monkeypatch) -> None:
    from app.application import zalo_auto_message_service as service

    calls = {}

    class FakeRepository:
        def save_zalo_message_capture(self, schedule_id, image_base64, mime_type, page_url="", created_by=""):
            calls["saved"] = {
                "schedule_id": schedule_id,
                "image_base64": image_base64,
                "mime_type": mime_type,
                "page_url": page_url,
                "created_by": created_by,
            }
            return {"capture_id": "CAPTEST", "public_token": "token"}

    def fake_capture_page_screenshot_bytes(repository, settings, page_url, selector=service.DASHBOARD_CAPTURE_SELECTOR):
        calls["capture"] = {"page_url": page_url, "selector": selector}
        return b"\x89PNG\r\n"

    monkeypatch.setenv("APP_PUBLIC_URL", "https://vnptcto.com")
    get_settings.cache_clear()
    monkeypatch.setattr(service, "capture_page_screenshot_bytes", fake_capture_page_screenshot_bytes)
    try:
        result = service.capture_schedule_page_image(
            FakeRepository(),
            get_settings(),
            {"schedule_id": "ZALO0001", "page_url": "/dashboardtest"},
        )
        assert result["ok"] is True
        assert calls["capture"] == {"page_url": "/dashboardtest", "selector": service.DASHBOARD_CAPTURE_SELECTOR}
        assert calls["saved"]["image_base64"] == "iVBORw0K"
        assert result["capture_url"] == "https://vnptcto.com/api/zalo/auto-message-captures/CAPTEST?token=token"
    finally:
        get_settings.cache_clear()


def test_zalo_auto_message_scheduler_sends_due_photo(monkeypatch) -> None:
    from app.application.task_scheduler import LOCAL_TIMEZONE, ZaloAutoMessageScheduler

    events = []
    sent_messages = []

    def fake_send_photo(self, chat_id, photo_url, caption=""):
        events.append(("send", photo_url))
        sent_messages.append((chat_id, photo_url, caption))
        return True

    def fake_refresh_schedule_data(repository, settings, schedule):
        events.append(("refresh", schedule.get("page_url")))
        return {"ok": True, "page_id": "DASHBOARD_KINH_DOANH"}

    def fake_capture_schedule_page_image(repository, settings, schedule):
        events.append(("capture", schedule.get("page_url")))
        return {
            "ok": True,
            "capture": {"capture_id": "CAPTEST", "public_token": "token"},
            "capture_url": "https://vnptcto.com/api/zalo/auto-message-captures/CAPTEST?token=token",
        }

    monkeypatch.setenv("ZALO_BOT_TOKEN", "123456:test-token")
    get_settings.cache_clear()
    monkeypatch.setattr("app.application.zalo_auto_message_service.ZaloBotClient.send_photo", fake_send_photo)
    monkeypatch.setattr("app.application.zalo_auto_message_service.refresh_schedule_data", fake_refresh_schedule_data)
    monkeypatch.setattr("app.application.zalo_auto_message_service.capture_schedule_page_image", fake_capture_schedule_page_image)
    try:
        with TestClient(app) as client:
            login(client)
            created = client.post(
                "/api/admin/zalo/auto-messages",
                json={
                    "name": "Bao cao sang",
                    "page_url": "/dashboard",
                    "schedule_type": "Daily",
                    "run_time": "07:05",
                    "target_type": "person",
                    "chat_id": "chat-auto-001",
                    "caption": "Bao cao sang",
                    "photo_url": "https://example.com/dashboard.png",
                    "is_active": True,
                },
            )
            assert created.status_code == 200

            scheduler = ZaloAutoMessageScheduler()
            scheduler.configure(routes.build_app_repository(), get_settings())
            now = datetime(2026, 1, 5, 7, 5, tzinfo=LOCAL_TIMEZONE)
            assert scheduler.check_due_messages(now) == 1
            assert [event[0] for event in events] == ["refresh", "capture", "send"]
            assert sent_messages[-1] == (
                "chat-auto-001",
                "https://vnptcto.com/api/zalo/auto-message-captures/CAPTEST?token=token",
                "Bao cao sang",
            )
            assert scheduler.check_due_messages(now) == 0
    finally:
        get_settings.cache_clear()


def test_zalo_auto_message_requires_explicit_chat_id(monkeypatch) -> None:
    from app.application import zalo_auto_message_service as service

    refresh_calls = []

    class FakeRepository:
        def list_audit_logs(self, limit=500):
            return [{"action": "zalo_message_received", "details": '{"chat_id":"latest-chat"}'}]

    def fake_refresh_schedule_data(repository, settings, schedule):
        refresh_calls.append(schedule)
        return {"ok": True}

    monkeypatch.setattr(service, "refresh_schedule_data", fake_refresh_schedule_data)
    result = service.send_zalo_auto_message(
        FakeRepository(),
        get_settings(),
        {"schedule_id": "ZALOEMPTY", "name": "Bao cao", "chat_id": "", "page_url": "/dashboard"},
    )

    assert result["ok"] is False
    assert result["chat_id"] == ""
    assert "chat_id" in result["message"]
    assert refresh_calls == []


def test_admin_can_manage_data_mining_schedules_and_run_now(monkeypatch) -> None:
    calls = []

    def fake_run_data_mining_schedule(repository, settings, schedule, **kwargs):
        calls.append({
            "schedule_id": schedule["schedule_id"],
            "otp": kwargs.get("otp"),
            "created_by": kwargs.get("created_by"),
            "parameter_overrides": kwargs.get("parameter_overrides"),
        })
        run = kwargs.get("existing_run") or repository.create_data_mining_run(
            schedule["schedule_id"],
            schedule.get("parameters"),
            created_by=kwargs.get("created_by") or "",
        )
        result = {
            "ok": True,
            "status": "success",
            "message": "Da tai bao cao OneBSS.",
            "file_name": "bien_dong_0700_08072026.xlsx",
            "file_path": "data/data_mining_downloads/bien_dong_0700_08072026.xlsx",
            "storage_status": "saved_local",
        }
        repository.finish_data_mining_run(run["run_id"], result)
        return {**result, "run_id": run["run_id"], "schedule_id": schedule["schedule_id"]}

    monkeypatch.setattr(routes, "run_data_mining_schedule", fake_run_data_mining_schedule)
    with TestClient(app) as client:
        login(client)
        report_url = "https://onebss.vnpt.vn/#/report/bi?path=PHATTRIENTHUEBAO%2FBIENDONGPHATTRIENTHUEBAO%2FRP_BSS_28429&name=Test"
        created = client.post(
            "/api/admin/data-mining/schedules",
            json={
                "name": "Bien dong thue bao",
                "report_url": report_url,
                "schedule_type": "Daily",
                "run_time": "07:00",
                "storage_link": "data/test_downloads",
                "file_name_template": "bien_dong",
                "parameters": {"P_TUNGAY": "01/07/2026", "P_DENNGAY": "08/07/2026"},
                "is_active": True,
            },
        )
        assert created.status_code == 200
        schedule = created.json()["schedule"]
        assert schedule["schedule_id"].startswith("MINE")
        assert schedule["parameters"]["P_TUNGAY"] == "01/07/2026"

        listed = client.get("/api/admin/data-mining/schedules")
        assert listed.status_code == 200
        assert any(item["schedule_id"] == schedule["schedule_id"] for item in listed.json()["schedules"])

        run_now = client.post(
            f"/api/admin/data-mining/schedules/{schedule['schedule_id']}/run-now",
            json={"otp": "123456", "allow_device_registration": True, "parameters": {"P_DENNGAY": "09/07/2026"}},
        )
        assert run_now.status_code == 200
        assert run_now.json()["ok"] is True
        assert run_now.json()["status"] == "queued"
        for _ in range(20):
            if calls:
                break
            time.sleep(0.05)
        assert calls[-1]["otp"] == "123456"
        assert calls[-1]["parameter_overrides"] == {"P_DENNGAY": "09/07/2026"}

        runs = []
        for _ in range(20):
            runs = client.get(f"/api/admin/data-mining/runs?schedule_id={schedule['schedule_id']}").json()["runs"]
            if runs and runs[0]["status"] == "success":
                break
            time.sleep(0.05)
        assert len(runs) == 1
        assert runs[0]["file_name"] == "bien_dong_0700_08072026.xlsx"

        assert client.delete(f"/api/admin/data-mining/schedules/{schedule['schedule_id']}").status_code == 200
        assert client.get(f"/api/admin/data-mining/runs?schedule_id={schedule['schedule_id']}").json()["runs"] == []


def test_google_drive_folder_link_and_storage_upload(monkeypatch, tmp_path) -> None:
    from app.application import onebss_data_mining_service as service
    from app.application.google_drive_service import extract_google_drive_folder_id

    uploaded_calls = []

    def fake_upload_file_to_google_drive(settings, local_path, file_name, folder_id, mime_type="", repository=None):
        uploaded_calls.append((str(local_path), file_name, folder_id))
        return {"file_id": "drive-file-001", "web_view_link": "https://drive.google.com/file/d/drive-file-001/view"}

    local_file = tmp_path / "report.xlsx"
    local_file.write_bytes(b"excel")
    monkeypatch.setattr(service, "upload_file_to_google_drive", fake_upload_file_to_google_drive)

    folder_url = "https://drive.google.com/drive/folders/1TJqLjq8OpZ_x_D-djxRk0w4HacUh4HmS"
    assert extract_google_drive_folder_id(folder_url) == "1TJqLjq8OpZ_x_D-djxRk0w4HacUh4HmS"
    result = service.save_downloaded_file(get_settings(), local_file, folder_url)
    assert result["ok"] is True
    assert result["storage_link"] == "https://drive.google.com/file/d/drive-file-001/view"
    assert result["storage_status"] == "uploaded_google_drive:drive-file-001"
    assert uploaded_calls == [(str(local_file), "report.xlsx", "1TJqLjq8OpZ_x_D-djxRk0w4HacUh4HmS")]


def test_google_drive_oauth_folder_config_uploads_without_env_folder(monkeypatch, tmp_path) -> None:
    from app.application import onebss_data_mining_service as service

    uploaded_calls = []

    class FakeRepository:
        def get_system_connection_by_code(self, code):
            assert code == "drive_storage"
            return {
                "config": {
                    "provider": "google_drive_oauth",
                    "folder": "oauth-folder-001",
                    "oauth_refresh_token_enc": "enc:stored-refresh-token",
                }
            }

    def fake_upload_file_to_google_drive(settings, local_path, file_name, folder_id, mime_type="", repository=None):
        uploaded_calls.append((str(local_path), file_name, folder_id, repository))
        return {"file_id": "drive-file-oauth", "web_view_link": "https://drive.google.com/file/d/drive-file-oauth/view"}

    local_file = tmp_path / "report.xlsx"
    local_file.write_bytes(b"excel")
    settings = get_settings().model_copy(update={"google_drive_folder_id": ""})
    repository = FakeRepository()
    monkeypatch.setattr(service, "upload_file_to_google_drive", fake_upload_file_to_google_drive)

    result = service.save_downloaded_file(settings, local_file, "", repository)

    assert result["ok"] is True
    assert result["storage_link"] == "https://drive.google.com/file/d/drive-file-oauth/view"
    assert result["storage_status"] == "uploaded_google_drive:drive-file-oauth"
    assert uploaded_calls == [(str(local_file), "report.xlsx", "oauth-folder-001", repository)]


def test_google_drive_oauth_start_and_protected_config(monkeypatch) -> None:
    monkeypatch.setenv("GOOGLE_DRIVE_OAUTH_CLIENT_ID", "client-id.apps.googleusercontent.com")
    monkeypatch.setenv("GOOGLE_DRIVE_OAUTH_CLIENT_SECRET", "client-secret")
    monkeypatch.setenv("APP_PUBLIC_URL", "https://vnptcto.com")
    get_settings.cache_clear()
    with TestClient(app) as client:
        login(client)
        repository = routes.build_app_repository()
        repository.upsert_system_connection(
            "drive_storage",
            "Google Drive",
            "drive",
            "Drive OAuth",
            {
                "provider": "google_drive_oauth",
                "folder": "folder-001",
                "oauth_email": "owner@example.com",
                "oauth_refresh_token_enc": "enc:stored-refresh-token",
            },
            True,
        )

        connections = client.get("/api/admin/connections")
        assert connections.status_code == 200
        drive = next(item for item in connections.json()["connections"] if item["code"] == "drive_storage")
        assert "oauth_refresh_token_enc" not in drive["config"]
        assert "oauth_refresh_token_enc" in drive["protected_config_keys"]

        saved = client.put(
            "/api/admin/connections/drive_storage",
            json={
                "name": "Google Drive",
                "connection_type": "drive",
                "description": "Drive OAuth",
                "config": {"provider": "google_drive_oauth", "folder": "folder-002"},
                "is_active": True,
            },
        )
        assert saved.status_code == 200
        stored = repository.get_system_connection_by_code("drive_storage")
        assert stored["config"]["folder"] == "folder-002"
        assert stored["config"]["oauth_refresh_token_enc"] == "enc:stored-refresh-token"

        folder = client.post(
            "/api/google-drive/oauth/folder",
            json={"folder_id": "https://drive.google.com/drive/folders/folder-003"},
        )
        assert folder.status_code == 200
        assert folder.json()["folder_id"] == "folder-003"
        stored = repository.get_system_connection_by_code("drive_storage")
        assert stored["config"]["folder"] == "folder-003"
        assert stored["config"]["folder_id"] == "folder-003"
        assert stored["config"]["oauth_refresh_token_enc"] == "enc:stored-refresh-token"

        start = client.post("/api/google-drive/oauth/start")
        assert start.status_code == 200
        body = start.json()
        assert body["redirect_uri"] == "https://vnptcto.com/api/google-drive/oauth/callback"
        assert "https://accounts.google.com/o/oauth2/v2/auth" in body["authorization_url"]
        assert "access_type=offline" in body["authorization_url"]
    monkeypatch.delenv("GOOGLE_DRIVE_OAUTH_CLIENT_ID", raising=False)
    monkeypatch.delenv("GOOGLE_DRIVE_OAUTH_CLIENT_SECRET", raising=False)
    monkeypatch.delenv("APP_PUBLIC_URL", raising=False)
    get_settings.cache_clear()


def test_internal_email_connection_password_is_configurable_and_protected() -> None:
    with TestClient(app) as client:
        login(client)
        repository = routes.build_app_repository()

        saved = client.put(
            "/api/admin/connections/internal_email",
            json={
                "name": "Email nội bộ VNPT",
                "connection_type": "internal_email",
                "description": "Đồng bộ hộp thư nội bộ qua IMAP.",
                "config": {
                    "host": "email.vnpt.vn",
                    "port": 993,
                    "use_ssl": True,
                    "username": "otp.user@vnpt.vn",
                    "password": "Mail@Test123!",
                    "mailbox": "INBOX",
                    "lookback_minutes": 20,
                    "max_messages": 25,
                    "timeout_seconds": 15,
                    "sync_interval_seconds": 45,
                },
                "is_active": True,
            },
        )
        assert saved.status_code == 200
        stored = repository.get_system_connection_by_code("internal_email")
        assert stored["config"]["password"] == "Mail@Test123!"
        assert stored["config"]["sync_interval_seconds"] == 45

        connections = client.get("/api/admin/connections")
        assert connections.status_code == 200
        email = next(item for item in connections.json()["connections"] if item["code"] == "internal_email")
        assert email["config"]["username"] == "otp.user@vnpt.vn"
        assert "password" not in email["config"]
        assert "password" in email["protected_config_keys"]

        updated = client.put(
            "/api/admin/connections/internal_email",
            json={
                "name": "Email nội bộ VNPT",
                "connection_type": "internal_email",
                "description": "Đồng bộ hộp thư nội bộ qua IMAP.",
                "config": {
                    "host": "email.vnpt.vn",
                    "port": 993,
                    "use_ssl": True,
                    "username": "otp.user@vnpt.vn",
                    "mailbox": "INBOX",
                    "lookback_minutes": 30,
                    "max_messages": 40,
                    "timeout_seconds": 20,
                    "sync_interval_seconds": 60,
                },
                "is_active": True,
            },
        )
        assert updated.status_code == 200
        stored_after_update = repository.get_system_connection_by_code("internal_email")
        assert stored_after_update["config"]["password"] == "Mail@Test123!"
        assert stored_after_update["config"]["sync_interval_seconds"] == 60


def test_data_mining_dynamic_date_parameters() -> None:
    from app.application.onebss_data_mining_service import LOCAL_TIMEZONE, resolve_dynamic_parameters

    now = datetime(2026, 7, 8, 9, 30, tzinfo=LOCAL_TIMEZONE)
    params = resolve_dynamic_parameters(
        {
            "P_TUNGAY": "{{month_start}}",
            "P_DENNGAY": "{{today}}",
            "P_HOMQUA": "{{yesterday}}",
            "P_CUOITHANG": "{{month_end}}",
            "P_OFFSET": "{{today-7d}}",
            "P_THANG": "{today;MM/yyyy}",
            "P_YMD": "{today;yyyyMMdd}",
            "P_OFFSET_FMT": "{today-7d;dd/MM/yyyy}",
            "P_LAST_MONTH": "{{last_month_start;MM/yyyy}}",
            "P_STATIC": "13",
            "P_UNKNOWN": "{not_a_date;MM/yyyy}",
        },
        now,
    )
    assert params["P_TUNGAY"] == "01/07/2026"
    assert params["P_DENNGAY"] == "08/07/2026"
    assert params["P_HOMQUA"] == "07/07/2026"
    assert params["P_CUOITHANG"] == "31/07/2026"
    assert params["P_OFFSET"] == "01/07/2026"
    assert params["P_THANG"] == "07/2026"
    assert params["P_YMD"] == "20260708"
    assert params["P_OFFSET_FMT"] == "01/07/2026"
    assert params["P_LAST_MONTH"] == "06/2026"
    assert params["P_STATIC"] == "13"
    assert params["P_UNKNOWN"] == "{not_a_date;MM/yyyy}"


def test_data_mining_run_resolves_parameters_before_download(monkeypatch) -> None:
    from app.application import onebss_data_mining_service as service

    class FixedDateTime(datetime):
        @classmethod
        def now(cls, tz=None):
            return datetime(2026, 7, 8, 9, 30, tzinfo=tz)

    class FakeRepository:
        def __init__(self):
            self.created_parameters = None
            self.finished_result = None

        def create_data_mining_run(self, schedule_id, parameters, created_by=""):
            self.created_parameters = parameters
            return {"run_id": "RUN001", "schedule_id": schedule_id, "parameters": parameters}

        def finish_data_mining_run(self, run_id, result):
            self.finished_result = result

    class FakeDownloader:
        def __init__(self, settings):
            self.settings = settings

        def download_report(self, schedule, **kwargs):
            return {
                "ok": True,
                "status": "success",
                "message": "ok",
                "parameters": schedule["parameters"],
            }

    monkeypatch.setattr(service, "datetime", FixedDateTime)
    monkeypatch.setattr(service, "OneBssReportDownloader", FakeDownloader)
    repository = FakeRepository()
    result = service.run_data_mining_schedule(
        repository,
        get_settings(),
        {
            "schedule_id": "MINE0001",
            "parameters": {"P_TUNGAY": "{{month_start}}", "P_DENNGAY": "{{today}}"},
        },
        parameter_overrides={"P_DENNGAY": "{{today-1d}}"},
        created_by="admin",
    )
    assert repository.created_parameters == {"P_TUNGAY": "01/07/2026", "P_DENNGAY": "07/07/2026"}
    assert result["parameters"] == repository.created_parameters


def test_data_mining_scheduler_runs_due_schedule_once(monkeypatch) -> None:
    from app.application.task_scheduler import DataMiningScheduler, LOCAL_TIMEZONE

    calls = []

    def fake_run_data_mining_schedule(repository, settings, schedule, **kwargs):
        calls.append((schedule["schedule_id"], kwargs.get("interactive")))
        return {
            "ok": True,
            "status": "success",
            "message": "Da tai bao cao OneBSS.",
            "file_name": "scheduler_0711_08072026.xlsx",
            "file_path": "data/data_mining_downloads/scheduler_0711_08072026.xlsx",
        }

    monkeypatch.setattr("app.application.task_scheduler.run_data_mining_schedule", fake_run_data_mining_schedule)
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/data-mining/schedules",
            json={
                "name": "Scheduler OneBSS",
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=PHATTRIENTHUEBAO%2FBIENDONGPHATTRIENTHUEBAO%2FRP_BSS_28429&name=Test",
                "schedule_type": "Daily",
                "run_time": "07:11",
                "file_name_template": "scheduler",
                "parameters": {},
                "is_active": True,
            },
        )
        assert created.status_code == 200
        schedule_id = created.json()["schedule"]["schedule_id"]

        scheduler = DataMiningScheduler()
        scheduler.configure(routes.build_app_repository(), get_settings())
        now = datetime(2026, 7, 8, 7, 11, tzinfo=LOCAL_TIMEZONE)
        assert scheduler.check_due_schedules(now) == 1
        assert calls == [(schedule_id, False)]
        assert scheduler.check_due_schedules(now) == 0

        refreshed = routes.build_app_repository().get_data_mining_schedule(schedule_id)
        assert refreshed["last_status"] == "success"
        assert refreshed["last_file_name"] == "scheduler_0711_08072026.xlsx"


def test_admin_can_manage_task_report_auto_and_queue_run() -> None:
    with TestClient(app) as client:
        login(client)
        source_code = f"BC_AUTO_SQL_{uuid.uuid4().hex[:8].upper()}"
        onebss_code = f"ONEBSS_AUTO_{uuid.uuid4().hex[:8].upper()}"
        ftp_code = f"FTP_AUTO_{uuid.uuid4().hex[:8].upper()}"
        assert client.post(
            "/api/admin/sql-reports",
            json={
                "ten_bao_cao": "Auto SQL source",
                "ma_bao_cao": source_code,
                "cau_lenh_sql": "SELECT :P_THANG AS P_THANG FROM dual;",
                "cac_tham_so": ["P_THANG"],
            },
        ).status_code == 200
        assert client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "Auto OneBSS source",
                "ma_bao_cao": onebss_code,
                "danh_sach_bien": ["P_TUNGAY", "P_DENNGAY"],
                "parameters": {"P_TUNGAY": "{{month_start}}", "P_DENNGAY": "{{today}}"},
                "otp_service_code": "onebss",
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test",
                "storage_link": "",
            },
        ).status_code == 200
        assert client.post(
            "/api/admin/ftp-reports",
            json={
                "ten_bao_cao": "Auto FTP source",
                "ma_bao_cao": ftp_code,
                "folder_path": "/reports",
                "file_name_template": "report_{yyyymmdd}.xlsx",
                "connection_code": "ftp_storage",
                "is_active": True,
            },
        ).status_code == 200

        source_configs = client.get("/api/admin/task-report-auto/source-configs")
        assert source_configs.status_code == 200
        source_payload = source_configs.json()["reports"]
        assert source_code in {report["ma_bao_cao"] for report in source_payload["sql"]}
        assert onebss_code in {report["ma_bao_cao"] for report in source_payload["onebss"]}
        assert ftp_code in {report["ma_bao_cao"] for report in source_payload["ftp"]}

        created = client.post(
            "/api/admin/task-report-auto/tasks",
            json={
                "name": "Auto SQL sang",
                "source_type": "sql",
                "source_code": source_code,
                "source_config": {"filters": {"P_THANG": "202608"}},
                "schedule_type": "TimeWindow",
                "time_slots": ["07:00", "11:30"],
                "run_time": "07:00",
                "spreadsheet_url": "https://docs.google.com/spreadsheets/d/1AbCdEfGhIjKlMnOpQrStUvWxYz1234567890/edit#gid=0",
                "sheet_name": "DATA",
                "public_url": "https://example.com/report",
                "public_wait_selector": "#report-root",
                "target_type": "group",
                "chat_id": "zalo-group-001",
                "chat_name": "Bao cao",
                "caption": "Bao cao auto",
                "retry_limit": 2,
                "is_active": True,
            },
        )
        assert created.status_code == 200
        task = created.json()["task"]
        assert task["task_id"].startswith("TRA")
        assert task["source_type"] == "sql"
        assert task["spreadsheet_id"] == "1AbCdEfGhIjKlMnOpQrStUvWxYz1234567890"
        assert task["time_slots"] == ["07:00", "11:30"]

        page = client.get("/taskreportauto")
        assert page.status_code == 200
        assert 'id="view-task-report-auto"' in page.text
        app_js = client.get("/static/app.js?v=227").text
        assert "/static/task-report-auto.js?v=2" in app_js
        task_auto_js = client.get("/static/task-report-auto.js?v=2")
        assert task_auto_js.status_code == 200
        assert "/api/admin/task-report-auto/source-configs" in task_auto_js.text
        assert 'name="source_code" required></select>' in task_auto_js.text
        assert "task-auto-advanced" in task_auto_js.text
        assert 'data-schedule-field="time_slots"' in task_auto_js.text

        queued = client.post(f"/api/admin/task-report-auto/tasks/{task['task_id']}/run-now", json={"source_config": {}})
        assert queued.status_code == 200
        assert queued.json()["run"]["status"] == "queued"

        runs = client.get(f"/api/admin/task-report-auto/runs?task_id={task['task_id']}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["task_id"] == task["task_id"]
        assert runs[0]["status"] == "queued"

        assert client.delete(f"/api/admin/task-report-auto/tasks/{task['task_id']}").status_code == 200
        assert client.get(f"/api/admin/task-report-auto/runs?task_id={task['task_id']}").json()["runs"] == []


def test_task_report_auto_runner_sql_writes_sheet_captures_and_sends_zalo(monkeypatch) -> None:
    from app.application import task_report_auto_service as service
    from app.application.task_report_auto_service import TaskReportAutoRunner

    class FakeDatabaseService:
        def __init__(self, *args, **kwargs):
            pass

        def export_dynamic_report(self, **kwargs):
            return {
                "ok": True,
                "message": "Da lay SQL.",
                "columns": ["MA_TB", "DOANH_THU"],
                "rows": [{"MA_TB": "TB001", "DOANH_THU": 120000}],
            }

    written_values = []
    sent_photos = []

    def fake_write_values(settings, repository, spreadsheet_id, sheet_name, values):
        written_values.append((spreadsheet_id, sheet_name, values))
        return {"updated_cells": sum(len(row) for row in values)}

    class FakeZaloBotClient:
        def __init__(self, settings):
            self.settings = settings

        def send_photo(self, chat_id, photo_url, caption):
            sent_photos.append((chat_id, photo_url, caption))
            return True

    monkeypatch.setattr(service, "DatabaseService", FakeDatabaseService)
    monkeypatch.setattr(service, "write_values_to_google_sheet", fake_write_values)
    monkeypatch.setattr(service, "capture_public_web_screenshot_bytes", lambda *args, **kwargs: b"\x89PNG\r\n\x1a\n" + b"x" * 200)
    monkeypatch.setattr(service, "ZaloBotClient", FakeZaloBotClient)

    repository = routes.build_app_repository()
    task_id = repository.generate_task_report_auto_task_id()
    repository.save_task_report_auto_task(
        {
            "task_id": task_id,
            "name": "Runner SQL",
            "source_type": "sql",
            "source_code": "BC_RUNNER_SQL",
            "source_config": {"filters": {"P": "1"}},
            "schedule_type": "Daily",
            "time_slots": [],
            "run_time": "07:00",
            "weekday": "",
            "month_day": 1,
            "spreadsheet_id": "sheet-runner-001",
            "spreadsheet_url": "",
            "sheet_name": "DATA",
            "public_url": "https://example.com/dashboard",
            "public_wait_selector": "",
            "target_type": "group",
            "chat_id": "zalo-runner-001",
            "chat_name": "Runner",
            "caption": "Caption runner",
            "retry_limit": 1,
            "is_active": True,
        }
    )
    run = repository.create_task_report_auto_run(task_id, "manual:test-runner", created_by="pytest", status="running")

    settings = get_settings().model_copy(update={"app_public_url": "https://vnptcto.com"})
    result = TaskReportAutoRunner(repository, settings, step_wait_seconds=0).run_claimed(run)

    assert result["ok"] is True
    assert written_values == [("sheet-runner-001", "DATA", [["MA_TB", "DOANH_THU"], ["TB001", 120000]])]
    assert sent_photos
    assert sent_photos[0][0] == "zalo-runner-001"
    assert sent_photos[0][1].startswith("https://vnptcto.com/api/task-report-auto/captures/")
    assert sent_photos[0][2] == "Caption runner"
    finished = repository.get_task_report_auto_run(run["run_id"])
    assert finished["status"] == "success"
    assert set(finished["step_results"]) == {"mine", "sheet", "capture", "zalo"}


def test_admin_can_manage_work_tasks() -> None:
    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_cong_viec": "Gia cuoc",
            "type": "Daily",
            "time": "07:00",
            "weekday": "",
            "once_date": "",
            "group": "ME",
            "check": False,
        }
        created = client.post("/api/admin/work-tasks", json=payload)
        assert created.status_code == 200
        task_id = created.json()["task"]["task_id"]
        assert task_id.startswith("TASK")

        tasks = client.get("/api/admin/work-tasks").json()["tasks"]
        assert any(task["task_id"] == task_id and task["check"] is False for task in tasks)

        completed = client.post(f"/api/admin/work-tasks/{task_id}/complete")
        assert completed.status_code == 200
        active_tasks = client.get("/api/admin/work-tasks").json()["tasks"]
        assert all(task["task_id"] != task_id for task in active_tasks)

        all_tasks = client.get("/api/admin/work-tasks?include_completed=true").json()["tasks"]
        assert any(task["task_id"] == task_id and task["check"] is True for task in all_tasks)


def test_database_health_requires_login_and_uses_mock_mode() -> None:
    with TestClient(app) as client:
        assert client.get("/api/health/database").status_code == 401
        login(client)
        response = client.get("/api/health/database")
        assert response.status_code == 200
        assert response.json()["details"]["mode"] == "mock"


def test_dashboard_datcoc_table_uses_internal_api() -> None:
    with TestClient(app) as client:
        assert client.get("/api/dashboard/datcoc-test").status_code == 401
        login(client)
        response = client.get("/api/dashboard/datcoc-test")
        assert response.status_code == 200
        payload = response.json()
        assert payload["sql"] == "select * from css_cto.db_datcoc where ma_tb = 'thanhbinh-omon'"
        assert payload["columns"]
        assert payload["rows"]


def test_dashboard_fiber_uses_internal_api() -> None:
    with TestClient(app) as client:
        assert client.get("/api/dashboard/fiber").status_code == 401
        login(client)
        response = client.get("/api/dashboard/fiber")
        assert response.status_code == 200
        payload = response.json()
        assert payload["groups"]["vnpt"]["rows"][0]["rank"] == 1
        assert len(payload["groups"]["vnpt"]["rows"]) == 13
        assert len(payload["groups"]["ttvt"]["rows"]) == 13
        assert payload["summary"]["production"]["fiber"] == payload["groups"]["vnpt"]["total"]


def test_system_status_requires_login_and_reports_internal_api_policy() -> None:
    with TestClient(app) as client:
        assert client.get("/api/system/status").status_code == 401
        login(client)
        response = client.get("/api/system/status")
        assert response.status_code == 200
        payload = response.json()
        assert payload["internal_api"]["mock_mode"] is True
        assert payload["internal_api"]["url"] == "http://10.92.17.88:8000/api/du-lieu-web"
        assert payload["query_policy"]["data_source"] == "internal_fastapi"
        assert payload["query_policy"]["page_size_max"] == 20


def test_api_ping_is_not_captured_by_frontend_route() -> None:
    with TestClient(app) as client:
        response = client.get("/api/ping")
        assert response.status_code == 200
        payload = response.json()
        assert payload["ok"] is True
        assert payload["status"] == "alive"
        assert "bootstrap" in payload


def test_admin_can_manage_sql_reports_and_run_dynamic_report() -> None:
    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Báo cáo thuê bao test",
            "ma_bao_cao": "BC_TEST_THUE_BAO",
            "cau_lenh_sql": "SELECT ma_tb, ten_tb FROM css_cto.db_thuebao WHERE trang_thai = :status;",
            "cac_tham_so": ["status"],
        }
        created = client.post("/api/admin/sql-reports", json=payload)
        assert created.status_code == 200

        reports = client.get("/api/admin/sql-reports")
        assert reports.status_code == 200
        assert any(report["ma_bao_cao"] == "BC_TEST_THUE_BAO" for report in reports.json()["reports"])

        public_configs = client.get("/api/reports/configs")
        assert public_configs.status_code == 200
        first_config = next(report for report in public_configs.json()["reports"] if report["ma_bao_cao"] == "BC_TEST_THUE_BAO")
        assert "cau_lenh_sql" not in first_config

        result = client.post(
            "/api/reports/run",
            json={"ma_bao_cao": "BC_TEST_THUE_BAO", "filters": {"status": "1"}, "page": 1, "page_size": 20},
        )
        assert result.status_code == 200
        body = result.json()
        assert body["columns"] == ["STT", "MA_BAO_CAO", "TEN_BAO_CAO", "THAM_SO"]
        assert body["pagination"]["page_size"] == 20


def test_dynamic_report_history_records_loaded_result(monkeypatch) -> None:
    def fake_run_sql_report(self, **kwargs):
        return {
            "ok": True,
            "columns": ["MA_TB", "TEN_TB"],
            "rows": [{"MA_TB": "tb-history", "TEN_TB": "Thue bao history"}],
            "total": 1,
            "page": kwargs["page"],
            "page_size": kwargs["page_size"],
            "message": "ok",
        }

    monkeypatch.setattr(routes.InternalApiClient, "run_sql_report", fake_run_sql_report)

    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Bao cao lich su",
            "ma_bao_cao": "BC_HISTORY_LOAD",
            "cau_lenh_sql": "SELECT ma_tb, ten_tb FROM css_cto.db_thuebao;",
            "cac_tham_so": [],
        }
        assert client.post("/api/admin/sql-reports", json=payload).status_code == 200
        result = client.post(
            "/api/reports/run",
            json={"ma_bao_cao": "BC_HISTORY_LOAD", "filters": {}, "page": 1, "page_size": 20},
        )
        assert result.status_code == 200

        history = client.get("/api/reports/history?limit=20")
        assert history.status_code == 200
        items = history.json()["items"]
        item = next(row for row in items if row["ma_bao_cao"] == "BC_HISTORY_LOAD")
        assert item["event_type"] == "load"
        assert item["status"] == "success"
        assert item["rows"] == 1
        assert item["total"] == 1


def test_dynamic_report_http_530_falls_back_to_sql_worker(monkeypatch, tmp_path) -> None:
    import httpx

    def raise_530(self, **kwargs):
        request = httpx.Request("POST", "https://api.vnptcto.com/api/du-lieu-web")
        response = httpx.Response(530, request=request, text="cloudflare tunnel error")
        raise httpx.HTTPStatusError("cloudflare tunnel error", request=request, response=response)

    monkeypatch.setattr(routes.InternalApiClient, "run_sql_report", raise_530)
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()

    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Bao cao fallback worker",
            "ma_bao_cao": "BC_SQL_WORKER_FALLBACK",
            "cau_lenh_sql": "SELECT ma_tb, ten_tb FROM css_cto.db_thuebao WHERE trang_thai = :status;",
            "cac_tham_so": ["status"],
        }
        assert client.post("/api/admin/sql-reports", json=payload).status_code == 200
        started = client.post(
            "/api/reports/run-jobs",
            json={"ma_bao_cao": "BC_SQL_WORKER_FALLBACK", "filters": {"status": "1"}, "page": 1, "page_size": 20},
        )
        assert started.status_code == 200
        job_id = started.json()["job_id"]

        job = {}
        for _ in range(20):
            job = client.get(f"/api/reports/run-jobs/{job_id}").json()
            if job["status"] == "queued_worker":
                break
            time.sleep(0.05)
        assert job["status"] == "queued_worker"
        assert "may tram" in job["message"].lower()

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-sql"}, headers=headers)
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["query"]["action"] == "run_sql_report"
        assert task["query"]["tham_so"] == {"status": "1"}
        assert task["query"]["pagination"] == {"page": 1, "page_size": 20}

        finished = client.post(
            f"/api/sql-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "May tram da tai du lieu SQL qua API local.",
                "columns": ["MA_TB", "TEN_TB"],
                "rows": [{"MA_TB": "tb-local", "TEN_TB": "Local API"}],
                "pagination": {"page": 1, "page_size": 20, "total": 1},
                "report": task["report"],
                "details": {"source": "local_api"},
            },
            headers=headers,
        )
        assert finished.status_code == 200
        completed = client.get(f"/api/reports/run-jobs/{job_id}").json()
        assert completed["status"] == "complete"
        assert completed["rows"] == [{"MA_TB": "tb-local", "TEN_TB": "Local API"}]
        assert completed["pagination"]["total"] == 1


def test_dynamic_report_direct_tunnel_endpoint_is_disabled(monkeypatch) -> None:
    calls = []

    def fake_run_sql_report(self, **kwargs):
        calls.append(kwargs)
        return {"ok": True, "columns": ["MA_TB"], "rows": [{"MA_TB": "should-not-run"}], "total": 1}

    monkeypatch.setattr(routes.InternalApiClient, "run_sql_report", fake_run_sql_report)
    monkeypatch.setattr(routes.get_settings(), "internal_api_url", "https://api.vnptcto.com/api/du-lieu-web")

    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Bao cao worker only",
            "ma_bao_cao": "BC_WORKER_ONLY",
            "cau_lenh_sql": "SELECT ma_tb FROM css_cto.db_thuebao;",
            "cac_tham_so": [],
        }
        assert client.post("/api/admin/sql-reports", json=payload).status_code == 200

        direct = client.post(
            "/api/reports/run",
            json={"ma_bao_cao": "BC_WORKER_ONLY", "filters": {}, "page": 1, "page_size": 20},
        )
        assert direct.status_code == 200
        direct_body = direct.json()
        assert direct_body["ok"] is False
        assert direct_body["status"] == "worker_only"
        assert direct_body["details"]["use_endpoint"] == "/api/reports/export-jobs"

        job = client.post(
            "/api/reports/run-jobs",
            json={"ma_bao_cao": "BC_WORKER_ONLY", "filters": {}, "page": 1, "page_size": 20},
        )
        assert job.status_code == 200
        job_body = job.json()
        assert job_body["status"] == "failed"
        assert job_body["details"]["disabled_endpoint"] is True
        assert calls == []


def test_database_service_disables_direct_tunnel_sql_calls() -> None:
    calls = []

    class FakeInternalApi:
        api_url = "https://api.vnptcto.com/api/du-lieu-web"

        def run_sql_report(self, **kwargs):
            calls.append(kwargs)
            raise AssertionError("direct tunnel SQL call should be blocked")

    class FakeRepository:
        def get_sql_report_by_id(self, report_id):
            return None

        def get_sql_report_by_code(self, code):
            return {
                "ten_bao_cao": "Bao cao worker only",
                "ma_bao_cao": "BC_WORKER_ONLY_SERVICE",
                "cau_lenh_sql": "SELECT ma_tb FROM css_cto.db_thuebao WHERE trang_thai = :STATUS;",
                "cac_tham_so": ["STATUS"],
            }

        def list_sql_reports(self):
            return []

    service = DatabaseService(FakeInternalApi(), FakeRepository())
    result = service.run_dynamic_report(
        ma_bao_cao="BC_WORKER_ONLY_SERVICE",
        filters={"STATUS": "1"},
        page=1,
        page_size=20,
    )
    fiber = service.run_dashboard_fiber()

    assert result["ok"] is False
    assert result["details"]["disabled_endpoint"] is True
    assert result["details"]["use_endpoint"] == "/api/reports/export-jobs"
    assert fiber["ok"] is False
    assert calls == []


def test_dynamic_report_search_and_excel_export_use_full_result_set(monkeypatch) -> None:
    rows = [
        {"MA_TB": "tb001", "TEN_TB": "Nguyen Van A", "DIACHI_LD": "Can Tho"},
        {"MA_TB": "tb002", "TEN_TB": "Tran Binh", "DIACHI_LD": "Soc Trang"},
        {"MA_TB": "tb003", "TEN_TB": "Phan Thuy Ngan", "DIACHI_LD": "Can Tho"},
    ]
    calls = []

    def fake_run_sql_report(self, **kwargs):
        calls.append(kwargs)
        result_rows = rows
        search_value = str(kwargs.get("tham_so", {}).get("SEARCH_TERM_1", "")).strip("%").lower()
        if search_value:
            result_rows = [
                row for row in rows
                if search_value in " ".join(str(value).lower() for value in row.values())
            ]
        return {
            "ok": True,
            "columns": ["MA_TB", "TEN_TB", "DIACHI_LD"],
            "rows": result_rows,
            "total": len(result_rows),
            "page": kwargs["page"],
            "page_size": kwargs["page_size"],
            "message": "ok",
        }

    monkeypatch.setattr(routes.InternalApiClient, "run_sql_report", fake_run_sql_report)

    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Bao cao search export",
            "ma_bao_cao": "BC_SEARCH_EXPORT",
            "cau_lenh_sql": "SELECT ma_tb, ten_tb, diachi_ld FROM css_cto.db_thuebao;",
            "cac_tham_so": [],
        }
        assert client.post("/api/admin/sql-reports", json=payload).status_code == 200

        result = client.post(
            "/api/reports/run",
            json={
                "ma_bao_cao": "BC_SEARCH_EXPORT",
                "filters": {},
                "search": "phan thuy",
                "search_columns": ["MA_TB", "TEN_TB", "DIACHI_LD"],
                "page": 1,
                "page_size": 20,
            },
        )
        assert result.status_code == 200
        body = result.json()
        assert body["pagination"]["total"] == 1
        assert body["rows"][0]["MA_TB"] == "tb003"
        assert calls[-1]["page_size"] == 20
        assert "WHERE" in calls[-1]["cau_lenh_sql"]

        export = client.post(
            "/api/reports/export",
            json={
                "ma_bao_cao": "BC_SEARCH_EXPORT",
                "filters": {},
                "search": "can tho",
                "search_columns": ["MA_TB", "TEN_TB", "DIACHI_LD"],
                "page": 1,
                "page_size": 20,
            },
        )
        assert export.status_code == 200
        assert export.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = openpyxl.load_workbook(BytesIO(export.content))
        sheet = workbook.active
        assert [cell.value for cell in sheet[1]] == ["MA_TB", "TEN_TB", "DIACHI_LD"]
        assert sheet.max_row == 3
        assert {sheet.cell(row=index, column=1).value for index in range(2, sheet.max_row + 1)} == {"tb001", "tb003"}

        loaded_export = client.post(
            "/api/reports/export-loaded",
            json={
                "ma_bao_cao": "BC_SEARCH_EXPORT",
                "columns": ["MA_TB", "TEN_TB", "DIACHI_LD"],
                "rows": [rows[2]],
                "search": "phan thuy",
            },
        )
        assert loaded_export.status_code == 200
        loaded_workbook = openpyxl.load_workbook(BytesIO(loaded_export.content))
        loaded_sheet = loaded_workbook.active
        assert loaded_sheet.max_row == 2
        assert loaded_sheet["A2"].value == "tb003"


def test_dynamic_report_direct_excel_export_downloads_full_result_set(monkeypatch) -> None:
    rows = [
        {"MA_TB": f"tb{index:04d}", "TEN_TB": f"Thue bao {index:04d}"}
        for index in range(5205)
    ]
    calls = []

    def fake_run_sql_report(self, **kwargs):
        calls.append(kwargs)
        page = int(kwargs["page"])
        page_size = int(kwargs["page_size"])
        start = (page - 1) * page_size
        return {
            "ok": True,
            "columns": ["MA_TB", "TEN_TB"],
            "rows": rows[start:start + page_size],
            "total": len(rows),
            "page": page,
            "page_size": page_size,
            "message": "ok",
        }

    monkeypatch.setattr(routes.InternalApiClient, "run_sql_report", fake_run_sql_report)

    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Bao cao export job",
            "ma_bao_cao": "BC_EXPORT_JOB",
            "cau_lenh_sql": "SELECT ma_tb, ten_tb FROM css_cto.db_thuebao;",
            "cac_tham_so": [],
        }
        assert client.post("/api/admin/sql-reports", json=payload).status_code == 200

        export = client.post(
            "/api/reports/export",
            json={"ma_bao_cao": "BC_EXPORT_JOB", "filters": {}, "page": 1, "page_size": 20},
        )
        assert export.status_code == 200
        assert [call["page"] for call in calls] == [1]
        assert all(call["page_size"] == 20000 for call in calls)
        assert all(call.get("timeout", 0) >= 20 for call in calls)

        workbook = openpyxl.load_workbook(BytesIO(export.content), read_only=True)
        sheet = workbook.active
        exported_rows = list(sheet.iter_rows(values_only=True))
        assert len(exported_rows) == len(rows) + 1
        assert list(exported_rows[0]) == ["MA_TB", "TEN_TB"]
        assert exported_rows[-1][0] == "tb5204"


def test_dynamic_report_export_worker_queue_can_cancel_waiting_job(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings, storage_link="", repository=None: "")

    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()

    with TestClient(app) as client:
        login(client)
        assert client.post(
            "/api/admin/sql-reports",
            json={
                "ten_bao_cao": "Queue worker",
                "ma_bao_cao": "QUEUE_WORKER",
                "cau_lenh_sql": "SELECT ma_tb FROM css_cto.db_thuebao;",
                "cac_tham_so": [],
            },
        ).status_code == 200

        started = client.post("/api/reports/export-jobs", json={"ma_bao_cao": "QUEUE_WORKER", "filters": {}, "page": 1, "page_size": 20})
        assert started.status_code == 200
        body = started.json()
        job_id = body["job_id"]
        assert body["status"] == "queued_worker"
        assert body["worker_state"]["status"] in {"ready", "no_online_worker"}

        queue = client.get("/api/reports/export-jobs?limit=10")
        assert queue.status_code == 200
        queued_job = next(job for job in queue.json()["jobs"] if job["job_id"] == job_id)
        assert queued_job["status"] == "queued_worker"
        assert queued_job["queue_position"] >= 1
        assert queued_job["can_cancel"] is True

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-sql"}, headers=headers)
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["task_type"] == "dynamic_report_export"
        assert task["query"]["drive_folder_id"] == ""
        finished = client.post(
            f"/api/sql-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "May tram da xuat Excel va upload Google Drive.",
                "drive_url": "https://drive.google.com/file/d/queue-worker/view",
                "file_name": "queue_worker.xlsx",
                "total": 1,
            },
            headers=headers,
        )
        assert finished.status_code == 200

        started_again = client.post("/api/reports/export-jobs", json={"ma_bao_cao": "QUEUE_WORKER", "filters": {}, "page": 1, "page_size": 20})
        assert started_again.status_code == 200
        cancel_job_id = started_again.json()["job_id"]
        cancelled = client.delete(f"/api/reports/export-jobs/{cancel_job_id}")
        assert cancelled.status_code == 200
        cancelled_body = cancelled.json()
        assert cancelled_body["status"] == "cancelled"
        assert cancelled_body["can_cancel"] is False


def test_running_sql_export_cancel_is_not_overwritten_by_worker_updates(monkeypatch, tmp_path: Path) -> None:
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings, storage_link="", repository=None: "")
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()

    with TestClient(app) as client:
        login(client)
        report_code = f"CANCEL_SQL_{uuid.uuid4().hex[:8].upper()}"
        assert client.post(
            "/api/admin/sql-reports",
            json={"ten_bao_cao": "Cancel SQL", "ma_bao_cao": report_code, "cau_lenh_sql": "SELECT 1 FROM dual;", "cac_tham_so": []},
        ).status_code == 200
        job_id = client.post(
            "/api/reports/export-jobs",
            json={"ma_bao_cao": report_code, "filters": {}, "page": 1, "page_size": 20},
        ).json()["job_id"]
        headers = {"Authorization": "Bearer test-worker-token"}
        assert client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-sql"}, headers=headers).status_code == 200
        cancelled = client.delete(f"/api/reports/export-jobs/{job_id}")
        assert cancelled.json()["status"] == "cancel_requested"

        heartbeat = client.post(
            f"/api/sql-worker/tasks/{job_id}/status",
            json={"status": "running_worker", "message": "still running", "worker_id": "ws-sql"},
            headers=headers,
        )
        assert heartbeat.status_code == 200
        assert heartbeat.json()["cancelled"] is True
        assert heartbeat.json()["status"] == "cancelled"
        assert heartbeat.json()["run"]["status"] == "cancelled"

        late_result = client.post(
            f"/api/sql-worker/tasks/{job_id}/result",
            json={"ok": True, "status": "success", "drive_url": "https://drive.google.com/file/d/late/view"},
            headers=headers,
        )
        assert late_result.status_code == 200
        assert late_result.json()["cancelled"] is True
        assert late_result.json()["run"]["status"] == "cancelled"


def test_dynamic_report_export_job_can_return_drive_link(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings, storage_link="", repository=None: "drive-folder-001")
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()

    with TestClient(app) as client:
        login(client)
        assert client.post(
            "/api/admin/sql-reports",
            json={
                "ten_bao_cao": "CRS",
                "ma_bao_cao": "BC_DRIVE_EXPORT",
                "cau_lenh_sql": "SELECT ma_tb, ten_tb FROM css_cto.db_thuebao;",
                "cac_tham_so": [],
            },
        ).status_code == 200
        started = client.post(
            "/api/reports/export-jobs",
            json={"ma_bao_cao": "BC_DRIVE_EXPORT", "filters": {}, "page": 1, "page_size": 20},
        )
        assert started.status_code == 200
        started_body = started.json()
        job_id = started_body["job_id"]
        assert started_body["status"] == "queued_worker"
        assert started_body["updated_at"]
        assert started_body["worker_state"]["status"] in {"ready", "no_online_worker"}
        assert any("Da gui lenh lay du lieu" in step["message"] for step in started_body["progress_steps"])
        history = client.get("/api/reports/history?limit=5")
        assert history.status_code == 200
        history_item = next(item for item in history.json()["items"] if item["history_id"] == job_id)
        assert history_item["status"] == "queued_worker"

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-sql"}, headers=headers)
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["task_type"] == "dynamic_report_export"
        assert task["query"]["action"] == "export_sql_report_to_drive"
        assert task["query"]["drive_folder_id"] == "drive-folder-001"
        assert task["query"]["ma_bao_cao"] == "BC_DRIVE_EXPORT"
        progress = client.post(
            f"/api/sql-worker/tasks/{job_id}/status",
            json={
                "status": "running_worker",
                "message": "Dang ket noi Oracle noi bo, xuat Excel va upload Google Drive.",
                "worker_id": "ws-sql",
                "details": {"step": "oracle_export_drive"},
            },
            headers=headers,
        )
        assert progress.status_code == 200
        assert progress.json()["run"]["worker_id"] == "ws-sql"
        assert progress.json()["run"]["updated_at"]
        assert any("Dang ket noi Oracle" in step["message"] for step in progress.json()["run"]["progress_steps"])

        finished = client.post(
            f"/api/sql-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "May tram da xuat Excel va upload Google Drive.",
                "columns": ["MA_TB", "TEN_TB"],
                "rows": [],
                "pagination": {"page": 1, "page_size": 5000, "total": 145433},
                "report": task["report"],
                "drive_url": "https://drive.google.com/file/d/export-file/view",
                "file_name": "crs_export.xlsx",
                "file_id": "export-file",
                "total": 145433,
                "details": {"source": "workstation"},
            },
            headers=headers,
        )
        assert finished.status_code == 200

        status_body = client.get(f"/api/reports/export-jobs/{job_id}").json()
        assert status_body["status"] == "complete"
        assert status_body["drive_url"] == "https://drive.google.com/file/d/export-file/view"
        assert status_body["download_url"] == status_body["drive_url"]
        assert status_body["rows"] == 145433
        messages = [step["message"] for step in status_body["progress_steps"]]
        assert any("May tram ws-sql da nhan lenh lay du lieu SQL" in message for message in messages)
        assert any("May tram da xuat Excel va upload Google Drive." in message for message in messages)


def test_dynamic_report_export_job_status_recovers_from_persisted_metadata(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings, storage_link="", repository=None: "drive-folder-001")
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()

    with TestClient(app) as client:
        login(client)
        assert client.post(
            "/api/admin/sql-reports",
            json={
                "ten_bao_cao": "CRS",
                "ma_bao_cao": "BC_DRIVE_RECOVERED",
                "cau_lenh_sql": "SELECT ma_tb FROM css_cto.db_thuebao;",
                "cac_tham_so": [],
            },
        ).status_code == 200
        started = client.post(
            "/api/reports/export-jobs",
            json={"ma_bao_cao": "BC_DRIVE_RECOVERED", "filters": {}, "page": 1, "page_size": 20},
        )
        assert started.status_code == 200
        job_id = started.json()["job_id"]
        assert started.json()["status"] == "queued_worker"

        with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
            routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()

        recovered = client.get(f"/api/reports/export-jobs/{job_id}")
        assert recovered.status_code == 200
        recovered_body = recovered.json()
        assert recovered_body["status"] == "queued_worker"
        assert any("Da gui lenh lay du lieu" in step["message"] for step in recovered_body["progress_steps"])

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-recovered"}, headers=headers)
        assert claim.status_code == 200
        assert claim.json()["task"]["run_id"] == job_id
        finished = client.post(
            f"/api/sql-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "May tram da xuat Excel va upload Google Drive.",
                "drive_url": "https://drive.google.com/file/d/recovered-export/view",
                "file_name": "crs_export.xlsx",
                "total": 12,
            },
            headers=headers,
        )
        assert finished.status_code == 200

        recovered_body = client.get(f"/api/reports/export-jobs/{job_id}").json()
        assert recovered_body["status"] == "complete"
        assert recovered_body["drive_url"] == "https://drive.google.com/file/d/recovered-export/view"
        assert recovered_body["download_url"] == recovered_body["drive_url"]


def test_dynamic_report_export_job_recovers_from_audit_history(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings, storage_link="", repository=None: "drive-folder-001")
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()

    with TestClient(app) as client:
        login(client)
        assert client.post(
            "/api/admin/sql-reports",
            json={
                "ten_bao_cao": "CRS audit recovery",
                "ma_bao_cao": "BC_DRIVE_AUDIT_RECOVERY",
                "cau_lenh_sql": "SELECT ma_tb FROM css_cto.db_thuebao WHERE trang_thai = :STATUS;",
                "cac_tham_so": ["STATUS"],
            },
        ).status_code == 200
        started = client.post(
            "/api/reports/export-jobs",
            json={"ma_bao_cao": "BC_DRIVE_AUDIT_RECOVERY", "filters": {"STATUS": "1"}, "page": 1, "page_size": 20},
        )
        assert started.status_code == 200
        job_id = started.json()["job_id"]
        assert started.json()["status"] == "queued_worker"

        with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
            routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
        routes._dynamic_report_export_job_path(job_id).unlink(missing_ok=True)

        recovered = client.get(f"/api/reports/export-jobs/{job_id}")
        assert recovered.status_code == 200
        assert recovered.json()["status"] == "queued_worker"
        assert any("Da gui lenh lay du lieu" in step["message"] for step in recovered.json()["progress_steps"])

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-audit"}, headers=headers)
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["task_type"] == "dynamic_report_export"
        assert task["query"]["drive_folder_id"] == "drive-folder-001"
        assert task["query"]["tham_so"] == {"STATUS": "1"}


def test_sql_worker_claim_recovers_export_job_from_audit_history(monkeypatch, tmp_path) -> None:
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings, storage_link="", repository=None: "drive-folder-001")
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()

    with TestClient(app) as client:
        login(client)
        assert client.post(
            "/api/admin/sql-reports",
            json={
                "ten_bao_cao": "CRS claim recovery",
                "ma_bao_cao": "BC_DRIVE_CLAIM_RECOVERY",
                "cau_lenh_sql": "SELECT ma_tb FROM css_cto.db_thuebao WHERE trang_thai = :STATUS;",
                "cac_tham_so": ["STATUS"],
            },
        ).status_code == 200
        started = client.post(
            "/api/reports/export-jobs",
            json={"ma_bao_cao": "BC_DRIVE_CLAIM_RECOVERY", "filters": {"STATUS": "1"}, "page": 1, "page_size": 20},
        )
        assert started.status_code == 200
        job_id = started.json()["job_id"]
        assert started.json()["status"] == "queued_worker"

        with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
            routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()
        routes._dynamic_report_export_job_path(job_id).unlink(missing_ok=True)
        with routes.DYNAMIC_REPORT_EXPORT_HISTORY_RECOVERY_LOCK:
            routes.DYNAMIC_REPORT_EXPORT_HISTORY_RECOVERY_LAST_TS = 0.0

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-direct-claim"}, headers=headers)
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["task_type"] == "dynamic_report_export"
        assert task["query"]["drive_folder_id"] == "drive-folder-001"
        assert task["query"]["tham_so"] == {"STATUS": "1"}


def test_dynamic_report_drive_export_sends_compiled_sql_to_internal_api() -> None:
    captured = {}

    class FakeInternalApi:
        settings = get_settings()

        def export_sql_report_to_drive(self, **kwargs):
            captured.update(kwargs)
            return {"ok": True, "drive_url": "https://drive.google.com/file/d/sql-export/view", "rows": 2, "total": 2}

    class FakeRepository:
        def get_sql_report_by_code(self, code):
            return {
                "ten_bao_cao": "CRS",
                "ma_bao_cao": "CRS",
                "cau_lenh_sql": "SELECT ma_tb, ten_tb FROM css_cto.db_thuebao WHERE trang_thai = :STATUS;",
                "cac_tham_so": ["STATUS"],
            }

        def get_sql_report_by_id(self, report_id):
            return None

        def list_sql_reports(self):
            return []

    service = DatabaseService(FakeInternalApi(), FakeRepository())
    result = service.export_dynamic_report_to_drive(
        ma_bao_cao="CRS",
        filters={"STATUS": "1", "IGNORED": "x"},
        search="nguyen",
        search_columns=["TEN_TB"],
        drive_folder_id="drive-folder-001",
        file_name="crs.xlsx",
    )

    assert result["ok"] is True
    assert captured["drive_folder_id"] == "drive-folder-001"
    assert captured["file_name"] == "crs.xlsx"
    assert captured["page_size"] == 20000
    assert captured["max_rows"] >= 1000000
    assert captured["tham_so"]["STATUS"] == "1"
    assert "SEARCH_TERM_1" in captured["tham_so"]
    assert "SELECT * FROM (" in captured["cau_lenh_sql"]
    assert result["ignored_filters"] == ["IGNORED"]


def test_sql_worker_claim_formats_oracle_date_binds_for_report_mask(monkeypatch, tmp_path) -> None:
    import httpx

    def raise_530(self, **kwargs):
        request = httpx.Request("POST", "https://api.vnptcto.com/api/du-lieu-web")
        response = httpx.Response(530, request=request, text="cloudflare tunnel error")
        raise httpx.HTTPStatusError("cloudflare tunnel error", request=request, response=response)

    monkeypatch.setattr(routes.InternalApiClient, "run_sql_report", raise_530)
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(routes, "DYNAMIC_REPORT_EXPORT_JOB_DIR", tmp_path / "exports" / "jobs")
    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()
    with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()

    try:
        with TestClient(app) as client:
            login(client)
            assert client.post(
                "/api/admin/sql-reports",
                json={
                    "ten_bao_cao": "Doanh thu CNTT theo OneBSS",
                    "ma_bao_cao": "BC_ORACLE_DATE_BIND",
                    "cau_lenh_sql": (
                        "SELECT ma_tb FROM css_cto.db_thuebao "
                        "WHERE ngay >= TO_DATE(:P_TUNGAY, 'YYYYMMDD') "
                        "AND ngay <= TO_DATE(:P_DENNGAY, 'YYYYMMDD');"
                    ),
                    "cac_tham_so": ["P_TUNGAY", "P_DENNGAY"],
                },
                ).status_code == 200

            started = client.post(
                "/api/reports/run-jobs",
                json={
                    "ma_bao_cao": "BC_ORACLE_DATE_BIND",
                    "filters": {"P_TUNGAY": "01/07/2026", "P_DENNGAY": "30/07/2026"},
                    "page": 1,
                    "page_size": 20,
                },
            )
            assert started.status_code == 200
            job_id = started.json()["job_id"]
            job = {}
            for _ in range(20):
                job = client.get(f"/api/reports/run-jobs/{job_id}").json()
                if job["status"] == "queued_worker":
                    break
                time.sleep(0.05)
            assert job["status"] == "queued_worker"
            with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
                target_job = routes.DYNAMIC_REPORT_RUN_JOBS.get(job_id)
                routes.DYNAMIC_REPORT_RUN_JOBS.clear()
                if target_job:
                    routes.DYNAMIC_REPORT_RUN_JOBS[job_id] = target_job
            with routes.DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
                routes.DYNAMIC_REPORT_EXPORT_JOBS.clear()

            headers = {"Authorization": "Bearer test-worker-token"}
            claim = client.post("/api/sql-worker/tasks/claim", json={"worker_id": "ws-date"}, headers=headers)
            assert claim.status_code == 200
            task = claim.json()["task"]
            assert task["query"]["tham_so"].get("P_TUNGAY") == "20260701", task
            assert task["query"]["tham_so"].get("P_DENNGAY") == "20260730", task
    finally:
        with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
            routes.DYNAMIC_REPORT_RUN_JOBS.clear()


def test_dynamic_report_define_date_uses_oracle_mask() -> None:
    class FakeInternalApi:
        settings = get_settings()

    class FakeRepository:
        def get_sql_report_by_code(self, code):
            return {
                "ten_bao_cao": "Date define",
                "ma_bao_cao": "DATE_DEFINE",
                "cau_lenh_sql": (
                    "DEFINE FROM_DATE = :P_TUNGAY\n"
                    "SELECT * FROM dual WHERE SYSDATE >= TO_DATE('&FROM_DATE', 'YYYYMMDD');"
                ),
                "cac_tham_so": ["P_TUNGAY"],
            }

        def get_sql_report_by_id(self, report_id):
            return None

        def list_sql_reports(self):
            return []

    service = DatabaseService(FakeInternalApi(), FakeRepository())
    prepared = service.prepare_dynamic_report_query(
        ma_bao_cao="DATE_DEFINE",
        filters={"P_TUNGAY": "01/07/2026"},
        page=1,
        page_size=20,
    )

    assert prepared["ok"] is True
    assert "TO_DATE('20260701', 'YYYYMMDD')" in prepared["cau_lenh_sql"]
    assert "01/07/2026" not in prepared["cau_lenh_sql"]


def test_dynamic_report_month_define_does_not_duplicate_literal_day() -> None:
    class FakeInternalApi:
        settings = get_settings()

    class FakeRepository:
        def get_sql_report_by_code(self, code):
            return {
                "ten_bao_cao": "Month define",
                "ma_bao_cao": "MONTH_DEFINE",
                "cau_lenh_sql": (
                    "DEFINE thang = :THANG\n"
                    "SELECT * FROM dual WHERE ngay >= TO_DATE('01/&thang', 'DD/MM/YYYY');"
                ),
                "cac_tham_so": ["THANG"],
            }

        def get_sql_report_by_id(self, report_id):
            return None

        def list_sql_reports(self):
            return []

    service = DatabaseService(FakeInternalApi(), FakeRepository())
    prepared = service.prepare_dynamic_report_query(
        ma_bao_cao="MONTH_DEFINE",
        filters={"THANG": "2026-08"},
        page=1,
        page_size=20,
    )

    assert prepared["ok"] is True
    assert "TO_DATE('01/08/2026', 'DD/MM/YYYY')" in prepared["cau_lenh_sql"]
    assert "01/01/08/2026" not in prepared["cau_lenh_sql"]


def test_dynamic_report_define_date_accepts_bare_filter_name_and_time_suffix() -> None:
    class FakeInternalApi:
        settings = get_settings()

    class FakeRepository:
        def get_sql_report_by_code(self, code):
            return {
                "ten_bao_cao": "Bare date define",
                "ma_bao_cao": "BARE_DATE_DEFINE",
                "cau_lenh_sql": (
                    "DEFINE p_tungay = :TUNGAY\n"
                    "DEFINE p_denngay = DENNGAY\n"
                    "SELECT * FROM dual "
                    "WHERE ngay >= TO_DATE('&p_tungay' || ' 00:00:00', 'DD/MM/YYYY HH24:MI:SS') "
                    "AND ngay < TO_DATE('&p_denngay' || ' 00:00:00', 'DD/MM/YYYY HH24:MI:SS');"
                ),
                "cac_tham_so": ["TUNGAY", "DENNGAY"],
            }

        def get_sql_report_by_id(self, report_id):
            return None

        def list_sql_reports(self):
            return []

    service = DatabaseService(FakeInternalApi(), FakeRepository())
    prepared = service.prepare_dynamic_report_query(
        ma_bao_cao="BARE_DATE_DEFINE",
        filters={"TUNGAY": "2026-08-01", "DENNGAY": "2026-08-31"},
        page=1,
        page_size=20,
    )

    assert prepared["ok"] is True
    assert "&p_tungay" not in prepared["cau_lenh_sql"].lower()
    assert "&p_denngay" not in prepared["cau_lenh_sql"].lower()
    assert "DENNGAY" not in prepared["cau_lenh_sql"]
    assert "'01/08/2026' || ' 00:00:00'" in prepared["cau_lenh_sql"]
    assert "'31/08/2026' || ' 00:00:00'" in prepared["cau_lenh_sql"]


def test_dynamic_report_prepare_keeps_large_worker_page_size() -> None:
    class FakeInternalApi:
        settings = get_settings()

    class FakeRepository:
        def get_sql_report_by_code(self, code):
            return {
                "ten_bao_cao": "Large page",
                "ma_bao_cao": "LARGE_PAGE",
                "cau_lenh_sql": "SELECT :P_THANG AS P_THANG FROM dual;",
                "cac_tham_so": ["P_THANG"],
            }

        def get_sql_report_by_id(self, report_id):
            return None

        def list_sql_reports(self):
            return []

    service = DatabaseService(FakeInternalApi(), FakeRepository())
    prepared = service.prepare_dynamic_report_query(
        ma_bao_cao="LARGE_PAGE",
        filters={"P_THANG": "202608"},
        page=1,
        page_size=20000,
    )
    capped = service.prepare_dynamic_report_query(
        ma_bao_cao="LARGE_PAGE",
        filters={"P_THANG": "202608"},
        page=1,
        page_size=50000,
    )

    assert prepared["ok"] is True
    assert prepared["page_size"] == 20000
    assert capped["page_size"] == 20000


def test_admin_can_manage_and_run_onebss_report(monkeypatch) -> None:
    monkeypatch.setattr(
        routes,
        "start_onebss_otp_mobile_gateway_request",
        lambda *args, **kwargs: {
            "ok": False,
            "status": "otp_required",
            "message": "May tram dang doi OTP.",
            "otp_request_id": "OTP-WORKER-001",
        },
    )
    monkeypatch.setattr(routes, "match_onebss_mobile_gateway_manual_otp", lambda *args, **kwargs: {"ok": True, "status": "matched"})
    monkeypatch.setattr(routes, "consume_onebss_mobile_gateway_otp", lambda *args, **kwargs: {"ok": True, "status": "matched", "otp": "123456"})
    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Bien dong PTTB",
            "danh_sach_bien": ["P_TUNGAY", "P_DENNGAY"],
            "parameters": {"P_TUNGAY": "{{month_start}}", "P_DENNGAY": "{{today}}"},
            "otp_service_code": "otp_onebss",
            "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test",
            "storage_link": "https://drive.google.com/drive/folders/test-folder",
        }
        created = client.post("/api/admin/onebss-reports", json=payload)
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]
        assert code.startswith("ONEBSS")

        configs = client.get("/api/onebss-reports/configs")
        assert configs.status_code == 200
        report = next(item for item in configs.json()["reports"] if item["ma_bao_cao"] == code)
        assert report["danh_sach_bien"] == ["P_TUNGAY", "P_DENNGAY"]
        assert report["parameters"] == {"P_TUNGAY": "{{month_start}}", "P_DENNGAY": "{{today}}"}
        assert report["otp_service_code"] == "otp_onebss"
        assert report["storage_link"] == ""

        first_run = client.post(
            "/api/onebss-reports/run",
            json={"ma_bao_cao": code},
        )
        assert first_run.status_code == 200
        assert first_run.json()["status"] == "queued"
        job_id = first_run.json()["job_id"]

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": "ws-01"}, headers=headers)
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["parameters"] == payload["parameters"]
        assert task["storage_link"] == ""

        waiting_otp = client.post(
            f"/api/onebss-worker/tasks/{job_id}/status",
            json={"status": "otp_required", "message": "Can OTP", "worker_id": "ws-01", "worker_session_id": "worker-session-001"},
            headers=headers,
        )
        assert waiting_otp.status_code == 200
        assert waiting_otp.json()["otp_request_id"] == "OTP-WORKER-001"

        otp_submit = client.post(
            f"/api/onebss-reports/jobs/{job_id}/otp",
            json={"otp": "123456", "otp_request_id": "OTP-WORKER-001", "otp_source": "manual"},
        )
        assert otp_submit.status_code == 200
        assert otp_submit.json()["ok"] is True

        worker_otp = client.get(f"/api/onebss-worker/tasks/{job_id}/otp", headers=headers)
        assert worker_otp.status_code == 200
        assert worker_otp.json()["otp"] == "123456"

        finished = client.post(
            f"/api/onebss-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "Da tai bao cao OneBSS va upload Google Drive.",
                "file_name": "onebss.xlsx",
                "storage_link": "https://drive.google.com/file/d/onebss-file/view",
                "storage_status": "uploaded_google_drive:onebss-file",
                "duration_ms": 1234,
            },
            headers=headers,
        )
        assert finished.status_code == 200
        assert finished.json()["run"]["status"] == "success"

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["storage_link"] == "https://drive.google.com/file/d/onebss-file/view"

        cleared = client.delete(f"/api/onebss-reports/runs?ma_bao_cao={code}")
        assert cleared.status_code == 200
        assert cleared.json()["deleted"] == 1
        runs_after_clear = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert runs_after_clear == []
        post_clear = client.post(f"/api/onebss-reports/runs/clear?ma_bao_cao={code}")
        assert post_clear.status_code == 200
        assert post_clear.json()["deleted"] == 0


def test_onebss_worker_get_otp_creates_request_when_status_update_timed_out(monkeypatch) -> None:
    started_requests: list[dict[str, str]] = []
    consumed_requests: list[str] = []

    def fake_start_otp(settings, session_label, parameters, **kwargs):
        started_requests.append(
            {
                "session_label": session_label,
                "otp_service_code": kwargs.get("otp_service_code") or "",
                "report_url": kwargs.get("report_url") or "",
            }
        )
        return {
            "ok": False,
            "status": "otp_required",
            "message": "May tram dang doi OTP.",
            "otp_request_id": "OTP-AUTO-GET-001",
        }

    def fake_consume_otp(settings, otp_request_id):
        consumed_requests.append(otp_request_id)
        return {"ok": True, "status": "matched", "otp": "654321"}

    monkeypatch.setattr(routes, "start_onebss_otp_mobile_gateway_request", fake_start_otp)
    monkeypatch.setattr(routes, "consume_onebss_mobile_gateway_otp", fake_consume_otp)

    with TestClient(app) as client:
        login(client)
        client.delete("/api/onebss-reports/runs")
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS OTP auto ensure",
                "danh_sach_bien": ["P_TUNGAY"],
                "parameters": {"P_TUNGAY": "01/07/2026"},
                "otp_service_code": "otp_onebss_auto",
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_OTP_AUTO&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        queued = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code})
        assert queued.status_code == 200
        job_id = queued.json()["job_id"]

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": "ws-otp-auto"}, headers=headers)
        assert claim.status_code == 200
        assert claim.json()["task"]["run_id"] == job_id

        worker_otp = client.get(f"/api/onebss-worker/tasks/{job_id}/otp", headers=headers)
        assert worker_otp.status_code == 200
        assert worker_otp.json()["status"] == "matched"
        assert worker_otp.json()["otp"] == "654321"
        assert started_requests == [
            {
                "session_label": job_id,
                "otp_service_code": "otp_onebss_auto",
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_OTP_AUTO&name=Test",
            }
        ]
        assert consumed_requests == ["OTP-AUTO-GET-001"]

        job = client.get(f"/api/onebss-reports/jobs/{job_id}")
        assert job.status_code == 200
        assert job.json()["otp_request_id"] == "OTP-AUTO-GET-001"


def test_admin_can_manage_and_run_ftp_report(monkeypatch) -> None:
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings_arg, storage_link="", repository=None: "")
    with TestClient(app) as client:
        login(client)
        ftp_connection = client.put(
            "/api/admin/connections/ftp_storage",
            json={
                "name": "FTP nội bộ",
                "connection_type": "ftp",
                "description": "Ket noi FTP noi bo",
                "config": {
                    "host": "10.159.23.100",
                    "port": 21,
                    "username": "thangph.cto",
                    "password": "$Phthang125125",
                    "passive": True,
                    "timeout_seconds": 60,
                    "secret_ref": "FTP_PASSWORD",
                },
                "is_active": True,
            },
        )
        assert ftp_connection.status_code == 200
        connections = client.get("/api/admin/connections").json()["connections"]
        ftp_public = next(item for item in connections if item["code"] == "ftp_storage")
        assert ftp_public["config"]["host"] == "10.159.23.100"
        assert ftp_public["config"]["username"] == "thangph.cto"
        assert "password" not in ftp_public["config"]
        assert "password" in ftp_public["protected_config_keys"]

        payload = {
            "ten_bao_cao": "FTP doanh thu",
            "folder_path": "/bao_cao/doanh_thu",
            "file_name_template": "doanh_thu_{yyyymmdd}.xlsx",
            "connection_code": "ftp_storage",
            "is_active": True,
        }
        created = client.post("/api/admin/ftp-reports", json=payload)
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]
        assert code.startswith("FTP")

        configs = client.get("/api/ftp-reports/configs")
        assert configs.status_code == 200
        report = next(item for item in configs.json()["reports"] if item["ma_bao_cao"] == code)
        assert report["folder_path"] == payload["folder_path"]
        assert report["file_name_template"] == payload["file_name_template"]

        run = client.post(
            "/api/ftp-reports/run",
            json={
                "ma_bao_cao": code,
                "folder_path": "/bao_cao/override",
                "file_name_template": "override_{ddmmyyyy}.xlsx",
            },
        )
        assert run.status_code == 200
        assert run.json()["status"] == "queued"
        job_id = run.json()["job_id"]

        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/ftp-worker/tasks/claim", json={"worker_id": "ws-ftp"}, headers=headers)
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["folder_path"] == "/bao_cao/override"
        assert task["file_name_template"] == "override_{ddmmyyyy}.xlsx"
        assert task["connection"]["config"]["password"] == "$Phthang125125"

        progress = client.post(
            f"/api/ftp-worker/tasks/{job_id}/status",
            json={"status": "running", "message": "Dang tai FTP", "worker_id": "ws-ftp", "resolved_file_name": "override_28072026.xlsx"},
            headers=headers,
        )
        assert progress.status_code == 200
        assert progress.json()["run"]["resolved_file_name"] == "override_28072026.xlsx"

        uploaded = client.post(
            f"/api/ftp-worker/tasks/{job_id}/file",
            files={"file": ("override_28072026.xlsx", BytesIO(b"ftp-data"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )
        assert uploaded.status_code == 200
        assert uploaded.json()["run"]["download_url"].endswith(f"/api/ftp-reports/runs/{job_id}/download")

        finished = client.post(
            f"/api/ftp-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "Da tai file FTP.",
                "file_name": "override_28072026.xlsx",
                "resolved_file_name": "override_28072026.xlsx",
                "duration_ms": 456,
            },
            headers=headers,
        )
        assert finished.status_code == 200
        assert finished.json()["run"]["status"] == "success"
        assert finished.json()["run"]["storage_status"] == "uploaded_worker_file"

        runs = client.get(f"/api/ftp-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["download_url"].endswith(f"/api/ftp-reports/runs/{job_id}/download")
        download = client.get(runs[0]["download_url"])
        assert download.status_code == 200
        assert download.content == b"ftp-data"

        cleared = client.delete(f"/api/ftp-reports/runs?ma_bao_cao={code}")
        assert cleared.status_code == 200
        assert cleared.json()["deleted"] == 1


def test_ftp_worker_web_upload_uses_global_drive_folder(monkeypatch, tmp_path) -> None:
    settings = get_settings().model_copy(
        update={
            "data_mining_download_dir": str(tmp_path),
            "google_drive_folder_id": "drive-folder-ftp",
        }
    )
    saved_calls = []

    def fake_save_downloaded_file(settings_arg, source_file, storage_link, repository=None):
        saved_calls.append((str(source_file), storage_link))
        return {
            "ok": True,
            "message": "Da upload FTP vao Drive.",
            "storage_link": "https://drive.google.com/file/d/ftp-drive-file/view",
            "storage_status": "uploaded_google_drive:ftp-drive-file",
        }

    monkeypatch.setattr(routes, "get_settings", lambda: settings)
    monkeypatch.setattr(routes, "save_downloaded_file", fake_save_downloaded_file)
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/ftp-reports",
            json={
                "ma_bao_cao": "FTP_DRIVE_WEB",
                "ten_bao_cao": "FTP Drive web upload",
                "folder_path": "/bao_cao/drive",
                "file_name_template": "drive_{yyyyMM}.xlsx",
                "connection_code": "ftp_storage",
                "is_active": True,
            },
        )
        assert created.status_code == 200

        response = client.post("/api/ftp-reports/run", json={"ma_bao_cao": "FTP_DRIVE_WEB"})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        uploaded = client.post(
            f"/api/ftp-worker/tasks/{job_id}/file",
            files={"file": ("ftp_result.xlsx", b"xlsx-bytes", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"Authorization": "Bearer test-worker-token"},
        )

        assert uploaded.status_code == 200
        file_payload = uploaded.json()["file"]
        assert file_payload["storage_link"] == "https://drive.google.com/file/d/ftp-drive-file/view"
        assert file_payload["storage_status"] == "uploaded_google_drive:ftp-drive-file"
        run_payload = uploaded.json()["run"]
        assert run_payload["file_url"] == "https://drive.google.com/file/d/ftp-drive-file/view"
        assert "download_url" not in run_payload
        assert saved_calls and saved_calls[0][1] == "drive-folder-ftp"
        cleared = client.delete("/api/ftp-reports/runs?ma_bao_cao=FTP_DRIVE_WEB")
        assert cleared.status_code == 200


def test_ftp_worker_claim_includes_drive_folder(monkeypatch) -> None:
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings_arg, storage_link="", repository=None: "drive-folder-ftp")
    with TestClient(app) as client:
        login(client)
        ftp_connection = client.put(
            "/api/admin/connections/ftp_storage",
            json={
                "name": "FTP noi bo",
                "connection_type": "ftp",
                "description": "Ket noi FTP noi bo",
                "config": {
                    "host": "10.159.23.100",
                    "port": 21,
                    "username": "thangph.cto",
                    "password": "$Phthang125125",
                    "passive": True,
                    "timeout_seconds": 60,
                    "secret_ref": "FTP_PASSWORD",
                },
                "is_active": True,
            },
        )
        assert ftp_connection.status_code == 200
        created = client.post(
            "/api/admin/ftp-reports",
            json={
                "ma_bao_cao": "FTP_DRIVE_CLAIM",
                "ten_bao_cao": "FTP Drive claim",
                "folder_path": "/bao_cao/drive",
                "file_name_template": "drive_{yyyyMM}.xlsx",
                "connection_code": "ftp_storage",
                "is_active": True,
            },
        )
        assert created.status_code == 200

        response = client.post("/api/ftp-reports/run", json={"ma_bao_cao": "FTP_DRIVE_CLAIM"})
        assert response.status_code == 200
        claim = client.post(
            "/api/ftp-worker/tasks/claim",
            json={"worker_id": "ws-ftp-drive"},
            headers={"Authorization": "Bearer test-worker-token"},
        )

        assert claim.status_code == 200
        assert claim.json()["task"]["drive_folder_id"] == "drive-folder-ftp"
        cleared = client.delete("/api/ftp-reports/runs?ma_bao_cao=FTP_DRIVE_CLAIM")
        assert cleared.status_code == 200


def test_ftp_report_run_accepts_variables_and_multi_source_config() -> None:
    with TestClient(app) as client:
        login(client)
        ftp_connection = client.put(
            "/api/admin/connections/ftp_storage",
            json={
                "name": "FTP noi bo",
                "connection_type": "ftp",
                "description": "Ket noi FTP noi bo",
                "config": {
                    "host": "10.159.23.100",
                    "port": 21,
                    "username": "thangph.cto",
                    "password": "$Phthang125125",
                    "passive": True,
                    "timeout_seconds": 60,
                    "secret_ref": "FTP_PASSWORD",
                },
                "is_active": True,
            },
        )
        assert ftp_connection.status_code == 200
        advanced_template = json.dumps({
            "version": 1,
            "variables": {"thang": "{yyyyMM}"},
            "output_file_name_template": "FiberPTM_{thang}.xlsx",
            "file_name_template": "FiberPTM_{thang}.xlsx",
            "sources": [
                {
                    "name": "CTO",
                    "folder_path": "/DATA_BILLING/CTO/FiberPTM/{thang}",
                    "file_name_template": "CTO_Fiber_PTM_LK_ngay_{last_dd}.xlsx",
                },
                {
                    "name": "HAG",
                    "folder_path": "/DATA_BILLING/HGA/FiberPTM/{thang}",
                    "file_name_template": "HGA_Fiber_PTM_LK_ngay_{last_dd}.xlsx",
                },
                {
                    "name": "STG",
                    "folder_path": "/DATA_BILLING/STG/FiberPTM/{thang}",
                    "file_name_template": "STG_Fiber_PTM_LK_ngay_{last_dd}.xlsx",
                },
            ],
        })
        created = client.post(
            "/api/admin/ftp-reports",
            json={
                "ma_bao_cao": "FTP_PTM_MULTI",
                "ten_bao_cao": "FTP PTM multi",
                "folder_path": "",
                "file_name_template": advanced_template,
                "connection_code": "ftp_storage",
                "is_active": True,
            },
        )
        assert created.status_code == 200

        run = client.post(
            "/api/ftp-reports/run",
            json={
                "ma_bao_cao": "FTP_PTM_MULTI",
                "variables": {"thang": "{202607}"},
            },
        )
        assert run.status_code == 200
        job_id = run.json()["job_id"]

        claim = client.post(
            "/api/ftp-worker/tasks/claim",
            json={"worker_id": "ws-ftp-multi"},
            headers={"Authorization": "Bearer test-worker-token"},
        )
        assert claim.status_code == 200
        task = claim.json()["task"]
        assert task["run_id"] == job_id
        assert task["folder_path"] == "/DATA_BILLING/CTO/FiberPTM/{thang}"
        task_template = json.loads(task["file_name_template"])
        assert task_template["variables"]["thang"] == "202607"
        assert len(task_template["sources"]) == 3
        assert task_template["sources"][0]["folder_path"] == "/DATA_BILLING/CTO/FiberPTM/{thang}"
        assert task_template["sources"][1]["name"] == "HAG"
        assert task_template["sources"][1]["folder_path"] == "/DATA_BILLING/HAG/FiberPTM/{thang}"
        assert task_template["sources"][1]["file_name_template"] == "HAG_Fiber_PTM_LK_ngay_{last_dd}.xlsx"


def test_save_onebss_report_ignores_audit_log_failure(monkeypatch) -> None:
    class FakeRepository:
        def get_user_by_id(self, user_id):
            return {"id": user_id, "username": "admin", "full_name": "Admin", "role": "admin", "is_active": True}

        def get_user_permissions(self, user_id):
            return []

        def generate_onebss_report_code(self):
            return "ONEBSS9999"

        def save_onebss_report(self, *args, **kwargs):
            self.saved_args = args
            return 123

        def add_audit_log(self, *args, **kwargs):
            raise RuntimeError("audit_logs unavailable")

    repository = FakeRepository()

    with TestClient(app) as client:
        login(client)
        monkeypatch.setattr(routes, "build_app_repository", lambda: repository)
        response = client.post(
            "/api/admin/onebss-reports",
            json={
                "ma_bao_cao": "MYTV_KTT",
                "ten_bao_cao": "DS MyTV không tương tác",
                "danh_sach_bien": ["p_phanvung_id", "p_nhanvienkd_id"],
                "parameters": {
                    "p_phanvung_id": {"$each": ["13", "47", "66"]},
                    "p_nhanvienkd_id": "0",
                    "$merge_excel": {"sheet": "DATA", "source_column": "p_phanvung_id"},
                },
                "otp_service_code": "onebss",
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=KHAC%2FBRCD%2FRP_BSS_107195&name=Test",
                "storage_link": "https://drive.google.com/drive/folders/test-folder",
            },
        )

    assert response.status_code == 200
    assert response.json()["id"] == 123
    assert repository.saved_args[1] == "MYTV_KTT"
    assert repository.saved_args[4]["p_phanvung_id"]["$each"] == ["13", "47", "66"]


def test_onebss_login_deviceid_screen_requests_otp() -> None:
    from app.application.onebss_report_service import (
        handle_onebss_otp_request,
        pop_onebss_session,
        close_browser_stack,
    )

    class FakeBodyLocator:
        def __init__(self, page):
            self.page = page

        def inner_text(self, timeout=0):
            return self.page.body_text

    class FakePage:
        url = "https://onebss.vnpt.vn/#/auth/login?username=quyennt.cto&deviceId=12345"

        def __init__(self):
            self.body_text = "Xac nhan gui yeu cau"
            self.waits = 0

        def locator(self, selector):
            assert selector == "body"
            return FakeBodyLocator(self)

        def wait_for_load_state(self, *args, **kwargs):
            self.waits += 1

        def wait_for_timeout(self, *args, **kwargs):
            self.waits += 1

    class FakeHelper:
        def __init__(self):
            self.clicks = 0

        def _click_button_text(self, page, texts):
            self.clicks += 1
            page.body_text = "Nhap ma OTP"
            return True

    class FakeClosable:
        def __init__(self):
            self.closed = False

        def close(self):
            self.closed = True

    class FakePlaywright:
        def __init__(self):
            self.stopped = False

        def stop(self):
            self.stopped = True

    page = FakePage()
    helper = FakeHelper()
    browser = FakeClosable()
    context = FakeClosable()
    playwright = FakePlaywright()

    result = handle_onebss_otp_request(
        page,
        helper,
        playwright,
        browser,
        context,
        {"ma_bao_cao": "TEST"},
        {"P_DENNGAY": "{{today}}"},
        "admin",
    )

    assert result is not None
    assert result["status"] == "otp_required"
    assert result["session_id"]
    assert helper.clicks == 1
    pending = pop_onebss_session(result["session_id"])
    assert pending is not None
    close_browser_stack(pending.browser, pending.context, pending.playwright)
    assert browser.closed is True
    assert context.closed is True
    assert playwright.stopped is True


def test_onebss_pending_browser_session_keeps_worker_state_path(monkeypatch, tmp_path) -> None:
    from app.application import onebss_report_service as service

    class FakeBodyLocator:
        def __init__(self, page):
            self.page = page

        def inner_text(self, timeout=0):
            return self.page.body_text

    class FakePage:
        url = "https://onebss.vnpt.vn/#/auth/login?username=quyennt.cto&deviceId=12345"

        def __init__(self):
            self.body_text = "Xac nhan gui yeu cau"

        def locator(self, selector):
            return FakeBodyLocator(self)

        def wait_for_load_state(self, *args, **kwargs):
            return None

        def wait_for_timeout(self, *args, **kwargs):
            return None

    class FakeHelper:
        def _click_button_text(self, page, texts):
            page.body_text = "Nhap ma OTP"
            return True

    class FakeClosable:
        def close(self):
            return None

    state_path = tmp_path / "slot-1-state.json"

    monkeypatch.setattr(
        service,
        "start_onebss_otp_mobile_gateway_request",
        lambda settings, session_id, parameters, **kwargs: {
            "ok": False,
            "status": "otp_required",
            "session_id": session_id,
            "parameters": parameters,
        },
    )

    result = service.handle_onebss_otp_request(
        get_settings(),
        FakePage(),
        FakeHelper(),
        FakeClosable(),
        FakeClosable(),
        FakeClosable(),
        {"ma_bao_cao": "TEST"},
        {"P_DENNGAY": "{{today}}"},
        "worker",
        state_path=state_path,
    )

    assert result is not None
    pending = service.pop_onebss_session(result["session_id"])
    assert pending is not None
    assert pending.state_path == str(state_path)
    service.close_browser_stack(pending.browser, pending.context, pending.playwright)


def test_onebss_api_parallel_login_waits_for_shared_token(monkeypatch) -> None:
    from app.application import onebss_report_service as service

    username = "parallel.onebss"
    with service.PENDING_ONEBSS_LOCK:
        service.PENDING_ONEBSS_API_SESSIONS.clear()
        service.ONEBSS_API_TOKENS.pop(username, None)
    pending = service.keep_onebss_api_session("secret", {"ma_bao_cao": "TEST"}, {}, username, "mobile", "device", "worker")

    sleeps = {"count": 0}

    def fake_sleep(seconds):
        sleeps["count"] += 1
        if sleeps["count"] == 1:
            service.remember_onebss_api_token(
                username,
                {"access_token": "token", "token_type": "Bearer", "expires_in": 300},
                mobile_id="mobile",
                device_id="device",
            )

    monkeypatch.setattr(service.time, "sleep", fake_sleep)
    token = service.wait_for_onebss_api_token_from_pending_login(username, timeout_seconds=5)

    assert token is not None
    assert token.access_token == "token"
    service.pop_onebss_api_session(pending.session_id)
    with service.PENDING_ONEBSS_LOCK:
        service.ONEBSS_API_TOKENS.pop(username, None)


def test_onebss_mobile_gateway_default_filter_matches_vnpt_sms() -> None:
    from app.modules.mobile_gateway.otp_service import OtpService
    from app.modules.mobile_gateway.repository import MobileGatewayRepository
    from app.modules.mobile_gateway.schemas import SmsMessageIn

    with TestClient(app) as client:
        login(client)
        repository = MobileGatewayRepository(routes.build_app_repository(), get_settings())
        config = repository.get_otp_configuration("onebss")
        onebss_filter = next(item for item in repository.list_otp_filters("onebss", enabled_only=True) if item["filter_id"] == "onebss")
        assert config["sender_pattern"] == "VNPT"
        assert onebss_filter["sender_pattern"] == "VNPT"
        assert onebss_filter["start_prefix"] == ""

        service = OtpService(repository)
        request = service.create_request("onebss", job_id="onebss-vnpt-test")
        inserted, skipped = repository.save_sms_messages(
            "test-device-onebss",
            [
                SmsMessageIn(
                    external_id=f"vnpt-{request['request_id']}",
                    sender="VNPT",
                    body="Ma OTP dang nhap OneBSS cua Quy khach la 654321. Tran trong.",
                    received_at=repository.now(),
                )
            ],
        )
        assert skipped == 0
        assert inserted
        matched = service.match_incoming_sms(inserted[0])
        assert matched is not None
        assert service.consume_code(request["request_id"]) == "654321"


def test_mobile_gateway_expires_latest_otp_with_one_bulk_update() -> None:
    from app.modules.mobile_gateway.repository import MobileGatewayRepository

    class FakeSupabaseRepository:
        def __init__(self) -> None:
            self.patches = []

        def _patch(self, table, params, payload):
            self.patches.append((table, params, payload))

    base_repository = FakeSupabaseRepository()
    repository = MobileGatewayRepository(base_repository, get_settings())

    assert repository.expire_otp_latest_values() == 0
    assert len(base_repository.patches) == 1
    table, params, payload = base_repository.patches[0]
    assert table == "otp_latest_values"
    assert params["status"] == "eq.valid"
    assert params["expires_at"].startswith("lt.")
    assert payload["status"] == "expired"


def test_onebss_mobile_gateway_request_uses_latest_otp_received_before_request() -> None:
    from app.modules.mobile_gateway.otp_service import OtpService
    from app.modules.mobile_gateway.repository import MobileGatewayRepository
    from app.modules.mobile_gateway.schemas import SmsMessageIn

    with TestClient(app) as client:
        login(client)
        repository = MobileGatewayRepository(routes.build_app_repository(), get_settings())
        service = OtpService(repository)
        inserted, skipped = repository.save_sms_messages(
            "test-device-onebss-latest",
            [
                SmsMessageIn(
                    external_id=f"vnpt-latest-before-request-{time.time()}",
                    sender="VNPT",
                    body="Ma OTP dang nhap OneBSS cua Quy khach la 987654. Tran trong.",
                    received_at=repository.now(),
                )
            ],
        )
        assert skipped == 0
        assert inserted
        latest = service.record_latest_from_sms(inserted[0])
        assert latest is not None

        request = service.create_request("onebss", job_id="onebss-latest-before-request")

        assert service.consume_code(request["request_id"]) == "987654"
        consumed = repository.get_otp_request(request["request_id"])
        assert consumed is not None
        assert consumed["status"] == "consumed"


def test_internal_email_parser_keeps_full_body_and_otp() -> None:
    from app.modules.internal_email.service import parse_email_message

    raw_message = (
        b"From: VNPT <noreply@vnpt.vn>\r\n"
        b"Subject: Ma OTP dang nhap\r\n"
        b"Date: Tue, 14 Jul 2026 07:00:00 +0700\r\n"
        b"Message-ID: <otp-test@example.vn>\r\n"
        b"\r\n"
        b"Ma OTP cua ban la 123456. Khong chia se ma nay.\r\n"
    )

    parsed = parse_email_message(raw_message, "42")

    assert parsed["metadata"]["uid"] == "42"
    assert parsed["metadata"]["sender_email"] == "noreply@vnpt.vn"
    assert "123456" in parsed["search_text"]
    assert parsed["metadata"]["is_otp_candidate"] is True
    assert parsed["metadata"]["otp_code"] == "123456"
    assert "123456" in parsed["metadata"]["body_masked"]


def test_internal_email_parser_ignores_digits_inside_email_addresses() -> None:
    from app.modules.internal_email.service import parse_email_message

    raw_message = (
        b"From: VNPT HKD <noreply@vnpt.vn>\r\n"
        b"Subject: Ma OTP thay doi email\r\n"
        b"Date: Mon, 27 Jul 2026 08:23:39 +0700\r\n"
        b"Message-ID: <otp-email-change@example.vn>\r\n"
        b"\r\n"
        b"Xin chao, chung toi da nhan duoc yeu cau doi email tu dangminhtri1234@gmail.com "
        b"sang gialinh210797@gmail.com. Nhap ma doi email sau day 654321 Tran trong cam on.\r\n"
    )

    parsed = parse_email_message(raw_message, "43")

    assert parsed["metadata"]["otp_code"] == "654321"
    assert parsed["metadata"]["otp_code"] != "210797"


def test_internal_email_parser_does_not_treat_plain_year_as_otp() -> None:
    from app.modules.internal_email.service import parse_email_message

    raw_message = (
        b"From: Tender <notice@example.vn>\r\n"
        b"Subject: Thong tin goi thau moi 25/07/2026\r\n"
        b"Date: Mon, 27 Jul 2026 08:23:39 +0700\r\n"
        b"Message-ID: <plain-year@example.vn>\r\n"
        b"\r\n"
        b"Day la thong tin cua cac goi thau nam 2026, chi de tham khao noi bo.\r\n"
    )

    parsed = parse_email_message(raw_message, "45")

    assert parsed["metadata"]["is_otp_candidate"] is False
    assert parsed["metadata"]["otp_code"] == ""


def test_internal_email_parser_uses_configured_otp_rule() -> None:
    from app.modules.internal_email.service import parse_email_message

    raw_message = (
        b"From: VNPT HKD <noreply@vnpt.vn>\r\n"
        b"Subject: Ma OTP thay doi email\r\n"
        b"Date: Mon, 27 Jul 2026 08:23:39 +0700\r\n"
        b"Message-ID: <otp-rule@example.vn>\r\n"
        b"\r\n"
        b"Ma tham chieu 111111. Dia chi moi la gialinh210797@gmail.com. Ma OTP 654321.\r\n"
    )

    parsed = parse_email_message(
        raw_message,
        "44",
        otp_rules=[
            {
                "sender_pattern": "VNPT HKD",
                "sender_match_type": "contains",
                "direction": "right_to_left",
                "occurrence_index": 1,
                "start_position": 1,
                "otp_length": 6,
                "enabled": True,
            }
        ],
    )

    assert parsed["metadata"]["otp_code"] == "654321"


def test_internal_email_otp_rule_api_saves_lists_and_deletes() -> None:
    with TestClient(app) as client:
        login(client)
        sender = f"VNPT HKD {uuid.uuid4().hex[:8]}"
        payload = {
            "sender_pattern": sender,
            "sender_match_type": "contains",
            "label": "OTP doi email",
            "direction": "right_to_left",
            "occurrence_index": 1,
            "start_position": 1,
            "otp_length": 6,
            "regex": "",
            "priority": 12,
            "enabled": True,
        }

        save_response = client.post("/api/admin/internal-email/otp-rules", json=payload)
        assert save_response.status_code == 200
        rule = save_response.json()["rule"]
        assert rule["sender_pattern"] == sender
        assert rule["direction"] == "right_to_left"
        assert rule["otp_length"] == 6

        list_response = client.get("/api/admin/internal-email/otp-rules")
        assert list_response.status_code == 200
        assert any(item["id"] == rule["id"] for item in list_response.json()["rules"])

        delete_response = client.delete(f"/api/admin/internal-email/otp-rules/{rule['id']}")
        assert delete_response.status_code == 200
        after_delete = client.get("/api/admin/internal-email/otp-rules")
        assert all(item["id"] != rule["id"] for item in after_delete.json()["rules"])


def test_internal_email_refresh_existing_endpoint_is_available() -> None:
    with TestClient(app) as client:
        login(client)
        repository = routes.build_app_repository()
        connection = repository.get_system_connection_by_code("internal_email") or {}
        repository.upsert_system_connection(
            "internal_email",
            connection.get("name") or "Email nội bộ VNPT",
            "internal_email",
            connection.get("description") or "Đồng bộ email nội bộ.",
            connection.get("config") if isinstance(connection.get("config"), dict) else {},
            False,
        )
        response = client.post("/api/admin/internal-email/refresh-existing")

    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["details"]["enabled"] is False


def test_internal_email_migration_upgrades_legacy_otp_columns(tmp_path) -> None:
    from app.modules.internal_email.migrations import ensure_internal_email_sqlite_schema

    database_path = tmp_path / "legacy_internal_email.db"
    with sqlite3.connect(database_path) as connection:
        connection.execute(
            """
            CREATE TABLE internal_email_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                account_key TEXT NOT NULL DEFAULT 'internal_email',
                mailbox TEXT NOT NULL DEFAULT 'INBOX',
                uid TEXT NOT NULL,
                message_id TEXT NOT NULL DEFAULT '',
                sender TEXT NOT NULL DEFAULT '',
                sender_email TEXT NOT NULL DEFAULT '',
                subject TEXT NOT NULL DEFAULT '',
                body_masked TEXT NOT NULL DEFAULT '',
                received_at TEXT NOT NULL,
                synced_at TEXT NOT NULL,
                is_otp_candidate INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(account_key, mailbox, uid)
            )
            """
        )
        ensure_internal_email_sqlite_schema(connection)
        columns = {row[1] for row in connection.execute("PRAGMA table_info(internal_email_messages)").fetchall()}

    assert {"otp_code", "otp_code_masked", "otp_service_code", "otp_request_id"}.issubset(columns)


def test_internal_email_repository_tolerates_supabase_legacy_otp_columns() -> None:
    from app.modules.internal_email.repository import InternalEmailRepository

    class LegacySupabaseLikeRepository:
        def __init__(self) -> None:
            self.rows = []
            self.patched_payload = {}

        def _get(self, table, params):
            assert table == "internal_email_messages"
            if params.get("uid") == "eq.legacy-otp":
                return list(self.rows)
            return []

        def _insert(self, table, payload):
            assert table == "internal_email_messages"
            if "otp_code_masked" in payload:
                raise RuntimeError("Could not find the 'otp_code_masked' column of 'internal_email_messages' in the schema cache")
            row = {**payload, "id": 1}
            self.rows = [row]
            return row

        def _patch(self, table, params, payload):
            assert table == "internal_email_messages"
            if "otp_code_masked" in payload:
                raise RuntimeError("Could not find the 'otp_code_masked' column of 'internal_email_messages' in the schema cache")
            self.patched_payload = payload

    base_repository = LegacySupabaseLikeRepository()
    repository = InternalEmailRepository(base_repository)

    saved, created = repository.save_message(
        {
            "account_key": "internal_email",
            "mailbox": "INBOX",
            "uid": "legacy-otp",
            "message_id": "<legacy@example.vn>",
            "sender": "VNPT",
            "sender_email": "noreply@vnpt.vn",
            "subject": "Ma OTP",
            "body_masked": "Ma OTP la ******.",
            "received_at": repository.now(),
            "synced_at": repository.now(),
            "otp_code": "123456",
        }
    )
    repository.mark_message_otp(saved["id"], "onebss", "123456", "******", "REQ1")

    assert created is True
    assert saved["otp_code_masked"] == ""
    assert base_repository.patched_payload["is_otp_candidate"] is True
    assert "otp_code_masked" not in base_repository.patched_payload


def test_internal_email_messages_return_full_otp_for_copy() -> None:
    from app.modules.internal_email.repository import InternalEmailRepository

    with TestClient(app) as client:
        login(client)
        repository = InternalEmailRepository(routes.build_app_repository())
        uid = f"copy-otp-test-{uuid.uuid4().hex}"
        saved, created = repository.save_message(
            {
                "account_key": "internal_email",
                "mailbox": "INBOX",
                "uid": uid,
                "message_id": "<copy-otp-test@example.vn>",
                "sender": "VNPT",
                "sender_email": "noreply@vnpt.vn",
                "subject": "Ma OTP dang nhap",
                "body_masked": "Ma OTP cua ban la ******.",
                "received_at": repository.now(),
                "synced_at": repository.now(),
            }
        )
        assert created is True
        repository.mark_message_otp(saved["id"], "onebss", "246810", "******")

        response = client.get("/api/admin/internal-email/messages?limit=50&otp_only=true")
        assert response.status_code == 200
        message = next(item for item in response.json()["messages"] if item["uid"] == uid)
        assert message["otp_code"] == "246810"
        assert message["otp_code_masked"] == "******"
        assert "246810" not in message["body_masked"]
        assert "246810" in message["body_preview"]


def test_public_messages_feed_uses_allowed_email_and_sms_senders() -> None:
    from app.modules.internal_email.repository import InternalEmailRepository
    from app.modules.mobile_gateway.repository import MobileGatewayRepository
    from app.modules.mobile_gateway.schemas import SmsMessageIn

    with TestClient(app) as client:
        login(client)
        base_repository = routes.build_app_repository()
        email_repository = InternalEmailRepository(base_repository)
        mobile_repository = MobileGatewayRepository(base_repository, get_settings())
        unique = uuid.uuid4().hex
        email_sender = f"public-{unique}@example.vn"
        sms_sender = f"PUBLIC{unique[:8].upper()}"

        saved_email, _ = email_repository.save_message(
            {
                "account_key": "internal_email",
                "mailbox": "INBOX",
                "uid": f"public-email-{unique}",
                "message_id": f"<public-{unique}@example.vn>",
                "sender": "Public Mail",
                "sender_email": email_sender,
                "subject": "Ma OTP public",
                "body_masked": "Ma OTP email la ******.",
                "received_at": email_repository.now(),
                "synced_at": email_repository.now(),
            }
        )
        email_repository.mark_message_otp(saved_email["id"], "public", "112233", "******")
        inserted_sms, skipped = mobile_repository.save_sms_messages(
            "public-feed-device",
            [
                SmsMessageIn(
                    external_id=f"public-sms-{unique}",
                    sender=sms_sender,
                    body="Ma OTP SMS public cua ban la 445566.",
                    received_at=mobile_repository.now(),
                )
            ],
        )
        assert skipped == 0
        assert inserted_sms

        before_rules = client.get("/api/admin/public-messages/feed?limit=100")
        assert before_rules.status_code == 200
        assert all(item["id"] not in {f"email:{saved_email['id']}", f"sms:{inserted_sms[0]['id']}"} for item in before_rules.json()["items"])

        email_rule = client.post(
            "/api/admin/public-messages/rules",
            json={"source_type": "email", "sender_pattern": email_sender, "label": "Email public", "is_active": True},
        )
        sms_rule = client.post(
            "/api/admin/public-messages/rules",
            json={"source_type": "sms", "sender_pattern": sms_sender, "label": "SMS public", "is_active": True},
        )
        assert email_rule.status_code == 200
        assert sms_rule.status_code == 200

        response = client.get("/api/admin/public-messages/feed?limit=100")
        assert response.status_code == 200
        assert response.headers["cache-control"] == "no-store"
        items = response.json()["items"]
        assert len(items) <= 10
        assert response.json()["cursor"] == items[0]["received_at"]
        email_item = next(item for item in items if item["id"] == f"email:{saved_email['id']}")
        sms_item = next(item for item in items if item["id"] == f"sms:{inserted_sms[0]['id']}")

        assert email_item["type_label"] == "Mail nội bộ"
        assert email_item["title"] == "Ma OTP public"
        assert email_item["otp"] == "112233"
        assert "112233" in email_item["content"]
        assert sms_item["type_label"] == "SMS"
        assert sms_item["title"] == ""
        assert sms_item["otp"] == "445566"
        assert "445566" in sms_item["content"]

        delta = client.get(
            "/api/admin/public-messages/feed",
            params={"limit": 10, "after": response.json()["cursor"]},
        )
        assert delta.status_code == 200
        assert len(delta.json()["items"]) <= 10
        assert delta.json()["cursor"]

        invalid_cursor = client.get(
            "/api/admin/public-messages/feed",
            params={"after": "not-a-datetime"},
        )
        assert invalid_cursor.status_code == 400


def test_public_messages_parent_permission_can_view_feed() -> None:
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/users",
            json={
                "username": "viewer_public_parent",
                "full_name": "Viewer Public Parent",
                "password": "Viewer@Public123",
                "role": "viewer",
            },
        )
        assert created.status_code == 200
        viewer_id = created.json()["user"]["id"]
        assert client.put(
            f"/api/admin/users/{viewer_id}/permissions",
            json={"feature_codes": ["publicmessages"]},
        ).status_code == 200

        client.post("/api/auth/logout")
        login(client, "viewer_public_parent", "Viewer@Public123")
        page = client.get("/publicmessages")
        assert page.status_code == 200
        assert 'id="view-public-messages"' in page.text
        feed = client.get("/api/admin/public-messages/feed?limit=20")
        assert feed.status_code == 200
        assert feed.json()["ok"] is True


def test_internal_email_status_and_email_otp_can_match_request() -> None:
    from app.modules.mobile_gateway.otp_service import OtpService
    from app.modules.mobile_gateway.repository import MobileGatewayRepository

    with TestClient(app) as client:
        login(client)
        status_response = client.get("/api/admin/internal-email/status")
        assert status_response.status_code == 200
        status_payload = status_response.json()
        assert status_payload["ok"] is True
        assert status_payload["details"]["host"] == "email.vnpt.vn"
        assert status_payload["details"]["mailbox"] == "INBOX"

        messages_response = client.get("/api/admin/internal-email/messages?limit=200")
        assert messages_response.status_code == 200
        assert messages_response.json()["ok"] is True

        repository = MobileGatewayRepository(routes.build_app_repository(), get_settings())
        repository.ensure_defaults()
        service = OtpService(repository)
        request = service.create_request("onebss", job_id="email-otp-test")
        latest = service.record_latest_from_email(
            {
                "id": "email-otp-test-message",
                "sender": "VNPT",
                "sender_email": "noreply@vnpt.vn",
                "subject": "Ma OTP dang nhap OneBSS",
                "body": "Ma OTP dang nhap OneBSS cua Quy khach la 135790.",
                "received_at": repository.now(),
            }
        )

        assert latest is not None
        assert latest["code"] == "135790"
        assert latest["request_id"] == request["request_id"]
        assert service.consume_code(request["request_id"]) == "135790"


def test_onebss_mobile_gateway_resolver_auto_submits_otp(monkeypatch) -> None:
    from app.application import onebss_report_service as service

    events = {}

    class FakeRepository:
        def __init__(self, base_repository, settings):
            events["repository_created"] = True

        def get_otp_configuration(self, service_code):
            assert service_code == "otp_onebss"
            return {"manual_fallback_enabled": True, "auto_fill_enabled": True, "wait_timeout_seconds": 3}

    class FakeOtpService:
        def __init__(self, repository):
            events["otp_service_created"] = True

        def create_request(self, service_code, job_id=""):
            events["service_code"] = service_code
            events["job_id"] = job_id
            return {"request_id": "OTP-AUTO-001"}

        def wait_for_code(self, request_id, timeout_seconds):
            events["request_id"] = request_id
            events["timeout_seconds"] = timeout_seconds
            return "654321"

    def fake_continue(settings, session_id, otp, parameters):
        events["continued"] = (session_id, otp, parameters)
        return {"ok": True, "status": "success", "message": "auto otp ok", "parameters": parameters}

    monkeypatch.setattr(service, "MobileGatewayRepository", FakeRepository)
    monkeypatch.setattr(service, "OtpService", FakeOtpService)
    monkeypatch.setattr(service, "build_repository", lambda settings=None: object())
    monkeypatch.setattr(service, "continue_onebss_api_session", fake_continue)

    result = service.resolve_onebss_otp_with_mobile_gateway(
        get_settings(),
        "api-session-001",
        {"P_DENNGAY": "11/07/2026"},
        otp_service_code="otp_onebss",
    )

    assert result["ok"] is True
    assert events["service_code"] == "otp_onebss"
    assert events["job_id"] == "api-session-001"
    assert events["request_id"] == "OTP-AUTO-001"
    assert events["timeout_seconds"] == 3
    assert events["continued"] == ("api-session-001", "654321", {"P_DENNGAY": "11/07/2026"})


def test_onebss_mobile_gateway_request_returns_without_blocking(monkeypatch) -> None:
    from app.application import onebss_report_service as service

    events = {}

    class FakeRepository:
        def __init__(self, base_repository, settings):
            events["repository_created"] = True

        def get_otp_configuration(self, service_code):
            assert service_code == "otp_onebss"
            return {"manual_fallback_enabled": True, "auto_fill_enabled": True, "wait_timeout_seconds": 90}

    class FakeOtpService:
        def __init__(self, repository):
            events["otp_service_created"] = True

        def create_request(self, service_code, job_id=""):
            events["service_code"] = service_code
            events["job_id"] = job_id
            return {"request_id": "OTP-POLL-001"}

        def wait_for_code(self, request_id, timeout_seconds):
            raise AssertionError("start_onebss_otp_mobile_gateway_request must not block waiting for OTP")

    monkeypatch.setattr(service, "MobileGatewayRepository", FakeRepository)
    monkeypatch.setattr(service, "OtpService", FakeOtpService)
    monkeypatch.setattr(service, "build_repository", lambda settings=None: object())

    result = service.start_onebss_otp_mobile_gateway_request(
        get_settings(),
        "api-session-002",
        {"P_DENNGAY": "12/07/2026"},
        otp_service_code="otp_onebss",
    )

    assert result["ok"] is False
    assert result["status"] == "otp_required"
    assert result["session_id"] == "api-session-002"
    assert result["otp_request_id"] == "OTP-POLL-001"
    assert events["service_code"] == "otp_onebss"
    assert events["job_id"] == "api-session-002"


def test_onebss_mobile_gateway_poll_consumes_matched_otp(monkeypatch) -> None:
    from app.application import onebss_report_service as service

    events = {}

    class FakeRepository:
        def __init__(self, base_repository, settings):
            events["repository_created"] = True

        def expire_otp_requests(self):
            events["expired_checked"] = True

        def get_otp_request(self, request_id):
            events["request_id"] = request_id
            return {
                "request_id": request_id,
                "status": "matched",
                "matched_source_type": "sms",
                "matched_source_id": "42",
                "matched_at": "2026-07-12T08:00:00",
            }

    class FakeOtpService:
        def __init__(self, repository):
            events["otp_service_created"] = True

        def consume_code(self, request_id):
            events["consumed"] = request_id
            return "654321"

    monkeypatch.setattr(service, "MobileGatewayRepository", FakeRepository)
    monkeypatch.setattr(service, "OtpService", FakeOtpService)
    monkeypatch.setattr(service, "build_repository", lambda settings=None: object())

    result = service.consume_onebss_mobile_gateway_otp(get_settings(), "OTP-POLL-001")

    assert result["ok"] is True
    assert result["status"] == "matched"
    assert result["otp"] == "654321"
    assert result["source_type"] == "sms"
    assert result["source_id"] == "42"
    assert events["expired_checked"] is True
    assert events["consumed"] == "OTP-POLL-001"


def test_onebss_mobile_gateway_poll_recovers_recent_consumed_otp(monkeypatch) -> None:
    from app.application import onebss_report_service as service

    settings = get_settings()
    encrypted = service.security.encrypt_text(settings, "654321", "otp")
    consumed_at = datetime.now(service.LOCAL_TIMEZONE).isoformat(timespec="seconds")

    class FakeRepository:
        def __init__(self, base_repository, settings_arg):
            self.settings = settings_arg

        def expire_otp_requests(self):
            return 0

        def get_otp_request(self, request_id):
            return {
                "request_id": request_id,
                "status": "consumed",
                "matched_source_type": "sms",
                "matched_source_id": "42",
                "matched_at": consumed_at,
                "consumed_at": consumed_at,
                "code_encrypted": encrypted,
                "code_masked": "******",
            }

    monkeypatch.setattr(service, "MobileGatewayRepository", FakeRepository)
    monkeypatch.setattr(service, "build_repository", lambda settings=None: object())

    result = service.consume_onebss_mobile_gateway_otp(settings, "OTP-POLL-CONSUMED")

    assert result["ok"] is True
    assert result["status"] == "consumed"
    assert result["otp"] == "654321"
    assert result["source_type"] == "sms"
    assert result["source_id"] == "42"


def test_onebss_otp_request_poll_route_reports_matched_without_consuming(monkeypatch) -> None:
    def fake_inspect(settings, request_id):
        assert request_id == "OTP-POLL-001"
        return {"ok": True, "status": "matched", "code_masked": "******", "source_type": "sms"}

    monkeypatch.setattr(routes, "inspect_onebss_mobile_gateway_otp", fake_inspect)

    with TestClient(app) as client:
        login(client)
        response = client.get("/api/onebss-reports/otp-requests/OTP-POLL-001")

    assert response.status_code == 200
    assert response.json()["status"] == "matched"
    assert "otp" not in response.json()


def test_onebss_worker_consumes_otp_bound_to_request_id(monkeypatch) -> None:
    calls = []

    def fake_consume(settings, request_id):
        assert request_id == "OTP-POLL-001"
        calls.append(("consume", request_id))
        return {"ok": True, "status": "matched", "otp": "654321", "source_type": "sms"}

    def fake_inspect(settings, request_id):
        assert request_id == "OTP-POLL-001"
        calls.append(("inspect", request_id))
        return {"ok": True, "status": "matched", "code_masked": "******", "source_type": "sms"}

    monkeypatch.setattr(routes, "inspect_onebss_mobile_gateway_otp", fake_inspect)
    monkeypatch.setattr(routes, "consume_onebss_mobile_gateway_otp", fake_consume)

    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "Bien dong PTTB",
                "parameters": {"P_DENNGAY": "12/07/2026"},
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test",
            },
        )
        assert created.status_code == 200
        queued = client.post(
            "/api/onebss-reports/run",
            json={
                "ma_bao_cao": created.json()["ma_bao_cao"],
            },
        )
        assert queued.status_code == 200
        job_id = queued.json()["job_id"]
        response = client.post(
            f"/api/onebss-reports/jobs/{job_id}/otp",
            json={"otp_request_id": "OTP-POLL-001", "otp_source": "auto"},
        )
        worker_response = client.get(
            f"/api/onebss-worker/tasks/{job_id}/otp",
            headers={"Authorization": "Bearer test-worker-token"},
        )

    assert response.status_code == 200
    assert response.json()["ok"] is True
    assert worker_response.status_code == 200
    assert worker_response.json()["otp"] == "654321"
    assert calls == [("inspect", "OTP-POLL-001"), ("consume", "OTP-POLL-001")]


def test_onebss_auth_transition_waits_for_delayed_otp() -> None:
    from app.application.onebss_report_service import page_contains, wait_for_onebss_auth_transition

    class FakeBodyLocator:
        def __init__(self, page):
            self.page = page

        def inner_text(self, timeout=0):
            self.page.reads += 1
            if self.page.reads < 4:
                return "Dang nhap"
            return "XAC THUC OTP"

    class FakePage:
        def __init__(self):
            self.reads = 0
            self.waits = 0

        def locator(self, selector):
            assert selector == "body"
            return FakeBodyLocator(self)

        def wait_for_timeout(self, timeout):
            self.waits += 1

    class FakeHelper:
        def _is_login_page(self, page):
            return True

    page = FakePage()
    wait_for_onebss_auth_transition(page, FakeHelper(), timeout_ms=3000)

    assert page.waits >= 1
    assert page_contains(page, ["OTP"]) is True


def test_onebss_each_parameter_builds_multiple_payloads() -> None:
    from app.application.onebss_report_service import build_onebss_parameter_runs

    runs, merge_config, each_keys = build_onebss_parameter_runs(
        {
            "P_PHANVUNG_ID": {"$each": ["13", "14", "15"]},
            "P_LOAI_NGAY": "1",
            "P_TUNGAY": "01/07/2026",
            "P_DENNGAY": "09/07/2026",
            "baocao_id": 41668,
            "$merge_excel": {"mode": "append", "sheet": "DATA", "source_column": "P_PHANVUNG_ID"},
        }
    )

    assert each_keys == ["P_PHANVUNG_ID"]
    assert merge_config["sheet"] == "DATA"
    assert [run.parameters["P_PHANVUNG_ID"] for run in runs] == ["13", "14", "15"]
    assert all("$merge_excel" not in run.parameters for run in runs)
    assert all("baocao_id" not in run.parameters for run in runs)
    assert all("$each" not in run.parameters["P_PHANVUNG_ID"] for run in runs)


def test_onebss_tinh_each_parameter_automatically_enables_excel_merge() -> None:
    from app.application.onebss_report_service import build_onebss_parameter_runs

    runs, merge_config, each_keys = build_onebss_parameter_runs(
        {
            "P_TINH": {"$each": ["13", "47", "66"]},
            "P_LOAIKHO": "0",
        }
    )

    assert each_keys == ["P_TINH"]
    assert merge_config == {"sheet": "DATA", "source_column": "P_TINH"}
    assert [run.parameters["P_TINH"] for run in runs] == ["13", "47", "66"]


def test_onebss_tinh_each_parameter_maps_username_per_region() -> None:
    from app.application.onebss_report_service import build_onebss_parameter_runs

    runs, _, _ = build_onebss_parameter_runs(
        {
            "p_tinh": {"$each": ["13", "47", "66"]},
            "USERNAME": {
                "$by": "p_tinh",
                "values": {"13": "quyennt.cto", "47": "quyennt.cto_47", "66": "quyennt.cto_66"},
                "default": "quyennt.cto",
            },
        }
    )

    assert [(run.parameters["p_tinh"], run.parameters["USERNAME"]) for run in runs] == [
        ("13", "quyennt.cto"),
        ("47", "quyennt.cto_47"),
        ("66", "quyennt.cto_66"),
    ]


def test_onebss_non_region_each_parameter_does_not_automatically_merge() -> None:
    from app.application.onebss_report_service import build_onebss_parameter_runs

    _, merge_config, each_keys = build_onebss_parameter_runs({"P_LOAIKHO": {"$each": ["0", "1"]}})

    assert each_keys == ["P_LOAIKHO"]
    assert merge_config == {}


def test_onebss_report_id_uses_configured_meta_value() -> None:
    from app.application.onebss_report_service import OneBssApiToken, onebss_export_parameters, onebss_report_id

    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )
    parameters = {"$baocao_id": 41668, "baocao_id": 123, "P_PHANVUNG_ID": "13"}

    assert onebss_report_id({"report_url": "https://onebss.vnpt.vn/#/report/bi?path=UNKNOWN"}, parameters, token) == 41668
    assert onebss_export_parameters(parameters) == {"P_PHANVUNG_ID": "13"}


def test_onebss_download_falls_back_when_grid_has_no_rows(monkeypatch, tmp_path) -> None:
    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssDownloadError, OneBssDownloadedFile, OneBssApiToken

    events = []
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )

    def fake_grid(*args, **kwargs):
        events.append("grid")
        raise OneBssDownloadError("OneBSS run_v7 grid khong co du lieu")

    def fake_export(settings, token, report, parameters, **kwargs):
        events.append("export")
        target = kwargs.get("target_file") or tmp_path / "fallback.xlsx"
        target.write_bytes(b"PK\x03\x04")
        return OneBssDownloadedFile(
            file_path=target,
            suggested_filename="fallback.xlsx",
            export_info={"params": parameters},
            parameters=parameters,
            source_values=kwargs.get("source_values") or {},
        )

    monkeypatch.setattr(service, "download_onebss_grid_file_api", fake_grid)
    monkeypatch.setattr(service, "download_onebss_export_file_api", fake_export)

    result = service.download_onebss_report_file_api(
        get_settings(),
        token,
        {"report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test"},
        {"P_PHANVUNG_ID": "13", "$download_source": "grid"},
    )

    assert events == ["grid", "export"]
    assert result.suggested_filename == "fallback.xlsx"


def test_onebss_download_falls_back_when_grid_processing_timeout_persists(monkeypatch, tmp_path) -> None:
    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssDownloadError, OneBssDownloadedFile, OneBssApiToken

    events = []
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )

    def fake_grid(*args, **kwargs):
        events.append("grid")
        raise OneBssDownloadError("Quá thời gian xử lý")

    def fake_export(settings, token, report, parameters, **kwargs):
        events.append("export")
        target = kwargs.get("target_file") or tmp_path / "fallback.xlsx"
        target.write_bytes(b"PK\x03\x04")
        return OneBssDownloadedFile(
            file_path=target,
            suggested_filename="fallback.xlsx",
            export_info={"params": parameters},
            parameters=parameters,
            source_values=kwargs.get("source_values") or {},
        )

    monkeypatch.setattr(service, "download_onebss_grid_file_api", fake_grid)
    monkeypatch.setattr(service, "download_onebss_export_file_api", fake_export)

    result = service.download_onebss_report_file_api(
        get_settings(),
        token,
        {"report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test"},
        {"P_PHANVUNG_ID": "13", "$download_source": "grid"},
    )

    assert events == ["grid", "export"]
    assert result.suggested_filename == "fallback.xlsx"


def test_onebss_download_uses_grid_by_default(monkeypatch, tmp_path) -> None:
    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssDownloadedFile, OneBssApiToken

    events = []
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )

    def fake_grid(settings, token, report, parameters, **kwargs):
        events.append("grid")
        target = kwargs.get("target_file") or tmp_path / "grid.xlsx"
        target.write_bytes(b"PK\x03\x04")
        return OneBssDownloadedFile(
            file_path=target,
            suggested_filename="grid.xlsx",
            export_info={"params": parameters},
            parameters=parameters,
            source_values=kwargs.get("source_values") or {},
        )

    def fake_export(*args, **kwargs):
        events.append("export")
        raise AssertionError("Default OneBSS download should try grid before Excel export")

    monkeypatch.setattr(service, "download_onebss_grid_file_api", fake_grid)
    monkeypatch.setattr(service, "download_onebss_export_file_api", fake_export)

    result = service.download_onebss_report_file_api(
        get_settings(),
        token,
        {"report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test"},
        {"P_PHANVUNG_ID": "13"},
    )

    assert events == ["grid"]
    assert result.suggested_filename == "grid.xlsx"


def test_onebss_grid_processing_timeout_retries_same_token(monkeypatch, tmp_path) -> None:
    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssDownloadError, OneBssDownloadedFile, OneBssApiToken

    calls = []
    progress = []
    sleeps = []
    settings = get_settings().model_copy(
        update={
            "onebss_processing_timeout_retry_attempts": 3,
            "onebss_processing_timeout_retry_delay_seconds": 0,
        }
    )
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )

    def fake_grid_once(settings_arg, token_arg, report, parameters, **kwargs):
        calls.append({"token": token_arg, "parameters": dict(parameters)})
        if len(calls) == 1:
            raise OneBssDownloadError("Quá thời gian xử lý")
        target = kwargs.get("target_file") or tmp_path / "grid.xlsx"
        target.write_bytes(b"PK\x03\x04")
        return OneBssDownloadedFile(
            file_path=target,
            suggested_filename="grid.xlsx",
            export_info={"params": parameters},
            parameters=parameters,
            source_values=kwargs.get("source_values") or {},
        )

    monkeypatch.setattr(service, "_download_onebss_grid_file_api_once", fake_grid_once)
    monkeypatch.setattr(service.time, "sleep", lambda seconds: sleeps.append(seconds))

    result = service.download_onebss_grid_file_api(
        settings,
        token,
        {"report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test"},
        {"P_PHANVUNG_ID": "13"},
        progress_callback=progress.append,
    )

    assert result.suggested_filename == "grid.xlsx"
    assert len(calls) == 2
    assert calls[0]["token"] is token
    assert calls[1]["token"] is token
    assert calls[0]["parameters"] == calls[1]["parameters"] == {"P_PHANVUNG_ID": "13"}
    assert sleeps == [0.0]
    assert any("cung phien dang nhap" in message for message in progress)


def test_onebss_grid_wait_timeout_falls_back_without_retry(monkeypatch) -> None:
    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssDownloadError, OneBssDownloadedFile, OneBssApiToken

    calls = []
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )
    settings = get_settings().model_copy(update={"onebss_processing_timeout_retry_attempts": 3})

    def fake_grid_once(*args, **kwargs):
        calls.append("grid")
        raise OneBssDownloadError("grid_timeout: OneBSS tra du lieu luoi qua lau")

    monkeypatch.setattr(service, "_download_onebss_grid_file_api_once", fake_grid_once)

    with pytest.raises(OneBssDownloadError) as error:
        service.download_onebss_grid_file_api(
            settings,
            token,
            {"report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test"},
            {"P_PHANVUNG_ID": "13"},
        )

    assert calls == ["grid"]
    assert service.should_fallback_to_onebss_excel_export(error.value) is True


def test_onebss_start_uses_valid_cached_token_without_otp(monkeypatch) -> None:
    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssApiToken

    token = OneBssApiToken(
        access_token="cached-token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )
    progress = []
    finished_calls = []

    def fake_finish(settings, token_arg, report, parameters, **kwargs):
        finished_calls.append({"token": token_arg, "parameters": dict(parameters)})
        return {"ok": True, "status": "success", "message": "cached session ok", "parameters": parameters}

    monkeypatch.setattr(service, "onebss_api_credentials", lambda settings: ("test@vnpt.vn", "secret"))
    monkeypatch.setattr(service, "get_valid_onebss_api_token", lambda username: token)
    monkeypatch.setattr(service, "onebss_validate_api_token", lambda settings, token_arg: True)
    monkeypatch.setattr(service, "finish_onebss_report_download_api", fake_finish)
    monkeypatch.setattr(
        service,
        "start_onebss_otp_mobile_gateway_request",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("valid cached token must not request OTP")),
    )

    result = service.start_onebss_api_session(
        get_settings(),
        {"report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test"},
        {"P_TUNGAY": "01/07/2026"},
        progress_callback=progress.append,
    )

    assert result["ok"] is True
    assert finished_calls == [{"token": token, "parameters": {"P_TUNGAY": "01/07/2026"}}]
    assert any("khong can OTP" in message for message in progress)


def test_onebss_excel_export_405_falls_back_to_grid(monkeypatch, tmp_path) -> None:
    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssDownloadError, OneBssDownloadedFile, OneBssApiToken

    events = []
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )

    def fake_export(*args, **kwargs):
        events.append("export")
        raise OneBssDownloadError("OneBSS khong tra file bao cao. HTTP 405.")

    def fake_grid(settings, token, report, parameters, **kwargs):
        events.append("grid")
        target = kwargs.get("target_file") or tmp_path / "grid.xlsx"
        target.write_bytes(b"PK\x03\x04")
        return OneBssDownloadedFile(
            file_path=target,
            suggested_filename="grid.xlsx",
            export_info={"params": parameters},
            parameters=parameters,
            source_values=kwargs.get("source_values") or {},
        )

    monkeypatch.setattr(service, "download_onebss_export_file_api", fake_export)
    monkeypatch.setattr(service, "download_onebss_grid_file_api", fake_grid)

    result = service.download_onebss_report_file_api(
        get_settings(),
        token,
        {"report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test"},
        {"P_PHANVUNG_ID": "13", "$download_source": "excel"},
    )

    assert events == ["export", "grid"]
    assert result.suggested_filename == "grid.xlsx"


def test_onebss_grid_rows_are_written_to_excel(tmp_path) -> None:
    from openpyxl import load_workbook

    from app.application.onebss_report_service import onebss_grid_rows, write_onebss_grid_excel

    rows = onebss_grid_rows({"error_code": "BSS-00000000", "data": [{"MA_TB": "TB1", "DOANH_THU": 10}, {"MA_TB": "TB2", "GOI": "FIBER"}]})
    target = tmp_path / "onebss_grid.xlsx"
    write_onebss_grid_excel(rows, target)

    workbook = load_workbook(target, data_only=True)
    try:
        sheet = workbook["DATA"]
        assert [sheet.cell(row=1, column=index).value for index in range(1, 4)] == ["MA_TB", "DOANH_THU", "GOI"]
        assert sheet.cell(row=2, column=1).value == "TB1"
        assert sheet.cell(row=3, column=3).value == "FIBER"
    finally:
        workbook.close()


def test_onebss_finish_api_splits_regions_and_merges_excel(monkeypatch, tmp_path) -> None:
    from openpyxl import Workbook, load_workbook

    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssApiToken, OneBssDownloadedFile

    settings = get_settings().model_copy(update={"data_mining_download_dir": str(tmp_path)})
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )
    regions = []
    progress = []
    expected_token = token

    def fake_download(settings, token_arg, report, parameters, **kwargs):
        assert token_arg is expected_token
        region = str(parameters["P_PHANVUNG_ID"])
        regions.append(region)
        target = kwargs.get("target_file") or tmp_path / f"part_{region}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "DATA"
        sheet.append(["MA_TB", "DOANH_THU"])
        sheet.append([f"TB{region}", int(region)])
        workbook.save(target)
        workbook.close()
        return OneBssDownloadedFile(
            file_path=target,
            suggested_filename=f"part_{region}.xlsx",
            export_info={"report_id": 41668, "title": "Bao cao phat trien moi", "params": parameters},
            parameters=parameters,
            source_values=kwargs.get("source_values") or {},
        )

    monkeypatch.setattr(service, "download_onebss_report_file_api", fake_download)
    monkeypatch.setattr(service, "save_downloaded_file", lambda settings, target, storage: {"ok": True, "storage_link": str(target), "storage_status": "local"})

    result = service.finish_onebss_report_download_api(
        settings,
        token,
        {
            "ma_bao_cao": "ONEBSS_PTM",
            "ten_bao_cao": "Bao cao phat trien moi",
            "report_url": "https://onebss.vnpt.vn/#/report/bi?path=PHATTRIENTHUEBAO%2FBIENDONGPHATTRIENTHUEBAO%2FRP_BSS_28429&name=Test",
        },
        {
            "P_PHANVUNG_ID": {"$each": ["13", "14", "15"]},
            "P_LOAI_NGAY": "1",
            "P_TUNGAY": "01/07/2026",
            "P_DENNGAY": "14/07/2026",
            "$merge_excel": {"sheet": "DATA", "source_column": "P_PHANVUNG_ID"},
        },
        progress_callback=progress.append,
    )

    assert result["ok"] is True
    assert result["merged_file_count"] == 3
    assert regions == ["13", "14", "15"]
    assert any("dung chung phien dang nhap" in message for message in progress)

    workbook = load_workbook(result["file_path"], data_only=True)
    try:
        sheet = workbook["DATA"]
        assert sheet.max_row == 4
        assert [cell.value for cell in sheet[1]] == ["MA_TB", "DOANH_THU", "P_PHANVUNG_ID"]
        assert [sheet.cell(row=row, column=3).value for row in range(2, 5)] == ["13", "14", "15"]
    finally:
        workbook.close()


def test_onebss_finish_api_splits_regions_to_zip_by_default(monkeypatch, tmp_path) -> None:
    import zipfile
    from openpyxl import Workbook

    from app.application import onebss_report_service as service
    from app.application.onebss_report_service import OneBssApiToken, OneBssDownloadedFile

    settings = get_settings().model_copy(update={"data_mining_download_dir": str(tmp_path)})
    token = OneBssApiToken(
        access_token="token",
        token_type="Bearer",
        username="test@vnpt.vn",
        mobile_id="mobile",
        device_id="device",
        expires_at=9999999999,
    )

    def fake_download(settings, token, report, parameters, **kwargs):
        region = str(parameters["P_PHANVUNG_ID"])
        target = kwargs.get("target_file") or tmp_path / f"part_{region}.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["MA_TB"])
        sheet.append([f"TB{region}"])
        workbook.save(target)
        workbook.close()
        return OneBssDownloadedFile(
            file_path=target,
            suggested_filename=f"part_{region}.xlsx",
            export_info={"report_id": 41668, "title": "Bao cao phat trien moi", "params": parameters},
            parameters=parameters,
            source_values=kwargs.get("source_values") or {},
        )

    monkeypatch.setattr(service, "download_onebss_report_file_api", fake_download)
    monkeypatch.setattr(service, "save_downloaded_file", lambda settings, target, storage: {"ok": True, "storage_link": str(target), "storage_status": "local"})

    result = service.finish_onebss_report_download_api(
        settings,
        token,
        {
            "ma_bao_cao": "ONEBSS_PTM",
            "ten_bao_cao": "Bao cao phat trien moi",
            "report_url": "https://onebss.vnpt.vn/#/report/bi?path=PHATTRIENTHUEBAO%2FBIENDONGPHATTRIENTHUEBAO%2FRP_BSS_28429&name=Test",
        },
        {"P_PHANVUNG_ID": {"$each": ["13", "14", "15"]}},
    )

    assert result["ok"] is True
    assert result["output_mode"] == "split_archive"
    assert result["split_file_count"] == 3
    assert result["file_name"].endswith(".zip")
    with zipfile.ZipFile(result["file_path"]) as archive:
        names = archive.namelist()
    assert len(names) == 3
    assert any("P_PHANVUNG_ID_13" in name for name in names)


def test_onebss_merge_excel_files_appends_rows_with_source_column(tmp_path) -> None:
    from openpyxl import Workbook, load_workbook

    from app.application.onebss_report_service import OneBssDownloadedFile, merge_onebss_excel_files

    files = []
    for region, amount in [("13", 100), ("14", 200), ("15", 300)]:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Bao cao"
        sheet.append(["Bao cao phat trien thue bao"])
        sheet.append(["MA_TB", "DOANH_THU"])
        sheet.append([f"TB{region}", amount])
        source = tmp_path / f"region_{region}.xlsx"
        workbook.save(source)
        files.append(
            OneBssDownloadedFile(
                file_path=source,
                suggested_filename=source.name,
                export_info={},
                parameters={"P_PHANVUNG_ID": region},
                source_values={"P_PHANVUNG_ID": region},
            )
        )

    target = tmp_path / "merged.xlsx"
    merge_onebss_excel_files(
        files,
        target,
        {"mode": "append", "sheet": "DATA", "source_column": "P_PHANVUNG_ID"},
        ["P_PHANVUNG_ID"],
    )

    merged = load_workbook(target)
    rows = list(merged["DATA"].values)
    assert rows == [
        ("Bao cao phat trien thue bao", None, None),
        ("MA_TB", "DOANH_THU", "P_PHANVUNG_ID"),
        ("TB13", 100, "13"),
        ("TB14", 200, "14"),
        ("TB15", 300, "15"),
    ]


def test_onebss_report_run_records_worker_errors() -> None:
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS failed run",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_FAIL&name=Test",
                "storage_link": "https://drive.google.com/drive/folders/test-folder",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        failed = client.post(
            f"/api/onebss-worker/tasks/{job_id}/result",
            json={"ok": False, "status": "failed", "message": "browser launch failed"},
            headers={"Authorization": "Bearer test-worker-token"},
        )
        assert failed.status_code == 200
        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["status"] == "failed"
        assert "browser launch failed" in runs[0]["message"]


def test_onebss_report_run_expires_stale_worker_task(monkeypatch) -> None:
    monkeypatch.setattr(routes, "ONEBSS_REPORT_STALE_ACTIVE_SECONDS", 1)
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS stale run",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_STALE&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": "ws-stale"}, headers=headers)
        assert claim.status_code == 200
        assert claim.json()["task"]["run_id"] == job_id

        stale_at = (datetime.now(UTC) - timedelta(seconds=5)).isoformat(timespec="seconds").replace("+00:00", "Z")
        repository = routes.build_app_repository()
        repository.update_onebss_report_run(
            job_id,
            {"status": "running", "message": "Dang ket noi OneBSS.", "claimed_at": stale_at, "updated_at": stale_at},
        )

        job = client.get(f"/api/onebss-reports/jobs/{job_id}")
        assert job.status_code == 200
        assert job.json()["status"] == "failed"
        assert "bi treo" in job.json()["message"]

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["status"] == "failed"
        assert "bi treo" in runs[0]["message"]
        assert runs[0]["can_cancel"] is False


def test_onebss_report_run_expires_claimed_task_without_progress(monkeypatch) -> None:
    monkeypatch.setattr(routes, "ONEBSS_REPORT_CLAIM_NO_PROGRESS_SECONDS", 1)
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS claimed no progress",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_CLAIMED&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": "ws-claimed"}, headers=headers)
        assert claim.status_code == 200
        assert claim.json()["task"]["run_id"] == job_id

        stale_at = (datetime.now(UTC) - timedelta(seconds=5)).isoformat(timespec="seconds").replace("+00:00", "Z")
        routes.build_app_repository().update_onebss_report_run(job_id, {"claimed_at": stale_at, "updated_at": stale_at})

        job = client.get(f"/api/onebss-reports/jobs/{job_id}")
        assert job.status_code == 200
        assert job.json()["status"] == "failed"
        assert "khong gui buoc xu ly tiep theo" in job.json()["message"]

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["status"] == "failed"
        assert "khong gui buoc xu ly tiep theo" in runs[0]["message"]
        assert runs[0]["can_cancel"] is False


def test_onebss_report_run_expires_queued_task_when_no_worker(monkeypatch) -> None:
    monkeypatch.setattr(routes, "ONEBSS_REPORT_QUEUED_NO_WORKER_SECONDS", 1)
    with routes.WORKSTATION_HEARTBEATS_LOCK:
        routes.WORKSTATION_HEARTBEATS.clear()
    try:
        with TestClient(app) as client:
            login(client)
            created = client.post(
                "/api/admin/onebss-reports",
                json={
                    "ten_bao_cao": "OneBSS no worker",
                    "danh_sach_bien": ["P_TUNGAY"],
                    "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_NO_WORKER&name=Test",
                    "storage_link": "",
                },
            )
            assert created.status_code == 200
            code = created.json()["ma_bao_cao"]

            response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
            assert response.status_code == 200
            body = response.json()
            assert body["status"] == "queued"
            assert "chua thay worker OneBSS online" in body["message"]

            job_id = body["job_id"]
            stale_at = (datetime.now(UTC) - timedelta(seconds=5)).isoformat(timespec="seconds").replace("+00:00", "Z")
            routes.build_app_repository().update_onebss_report_run(job_id, {"updated_at": stale_at})

            job = client.get(f"/api/onebss-reports/jobs/{job_id}")
            assert job.status_code == 200
            assert job.json()["status"] == "failed"
            assert "Chua thay worker OneBSS online" in job.json()["message"]
            assert job.json()["can_cancel"] is False
    finally:
        with routes.WORKSTATION_HEARTBEATS_LOCK:
            routes.WORKSTATION_HEARTBEATS.clear()


def test_onebss_report_run_can_be_cancelled_without_worker_overwrite() -> None:
    with TestClient(app) as client:
        login(client)
        client.delete("/api/onebss-reports/runs")
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS cancellable run",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_CANCEL&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        headers = {"Authorization": "Bearer test-worker-token"}
        claim = client.post("/api/onebss-worker/tasks/claim", json={"worker_id": "ws-cancel"}, headers=headers)
        assert claim.status_code == 200
        assert claim.json()["task"]["run_id"] == job_id

        cancelled = client.post(f"/api/onebss-reports/runs/{job_id}/cancel")
        assert cancelled.status_code == 200
        assert cancelled.json()["status"] == "cancelled"
        assert cancelled.json()["can_cancel"] is False

        status_update = client.post(
            f"/api/onebss-worker/tasks/{job_id}/status",
            json={"status": "running", "message": "Still running", "worker_id": "ws-cancel"},
            headers=headers,
        )
        assert status_update.status_code == 200
        assert status_update.json()["cancelled"] is True

        finished = client.post(
            f"/api/onebss-worker/tasks/{job_id}/result",
            json={"ok": True, "status": "success", "message": "Should not overwrite cancel"},
            headers=headers,
        )
        assert finished.status_code == 200
        assert finished.json()["run"]["status"] == "cancelled"

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["status"] == "cancelled"
        assert runs[0]["can_cancel"] is False

        rerun = client.post(
            "/api/onebss-reports/run",
            json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "02/07/2026"}, "job_id": job_id},
        )
        assert rerun.status_code == 200
        assert rerun.json()["status"] == "queued"
        assert rerun.json()["job_id"] != job_id

        runs_after_rerun = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs_after_rerun) == 2
        assert {run["status"] for run in runs_after_rerun} == {"cancelled", "queued"}


def test_onebss_report_run_deduplicates_same_submit_and_uses_click_time_label() -> None:
    with TestClient(app) as client:
        login(client)
        client.delete("/api/onebss-reports/runs")
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS dedupe submit",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_DEDUPE&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        payload = {
            "ma_bao_cao": code,
            "parameters": {"P_TUNGAY": "01/07/2026"},
            "client_request_id": "same-submit-001",
        }
        first = client.post("/api/onebss-reports/run", json=payload)
        assert first.status_code == 200
        first_body = first.json()
        assert first_body["job_id"] == "same-submit-001"
        assert first_body["status"] == "queued"
        assert ":" in first_body["started_at_label"]
        assert "/" in first_body["started_at_label"]

        duplicate = client.post("/api/onebss-reports/run", json=payload)
        assert duplicate.status_code == 200
        assert duplicate.json()["job_id"] == first_body["job_id"]

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["run_id"] == "same-submit-001"
        assert runs[0]["started_at_label"] == first_body["started_at_label"]

        second = client.post(
            "/api/onebss-reports/run",
            json={
                "ma_bao_cao": code,
                "parameters": {"P_TUNGAY": "02/07/2026"},
                "client_request_id": "same-submit-002",
            },
        )
        assert second.status_code == 200
        assert second.json()["job_id"] == "same-submit-002"

        runs_after_second = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs_after_second) == 2
        assert {run["run_id"] for run in runs_after_second} == {"same-submit-001", "same-submit-002"}


def test_onebss_workstation_worker_updates_existing_status_message(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []

    class FakeResponse:
        def __init__(self, payload: dict) -> None:
            self.payload = payload

        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return self.payload

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append((method, path, kwargs.get("json") or {}))
            return FakeResponse({"ok": True, "run": {"status": "running"}})

    def fake_run_onebss_report_request(settings, report, parameters, **kwargs):
        progress_callback = kwargs["progress_callback"]
        progress_callback("Da dien tai khoan OneBSS.")
        progress_callback("Da dien mat khau OneBSS.")
        progress_callback("Da gui OTP ve dien thoai.")
        progress_callback("Da di den bao cao OneBSS.")
        return {
            "ok": True,
            "status": "success",
            "message": "Da tai bao cao OneBSS.",
            "storage_status": "uploaded_google_drive:test-file",
        }

    monkeypatch.setattr(worker, "run_onebss_report_request", fake_run_onebss_report_request)

    worker.process_task(FakeClient(), {"run_id": "RUN-PROGRESS", "report": {}, "parameters": {}}, "ws-progress", 0)

    messages = [payload.get("message") for _, path, payload in calls if path.endswith("/status")]
    assert "Da dien tai khoan OneBSS." in messages
    assert "Da dien mat khau OneBSS." in messages
    assert "Da gui OTP ve dien thoai." in messages
    assert "Da di den bao cao OneBSS." in messages


def test_ftp_workstation_worker_renders_date_file_template() -> None:
    from scripts import onebss_workstation_worker as worker

    now = datetime(2026, 7, 8, 9, 5)
    assert worker.render_ftp_file_template("bao_cao_{yyyymmdd}.xlsx", now) == "bao_cao_20260708.xlsx"
    assert worker.render_ftp_file_template("bao_cao_{ddmmyyyy}.xlsx", now) == "bao_cao_08072026.xlsx"
    assert worker.render_ftp_file_template("bao_cao_{{yesterday}}.xlsx", now) == "bao_cao_20260707.xlsx"
    assert worker.render_ftp_file_template("/DATA_BILLING/CTO/FiberPTM/{yyyyMM}", now) == "/DATA_BILLING/CTO/FiberPTM/202607"
    assert worker.render_ftp_file_template("CTO_DTTS_HOAMANG_{thang}01.CSV", now, {"thang": "202607"}) == "CTO_DTTS_HOAMANG_20260701.CSV"
    assert worker.render_ftp_file_template("CTO_DTTS_HOAMANG_{thang}01.CSV", now, {"thang": "{202607}"}) == "CTO_DTTS_HOAMANG_20260701.CSV"
    assert worker.render_ftp_file_template("CTO_Fiber_PTM_LK_ngay_{last_dd}.xlsx", now, {"thang": "202607"}) == "CTO_Fiber_PTM_LK_ngay_31.xlsx"

    config, folder_path, file_template = worker.parse_ftp_task({
        "folder_path": "ftp://ftp-user:ftp-pass@10.159.23.100:2121/reports/doanh_thu_{ddmmyyyy}.xlsx",
        "file_name_template": "",
        "connection": {"config": {"passive": False, "timeout_seconds": 90}},
    })
    assert config["host"] == "10.159.23.100"
    assert config["port"] == 2121
    assert config["username"] == "ftp-user"
    assert config["password"] == "ftp-pass"
    assert config["passive"] is False
    assert config["timeout_seconds"] == 90
    assert folder_path == "/reports"
    assert file_template == "doanh_thu_{ddmmyyyy}.xlsx"


def test_ftp_workstation_worker_plans_and_merges_multi_source_files(tmp_path) -> None:
    from scripts import onebss_workstation_worker as worker

    now = datetime(2026, 7, 8, 9, 5)
    advanced_template = json.dumps({
        "version": 1,
        "variables": {"thang": "{yyyyMM}"},
        "output_file_name_template": "DTTS_HOAMANG_{thang}.xlsx",
        "sources": [
            {"name": "CTO", "folder_path": "/DATA_BILLING/CTO/SUBS", "file_name_template": "CTO_DTTS_HOAMANG_{thang}01.CSV"},
            {"name": "HAG", "folder_path": "/DATA_BILLING/HGA/SUBS", "file_name_template": "HGA_DTTS_HOAMANG_{thang}01.CSV"},
            {"name": "STG", "folder_path": "/DATA_BILLING/STG/SUBS", "file_name_template": "STG_DTTS_HOAMANG_{thang}01.CSV"},
        ],
    })
    plan = worker.build_ftp_download_plan({
        "ma_bao_cao": "FTP_SUBS",
        "folder_path": "/DATA_BILLING/CTO/SUBS",
        "file_name_template": advanced_template,
        "variables": {"thang": "{202607}"},
        "connection": {"config": {"host": "10.159.23.100", "username": "u", "password": "p"}},
    }, now)
    assert plan["is_multi_source"] is True
    assert plan["output_file_name"] == "DTTS_HOAMANG_202607.xlsx"
    assert plan["sources"][0]["folder_path"] == "/DATA_BILLING/CTO/SUBS"
    assert plan["sources"][0]["file_name_template"] == "CTO_DTTS_HOAMANG_20260701.CSV"
    assert plan["sources"][1]["name"] == "HAG"
    assert plan["sources"][1]["folder_path"] == "/DATA_BILLING/HAG/SUBS"
    assert plan["sources"][1]["file_name_template"] == "HAG_DTTS_HOAMANG_20260701.CSV"

    cto = tmp_path / "cto.csv"
    hag = tmp_path / "hag.csv"
    stg = tmp_path / "stg.csv"
    cto.write_text("ma_tb,doanh_thu\nA,10\n", encoding="utf-8")
    hag.write_text("ma_tb,doanh_thu\nB,20\n", encoding="utf-8")
    stg.write_text("ma_tb,doanh_thu\nC,30\n", encoding="utf-8")
    target = tmp_path / "merged.xlsx"
    worker._merge_ftp_downloaded_files([
        {"source": "CTO", "file_path": str(cto), "resolved_file_name": "cto.csv"},
        {"source": "HAG", "file_path": str(hag), "resolved_file_name": "hag.csv"},
        {"source": "STG", "file_path": str(stg), "resolved_file_name": "stg.csv"},
    ], target)

    workbook = openpyxl.load_workbook(target, read_only=True)
    rows = list(workbook["TongHop"].iter_rows(values_only=True))
    workbook.close()
    assert rows == [
        ("Nguon", "ma_tb", "doanh_thu"),
        ("CTO", "A", "10"),
        ("HAG", "B", "20"),
        ("STG", "C", "30"),
    ]


def test_ftp_workstation_worker_downloads_each_run_in_own_directory(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    from scripts import onebss_workstation_worker as worker

    settings = get_settings().model_copy(update={"data_mining_download_dir": str(tmp_path)})
    captured_dirs = []

    def fake_download_source(source, local_dir, progress_callback=None):
        captured_dirs.append(local_dir)
        local_path = local_dir / "same_report.csv"
        local_path.write_text("ma_tb\nTB001\n", encoding="utf-8")
        return {
            "source": source.get("name") or "FTP",
            "resolved_file_name": "same_report.csv",
            "file_name": local_path.name,
            "file_path": str(local_path),
        }

    monkeypatch.setattr(worker, "get_settings", lambda: settings)
    monkeypatch.setattr(worker, "_download_ftp_source_file", fake_download_source)

    task = {
        "run_id": "FTP-RUN-001",
        "folder_path": "/reports",
        "file_name_template": "same_report.csv",
        "connection": {"config": {"host": "10.159.23.100", "username": "u", "password": "p"}},
    }
    result = worker.download_ftp_report_file(task)

    assert result["ok"] is True
    assert captured_dirs == [tmp_path / "ftp" / "runs" / "FTP-RUN-001"]
    assert Path(result["file_path"]).parent == captured_dirs[0]


def test_ftp_workstation_worker_uploads_drive_before_web_file(monkeypatch, tmp_path) -> None:
    from scripts import onebss_workstation_worker as worker

    source = tmp_path / "ftp_result.xlsx"
    source.write_bytes(b"xlsx-bytes")
    drive_calls = []
    web_upload_calls = []
    progress_messages = []

    def fake_upload_result_file_to_internal_drive(file_path, drive_folder_id, **kwargs):
        drive_calls.append((file_path, drive_folder_id, kwargs))
        return {
            "file_name": "ftp_result.xlsx",
            "storage_link": "https://drive.google.com/file/d/ftp-worker-file/view",
            "storage_status": "uploaded_google_drive:ftp-worker-file",
            "message": "Da upload file FTP len Google Drive qua API trung gian.",
        }

    def fake_upload_ftp_result_file(client, run_id, file_path):
        web_upload_calls.append((run_id, file_path))
        return {}

    monkeypatch.setattr(worker, "upload_result_file_to_internal_drive", fake_upload_result_file_to_internal_drive)
    monkeypatch.setattr(worker, "upload_ftp_result_file", fake_upload_ftp_result_file)

    result = worker.attach_ftp_file_if_needed(
        object(),
        "RUN-FTP-DRIVE",
        {"ok": True, "status": "success", "file_name": "ftp_result.xlsx", "file_path": str(source)},
        "drive-folder-ftp",
        lambda message, *args: progress_messages.append(message),
    )

    assert result["storage_link"] == "https://drive.google.com/file/d/ftp-worker-file/view"
    assert result["storage_status"] == "uploaded_google_drive:ftp-worker-file"
    assert result["message"] == "Da upload file FTP len Google Drive qua API trung gian."
    assert drive_calls == [(
        str(source),
        "drive-folder-ftp",
        {
            "request_source": "ftp-worker",
            "default_message": "Da upload file FTP len Google Drive qua API trung gian.",
            "job_id": "RUN-FTP-DRIVE",
        },
    )]
    assert web_upload_calls == []
    assert "Da upload file FTP len Google Drive." in progress_messages


def test_ftp_workstation_worker_cwd_error_names_source_and_folder(monkeypatch, tmp_path) -> None:
    from scripts import onebss_workstation_worker as worker

    class FailingFTP:
        def connect(self, **kwargs) -> None:
            return None

        def login(self, **kwargs) -> None:
            return None

        def set_pasv(self, passive: bool) -> None:
            return None

        def cwd(self, path: str) -> None:
            raise worker.ftplib.error_perm("550 Failed to change directory.")

        def quit(self) -> None:
            return None

    monkeypatch.setattr(worker.ftplib, "FTP", FailingFTP)

    with pytest.raises(RuntimeError) as error:
        worker._download_ftp_source_file({
            "name": "HAG",
            "folder_path": "/DATA_BILLING/HAG/SUBS",
            "file_name_template": "HAG_DTTS_HOAMANG_20260701.CSV",
            "config": {"host": "10.159.23.100", "username": "u", "password": "p"},
        }, tmp_path)

    message = str(error.value)
    assert "HAG" in message
    assert "/DATA_BILLING/HAG/SUBS" in message
    assert "550 Failed to change directory" in message


def test_onebss_workstation_worker_retries_transient_web_errors(monkeypatch) -> None:
    import httpx
    from scripts import onebss_workstation_worker as worker

    attempts = {"count": 0}
    monkeypatch.setattr(worker.time, "sleep", lambda seconds: None)

    class FakeResponse:
        def __init__(self, status_code: int, payload: dict | None = None) -> None:
            self.status_code = status_code
            self.payload = payload or {}
            self.request = httpx.Request("POST", "https://vnptcto.com/api/onebss-worker/tasks/claim")

        def raise_for_status(self) -> None:
            if self.status_code >= 400:
                response = httpx.Response(self.status_code, request=self.request)
                raise httpx.HTTPStatusError("temporary error", request=self.request, response=response)

        def json(self) -> dict:
            return self.payload

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            attempts["count"] += 1
            if attempts["count"] == 1:
                return FakeResponse(502)
            return FakeResponse(200, {"ok": True, "task": None})

    data = worker.request_json(FakeClient(), "POST", "/api/onebss-worker/tasks/claim", json={"worker_id": "ws"})

    assert data == {"ok": True, "task": None}
    assert attempts["count"] == 2


def test_onebss_workstation_worker_continues_when_progress_status_is_transient(monkeypatch) -> None:
    import httpx
    from scripts import onebss_workstation_worker as worker

    calls = []
    ran = {"onebss": False}
    monkeypatch.setattr(worker.time, "sleep", lambda seconds: None)

    class FakeResponse:
        def __init__(self, status_code: int, payload: dict | None = None, path: str = "") -> None:
            self.status_code = status_code
            self.payload = payload or {"ok": True}
            self.request = httpx.Request("POST", f"https://vnptcto.com{path}")

        def raise_for_status(self) -> None:
            if self.status_code >= 400:
                response = httpx.Response(self.status_code, request=self.request)
                raise httpx.HTTPStatusError("temporary error", request=self.request, response=response)

        def json(self) -> dict:
            return self.payload

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})
            if path.endswith("/status"):
                return FakeResponse(502, path=path)
            return FakeResponse(200, {"ok": True}, path=path)

    def fake_onebss(*args, **kwargs):
        ran["onebss"] = True
        return {"ok": True, "status": "success", "message": "done"}

    monkeypatch.setattr(worker, "run_onebss_report_request", fake_onebss)

    worker.process_task(
        FakeClient(),
        {"run_id": "RUN-PROGRESS-502", "report": {"ma_bao_cao": "ONEBSS"}, "parameters": {}},
        "ws-progress-502",
        0,
    )

    assert ran["onebss"] is True
    result_calls = [call for call in calls if call["path"] == "/api/onebss-worker/tasks/RUN-PROGRESS-502/result"]
    assert len(result_calls) == 1
    assert result_calls[0]["json"]["ok"] is True


def test_onebss_workstation_worker_relogs_when_otp_session_expired(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []
    run_calls = []
    otp_values = iter(["111111", "222222"])
    results = iter(
        [
            {"ok": False, "status": "otp_required", "session_id": "api-old", "message": "Can OTP"},
            {"ok": False, "status": "otp_session_expired", "message": "Phien OTP OneBSS da het han"},
            {"ok": False, "status": "otp_required", "session_id": "api-new", "message": "Can OTP moi"},
            {"ok": True, "status": "success", "message": "done"},
        ]
    )

    def fake_request_json(client, method: str, path: str, **kwargs):
        calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})
        return {"ok": True}

    def fake_run(client, worker_id, run_id, report, parameters, **kwargs):
        run_calls.append({"session_id": kwargs.get("session_id") or "", "otp": kwargs.get("otp") or ""})
        return next(results)

    monkeypatch.setattr(worker, "request_json", fake_request_json)
    monkeypatch.setattr(worker, "run_onebss_report_request_guarded", fake_run)
    monkeypatch.setattr(worker, "wait_for_otp", lambda *args, **kwargs: next(otp_values))
    monkeypatch.setattr(worker, "attach_worker_file_if_needed", lambda client, run_id, result, drive_folder_id, progress: result)
    monkeypatch.setattr(worker, "send_heartbeat", lambda *args, **kwargs: None)

    worker.process_task(
        object(),
        {"run_id": "RUN-OTP-EXPIRED", "report": {"ma_bao_cao": "ONEBSS_OTP"}, "parameters": {}},
        "ws-otp",
        0,
    )

    assert run_calls == [
        {"session_id": "", "otp": ""},
        {"session_id": "api-old", "otp": "111111"},
        {"session_id": "", "otp": ""},
        {"session_id": "api-new", "otp": "222222"},
    ]
    result_calls = [call for call in calls if call["path"] == "/api/onebss-worker/tasks/RUN-OTP-EXPIRED/result"]
    assert len(result_calls) == 1
    assert result_calls[0]["json"]["ok"] is True
    status_messages = [call["json"].get("message", "") for call in calls if call["path"].endswith("/status")]
    assert any("dang nhap lai" in message for message in status_messages)


def test_onebss_workstation_worker_reports_unexpected_failure(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {"ok": True}

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})
            return FakeResponse()

    def fail_onebss(*args, **kwargs):
        raise RuntimeError("selenium timeout")

    monkeypatch.setattr(worker, "run_onebss_report_request", fail_onebss)

    worker.process_task(
        FakeClient(),
        {"run_id": "RUN-ERR", "report": {"ma_bao_cao": "ONEBSS_ERR"}, "parameters": {}},
        "ws-err",
        0,
    )

    result_calls = [call for call in calls if call["path"] == "/api/onebss-worker/tasks/RUN-ERR/result"]
    assert len(result_calls) == 1
    payload = result_calls[0]["json"]
    assert payload["ok"] is False
    assert payload["status"] == "failed"
    assert "selenium timeout" in payload["message"]
    assert payload["details"]["error_type"] == "RuntimeError"


def test_onebss_worker_task_guard_is_disabled_by_default(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    monkeypatch.delenv("ONEBSS_WORKER_DISABLE_TASK_GUARD", raising=False)
    monkeypatch.delenv("ONEBSS_WORKER_ENABLE_TASK_GUARD", raising=False)

    class FakeClient:
        base_url = "https://vnptcto.com"

    def fail_process(*args, **kwargs):
        raise AssertionError("OneBSS OTP flow must not use multiprocessing guard by default")

    monkeypatch.setattr(worker.mp, "Process", fail_process)
    monkeypatch.setattr(worker, "run_onebss_report_request", lambda *args, **kwargs: {"ok": True, "status": "success"})

    result = worker.run_onebss_report_request_guarded(
        FakeClient(),
        "ws-otp",
        "RUN-OTP",
        {"ma_bao_cao": "OTP"},
        {},
        otp="123456",
        session_id="SESSION-1",
    )

    assert result == {"ok": True, "status": "success"}


def test_onebss_worker_wait_for_otp_times_out(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    times = iter([0.0, 0.0, 31.0])
    monkeypatch.setenv("ONEBSS_WORKER_OTP_WAIT_SECONDS", "30")
    monkeypatch.setattr(worker.time, "monotonic", lambda: next(times))
    monkeypatch.setattr(worker.time, "sleep", lambda seconds: None)
    monkeypatch.setattr(worker, "request_json", lambda *args, **kwargs: {"ok": False, "message": "Chua thay OTP."})

    with pytest.raises(TimeoutError) as error:
        worker.wait_for_otp(object(), "RUN-OTP", 0)

    assert "Khong nhan duoc OTP OneBSS" in str(error.value)


def test_onebss_worker_prefers_local_drive_upload_api_when_public_is_configured(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    monkeypatch.setenv("ONEBSS_DRIVE_UPLOAD_API_URL", "https://api.vnptcto.com/api/du-lieu-web")
    monkeypatch.setenv("INTERNAL_API_URL", "https://api.vnptcto.com/api/du-lieu-web")

    urls = worker.internal_drive_upload_api_urls()

    assert urls[0] == "http://127.0.0.1:8000/api/du-lieu-web"
    assert urls.count("https://api.vnptcto.com/api/du-lieu-web") == 1


def test_sql_worker_prefers_local_internal_api_when_public_is_configured(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    monkeypatch.setenv("INTERNAL_API_URL", "https://api.vnptcto.com/api/du-lieu-web")
    monkeypatch.delenv("SQL_WORKER_API_URL", raising=False)

    urls = worker.internal_sql_api_urls()

    assert urls[0] == "http://127.0.0.1:8000/api/du-lieu-web"
    assert urls.count("https://api.vnptcto.com/api/du-lieu-web") == 1


def test_sql_worker_posts_result_to_web(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {"ok": True}

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})
            return FakeResponse()

    monkeypatch.setattr(
        worker,
        "run_sql_worker_query",
        lambda task: {
            "ok": True,
            "columns": ["MA_TB"],
            "rows": [{"MA_TB": "tb-local"}],
            "total": 1,
            "page": 1,
            "page_size": 20,
            "message": "ok local",
        },
    )

    worker.process_sql_task(
        FakeClient(),
        {"run_id": "SQL-1", "report_code": "BC_SQL", "query": {"pagination": {"page": 1, "page_size": 20}}},
        "ws-sql",
    )

    result_calls = [call for call in calls if call["path"] == "/api/sql-worker/tasks/SQL-1/result"]
    assert len(result_calls) == 1
    payload = result_calls[0]["json"]
    assert payload["ok"] is True
    assert payload["columns"] == ["MA_TB"]
    assert payload["rows"] == [{"MA_TB": "tb-local"}]
    assert payload["pagination"]["total"] == 1


def test_sql_worker_collects_all_pages_before_posting_result(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []
    queried_pages = []

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {"ok": True}

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})
            return FakeResponse()

    def fake_run_sql_worker_query(task: dict) -> dict:
        pagination = task["query"]["pagination"]
        page = int(pagination["page"])
        queried_pages.append(page)
        rows_by_page = {
            1: [{"MA_TB": "TB001"}, {"MA_TB": "TB002"}],
            2: [{"MA_TB": "TB003"}, {"MA_TB": "TB004"}],
            3: [{"MA_TB": "TB005"}],
        }
        return {
            "ok": True,
            "columns": ["MA_TB"],
            "rows": rows_by_page.get(page, []),
            "total": 5,
            "pagination": {"page": page, "page_size": 2, "total": 5},
            "message": f"page {page}",
        }

    monkeypatch.setattr(worker, "run_sql_worker_query", fake_run_sql_worker_query)

    worker.process_sql_task(
        FakeClient(),
        {
            "run_id": "SQL-ALL-PAGES",
            "report_code": "BC_SQL_ALL",
            "query": {
                "action": "run_sql_report",
                "collect_all_pages": True,
                "max_rows": 50000,
                "pagination": {"page": 1, "page_size": 2},
            },
        },
        "ws-sql",
    )

    result_calls = [call for call in calls if call["path"] == "/api/sql-worker/tasks/SQL-ALL-PAGES/result"]
    assert queried_pages == [1, 2, 3]
    assert len(result_calls) == 1
    payload = result_calls[0]["json"]
    assert payload["ok"] is True
    assert payload["rows"] == [{"MA_TB": "TB001"}, {"MA_TB": "TB002"}, {"MA_TB": "TB003"}, {"MA_TB": "TB004"}, {"MA_TB": "TB005"}]
    assert payload["pagination"]["total"] == 5
    assert payload["pagination"]["fetched_rows"] == 5
    assert payload["pagination"]["truncated"] is False
    assert payload["details"]["collect_all_pages"] is True


def test_workstation_worker_claims_onebss_before_sql(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []
    processed = []

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})

            class FakeResponse:
                status_code = 200

                def raise_for_status(self) -> None:
                    return None

                def json(self) -> dict:
                    if path == "/api/onebss-worker/tasks/claim":
                        return {
                            "ok": True,
                            "task": {
                                "run_id": "ONEBSS-FIRST",
                                "report": {"ma_bao_cao": "BC_ONEBSS"},
                                "parameters": {},
                            },
                        }
                    if path == "/api/sql-worker/tasks/claim":
                        return {"ok": True, "task": {"run_id": "SQL-SHOULD-WAIT", "report_code": "BC_SQL", "query": {}}}
                    return {"ok": True, "task": None}

            return FakeResponse()

    def fake_onebss_task(client, task, worker_id, poll_seconds):
        processed.append(("onebss", task["run_id"], worker_id))

    monkeypatch.setattr(worker, "process_task", fake_onebss_task)
    monkeypatch.setattr(worker, "process_sql_task", lambda *args, **kwargs: processed.append(("sql",)))
    monkeypatch.setattr(worker, "send_heartbeat", lambda *args, **kwargs: None)

    assert worker.poll_worker_once(FakeClient(), "ws-priority", 0) is True
    assert processed == [("onebss", "ONEBSS-FIRST", "ws-priority")]
    assert calls[0]["path"] == "/api/onebss-worker/tasks/claim"
    assert calls[0]["json"]["version"] == worker.WORKER_VERSION
    assert all(call["path"] != "/api/sql-worker/tasks/claim" for call in calls)


def test_workstation_worker_idle_claims_include_process_details() -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})

            class FakeResponse:
                status_code = 200

                def raise_for_status(self) -> None:
                    return None

                def json(self) -> dict:
                    return {"ok": True, "task": None}

            return FakeResponse()

    assert worker.poll_worker_once(FakeClient(), "ws-idle", 0) is False
    claim_calls = [call for call in calls if call["path"].endswith("/tasks/claim")]
    assert [call["path"] for call in claim_calls] == [
        "/api/onebss-worker/tasks/claim",
        "/api/sql-worker/tasks/claim",
        "/api/ftp-worker/tasks/claim",
    ]
    for call in claim_calls:
        assert call["json"]["version"] == worker.WORKER_VERSION
        assert call["json"]["details"]["pid"]
        assert call["json"]["details"]["worker_version"] == worker.WORKER_VERSION
        assert "dang chay nen" in call["json"]["details"]["worker_process"]


def test_workstation_worker_concurrency_limits_and_tracker(monkeypatch: pytest.MonkeyPatch) -> None:
    from scripts import onebss_workstation_worker as worker

    for name in (
        "VNPTCTO_WORKER_MAX_CONCURRENT_TASKS",
        "ONEBSS_WORKER_MAX_CONCURRENT_TASKS",
        "ONEBSS_WORKER_MAX_ONEBSS_TASKS",
        "SQL_WORKER_MAX_CONCURRENT_TASKS",
        "SQL_WORKER_MAX_TASKS",
        "FTP_WORKER_MAX_CONCURRENT_TASKS",
        "FTP_WORKER_MAX_TASKS",
    ):
        monkeypatch.delenv(name, raising=False)

    assert worker.worker_concurrency_limits() == {
        "total": 4,
        worker.TASK_KIND_ONEBSS: 2,
        worker.TASK_KIND_SQL: 2,
        worker.TASK_KIND_FTP: 2,
    }

    monkeypatch.setenv("VNPTCTO_WORKER_MAX_CONCURRENT_TASKS", "2")
    monkeypatch.setenv("ONEBSS_WORKER_MAX_ONEBSS_TASKS", "1")
    monkeypatch.setenv("SQL_WORKER_MAX_CONCURRENT_TASKS", "2")
    monkeypatch.setenv("FTP_WORKER_MAX_CONCURRENT_TASKS", "2")
    tracker = worker.WorkerConcurrencyTracker()
    assert tracker.try_start(worker.TASK_KIND_ONEBSS, "ONEBSS-1", "BC_ONEBSS") is True
    assert tracker.try_start(worker.TASK_KIND_ONEBSS, "ONEBSS-2", "BC_ONEBSS") is False
    assert tracker.try_start(worker.TASK_KIND_SQL, "SQL-1", "BC_SQL") is True
    assert tracker.try_start(worker.TASK_KIND_FTP, "FTP-1", "BC_FTP") is False
    assert tracker.counts() == {
        "total": 2,
        worker.TASK_KIND_ONEBSS: 1,
        worker.TASK_KIND_SQL: 1,
        worker.TASK_KIND_FTP: 0,
    }
    tracker.finish("ONEBSS-1")
    assert tracker.try_start(worker.TASK_KIND_FTP, "FTP-1", "BC_FTP") is True


def test_workstation_worker_parallel_poll_skips_full_onebss_slot(monkeypatch: pytest.MonkeyPatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []
    started = []

    class FakeDispatcher:
        def prune_threads(self) -> None:
            return None

        def active_details(self) -> dict:
            return {"active_counts": {"total": 1, "onebss": 1, "sql": 0, "ftp": 0}}

        def can_start(self, kind: str) -> bool:
            return kind != worker.TASK_KIND_ONEBSS

        def start_task(self, kind: str, task: dict) -> bool:
            started.append((kind, task.get("run_id")))
            return True

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})

            class FakeResponse:
                status_code = 200

                def raise_for_status(self) -> None:
                    return None

                def json(self) -> dict:
                    if path == "/api/sql-worker/tasks/claim":
                        return {"ok": True, "task": {"run_id": "SQL-1", "report_code": "BC_SQL", "query": {}}}
                    return {"ok": True, "task": None}

            return FakeResponse()

    monkeypatch.setattr(worker, "send_heartbeat", lambda *args, **kwargs: None)

    assert worker.poll_worker_once(FakeClient(), "ws-parallel", 0, dispatcher=FakeDispatcher()) is True
    assert [call["path"] for call in calls] == ["/api/sql-worker/tasks/claim"]
    assert started == [(worker.TASK_KIND_SQL, "SQL-1")]


def test_workstation_worker_assigns_separate_onebss_state_slots(monkeypatch: pytest.MonkeyPatch, tmp_path) -> None:
    from scripts import onebss_workstation_worker as worker

    started = []
    release = threading.Event()

    def fake_run_task(self, kind, task, run_id):
        started.append((kind, run_id, dict(task)))
        release.wait(timeout=2)
        self.tracker.finish(run_id)

    monkeypatch.setenv("VNPTCTO_WORKER_MAX_CONCURRENT_TASKS", "4")
    monkeypatch.setenv("ONEBSS_WORKER_MAX_ONEBSS_TASKS", "2")
    monkeypatch.setenv("VNPTCTO_WORKSTATION_ROOT", str(tmp_path))
    monkeypatch.setattr(worker.WorkerTaskDispatcher, "_run_task", fake_run_task)

    dispatcher = worker.WorkerTaskDispatcher("https://vnptcto.com", {"Authorization": "Bearer test"}, "ws-parallel", 0)
    assert dispatcher.start_task(worker.TASK_KIND_ONEBSS, {"run_id": "ONEBSS-1", "report": {"ma_bao_cao": "BC1"}}) is True
    deadline = time.monotonic() + 2
    while len(started) < 1 and time.monotonic() < deadline:
        time.sleep(0.01)
    assert dispatcher.start_task(worker.TASK_KIND_ONEBSS, {"run_id": "ONEBSS-2", "report": {"ma_bao_cao": "BC2"}}) is True
    deadline = time.monotonic() + 2
    while len(started) < 2 and time.monotonic() < deadline:
        time.sleep(0.01)
    release.set()
    dispatcher.wait_until_idle()

    started_by_run = {run_id: task for _, run_id, task in started}
    assert started_by_run["ONEBSS-1"]["_worker_slot"] == 1
    assert started_by_run["ONEBSS-2"]["_worker_slot"] == 2
    state_paths = [started_by_run["ONEBSS-1"]["_worker_state_path"], started_by_run["ONEBSS-2"]["_worker_state_path"]]
    assert state_paths[0] != state_paths[1]
    assert state_paths[0].endswith("ws-parallel-slot-1.json")
    assert state_paths[1].endswith("ws-parallel-slot-2.json")


def test_workstation_worker_can_skip_slow_secondary_claims() -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append(path)

            class FakeResponse:
                status_code = 200

                def raise_for_status(self) -> None:
                    return None

                def json(self) -> dict:
                    return {"ok": True, "task": None}

            return FakeResponse()

    assert worker.poll_worker_once(FakeClient(), "ws-idle", 0, include_sql=False, include_ftp=False) is False
    assert calls == ["/api/onebss-worker/tasks/claim"]


def test_sql_worker_forwards_run_id_to_local_api(monkeypatch: pytest.MonkeyPatch) -> None:
    from scripts import onebss_workstation_worker as worker

    posts = []

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {"ok": True, "rows": [], "total": 0}

    class FakeInternalClient:
        def __init__(self, *args, **kwargs) -> None:
            return None

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb) -> None:
            return None

        def post(self, api_url: str, **kwargs):
            posts.append({"api_url": api_url, "json": kwargs.get("json") or {}})
            return FakeResponse()

    monkeypatch.setattr(worker, "internal_sql_api_urls", lambda: ["http://127.0.0.1:8000/api/du-lieu-web"])
    monkeypatch.setattr(worker.httpx, "Client", FakeInternalClient)

    result = worker.run_sql_worker_query({
        "run_id": "SQL-RUN-001",
        "query": {"action": "export_sql_report_to_drive", "file_name": "crs.xlsx"},
    })

    assert result["ok"] is True
    assert posts[0]["json"]["job_id"] == "SQL-RUN-001"
    assert posts[0]["json"]["run_id"] == "SQL-RUN-001"
    assert posts[0]["json"]["worker_task_id"] == "SQL-RUN-001"


def test_workstation_worker_ftp_claim_transient_returns_to_poll(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []

    def fake_request_json(client, method: str, path: str, **kwargs):
        calls.append({"method": method, "path": path, "kwargs": dict(kwargs)})
        if path == "/api/ftp-worker/tasks/claim":
            return {"ok": False, "task": None, "transient_error": "read timeout"}
        return {"ok": True, "task": None}

    monkeypatch.setattr(worker, "request_json", fake_request_json)

    assert worker.poll_worker_once(object(), "ws-idle", 0) is False
    ftp_claim = [call for call in calls if call["path"] == "/api/ftp-worker/tasks/claim"][0]
    assert ftp_claim["kwargs"]["timeout"] == 10.0
    assert ftp_claim["kwargs"]["_retry_forever"] is False


def test_workstation_worker_main_continues_after_poll_exception(monkeypatch, capsys) -> None:
    from scripts import onebss_workstation_worker as worker

    monkeypatch.setattr(worker.sys, "argv", ["onebss_workstation_worker.py", "--token", "token", "--once"])
    monkeypatch.setattr(worker, "send_heartbeat", lambda *args, **kwargs: None)
    monkeypatch.setattr(worker.time, "sleep", lambda seconds: None)

    def fail_poll(*args, **kwargs):
        raise RuntimeError("boom")

    monkeypatch.setattr(worker, "poll_worker_once", fail_poll)

    assert worker.main() == 0
    assert "Vong lap worker loi" in capsys.readouterr().err


def test_sql_worker_posts_drive_export_result_to_web(monkeypatch) -> None:
    from scripts import onebss_workstation_worker as worker

    calls = []

    class FakeResponse:
        def raise_for_status(self) -> None:
            return None

        def json(self) -> dict:
            return {"ok": True}

    class FakeClient:
        def request(self, method: str, path: str, **kwargs):
            calls.append({"method": method, "path": path, "json": kwargs.get("json") or {}})
            return FakeResponse()

    monkeypatch.setattr(
        worker,
        "run_sql_worker_query",
        lambda task: {
            "ok": True,
            "columns": ["MA_TB"],
            "rows": 145433,
            "total": 145433,
            "message": "Da upload Drive",
            "file_id": "drive-file-001",
            "file_name": "crs_export.xlsx",
            "drive_url": "https://drive.google.com/file/d/drive-file-001/view",
        },
    )
    monkeypatch.setattr(worker, "run_sql_worker_query_cancellable", lambda task, progress_callback: worker.run_sql_worker_query(task))

    worker.process_sql_task(
        FakeClient(),
        {
            "run_id": "SQL-EXPORT-1",
            "report_code": "CRS",
            "query": {"action": "export_sql_report_to_drive", "file_name": "crs_export.xlsx"},
        },
        "ws-sql",
    )

    result_calls = [call for call in calls if call["path"] == "/api/sql-worker/tasks/SQL-EXPORT-1/result"]
    assert len(result_calls) == 1
    status_messages = [call["json"]["message"] for call in calls if call["path"] == "/api/sql-worker/tasks/SQL-EXPORT-1/status"]
    assert any("Dang ket noi Oracle noi bo" in message for message in status_messages)
    assert any("Dang cap nhat file va link Drive len web" in message for message in status_messages)
    payload = result_calls[0]["json"]
    assert payload["ok"] is True
    assert payload["rows"] == []
    assert payload["pagination"]["total"] == 145433
    assert payload["drive_url"] == "https://drive.google.com/file/d/drive-file-001/view"
    assert payload["file_name"] == "crs_export.xlsx"
    assert payload["total"] == 145433


def test_onebss_worker_retries_transient_file_upload(monkeypatch, tmp_path) -> None:
    import httpx
    from scripts import onebss_workstation_worker as worker

    attempts = {"count": 0}
    source = tmp_path / "result.xlsx"
    source.write_bytes(b"xlsx")
    monkeypatch.setattr(worker.time, "sleep", lambda seconds: None)

    class FakeResponse:
        def __init__(self, status_code: int, payload: dict | None = None) -> None:
            self.status_code = status_code
            self.payload = payload or {}
            self.request = httpx.Request("POST", "https://vnptcto.com/api/onebss-worker/tasks/run-1/file")

        def raise_for_status(self) -> None:
            if self.status_code >= 400:
                response = httpx.Response(self.status_code, request=self.request)
                raise httpx.HTTPStatusError("temporary upload error", request=self.request, response=response)

        def json(self) -> dict:
            return self.payload

    class FakeClient:
        def post(self, path: str, **kwargs):
            attempts["count"] += 1
            assert path == "/api/onebss-worker/tasks/run-1/file"
            assert "files" in kwargs
            if attempts["count"] == 1:
                return FakeResponse(502)
            return FakeResponse(200, {"ok": True, "file": {"file_name": "result.xlsx", "storage_status": "uploaded_worker_file"}})

    uploaded = worker.upload_task_file(FakeClient(), "/api/onebss-worker/tasks/run-1/file", str(source))

    assert attempts["count"] == 2
    assert uploaded["file_name"] == "result.xlsx"


def test_onebss_worker_uploads_result_file_for_download(monkeypatch, tmp_path) -> None:
    settings = get_settings().model_copy(update={"data_mining_download_dir": str(tmp_path)})
    monkeypatch.setattr(routes, "get_settings", lambda: settings)
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings_arg, storage_link="", repository=None: "")
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS worker upload",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_UPLOAD&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        headers = {"Authorization": "Bearer test-worker-token"}
        upload = client.post(
            f"/api/onebss-worker/tasks/{job_id}/file",
            files={"file": ("result.zip", b"zip-bytes", "application/zip")},
            headers=headers,
        )
        assert upload.status_code == 200
        uploaded = upload.json()["file"]
        assert uploaded["file_name"] == "result.zip"
        assert Path(uploaded["file_path"]).read_bytes() == b"zip-bytes"

        finished = client.post(
            f"/api/onebss-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "Da gui file ve web.",
                "file_name": uploaded["file_name"],
                "file_path": uploaded["file_path"],
                "storage_status": uploaded["storage_status"],
            },
            headers=headers,
        )
        assert finished.status_code == 200
        assert finished.json()["run"]["download_url"]

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        download = client.get(runs[0]["download_url"])
        assert download.status_code == 200
        assert download.content == b"zip-bytes"


def test_onebss_worker_web_upload_uses_global_drive_folder(monkeypatch, tmp_path) -> None:
    settings = get_settings().model_copy(
        update={
            "data_mining_download_dir": str(tmp_path),
            "google_drive_folder_id": "drive-folder-global",
        }
    )
    saved_calls = []

    def fake_save_downloaded_file(settings_arg, source_file, storage_link, repository=None):
        saved_calls.append((str(source_file), storage_link))
        return {
            "ok": True,
            "message": "Da upload vao thu muc Drive chung.",
            "storage_link": "https://drive.google.com/file/d/global-drive-file/view",
            "storage_status": "uploaded_google_drive:global-drive-file",
        }

    monkeypatch.setattr(routes, "get_settings", lambda: settings)
    monkeypatch.setattr(routes, "save_downloaded_file", fake_save_downloaded_file)
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS global Drive upload",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_GLOBAL_DRIVE&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        upload = client.post(
            f"/api/onebss-worker/tasks/{job_id}/file",
            files={"file": ("result.xlsx", b"xlsx-bytes", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers={"Authorization": "Bearer test-worker-token"},
        )

        assert upload.status_code == 200
        uploaded = upload.json()["file"]
        assert uploaded["storage_link"] == "https://drive.google.com/file/d/global-drive-file/view"
        assert uploaded["storage_status"] == "uploaded_google_drive:global-drive-file"
        assert saved_calls and saved_calls[0][1] == "drive-folder-global"


def test_onebss_worker_result_preserves_uploaded_web_file(monkeypatch, tmp_path) -> None:
    settings = get_settings().model_copy(update={"data_mining_download_dir": str(tmp_path)})
    monkeypatch.setattr(routes, "get_settings", lambda: settings)
    monkeypatch.setattr(routes, "google_drive_folder_id", lambda settings_arg, storage_link="", repository=None: "")
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS preserve web file",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_PRESERVE&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        headers = {"Authorization": "Bearer test-worker-token"}
        upload = client.post(
            f"/api/onebss-worker/tasks/{job_id}/file",
            files={"file": ("result.xlsx", b"xlsx-bytes", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )
        assert upload.status_code == 200

        finished = client.post(
            f"/api/onebss-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "Worker finished after web upload.",
                "file_name": "",
                "file_path": "",
                "storage_status": "",
            },
            headers=headers,
        )
        assert finished.status_code == 200
        run = finished.json()["run"]
        assert run["file_name"] == "result.xlsx"
        assert run["storage_status"] == "uploaded_worker_file"
        assert run["download_url"]
        download = client.get(run["download_url"])
        assert download.status_code == 200
        assert download.content == b"xlsx-bytes"


def test_onebss_worker_drive_link_does_not_expose_missing_local_download() -> None:
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS Drive link",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_DRIVE_LINK&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        finished = client.post(
            f"/api/onebss-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "Da upload file len Google Drive.",
                "file_name": "result.xlsx",
                "file_path": "C:/VNPTCTO/onebss/result.xlsx",
                "storage_link": "https://drive.google.com/open?id=drive-file-002",
                "storage_status": "uploaded_google_drive",
            },
            headers={"Authorization": "Bearer test-worker-token"},
        )
        assert finished.status_code == 200
        run = finished.json()["run"]
        assert run["storage_link"] == "https://drive.google.com/open?id=drive-file-002"
        assert run["file_url"] == "https://drive.google.com/open?id=drive-file-002"
        assert "download_url" not in run

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["storage_link"] == "https://drive.google.com/open?id=drive-file-002"
        assert runs[0]["file_url"] == "https://drive.google.com/open?id=drive-file-002"
        assert "download_url" not in runs[0]


def test_onebss_run_derives_drive_link_from_storage_status() -> None:
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/onebss-reports",
            json={
                "ten_bao_cao": "OneBSS Drive status only",
                "danh_sach_bien": ["P_TUNGAY"],
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_DRIVE_STATUS&name=Test",
                "storage_link": "",
            },
        )
        assert created.status_code == 200
        code = created.json()["ma_bao_cao"]

        response = client.post("/api/onebss-reports/run", json={"ma_bao_cao": code, "parameters": {"P_TUNGAY": "01/07/2026"}})
        assert response.status_code == 200
        job_id = response.json()["job_id"]
        finished = client.post(
            f"/api/onebss-worker/tasks/{job_id}/result",
            json={
                "ok": True,
                "status": "success",
                "message": "Da upload file len Google Drive.",
                "file_name": "result.xlsx",
                "file_path": "C:/VNPTCTO/onebss/result.xlsx",
                "storage_link": "",
                "storage_status": "uploaded_google_drive:driveFile_003-ABC",
            },
            headers={"Authorization": "Bearer test-worker-token"},
        )
        assert finished.status_code == 200
        run = finished.json()["run"]
        assert run["storage_link"] == "https://drive.google.com/file/d/driveFile_003-ABC/view"
        assert run["file_url"] == "https://drive.google.com/file/d/driveFile_003-ABC/view"
        assert "download_url" not in run

        runs = client.get(f"/api/onebss-reports/runs?ma_bao_cao={code}").json()["runs"]
        assert len(runs) == 1
        assert runs[0]["file_url"] == "https://drive.google.com/file/d/driveFile_003-ABC/view"


def test_supabase_requests_reuse_shared_http_client(monkeypatch) -> None:
    from app.data_access import supabase_repository as supabase_module

    class FakeResponse:
        status_code = 200
        text = "[]"

        @staticmethod
        def json():
            return []

    class FakeClient:
        def __init__(self) -> None:
            self.calls = []

        def request(self, method, url, **kwargs):
            self.calls.append((method, url, kwargs))
            return FakeResponse()

    client = FakeClient()
    monkeypatch.setattr(supabase_module, "SUPABASE_HTTP_CLIENT", client)
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    repository._request("GET", "mobile_devices")
    repository._request("GET", "mobile_sms_messages")

    assert len(client.calls) == 2
    assert client.calls[0][1].endswith("/mobile_devices")
    assert client.calls[1][1].endswith("/mobile_sms_messages")


def test_supabase_error_message_keeps_column_name_after_long_details(monkeypatch) -> None:
    from app.data_access import supabase_repository as supabase_module

    class FakeResponse:
        status_code = 400

        def __init__(self) -> None:
            self.body = {
                "code": "23502",
                "details": f"Failing row contains ({'x' * 900})",
                "hint": None,
                "message": 'null value in column "finished_at" of relation "onebss_report_runs" violates not-null constraint',
            }
            self.text = json.dumps(self.body)

        def json(self):
            return self.body

    class FakeClient:
        def request(self, method, url, **kwargs):
            return FakeResponse()

    monkeypatch.setattr(supabase_module, "SUPABASE_HTTP_CLIENT", FakeClient())
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    with pytest.raises(RuntimeError) as exc_info:
        repository._request("POST", "onebss_report_runs", json={"finished_at": None})

    text = str(exc_info.value)
    assert "23502" in text
    assert "finished_at" in text
    assert text.index("message") < text.index("details")


def test_supabase_onebss_run_retries_formatted_not_null_rest_error(monkeypatch) -> None:
    from app.data_access import supabase_repository as supabase_module

    class FakeResponse:
        def __init__(self, status_code, body) -> None:
            self.status_code = status_code
            self.body = body
            self.text = json.dumps(body)

        def json(self):
            return self.body

    class FakeClient:
        def __init__(self) -> None:
            self.payloads = []

        def request(self, method, url, **kwargs):
            payload = kwargs.get("json") or {}
            self.payloads.append(dict(payload))
            if len(self.payloads) == 1:
                return FakeResponse(
                    400,
                    {
                        "code": "23502",
                        "details": f"Failing row contains ({'x' * 900})",
                        "message": 'null value in column "finished_at" of relation "onebss_report_runs" violates not-null constraint',
                    },
                )
            return FakeResponse(201, [payload])

    client = FakeClient()
    monkeypatch.setattr(supabase_module, "SUPABASE_HTTP_CLIENT", client)
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    run = repository.save_onebss_report_run({
        "ma_bao_cao": "PTM_CHUAHT_NGAYYC",
        "ten_bao_cao": "Phat trien moi",
        "status": "queued",
        "parameters": {"P_TUNGAY": "{{month_start}}"},
    })

    assert len(client.payloads) == 2
    assert client.payloads[0]["finished_at"] is None
    assert client.payloads[1]["finished_at"] == client.payloads[0]["started_at"]
    assert run["status"] == "queued"


def test_supabase_onebss_run_uses_parameters_json_column(monkeypatch) -> None:
    captured = {}
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    def fake_insert(table, payload):
        captured["table"] = table
        captured["payload"] = payload
        return payload

    monkeypatch.setattr(repository, "_insert", fake_insert)
    run = repository.save_onebss_report_run({
        "ma_bao_cao": "TEST",
        "ten_bao_cao": "Test",
        "status": "failed",
        "parameters": {"P_TUNGAY": "01/07/2026"},
    })
    assert captured["table"] == "onebss_report_runs"
    assert "parameters_json" in captured["payload"]
    assert "parameters" not in captured["payload"]
    assert run["parameters"] == {"P_TUNGAY": "01/07/2026"}


def test_supabase_onebss_queued_run_sends_null_optional_timestamps(monkeypatch) -> None:
    captured = {}
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    def fake_insert(table, payload):
        captured["table"] = table
        captured["payload"] = payload
        return payload

    monkeypatch.setattr(repository, "_insert", fake_insert)
    run = repository.save_onebss_report_run({
        "ma_bao_cao": "TEST",
        "ten_bao_cao": "Test",
        "status": "queued",
        "parameters": {"P_TUNGAY": "01/07/2026"},
    })

    assert captured["table"] == "onebss_report_runs"
    assert captured["payload"]["finished_at"] is None
    assert captured["payload"]["claimed_at"] is None
    assert run["finished_at"] == ""
    assert run["claimed_at"] == ""


def test_supabase_onebss_run_retries_legacy_not_null_timestamps(monkeypatch) -> None:
    attempts = []
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    def fake_insert(table, payload):
        attempts.append((table, dict(payload)))
        if len(attempts) == 1:
            raise RuntimeError('Supabase REST loi 400: {"code":"23502","message":"null value in column \\"finished_at\\" of relation \\"onebss_report_runs\\" violates not-null constraint"}')
        if len(attempts) == 2:
            raise RuntimeError('Supabase REST loi 400: {"code":"23502","message":"null value in column \\"claimed_at\\" of relation \\"onebss_report_runs\\" violates not-null constraint"}')
        if len(attempts) == 3:
            raise RuntimeError('Supabase REST loi 400: {"code":"23502","message":"null value in column \\"created_at\\" of relation \\"onebss_report_runs\\" violates not-null constraint"}')
        return payload

    monkeypatch.setattr(repository, "_insert", fake_insert)
    run = repository.save_onebss_report_run({
        "ma_bao_cao": "TEST",
        "ten_bao_cao": "Test",
        "status": "queued",
        "parameters": {"P_TUNGAY": "{{month_start}}"},
    })

    assert [item[0] for item in attempts] == ["onebss_report_runs"] * 4
    started_at = attempts[0][1]["started_at"]
    assert attempts[0][1]["finished_at"] is None
    assert attempts[1][1]["finished_at"] == started_at
    assert attempts[2][1]["claimed_at"] == started_at
    assert attempts[3][1]["created_at"] == started_at
    assert run["status"] == "queued"
    assert run["finished_at"] == started_at
    assert run["claimed_at"] == started_at


def test_supabase_ftp_report_run_uses_ftp_table(monkeypatch) -> None:
    captured = {}
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    def fake_insert(table, payload):
        captured["table"] = table
        captured["payload"] = payload
        return payload

    monkeypatch.setattr(repository, "_insert", fake_insert)
    run = repository.save_ftp_report_run({
        "run_id": "FTP-RUN-001",
        "ma_bao_cao": "FTP0001",
        "ten_bao_cao": "FTP test",
        "status": "queued",
        "folder_path": "/reports",
        "file_name_template": "test_{yyyymmdd}.xlsx",
    })
    assert captured["table"] == "ftp_report_runs"
    assert captured["payload"]["run_id"] == "FTP-RUN-001"
    assert captured["payload"]["folder_path"] == "/reports"
    assert captured["payload"]["finished_at"] is None
    assert captured["payload"]["claimed_at"] is None
    assert run["file_name_template"] == "test_{yyyymmdd}.xlsx"


def test_supabase_ftp_run_retries_legacy_not_null_finished_at(monkeypatch) -> None:
    attempts = []
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    def fake_insert(table, payload):
        attempts.append((table, dict(payload)))
        if len(attempts) == 1:
            raise RuntimeError('Supabase REST loi 400: {"code":"23502","message":"null value in column \\"finished_at\\" of relation \\"ftp_report_runs\\" violates not-null constraint"}')
        return payload

    monkeypatch.setattr(repository, "_insert", fake_insert)
    run = repository.save_ftp_report_run({
        "run_id": "FTP-RUN-001",
        "ma_bao_cao": "FTP0001",
        "ten_bao_cao": "FTP test",
        "status": "queued",
    })

    assert [item[0] for item in attempts] == ["ftp_report_runs", "ftp_report_runs"]
    assert attempts[0][1]["finished_at"] is None
    assert attempts[1][1]["finished_at"] == attempts[0][1]["started_at"]
    assert run["finished_at"] == attempts[0][1]["started_at"]


def test_supabase_task_report_auto_run_retries_legacy_not_null_finished_at(monkeypatch) -> None:
    attempts = []
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    def fake_insert(table, payload):
        attempts.append((table, dict(payload)))
        if len(attempts) == 1:
            raise RuntimeError('Supabase REST loi 400: {"code":"23502","message":"null value in column \\"finished_at\\" of relation \\"task_report_auto_runs\\" violates not-null constraint"}')
        return payload

    monkeypatch.setattr(repository, "get_task_report_auto_task", lambda task_id: {"task_id": task_id, "source_type": "onebss"})
    monkeypatch.setattr(repository, "_insert", fake_insert)

    run = repository.create_task_report_auto_run(
        "TASK01",
        "2026-08-14",
        status="queued",
        message="Cho chay",
        created_by="admin",
    )

    assert [item[0] for item in attempts] == ["task_report_auto_runs", "task_report_auto_runs"]
    assert attempts[0][1]["finished_at"] is None
    assert attempts[1][1]["finished_at"] == attempts[0][1]["started_at"]
    assert run["status"] == "queued"


def test_supabase_run_updates_send_null_for_blank_optional_timestamps(monkeypatch) -> None:
    patches = []
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")

    def fake_patch(table, params, payload):
        patches.append((table, params, payload))

    def fake_get(table, params):
        if table == "onebss_report_runs":
            return [{"run_id": "ONEBSS-RUN-001", "parameters_json": {}, "updated_at": patches[-1][2]["updated_at"]}]
        if table == "ftp_report_runs":
            return [{"run_id": "FTP-RUN-001", "updated_at": patches[-1][2]["updated_at"]}]
        return []

    monkeypatch.setattr(repository, "_patch", fake_patch)
    monkeypatch.setattr(repository, "_get", fake_get)

    onebss = repository.update_onebss_report_run("ONEBSS-RUN-001", {"finished_at": "", "claimed_at": "", "updated_at": ""})
    ftp = repository.update_ftp_report_run("FTP-RUN-001", {"finished_at": "", "claimed_at": "", "updated_at": ""})

    assert patches[0][0] == "onebss_report_runs"
    assert patches[0][2]["finished_at"] is None
    assert patches[0][2]["claimed_at"] is None
    assert patches[0][2]["updated_at"]
    assert onebss["finished_at"] == ""
    assert onebss["claimed_at"] == ""
    assert patches[1][0] == "ftp_report_runs"
    assert patches[1][2]["finished_at"] is None
    assert patches[1][2]["claimed_at"] is None
    assert patches[1][2]["updated_at"]
    assert ftp["finished_at"] == ""
    assert ftp["claimed_at"] == ""


def test_supabase_ftp_fallback_uses_protected_connection_config(monkeypatch) -> None:
    from app.data_access.supabase_repository import FTP_FALLBACK_STORE_KEY

    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")
    connection = {
        "id": 1,
        "code": "ftp_storage",
        "name": "FTP",
        "connection_type": "ftp",
        "description": "FTP fallback",
        "config": {"host": "10.159.23.100", "username": "thangph.cto", "password": "$Phthang125125"},
        "is_active": True,
    }

    def missing_table(*args, **kwargs):
        raise RuntimeError('Supabase REST loi 404: {"code":"PGRST205","message":"Could not find the table public.ftp_reports in the schema cache"}')

    def fake_get_connection(code):
        assert code == "ftp_storage"
        return connection

    def fake_upsert_connection(code, name, connection_type, description, config, is_active):
        connection.update({
            "code": code,
            "name": name,
            "connection_type": connection_type,
            "description": description,
            "config": config,
            "is_active": is_active,
        })
        return 1

    monkeypatch.setattr(repository, "_get", missing_table)
    monkeypatch.setattr(repository, "_insert", missing_table)
    monkeypatch.setattr(repository, "_patch", missing_table)
    monkeypatch.setattr(repository, "_delete", missing_table)
    monkeypatch.setattr(repository, "get_system_connection_by_code", fake_get_connection)
    monkeypatch.setattr(repository, "upsert_system_connection", fake_upsert_connection)

    assert repository.generate_ftp_report_code() == "FTP0001"
    report_id = repository.save_ftp_report(None, "FTP0001", "FTP fallback", "/reports", "file_{yyyymmdd}.xlsx")
    assert report_id == 1
    reports = repository.list_ftp_reports()
    assert reports[0]["ma_bao_cao"] == "FTP0001"
    assert repository.get_ftp_report_by_code("FTP0001")["folder_path"] == "/reports"

    public_config, protected_keys = routes.public_connection_config(connection["config"])
    assert FTP_FALLBACK_STORE_KEY not in public_config
    assert FTP_FALLBACK_STORE_KEY in protected_keys

    run = repository.save_ftp_report_run({
        "run_id": "FTP-RUN-FALLBACK",
        "ma_bao_cao": "FTP0001",
        "ten_bao_cao": "FTP fallback",
        "status": "queued",
        "folder_path": "/reports",
        "file_name_template": "file_{yyyymmdd}.xlsx",
    })
    assert run["status"] == "queued"
    claimed = repository.claim_next_ftp_report_run("ws-fallback")
    assert claimed["run_id"] == "FTP-RUN-FALLBACK"
    assert claimed["status"] == "running"
    assert claimed["worker_id"] == "ws-fallback"

    updated = repository.update_ftp_report_run("FTP-RUN-FALLBACK", {"status": "success", "file_name": "file.xlsx"})
    assert updated["status"] == "success"
    assert repository.list_ftp_report_runs("FTP0001")[0]["file_name"] == "file.xlsx"
    assert repository.clear_ftp_report_runs("FTP0001") == 1
    assert repository.list_ftp_report_runs("FTP0001") == []


def test_supabase_onebss_report_save_falls_back_without_otp_service_code(monkeypatch) -> None:
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")
    payloads = []

    def fake_insert(table, payload):
        payloads.append(payload)
        if "otp_service_code" in payload:
            raise RuntimeError(
                'Supabase REST loi 400: {"code":"PGRST204","message":"Could not find the '
                "'otp_service_code' column of 'onebss_reports' in the schema cache\"}"
            )
        return {"id": 77, **payload}

    monkeypatch.setattr(repository, "_insert", fake_insert)
    report_id = repository.save_onebss_report(
        None,
        "MYTV_KTT",
        "DS MyTV",
        ["p_phanvung_id"],
        {"p_phanvung_id": {"$each": ["13", "47", "66"]}},
        "https://onebss.vnpt.vn/#/report/bi?path=TEST&name=Test",
        "https://drive.google.com/drive/folders/test",
        "onebss",
    )

    assert report_id == 77
    assert "otp_service_code" in payloads[0]
    assert "otp_service_code" not in payloads[1]
    assert payloads[1]["parameters"]["p_phanvung_id"]["$each"] == ["13", "47", "66"]


def test_supabase_clear_onebss_report_runs_uses_run_id(monkeypatch) -> None:
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")
    calls = []

    def fake_get(table, params):
        calls.append(("get", table, params))
        return [{"run_id": "RUN1"}, {"run_id": "RUN2"}]

    def fake_delete(table, params):
        calls.append(("delete", table, params))

    monkeypatch.setattr(repository, "_get", fake_get)
    monkeypatch.setattr(repository, "_delete", fake_delete)

    assert repository.clear_onebss_report_runs("ONEBSS01") == 2
    assert calls[0] == ("get", "onebss_report_runs", {"select": "run_id", "ma_bao_cao": "eq.ONEBSS01"})
    assert calls[1] == ("delete", "onebss_report_runs", {"ma_bao_cao": "eq.ONEBSS01"})

    calls.clear()
    assert repository.clear_onebss_report_runs() == 2
    assert calls[0] == ("get", "onebss_report_runs", {"select": "run_id"})
    assert calls[1] == ("delete", "onebss_report_runs", {"run_id": "not.is.null"})


def test_supabase_clear_ftp_report_runs_uses_run_id(monkeypatch) -> None:
    repository = SupabaseRepository("https://example.supabase.co/rest/v1", "secret")
    calls = []

    def fake_get(table, params):
        calls.append(("get", table, params))
        return [{"run_id": "FTP1"}, {"run_id": "FTP2"}]

    def fake_delete(table, params):
        calls.append(("delete", table, params))

    monkeypatch.setattr(repository, "_get", fake_get)
    monkeypatch.setattr(repository, "_delete", fake_delete)

    assert repository.clear_ftp_report_runs("FTP0001") == 2
    assert calls[0] == ("get", "ftp_report_runs", {"select": "run_id", "ma_bao_cao": "eq.FTP0001"})
    assert calls[1] == ("delete", "ftp_report_runs", {"ma_bao_cao": "eq.FTP0001"})

    calls.clear()
    assert repository.clear_ftp_report_runs() == 2
    assert calls[0] == ("get", "ftp_report_runs", {"select": "run_id"})
    assert calls[1] == ("delete", "ftp_report_runs", {"run_id": "not.is.null"})


def test_dynamic_report_expands_comma_values_for_in_bind_params() -> None:
    with TestClient(app) as client:
        login(client)
        payload = {
            "ten_bao_cao": "Loai hinh IN test",
            "ma_bao_cao": "BC_TEST_LOAIHINH_IN",
            "cau_lenh_sql": "SELECT COUNT(*) AS thuebao FROM V_CHITIET_PTM WHERE loaitb_id in (:LOAIHINH);",
            "cac_tham_so": ["LOAIHINH"],
        }
        assert client.post("/api/admin/sql-reports", json=payload).status_code == 200

        result = client.post(
            "/api/reports/run",
            json={"ma_bao_cao": "BC_TEST_LOAIHINH_IN", "filters": {"LOAIHINH": "61,171,271"}, "page": 1, "page_size": 20},
        )

        assert result.status_code == 200
        assert result.json()["rows"][0]["THAM_SO"] == "LOAIHINH_1=61, LOAIHINH_2=171, LOAIHINH_3=271"


def test_define_sql_is_compiled_with_raw_filter_values() -> None:
    sql = """
DEFINE p_loaihinh = :LOAIHINH
DEFINE p_thang = :MONTH
DEFINE p_donvi = :DONVI
SELECT *
FROM css_cto.db_thuebao
WHERE loaitb_id = '&p_loaihinh'
  AND ('&p_thang' IS NULL OR '&p_thang' = '')
  AND ten_donvi_cha LIKE '&p_donvi';
"""
    compiled, details = DatabaseService._compile_define_sql(
        sql,
        {"LOAIHINH": "58", "MONTH ": "", "DONVI": "VNPT%"},
    )

    assert "DEFINE" not in compiled.upper()
    assert "loaitb_id = '58'" in compiled
    assert "LIKE 'VNPT%'" in compiled
    assert "'' IS NULL OR '' = ''" in compiled
    assert not compiled.endswith(";")
    assert details["define_params"] == ["p_loaihinh", "p_thang", "p_donvi"]
    assert DatabaseService._filters_for_compiled_sql(compiled, {"LOAIHINH": "58", "MONTH": "", "DONVI": "VNPT%"}) == {}


def test_compiled_sql_keeps_only_remaining_bind_params() -> None:
    sql = "SELECT * FROM css_cto.db_thuebao WHERE ngay >= :FROM_DATE AND ten_donvi_cha LIKE '&p_donvi';"
    params = {"FROM_DATE": "2026-05-01", "DONVI": "VNPT%"}

    assert DatabaseService._filters_for_compiled_sql(sql, params) == {"FROM_DATE": "2026-05-01"}


def test_in_bind_param_expands_comma_values() -> None:
    sql = "SELECT COUNT(*) FROM V_CHITIET_PTM WHERE loaitb_id in (:LOAIHINH) AND trangthaitb_id = :STATUS;"
    expanded_sql, filters = DatabaseService._expand_in_list_bind_params(
        sql,
        {"LOAIHINH": "61,171,271", "STATUS": "1"},
    )

    assert "loaitb_id IN (:LOAIHINH_1, :LOAIHINH_2, :LOAIHINH_3)" in expanded_sql
    assert DatabaseService._filters_for_compiled_sql(expanded_sql, filters) == {
        "LOAIHINH_1": "61",
        "LOAIHINH_2": "171",
        "LOAIHINH_3": "271",
        "STATUS": "1",
    }


def test_admin_can_manage_dashboard_layout_and_lazy_load_tab_data(monkeypatch) -> None:
    with TestClient(app) as client:
        login(client)
        report_payload = {
            "ten_bao_cao": "Báo cáo Builder test",
            "ma_bao_cao": "BC_BUILDER_TEST",
            "cau_lenh_sql": "SELECT don_vi, so_luong FROM css_cto.builder_test WHERE trang_thai = :status;",
            "cac_tham_so": ["status"],
        }
        assert client.post("/api/admin/sql-reports", json=report_payload).status_code == 200

        layout_payload = {
            "page_id": "DASHBOARD_TEST_BUILDER",
            "page_name": "Dashboard Test Builder",
            "layout": {
                "page_id": "DASHBOARD_TEST_BUILDER",
                "tabs": [
                    {
                        "tab_id": "tab_a",
                        "tab_name": "Tab A",
                        "order": 1,
                        "grid_layout": [
                            {
                                "row_id": 1,
                                "layout_type": "2_columns",
                                "widgets": [
                                    {
                                        "position": 1,
                                        "type": "bar_chart",
                                        "title": "Widget A",
                                        "sql_code": "BC_BUILDER_TEST",
                                        "filters": {"status": "1"},
                                        "chart_config": {"orientation": "horizontal", "label_column": "don_vi", "value_column": "so_luong"},
                                    }
                                ],
                            }
                        ],
                    },
                    {
                        "tab_id": "tab_b",
                        "tab_name": "Tab B",
                        "order": 2,
                        "grid_layout": [
                            {
                                "row_id": 1,
                                "layout_type": "1_column",
                                "widgets": [
                                    {
                                        "position": 1,
                                        "type": "text_title",
                                        "title": "Tiêu đề thiết kế",
                                        "text_content": "Nội dung giới thiệu tab",
                                    }
                                ],
                            },
                            {
                                "row_id": 2,
                                "layout_type": "3_columns",
                                "widgets": [
                                    {
                                        "position": 1,
                                        "type": "combo_chart",
                                        "title": "Biểu đồ kết hợp",
                                        "sql_code": "BC_BUILDER_TEST",
                                        "chart_config": {
                                            "label_column": "don_vi",
                                            "bar_column": "so_luong",
                                            "line_column": "ty_le",
                                        },
                                    },
                                    {
                                        "position": 2,
                                        "type": "data_card",
                                        "title": "Thẻ dữ liệu",
                                        "sql_code": "BC_BUILDER_TEST",
                                        "icon_url": "https://example.vn/icon.png",
                                        "text_content": "Ghi chú thẻ",
                                    },
                                    {
                                        "position": 3,
                                        "type": "total_data_card",
                                        "title": "Tổng hoàn thành",
                                        "sql_code": "BC_BUILDER_TEST",
                                        "chart_config": {
                                            "actual_column": "TH",
                                            "target_column": "KH",
                                            "completion_column": "TLHT",
                                        },
                                    },
                                ],
                            }
                        ],
                    },
                ],
            },
        }
        saved = client.post("/api/admin/dashboard-layouts", json=layout_payload)
        assert saved.status_code == 200
        saved_layout = saved.json()["layout"]
        assert saved_layout["tabs"][0]["tab_id"] == "tab_a"
        assert saved_layout["tabs"][0]["grid_layout"][0]["widgets"][0]["chart_config"]["orientation"] == "horizontal"
        assert saved_layout["tabs"][1]["grid_layout"][0]["layout_type"] == "1_column"
        assert saved_layout["tabs"][1]["grid_layout"][0]["widgets"][0]["type"] == "text_title"
        assert saved_layout["tabs"][1]["grid_layout"][1]["layout_type"] == "3_columns"
        assert saved_layout["tabs"][1]["grid_layout"][1]["widgets"][1]["icon_url"] == "https://example.vn/icon.png"
        assert saved_layout["tabs"][1]["grid_layout"][1]["widgets"][2]["type"] == "total_data_card"
        assert saved_layout["tabs"][1]["grid_layout"][1]["widgets"][2]["chart_config"]["completion_column"] == "TLHT"

        layouts = client.get("/api/admin/dashboard-layouts")
        assert layouts.status_code == 200
        assert any(item["page_id"] == "DASHBOARD_TEST_BUILDER" for item in layouts.json()["layouts"])

        pages = client.get("/api/admin/dashboard-layout-pages")
        assert pages.status_code == 200
        builder_page = next(page for page in pages.json()["pages"] if page["page_id"] == "DASHBOARD_TEST_BUILDER")
        assert builder_page["feature_code"] == "dashboardtestbuilder"
        assert builder_page["feature_name"] == "Dashboard Test Builder"
        assert builder_page["saved"] is True

        me = client.get("/api/auth/me")
        assert me.status_code == 200
        assert "dashboardtestbuilder" in me.json()["user"]["permissions"]

        tab_a = client.get("/api/admin/dashboard-layouts/DASHBOARD_TEST_BUILDER/tabs/tab_a/data")
        assert tab_a.status_code == 200
        tab_payload = tab_a.json()
        assert len(tab_payload["widgets"]) == 1
        assert tab_payload["widgets"][0]["sql_code"] == "BC_BUILDER_TEST"
        assert tab_payload["widgets"][0]["data"]["columns"] == ["STT", "MA_BAO_CAO", "TEN_BAO_CAO", "THAM_SO"]
        assert tab_payload["widgets"][0]["data"]["rows"][0]["THAM_SO"] == "status=1"

        api_calls = []
        original_run_sql_report = routes.InternalApiClient.run_sql_report

        def counting_run_sql_report(self, **kwargs):
            api_calls.append((kwargs["ma_bao_cao"], kwargs["tham_so"]))
            return original_run_sql_report(self, **kwargs)

        monkeypatch.setattr(routes.InternalApiClient, "run_sql_report", counting_run_sql_report)
        tab_b = client.get("/api/admin/dashboard-layouts/DASHBOARD_TEST_BUILDER/tabs/tab_b/data")
        assert tab_b.status_code == 200
        assert [widget["type"] for widget in tab_b.json()["widgets"]] == ["combo_chart", "data_card", "total_data_card"]
        assert api_calls == [("BC_BUILDER_TEST", {})]

        inverted_report_payload = {
            "ten_bao_cao": "Check_Job",
            "ma_bao_cao": "CHECK JOB DU LIEU",
            "cau_lenh_sql": "SELECT job_name, status FROM css_cto.check_job;",
            "cac_tham_so": [],
        }
        inverted_created = client.post("/api/admin/sql-reports", json=inverted_report_payload)
        assert inverted_created.status_code == 200
        inverted_report_id = inverted_created.json()["id"]
        table_layout_payload = {
            "page_id": "DASHBOARD_CHECK_JOB",
            "page_name": "CHECK_JOB",
            "layout": {
                "page_id": "DASHBOARD_CHECK_JOB",
                "tabs": [
                    {
                        "tab_id": "tab_check",
                        "tab_name": "Tab moi",
                        "order": 1,
                        "grid_layout": [
                            {
                                "row_id": 1,
                                "layout_type": "1_column",
                                "widgets": [
                                    {
                                        "position": 1,
                                        "type": "data_table",
                                        "title": "Check job",
                                        "sql_code": "Check_Job (CHECK JOB DU LIEU)",
                                    }
                                ],
                            }
                        ],
                    }
                ],
            },
        }
        saved_table_layout = client.post("/api/admin/dashboard-layouts", json=table_layout_payload)
        assert saved_table_layout.status_code == 200
        saved_widget = saved_table_layout.json()["layout"]["tabs"][0]["grid_layout"][0]["widgets"][0]
        assert saved_widget["sql_code"] == "CHECK_JOB"
        tab_check = client.get("/api/admin/dashboard-layouts/DASHBOARD_CHECK_JOB/tabs/tab_check/data")
        assert tab_check.status_code == 200
        tab_check_payload = tab_check.json()
        assert tab_check_payload["widgets"][0]["sql_code"] == "CHECK_JOB"
        assert tab_check_payload["widgets"][0]["data"]["ok"] is True
        assert api_calls[-1] == ("CHECK_JOB", {})
        calls_after_cache_fill = len(api_calls)
        tab_check_cached = client.get("/api/admin/dashboard-layouts/DASHBOARD_CHECK_JOB/tabs/tab_check/data")
        assert tab_check_cached.status_code == 200
        cached_data = tab_check_cached.json()["widgets"][0]["data"]
        assert cached_data["ok"] is True
        assert cached_data["details"]["dashboard_cache"]["hit"] is True
        assert len(api_calls) == calls_after_cache_fill
        updated_inverted_report_payload = {
            **inverted_report_payload,
            "id": inverted_report_id,
            "cau_lenh_sql": "SELECT job_name, status, run_time FROM css_cto.check_job;",
        }
        assert client.post("/api/admin/sql-reports", json=updated_inverted_report_payload).status_code == 200
        tab_check_after_sql_update = client.get("/api/admin/dashboard-layouts/DASHBOARD_CHECK_JOB/tabs/tab_check/data")
        assert tab_check_after_sql_update.status_code == 200
        assert len(api_calls) == calls_after_cache_fill + 1
        assert "dashboard_cache" not in tab_check_after_sql_update.json()["widgets"][0]["data"].get("details", {})
        refresh_result = DatabaseService(routes.InternalApiClient(routes.get_settings()), routes.build_app_repository()).refresh_dashboard_chart_cache(page_id="DASHBOARD_CHECK_JOB")
        assert refresh_result["deleted_stale"] == 0

        empty_check_layout_payload = {
            **table_layout_payload,
            "layout": {
                "page_id": "DASHBOARD_CHECK_JOB",
                "tabs": [
                    {
                        "tab_id": "tab_check",
                        "tab_name": "Tab moi",
                        "order": 1,
                        "grid_layout": [],
                    }
                ],
            },
        }
        assert client.post("/api/admin/dashboard-layouts", json=empty_check_layout_payload).status_code == 200
        refresh_after_delete = DatabaseService(routes.InternalApiClient(routes.get_settings()), routes.build_app_repository()).refresh_dashboard_chart_cache(page_id="DASHBOARD_CHECK_JOB")
        assert refresh_after_delete["deleted_stale"] == 1

        short_code_report_payload = {
            "ten_bao_cao": "Check_Job_Table",
            "ma_bao_cao": "CHECK",
            "cau_lenh_sql": "SELECT job_name, status FROM css_cto.check_job;",
            "cac_tham_so": [],
        }
        assert client.post("/api/admin/sql-reports", json=short_code_report_payload).status_code == 200
        legacy_layout_payload = {
            "page_id": "DASHBOARD_CHECK_JOB_LEGACY",
            "page_name": "CHECK_JOB",
            "layout": {
                "page_id": "DASHBOARD_CHECK_JOB_LEGACY",
                "tabs": [
                    {
                        "tab_id": "tab_check",
                        "tab_name": "DANH SACH JOB",
                        "order": 1,
                        "grid_layout": [
                            {
                                "row_id": 1,
                                "layout_type": "1_column",
                                "widgets": [
                                    {
                                        "position": 1,
                                        "type": "data_table",
                                        "title": "Check_Job_Table",
                                        "sql_code": "CHECK_JOB_TABLE",
                                    }
                                ],
                            }
                        ],
                    }
                ],
            },
        }
        assert client.post("/api/admin/dashboard-layouts", json=legacy_layout_payload).status_code == 200
        legacy_tab = client.get("/api/admin/dashboard-layouts/DASHBOARD_CHECK_JOB_LEGACY/tabs/tab_check/data")
        assert legacy_tab.status_code == 200
        legacy_payload = legacy_tab.json()
        assert legacy_payload["ok"] is True
        assert legacy_payload["widgets"][0]["data"]["ok"] is True
        assert api_calls[-1] == ("CHECK", {})

        group_payload = {
            "page_id": "DASHBOARD_EMPTY_GROUP",
            "page_name": "NhÃ³m Dashboard rá»—ng",
            "layout": {
                "page_id": "DASHBOARD_EMPTY_GROUP",
                "tabs": [
                    {
                        "tab_id": "tab_group",
                        "tab_name": "NhÃ³m",
                        "order": 1,
                        "grid_layout": [],
                    }
                ],
            },
        }
        group_saved = client.post("/api/admin/dashboard-layouts", json=group_payload)
        assert group_saved.status_code == 200
        assert group_saved.json()["layout"]["tabs"][0]["grid_layout"] == []
        group_tab = client.get("/api/admin/dashboard-layouts/DASHBOARD_EMPTY_GROUP/tabs/tab_group/data")
        assert group_tab.status_code == 200
        assert group_tab.json()["widgets"] == []

        fiber_report = {
            "ten_bao_cao": "Fiber PTM",
            "ma_bao_cao": "FIBER_PTM",
            "cau_lenh_sql": "SELECT * FROM css_cto.fiber WHERE loaihinh = :LOAIHINH AND ngay = :SYSDATE AND donvi LIKE :DONVI;",
            "cac_tham_so": ["LOAIHINH", "SYSDATE", "DONVI"],
        }
        assert client.post("/api/admin/sql-reports", json=fiber_report).status_code == 200
        result = client.post(
            "/api/reports/run",
            json={
                "ma_bao_cao": "FIBER_PTM",
                "filters": {"loaihinh": "58", "sysdate": "SYSDATE", "donvi": "VNPT%"},
                "page": 1,
                "page_size": 20,
            },
        )
        assert result.status_code == 200
        assert result.json()["rows"][0]["THAM_SO"] == "LOAIHINH=58, SYSDATE=SYSDATE, DONVI=VNPT%"


def test_dashboard_refresh_uses_isolated_worker_queue(monkeypatch: pytest.MonkeyPatch) -> None:
    settings = routes.get_settings().model_copy(update={"internal_api_mock_mode": False})
    monkeypatch.setattr(routes, "get_settings", lambda: settings)
    monkeypatch.setattr(routes, "_next_dynamic_report_export_worker_job", lambda *args, **kwargs: None)
    worker_id = f"ws-dashboard-{uuid.uuid4().hex[:8]}"
    page_id = f"DASHBOARD_QUEUE_{uuid.uuid4().hex[:8].upper()}"
    report_code = f"DASH_REFRESH_{uuid.uuid4().hex[:8].upper()}"
    widget_sql_code = f"{report_code}_ALIAS"
    legacy_job_id = "dashboard_legacy_" + uuid.uuid4().hex[:8]
    headers = {"Authorization": "Bearer test-worker-token"}

    with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
        routes.DYNAMIC_REPORT_RUN_JOBS.clear()
        routes.DYNAMIC_REPORT_RUN_JOBS[legacy_job_id] = {
            "job_id": legacy_job_id,
            "status": "queued_worker",
            "created_at": time.time(),
            "updated_at": time.time(),
            "payload": {"ma_bao_cao": "LEGACY_DASHBOARD", "filters": {}, "page": 1, "page_size": 20},
        }
    with routes.DASHBOARD_REFRESH_JOBS_LOCK:
        routes.DASHBOARD_REFRESH_JOBS.clear()
    routes.invalidate_worker_claim_empty_cache("sql")

    try:
        with TestClient(app) as client:
            login(client)
            created = client.post(
                "/api/admin/sql-reports",
                json={
                    "ten_bao_cao": "Dashboard queue refresh",
                    "ma_bao_cao": report_code,
                    "cau_lenh_sql": "SELECT :P_NGAY AS P_NGAY FROM dual;",
                    "cac_tham_so": ["P_NGAY"],
                },
            )
            assert created.status_code == 200
            report_id = created.json()["id"]
            saved = client.post(
                "/api/admin/dashboard-layouts",
                json={
                    "page_id": page_id,
                    "page_name": "Dashboard queue",
                    "layout": {
                        "page_id": page_id,
                        "tabs": [
                            {
                                "tab_id": "tab_queue",
                                "tab_name": "Queue",
                                "order": 1,
                                "grid_layout": [
                                    {
                                        "row_id": 1,
                                        "layout_type": "1_column",
                                        "widgets": [
                                            {
                                                "position": 1,
                                                "type": "bar_chart",
                                                "title": "Dashboard queue refresh",
                                                "sql_code": widget_sql_code,
                                                "report_id": report_id,
                                                "filters": {"P_NGAY": "2026-08-08"},
                                            }
                                        ],
                                    }
                                ],
                            }
                        ],
                    },
                },
            )
            assert saved.status_code == 200

            queued = client.get(f"/api/admin/dashboard-layouts/{page_id}/tabs/tab_queue/data")
            assert queued.status_code == 200
            queued_body = queued.json()
            assert queued_body["refreshing"] is True
            assert queued_body["refresh_job_ids"]
            job_id = queued_body["refresh_job_ids"][0]
            assert job_id.startswith("dashboard_")
            assert queued_body["widgets"][0]["data"]["status"] == "refreshing"

            with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
                assert job_id not in routes.DYNAMIC_REPORT_RUN_JOBS
                assert legacy_job_id in routes.DYNAMIC_REPORT_RUN_JOBS
            with routes.DASHBOARD_REFRESH_JOBS_LOCK:
                assert job_id in routes.DASHBOARD_REFRESH_JOBS

            claim = client.post(
                "/api/sql-worker/tasks/claim",
                json={"worker_id": worker_id, "version": "test-dashboard-worker"},
                headers=headers,
            )
            assert claim.status_code == 200
            task = claim.json()["task"]
            assert task["run_id"] == job_id
            assert task["task_type"] == "dynamic_report_dashboard_refresh"
            assert task["query"]["action"] == "run_sql_report"
            assert task["query"]["ma_bao_cao"] == report_code
            assert task["query"]["tham_so"] == {"P_NGAY": "2026-08-08"}

            finished = client.post(
                f"/api/sql-worker/tasks/{job_id}/result",
                json={
                    "ok": True,
                    "status": "success",
                    "message": "Dashboard cache updated.",
                    "columns": ["P_NGAY"],
                    "rows": [{"P_NGAY": "2026-08-08"}],
                    "pagination": {"page": 1, "page_size": 50, "total": 1},
                    "details": {"worker_id": worker_id},
                },
                headers=headers,
            )
            assert finished.status_code == 200
            assert finished.json()["run"]["status"] == "complete"

            cached = client.get(f"/api/admin/dashboard-layouts/{page_id}/tabs/tab_queue/data")
            assert cached.status_code == 200
            cached_body = cached.json()
            assert cached_body["refreshing"] is False
            data = cached_body["widgets"][0]["data"]
            assert data["ok"] is True
            assert data["rows"] == [{"P_NGAY": "2026-08-08"}]
            assert data["details"]["dashboard_cache"]["hit"] is True
    finally:
        with routes.DYNAMIC_REPORT_RUN_JOBS_LOCK:
            routes.DYNAMIC_REPORT_RUN_JOBS.clear()
        with routes.DASHBOARD_REFRESH_JOBS_LOCK:
            routes.DASHBOARD_REFRESH_JOBS.clear()
        routes.invalidate_worker_claim_empty_cache("sql")


def test_dashboard_refresh_queue_persists_and_recovers(monkeypatch: pytest.MonkeyPatch) -> None:
    class DurableDashboardRepository:
        def __init__(self) -> None:
            self.saved: dict[str, dict] = {}

        def save_report_run(self, payload: dict) -> None:
            run_id = str(payload.get("run_id") or payload.get("job_id") or "")
            self.saved[run_id] = dict(payload)

        def get_report_run(self, run_id: str) -> dict | None:
            return dict(self.saved.get(run_id) or {}) or None

        def list_report_runs(self, run_type: str = "", statuses: list[str] | None = None, limit: int = 200) -> list[dict]:
            status_set = set(statuses or [])
            rows = []
            for payload in self.saved.values():
                if run_type and payload.get("run_type") != run_type:
                    continue
                if status_set and payload.get("status") not in status_set:
                    continue
                rows.append(dict(payload))
            return rows[:limit]

    repository = DurableDashboardRepository()
    job_id = "dashboard_" + uuid.uuid4().hex[:24]
    monkeypatch.setattr(routes, "build_app_repository", lambda: repository)
    with routes.DASHBOARD_REFRESH_JOBS_LOCK:
        routes.DASHBOARD_REFRESH_JOBS.clear()
    routes.invalidate_worker_claim_empty_cache("sql")

    try:
        routes._set_dashboard_refresh_job(
            job_id,
            status="queued_worker",
            message="Da gui lenh lam moi cache dashboard cho may tram.",
            report_code="DASHBOARD_QUEUE_TEST",
            report_name="Dashboard queue test",
            payload={
                "ma_bao_cao": "DASHBOARD_QUEUE_TEST",
                "filters": {},
                "page": 1,
                "page_size": 50,
                "dashboard_cache_metadata": {"chart_key": "chart:test"},
            },
        )
        assert repository.saved[job_id]["run_type"] == "dashboard_refresh"
        assert routes.worker_claim_empty_cached("sql") is False

        with routes.DASHBOARD_REFRESH_JOBS_LOCK:
            routes.DASHBOARD_REFRESH_JOBS.clear()
        recovered = routes._next_dashboard_refresh_worker_job()
        assert recovered is not None
        recovered_id, recovered_job = recovered
        assert recovered_id == job_id
        assert recovered_job["run_type"] == "dashboard_refresh"
        assert recovered_job["payload"]["ma_bao_cao"] == "DASHBOARD_QUEUE_TEST"
    finally:
        with routes.DASHBOARD_REFRESH_JOBS_LOCK:
            routes.DASHBOARD_REFRESH_JOBS.clear()
        routes.invalidate_worker_claim_empty_cache("sql")


def test_dashboard_layout_tab_uses_bulk_chart_cache_for_cached_widgets() -> None:
    class FakeSettings:
        dashboard_chart_cache_enabled = True
        dashboard_chart_cache_report_ids = "*"
        dashboard_chart_cache_report_codes = "*"
        dashboard_chart_cache_ttl_seconds = 300
        dashboard_tab_max_workers = 10

    class FakeInternalApi:
        settings = FakeSettings()

        def __init__(self) -> None:
            self.calls = []

        def run_sql_report(self, **kwargs):
            self.calls.append(kwargs)
            return {"ok": True, "columns": [], "rows": []}

    class FakeRepository:
        def __init__(self) -> None:
            self.bulk_calls = []
            self.single_reads = []
            self.layout = {
                "tabs": [
                    {
                        "tab_id": "tab_cache",
                        "grid_layout": [
                            {
                                "row_id": 1,
                                "widgets": [
                                    {
                                        "position": 1,
                                        "type": "bar_chart",
                                        "title": "Report A",
                                        "sql_code": "REPORT_A",
                                        "report_id": 1,
                                        "filters": {},
                                    },
                                    {
                                        "position": 2,
                                        "type": "metric",
                                        "title": "Report B",
                                        "sql_code": "REPORT_B",
                                        "report_id": 2,
                                        "filters": {"status": "1"},
                                    },
                                ],
                            }
                        ],
                    }
                ],
            }
            self.reports = [
                {"id": 1, "ma_bao_cao": "REPORT_A", "ten_bao_cao": "Report A"},
                {"id": 2, "ma_bao_cao": "REPORT_B", "ten_bao_cao": "Report B"},
            ]
            self.cache_by_key = {}
            for report, filters in ((self.reports[0], {}), (self.reports[1], {"status": "1"})):
                chart_key = DatabaseService.dashboard_chart_cache_key(
                    report_id=report["id"],
                    sql_code=report["ma_bao_cao"],
                    filters=filters,
                    report_code=report["ma_bao_cao"],
                )
                self.cache_by_key[chart_key] = {
                    "chart_key": chart_key,
                    "status": "success",
                    "payload": {
                        "ok": True,
                        "columns": ["TEN_BAO_CAO"],
                        "rows": [{"TEN_BAO_CAO": report["ten_bao_cao"]}],
                    },
                    "refreshed_at": "2026-06-26T00:00:00+00:00",
                    "expires_at": "2026-06-26T00:05:00+00:00",
                }

        def get_dashboard_layout(self, page_id):
            return {"page_id": page_id, "layout": self.layout}

        def list_sql_reports(self):
            return self.reports

        def get_dashboard_chart_cache_many(self, chart_keys):
            self.bulk_calls.append(list(chart_keys))
            return [self.cache_by_key[key] for key in chart_keys if key in self.cache_by_key]

        def get_dashboard_chart_cache(self, chart_key):
            self.single_reads.append(chart_key)
            return self.cache_by_key.get(chart_key)

    internal_api = FakeInternalApi()
    repository = FakeRepository()
    result = DatabaseService(internal_api, repository).run_dashboard_layout_tab(page_id="DASHBOARD_CACHE", tab_id="tab_cache")

    assert result["ok"] is True
    assert len(result["widgets"]) == 2
    assert len(repository.bulk_calls) == 1
    assert len(repository.bulk_calls[0]) == 2
    assert repository.single_reads == []
    assert internal_api.calls == []
    assert all(widget["data"]["details"]["dashboard_cache"]["hit"] is True for widget in result["widgets"])


def test_dashboard_layout_pages_include_overview_and_reports_not_web_admin() -> None:
    with TestClient(app) as client:
        login(client)
        web_layout = {
            "page_id": "ADMIN_USERS",
            "page_name": "Quan tri nguoi dung",
            "layout": {
                "page_id": "ADMIN_USERS",
                "tabs": [
                    {
                        "tab_id": "tab_admin_users",
                        "tab_name": "Admin users",
                        "order": 1,
                        "grid_layout": [
                            {"row_id": 1, "layout_type": "2_columns", "widgets": []}
                        ],
                    }
                ],
            },
        }
        assert client.post("/api/admin/dashboard-layouts", json=web_layout).status_code == 200

        response = client.get("/api/admin/dashboard-layout-pages")
        assert response.status_code == 200
        pages = response.json()["pages"]
        page_ids = [page["page_id"] for page in pages]

        assert "DASHBOARD_KINH_DOANH" not in page_ids
        assert "REPORTS" not in page_ids
        assert "ADMIN_USERS" in page_ids

        generated_admin_page = next(page for page in pages if page["page_id"] == "ADMIN_USERS")
        assert generated_admin_page["feature_code"] == "adminusers"
        assert generated_admin_page["saved"] is True
        assert not any(page["feature_code"] == "admin.users" for page in pages)
        assert not any(page["feature_code"] == "admin_users" for page in pages)

        features = client.get("/api/admin/features").json()["features"]
        reports_feature = next(feature for feature in features if feature["code"] == "truyvansql")
        new_reports_feature = next(feature for feature in features if feature["code"] == "baocaomoi")
        builder_feature = next(feature for feature in features if feature["code"] == "thietkelayoutbaocao")
        generated_feature = next(feature for feature in features if feature["code"] == "adminusers")
        assert reports_feature["name"] == "Đào dữ liệu SQL"
        assert new_reports_feature["name"] == "Báo cáo mới"
        assert builder_feature["parent_code"] == "baocaomoi"
        assert generated_feature["parent_code"] == "baocaomoi"
        moved_features = []
        for feature in features:
            item = {
                "code": feature["code"],
                "name": feature["name"],
                "parent_code": feature.get("parent_code"),
                "sort_order": feature.get("sort_order") or 0,
            }
            if item["code"] == "adminusers":
                item["parent_code"] = "quantriweb"
                item["sort_order"] = 999
            moved_features.append(item)
        assert client.put("/api/admin/features/layout", json={"features": moved_features}).status_code == 200

        moved_pages = client.get("/api/admin/dashboard-layout-pages").json()["pages"]
        moved_admin_page = next(page for page in moved_pages if page["page_id"] == "ADMIN_USERS")
        assert moved_admin_page["feature_code"] == "adminusers"
        assert moved_admin_page["saved"] is True

        assert client.post("/api/admin/dashboard-layouts", json=web_layout).status_code == 200
        refreshed_features = client.get("/api/admin/features").json()["features"]
        refreshed_admin_feature = next(feature for feature in refreshed_features if feature["code"] == "adminusers")
        assert refreshed_admin_feature["parent_code"] == "quantriweb"


def test_admin_can_create_root_menu_and_assign_dashboard_layout_to_it() -> None:
    with TestClient(app) as client:
        login(client)
        menu_response = client.post("/api/admin/features/menu", json={"name": "Menu doanh thu"})
        assert menu_response.status_code == 200
        menu_feature = menu_response.json()["feature"]
        assert menu_feature["code"] == "menudoanhthu"
        assert menu_feature["parent_code"] is None

        layout_payload = {
            "page_id": "DASHBOARD_MENU_CHILD",
            "page_name": "Dashboard menu con",
            "parent_code": menu_feature["code"],
            "layout": {
                "tabs": [
                    {
                        "tab_id": "tab_menu_child",
                        "tab_name": "Menu con",
                        "grid_layout": [
                            {"row_id": 1, "layout_type": "2_columns", "widgets": []},
                        ],
                    }
                ],
            },
        }
        saved = client.post("/api/admin/dashboard-layouts", json=layout_payload)
        assert saved.status_code == 200
        assert saved.json()["parent_code"] == menu_feature["code"]

        features = client.get("/api/admin/features").json()["features"]
        layout_feature = next(feature for feature in features if feature["code"] == "dashboardmenuchild")
        assert layout_feature["parent_code"] == menu_feature["code"]

        pages = client.get("/api/admin/dashboard-layout-pages").json()["pages"]
        saved_page = next(page for page in pages if page["page_id"] == "DASHBOARD_MENU_CHILD")
        assert saved_page["parent_code"] == menu_feature["code"]

        detail = client.get("/api/admin/dashboard-layouts/DASHBOARD_MENU_CHILD")
        assert detail.status_code == 200
        assert detail.json()["parent_code"] == menu_feature["code"]


def test_dashboard_layout_delete_keeps_page_as_unsaved_and_aliases_duplicate_codes() -> None:
    with TestClient(app) as client:
        login(client)
        layout_payload = {
            "page_id": "DASHBOARD_DELETE_ME",
            "page_name": "Dashboard Delete Me",
            "layout": {
                "page_id": "DASHBOARD_DELETE_ME",
                "tabs": [
                    {
                        "tab_id": "tab_delete",
                        "tab_name": "Tab delete",
                        "order": 1,
                        "grid_layout": [],
                    }
                ],
            },
        }
        assert client.post("/api/admin/dashboard-layouts", json=layout_payload).status_code == 200
        assert client.delete("/api/admin/dashboard-layouts/DASHBOARD_DELETE_ME").status_code == 200
        pages = client.get("/api/admin/dashboard-layout-pages").json()["pages"]
        deleted_page = next(page for page in pages if page["feature_code"] == "dashboarddeleteme")
        assert deleted_page["saved"] is False
        assert deleted_page["unsaved"] is True
        assert client.delete(f"/api/admin/dashboard-layout-pages/{deleted_page['feature_code']}").status_code == 200
        purged_pages = client.get("/api/admin/dashboard-layout-pages").json()["pages"]
        assert not any(page["feature_code"] == "dashboarddeleteme" for page in purged_pages)

        assert client.post("/api/admin/dashboard-layouts", json=layout_payload).status_code == 200
        saved_pages = client.get("/api/admin/dashboard-layout-pages").json()["pages"]
        saved_page = next(page for page in saved_pages if page["page_id"] == "DASHBOARD_DELETE_ME")
        assert client.delete(f"/api/admin/dashboard-layout-pages/{saved_page['feature_code']}").status_code == 400

    pages = routes.build_dashboard_layout_pages(
        [
            {"code": "baocaomoi", "name": "Bao cao moi", "parent_code": None, "sort_order": 1},
            {"code": "dashboard_tong_quan", "name": "Tong quan cu", "parent_code": "baocaomoi", "sort_order": 2},
            {"code": "dashboardtongquan", "name": "Tong quan moi", "parent_code": "baocaomoi", "sort_order": 3},
        ],
        [
            {"page_id": "DASHBOARD_TONG_QUAN", "page_name": "Tong quan", "created_at": None, "updated_at": None},
        ],
    )
    assert [page["page_id"] for page in pages].count("DASHBOARD_TONG_QUAN") == 1


def test_viewer_cannot_access_dashboard_builder_api_or_report_runner() -> None:
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/users",
            json={
                "username": "viewer_builder",
                "full_name": "Viewer Builder",
                "password": "Viewer@Builder123",
                "role": "viewer",
            },
        )
        assert created.status_code == 200
        client.post("/api/auth/logout")
        login(client, "viewer_builder", "Viewer@Builder123")
        home = client.get("/")
        assert home.status_code == 200
        assert "app-shell-placeholder" in home.text
        assert "/static/shell.js?v=41" in home.text
        assert "/static/app.js?v=227" not in home.text
        shell_js = client.get("/static/shell.js?v=41")
        assert shell_js.status_code == 200
        assert "function collapseNavigationTree" in shell_js.text
        assert "function dedupeFeaturesForDisplay" in shell_js.text
        assert "function readCachedNavigation" in shell_js.text
        assert "async function logoutFromClient" in shell_js.text
        assert 'window.location.replace("/login")' in shell_js.text
        assert "/static/app.js?v=227" in shell_js.text
        assert "dashboard-designed-section" not in home.text
        assert "create-user-dialog" not in home.text

        navigation = client.get("/api/navigation")
        assert navigation.status_code == 200
        assert "dashboard" not in {feature["code"] for feature in navigation.json()["features"]}
        blocked_dashboard = client.get("/dashboard", follow_redirects=False)
        assert blocked_dashboard.status_code == 303
        assert blocked_dashboard.headers["location"] == "/"
        dashboard = client.get("/dashboard")
        assert dashboard.status_code == 200
        assert "app-shell-placeholder" in dashboard.text
        assert "/static/shell.js?v=41" in dashboard.text
        assert "/static/app.js?v=227" not in dashboard.text
        assert "view-dashboard-builder" not in dashboard.text
        assert "dashboard-designed-section" not in dashboard.text

        client.post("/api/auth/logout")
        login(client)
        reports = client.get("/truyvansql")
        assert reports.status_code == 200
        assert "view-reports" in reports.text
        assert "Đào dữ liệu SQL" in reports.text
        assert "view-mobile-gateway" not in reports.text
        assert "sql-report-dialog" not in reports.text
        assert "dynamic-report-search" not in reports.text
        assert "search-dynamic-report" not in reports.text
        assert "export-dynamic-report" not in reports.text
        assert "dynamic-report-page-size" not in reports.text
        assert "dynamic-report-head" not in reports.text
        assert "dynamic-report-body" not in reports.text
        assert "dynamic-report-prev" not in reports.text
        assert "dynamic-report-next" not in reports.text
        assert "/static/app.js?v=227" in reports.text
        assert "/static/reports-runtime.js" not in reports.text
        assert reports.text.count('class="app-view') == 1

        workstation = client.get("/maytram")
        assert workstation.status_code == 200
        assert "view-workstation" in workstation.text
        assert "/static/app.js?v=227" in workstation.text
        assert "/static/workstation.js" not in workstation.text
        assert workstation.text.count('class="app-view') == 1

        work_tasks = client.get("/quanlycongviec")
        assert work_tasks.status_code == 200
        assert "view-work-tasks" in work_tasks.text
        assert "/static/app.js?v=227" in work_tasks.text
        assert "/static/work-tasks.js" not in work_tasks.text
        assert work_tasks.text.count('class="app-view') == 1

        report_links = client.get("/linkbaocao")
        assert report_links.status_code == 200
        assert "view-report-links" in report_links.text
        assert "/static/app.js?v=227" in report_links.text
        assert "/static/report-links.js" not in report_links.text
        assert report_links.text.count('class="app-view') == 1

        system = client.get("/quantriketnoi")
        assert system.status_code == 200
        assert "view-system" in system.text
        assert "/static/app.js?v=227" in system.text
        assert "/static/data-mining.js" not in system.text
        assert system.text.count('class="app-view') == 1

        onebss_mining = client.get("/daodulieuonebss")
        assert onebss_mining.status_code == 200
        assert "view-onebss-mining" in onebss_mining.text
        assert "/static/app.js?v=227" in onebss_mining.text
        assert "/static/reports-runtime.js" not in onebss_mining.text
        assert onebss_mining.text.count('class="app-view') == 1

        ftp_mining = client.get("/daodulieuftp")
        assert ftp_mining.status_code == 200
        assert "view-ftp-mining" in ftp_mining.text
        assert "/static/app.js?v=227" in ftp_mining.text
        assert "/static/ftp-mining.js" not in ftp_mining.text
        assert ftp_mining.text.count('class="app-view') == 1

        client.post("/api/auth/logout")
        login(client, "viewer_builder", "Viewer@Builder123")
        forbidden_urls = [
            "/api/admin/dashboard-layouts",
            "/api/admin/dashboard-layout-pages",
            "/api/admin/dashboard-layouts/DASHBOARD_TEST_BUILDER",
            "/api/admin/dashboard-layouts/DASHBOARD_TEST_BUILDER/tabs/tab_a/data",
            "/api/reports/configs",
        ]
        for url in forbidden_urls:
            response = client.get(url)
            assert response.status_code == 403
            assert response.json()["detail"] == "Bạn không có quyền truy cập chức năng này"

        run_response = client.post(
            "/api/reports/run",
            json={"ma_bao_cao": "BC_BUILDER_TEST", "filters": {}, "page": 1, "page_size": 20},
        )
        assert run_response.status_code == 403
        assert run_response.json()["detail"] == "Bạn không có quyền truy cập chức năng này"


def test_viewer_feature_permissions_unlock_mining_pages_and_runtime_apis() -> None:
    suffix = uuid.uuid4().hex[:8]
    username = f"viewer_mining_{suffix}"
    password = "Viewer@Mining123"
    sql_code = f"BC_MINING_{suffix}".upper()
    onebss_code = f"OB_MINING_{suffix}".upper()
    ftp_code = f"FTP_MINING_{suffix}".upper()
    with TestClient(app) as client:
        login(client)
        assert client.post(
            "/api/admin/sql-reports",
            json={
                "ten_bao_cao": "SQL mining viewer",
                "ma_bao_cao": sql_code,
                "cau_lenh_sql": "SELECT ma_tb FROM css_cto.db_thuebao WHERE trang_thai = :STATUS;",
                "cac_tham_so": ["STATUS"],
            },
        ).status_code == 200
        assert client.post(
            "/api/admin/onebss-reports",
            json={
                "ma_bao_cao": onebss_code,
                "ten_bao_cao": "OneBSS mining viewer",
                "danh_sach_bien": ["P_TUNGAY"],
                "parameters": {"P_TUNGAY": "01/07/2026"},
                "report_url": "https://onebss.vnpt.vn/#/report/bi?path=TEST_VIEWER_MINING&name=Test",
                "storage_link": "",
            },
        ).status_code == 200
        assert client.post(
            "/api/admin/ftp-reports",
            json={
                "ma_bao_cao": ftp_code,
                "ten_bao_cao": "FTP mining viewer",
                "folder_path": "/reports/viewer",
                "file_name_template": "viewer_{yyyymmdd}.xlsx",
                "connection_code": "ftp_storage",
                "is_active": True,
            },
        ).status_code == 200
        created = client.post(
            "/api/admin/users",
            json={
                "username": username,
                "full_name": "Viewer Mining",
                "password": password,
                "role": "viewer",
            },
        )
        assert created.status_code == 200
        viewer_id = created.json()["user"]["id"]
        granted = client.put(
            f"/api/admin/users/{viewer_id}/permissions",
            json={"feature_codes": ["truyvansql", "daodulieuonebss", "daodulieuftp"]},
        )
        assert granted.status_code == 200

        client.post("/api/auth/logout")
        login(client, username, password)
        navigation = client.get("/api/navigation")
        assert navigation.status_code == 200
        feature_codes = {feature["code"] for feature in navigation.json()["features"]}
        assert {"truyvansql", "daodulieuonebss", "daodulieuftp"} <= feature_codes

        for path, marker in [
            ("/truyvansql", "view-reports"),
            ("/daodulieuonebss", "view-onebss-mining"),
            ("/daodulieuftp", "view-ftp-mining"),
        ]:
            page = client.get(path)
            assert page.status_code == 200
            assert marker in page.text
            assert "app-shell-placeholder" not in page.text
            assert "/static/app.js?v=227" in page.text

        report_configs = client.get("/api/reports/configs")
        assert report_configs.status_code == 200
        assert sql_code in {report["ma_bao_cao"] for report in report_configs.json()["reports"]}
        assert client.get("/api/reports/history").status_code == 200
        assert client.get("/api/reports/export-jobs").status_code == 200

        onebss_configs = client.get("/api/onebss-reports/configs")
        assert onebss_configs.status_code == 200
        assert onebss_code in {report["ma_bao_cao"] for report in onebss_configs.json()["reports"]}
        assert client.get("/api/onebss-reports/runs?limit=5").status_code == 200

        ftp_configs = client.get("/api/ftp-reports/configs")
        assert ftp_configs.status_code == 200
        assert ftp_code in {report["ma_bao_cao"] for report in ftp_configs.json()["reports"]}
        assert client.get("/api/ftp-reports/runs?limit=5").status_code == 200


def test_auto_module_is_removed_from_dashboard() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.get("/")
        assert response.status_code == 200
        assert "data-feature-code=\"auto\"" not in response.text
        assert "attt" not in response.text.lower()


def test_admin_can_create_viewer_and_viewer_cannot_access_admin_api() -> None:
    with TestClient(app) as client:
        login(client)
        response = client.post(
            "/api/admin/users",
            json={
                "username": "viewer_test",
                "full_name": "Người xem thử nghiệm",
                "password": "Viewer@Test123",
                "role": "viewer",
            },
        )
        assert response.status_code == 200
        client.post("/api/auth/logout")
        login(client, "viewer_test", "Viewer@Test123")
        assert client.get("/api/admin/users").status_code == 403


def test_viewer_with_user_management_feature_can_manage_users() -> None:
    suffix = uuid.uuid4().hex[:8]
    manager_username = f"viewer_user_mgr_{suffix}"
    managed_username = f"viewer_managed_{suffix}"
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/users",
            json={
                "username": manager_username,
                "full_name": "Viewer User Manager",
                "password": "Viewer@Manager123",
                "role": "viewer",
            },
        )
        assert created.status_code == 200
        manager_id = created.json()["user"]["id"]
        granted = client.put(
            f"/api/admin/users/{manager_id}/permissions",
            json={"feature_codes": ["quantringuoidung"]},
        )
        assert granted.status_code == 200

        client.post("/api/auth/logout")
        login(client, manager_username, "Viewer@Manager123")
        page = client.get("/quantringuoidung")
        assert page.status_code == 200
        assert 'id="view-users"' in page.text
        assert "/static/app.js?v=227" in page.text
        assert client.get("/api/admin/users").status_code == 200
        managed = client.post(
            "/api/admin/users",
            json={
                "username": managed_username,
                "full_name": "Managed Viewer",
                "password": "Viewer@Managed123",
                "role": "viewer",
            },
        )
        assert managed.status_code == 200
        assert managed.json()["user"]["username"] == managed_username


def test_admin_can_generate_one_time_password_and_viewer_cannot() -> None:
    username = f"viewer_password_{uuid.uuid4().hex[:8]}"
    primary_password = "Viewer@Primary123"
    with TestClient(app) as client:
        login(client)
        created = client.post(
            "/api/admin/users",
            json={
                "username": username,
                "full_name": "Viewer Password",
                "password": primary_password,
                "role": "viewer",
            },
        )
        assert created.status_code == 200
        viewer_id = created.json()["user"]["id"]
        routes.build_app_repository().change_password(viewer_id, primary_password, must_change=False)

        generated = client.post(f"/api/admin/users/{viewer_id}/generate-password")
        assert generated.status_code == 200
        payload = generated.json()
        one_time_password = payload["password"]
        assert len(one_time_password) >= 14
        assert any(character.isupper() for character in one_time_password)
        assert any(character.islower() for character in one_time_password)
        assert any(character.isdigit() for character in one_time_password)
        assert payload["expires_at"]
        assert "password_hash" not in payload["user"]
        assert payload["user"]["must_change_password"] in (0, False)

        client.post("/api/auth/logout")
        primary_login = client.post("/api/auth/login", json={"username": username, "password": primary_password})
        assert primary_login.status_code == 200
        client.post("/api/auth/logout")
        one_time_login = client.post("/api/auth/login", json={"username": username, "password": one_time_password})
        assert one_time_login.status_code == 200
        client.post("/api/auth/logout")
        reused_one_time_login = client.post("/api/auth/login", json={"username": username, "password": one_time_password})
        assert reused_one_time_login.status_code == 401
        primary_login_after_one_time = client.post("/api/auth/login", json={"username": username, "password": primary_password})
        assert primary_login_after_one_time.status_code == 200
        assert client.get("/api/admin/users").status_code == 403
        forbidden = client.post(f"/api/admin/users/{viewer_id}/generate-password")
        assert forbidden.status_code == 403


def test_admin_can_manage_catalog_and_encrypted_web_credentials() -> None:
    with TestClient(app) as client:
        login(client)
        website = client.post(
            "/api/admin/websites",
            json={"name": "VNPT Test", "url": "https://example.vn", "requires_otp": True, "is_active": True},
        )
        assert website.status_code == 200
        website_id = website.json()["website"]["id"]
        saved = client.post(
            "/api/credentials",
            json={"website_id": website_id, "login_username": "user01", "password": "Secret@Test123", "notes": ""},
        )
        assert saved.status_code == 200
        credentials = client.get("/api/credentials").json()["credentials"]
        assert credentials[0]["requires_otp"] == 1
        assert "encrypted_password" not in credentials[0]
        revealed = client.post(f"/api/credentials/{credentials[0]['id']}/reveal")
        assert revealed.json()["password"] == "Secret@Test123"


def test_viewer_needs_feature_permission_for_vault() -> None:
    with TestClient(app) as client:
        login(client)
        users = client.get("/api/admin/users").json()["users"]
        viewer = next(user for user in users if user["username"] == "viewer_test")
        client.post("/api/auth/logout")
        login(client, "viewer_test", "Viewer@Test123")
        assert client.get("/api/credentials").status_code == 403
        client.post("/api/auth/logout")
        login(client)
        response = client.put(
            f"/api/admin/users/{viewer['id']}/permissions",
            json={"feature_codes": ["taikhoanweb", "xemdanhsachtaikhoan", "themvasuataikhoan", "xemmatkhaudaluu"]},
        )
        assert response.status_code == 200
        client.post("/api/auth/logout")
        login(client, "viewer_test", "Viewer@Test123")
        assert client.get("/api/credentials").status_code == 200


def test_dashboard_layout_preserves_google_sheet_embed_without_sql() -> None:
    with TestClient(app) as client:
        login(client)
        layout_payload = {
            "page_id": "DASHBOARD_GOOGLE_SHEET",
            "page_name": "Google Sheet Dashboard",
            "layout": {
                "page_id": "DASHBOARD_GOOGLE_SHEET",
                "tabs": [
                    {
                        "tab_id": "tab_sheet",
                        "tab_name": "Sheet",
                        "order": 1,
                        "grid_layout": [
                            {
                                "row_id": 1,
                                "layout_type": "1_column",
                                "widgets": [
                                    {
                                        "position": 1,
                                        "type": "google_sheet_embed",
                                        "title": "Published Sheet",
                                        "chart_config": {
                                            "embed_url": "https://docs.google.com/spreadsheets/d/e/2PACX-test/pubhtml",
                                            "embed_height": "560",
                                        },
                                    }
                                ],
                            }
                        ],
                    }
                ],
            },
        }

        saved = client.post("/api/admin/dashboard-layouts", json=layout_payload)
        assert saved.status_code == 200
        widget = saved.json()["layout"]["tabs"][0]["grid_layout"][0]["widgets"][0]
        assert widget["type"] == "google_sheet_embed"
        assert widget["sql_code"] == ""
        assert widget["chart_config"]["embed_url"].startswith("https://docs.google.com/spreadsheets/")

        reopened = client.get("/api/admin/dashboard-layouts/DASHBOARD_GOOGLE_SHEET")
        assert reopened.status_code == 200
        reopened_widget = reopened.json()["layout"]["tabs"][0]["grid_layout"][0]["widgets"][0]
        assert reopened_widget["type"] == "google_sheet_embed"
        assert reopened_widget["chart_config"]["embed_height"] == "560"

        tab_data = client.get("/api/admin/dashboard-layouts/DASHBOARD_GOOGLE_SHEET/tabs/tab_sheet/data")
        assert tab_data.status_code == 200
        assert tab_data.json()["ok"] is True
        assert tab_data.json()["widgets"] == []


def test_google_sheet_table_extractor_removes_sheet_headers() -> None:
    extractor = routes.GoogleSheetTableExtractor()
    extractor.feed(
        """
        <style>.s0{background:#fee;color:#111}.row-headers-background{background:#073763;color:#fff}</style>
        <table class="waffle">
          <tbody>
            <tr><th class="row-headers-background">1</th><td class="s0">A</td></tr>
            <tr><th class="row-headers-background">2</th><td class="s0">B</td></tr>
          </tbody>
        </table>
        """
    )
    html = extractor.sanitized_html()
    assert 'class="google-sheet-table-source ritz grid-container"' in html
    assert '<th class="row-headers-background"' not in html
    assert ">1<" not in html
    assert ">2<" not in html
    assert ">A<" in html
    assert ">B<" in html
