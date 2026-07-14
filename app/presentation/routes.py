from pathlib import Path
import base64
import binascii
import hmac
import json
import logging
import tempfile
import threading
import time
import uuid
from datetime import datetime
from html import escape as html_escape
from html.parser import HTMLParser
import re
import sqlite3
from io import BytesIO
from typing import Any, Literal
from urllib.parse import parse_qsl, quote, urlencode, urlparse, urlunparse

import httpx
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, File, HTTPException, Request, UploadFile, status
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from pydantic import BaseModel, Field

from app.application.auth_service import AuthService
from app.application.database_service import DatabaseService
from app.application.google_drive_service import (
    GOOGLE_DRIVE_OAUTH_SCOPES,
    GoogleDriveConfigurationError,
    clear_google_drive_oauth_tokens,
    google_drive_folder_id,
    google_drive_oauth_client_configured,
    google_drive_oauth_status,
    save_google_drive_oauth_tokens,
)
from app.application.vault_service import VaultService
from app.application.connection_service import ConnectionService
from app.application.telegram_notifier import TelegramNotifier
from app.application.onebss_data_mining_service import OneBssDownloadError, normalize_onebss_report_url, run_data_mining_schedule
from app.application.onebss_report_service import (
    cancel_onebss_mobile_gateway_otp,
    consume_onebss_mobile_gateway_otp,
    inspect_onebss_mobile_gateway_otp,
    match_onebss_mobile_gateway_manual_otp,
    run_onebss_report_request,
    start_onebss_otp_mobile_gateway_request,
)
from app.application.zalo_auto_message_service import capture_page_screenshot_bytes, capture_public_url, send_zalo_auto_message
from app.application.zalo_bot import ZaloBotClient
from app.data_access.app_repository import (
    AppRepository,
    DEFAULT_DASHBOARD_PAGE_ID,
    dashboard_feature_code_for_page,
    normalize_feature_code,
)
from app.data_access.internal_api_client import InternalApiClient
from app.data_access.repository_factory import build_repository
from app.settings import get_settings


router = APIRouter()
templates = Jinja2Templates(directory=Path("app/presentation/templates"))
logger = logging.getLogger(__name__)
FAILED_LOGIN_COUNTS: dict[str, int] = {}
MAX_USER_IMPORT_BYTES = 2 * 1024 * 1024
DYNAMIC_REPORT_EXPORT_JOBS: dict[str, dict[str, Any]] = {}
DYNAMIC_REPORT_EXPORT_JOBS_LOCK = threading.Lock()
DYNAMIC_REPORT_EXPORT_JOB_FILE_LOCK = threading.Lock()
DYNAMIC_REPORT_EXPORT_SEMAPHORE_LOCK = threading.Lock()
DYNAMIC_REPORT_EXPORT_SEMAPHORE: threading.Semaphore | None = None
DYNAMIC_REPORT_EXPORT_SEMAPHORE_WORKERS = 0
DYNAMIC_REPORT_EXPORT_JOB_TTL_SECONDS = 60 * 60
DYNAMIC_REPORT_EXPORT_DIR = Path(tempfile.gettempdir()) / "vnptcto_dynamic_report_exports"
DYNAMIC_REPORT_EXPORT_JOB_DIR = DYNAMIC_REPORT_EXPORT_DIR / "jobs"
DYNAMIC_REPORT_HISTORY_ACTION = "dynamic_report_history"
DYNAMIC_REPORT_EXPORT_ACTIVE_STATUSES = {"queued", "running", "cancel_requested"}
DYNAMIC_REPORT_EXPORT_FINAL_STATUSES = {"complete", "failed", "cancelled"}
ONEBSS_REPORT_JOBS: dict[str, dict[str, Any]] = {}
ONEBSS_REPORT_JOBS_LOCK = threading.Lock()
ONEBSS_REPORT_JOB_FILE_LOCK = threading.Lock()
ONEBSS_REPORT_SEMAPHORE = threading.Semaphore(1)
ONEBSS_REPORT_JOB_TTL_SECONDS = 24 * 60 * 60
ONEBSS_REPORT_JOB_DIR = Path(tempfile.gettempdir()) / "vnptcto_onebss_report_jobs"
ONEBSS_REPORT_ACTIVE_STATUSES = {"queued", "running", "otp_required", "otp_invalid", "manual_otp_required"}
ONEBSS_REPORT_FINAL_STATUSES = {"success", "failed", "cancelled", "storage_failed", "google_drive_not_configured", "google_drive_upload_failed"}
EXCEL_MAX_ROWS_PER_SHEET = 1_048_576
ADMIN_ONLY_MESSAGE = "Bạn không có quyền truy cập chức năng này"
DASHBOARD_LAYOUT_TYPES = {
    "1_column": 1,
    "2_columns": 2,
    "3_columns": 3,
    "4_columns": 4,
    "5_columns": 5,
    "6_columns": 6,
    "4_columns_1_3": 2,
    "4_columns_3_1": 2,
    "5_columns_1_4": 2,
    "5_columns_4_1": 2,
    "5_columns_2_3": 2,
    "5_columns_3_2": 2,
    "6_columns_1_5": 2,
    "6_columns_5_1": 2,
    "6_columns_2_4": 2,
    "6_columns_4_2": 2,
}
DASHBOARD_WIDGET_TYPES = {"bar_chart", "pie_chart", "line_chart", "combo_chart", "multi_bar_chart", "horizontal_multi_bar_chart", "multi_line_chart", "data_table", "metric", "data_card", "google_sheet_embed", "text_title"}
DASHBOARD_NON_SQL_WIDGET_TYPES = {"google_sheet_embed", "text_title"}
DASHBOARD_LAYOUT_EXCLUDED_FEATURE_CODES = {
    "dashboard",
    "truyvansql",
    "baocaomoi",
    "thietkelayoutbaocao",
    "reports",
    "new_reports",
    "admin.dashboard_builder",
}
DASHBOARD_LAYOUT_PARENT_EXCLUDED_FEATURE_CODES = {
    "dashboard",
    "quanlycongviec",
    "truyvansql",
    "reports",
    "new_reports",
}


class DynamicReportExportCancelled(Exception):
    pass


class GoogleSheetTableExtractor(HTMLParser):
    allowed_tags = {"table", "thead", "tbody", "tfoot", "tr", "td", "th", "colgroup", "col"}
    allowed_attrs = {"class", "style", "colspan", "rowspan", "width", "height"}

    def __init__(self) -> None:
        super().__init__(convert_charrefs=False)
        self.capture_table = False
        self.skip_table_node_depth = 0
        self.table_depth = 0
        self.table_chunks: list[str] = []
        self.capture_style = False
        self.style_chunks: list[str] = []
        self.found_table = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "style":
            self.capture_style = True
            return
        if tag == "table" and not self.found_table:
            self.capture_table = True
            self.found_table = True
        if not self.capture_table or tag not in self.allowed_tags:
            return
        if self.skip_table_node_depth:
            self.skip_table_node_depth += 1
            return
        class_value = " ".join(value or "" for name, value in attrs if (name or "").lower() == "class").lower()
        if tag in {"td", "th", "col"} and (
            "row-header" in class_value
            or "row-headers" in class_value
            or "column-header" in class_value
            or "column-headers" in class_value
            or "freezebar" in class_value
        ):
            if tag != "col":
                self.skip_table_node_depth = 1
            return
        if tag == "table":
            self.table_depth += 1
        safe_attrs = []
        for name, value in attrs:
            name = (name or "").lower()
            if name in self.allowed_attrs and value is not None:
                safe_attrs.append(f'{name}="{html_escape(value, quote=True)}"')
        suffix = f" {' '.join(safe_attrs)}" if safe_attrs else ""
        self.table_chunks.append(f"<{tag}{suffix}>")

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag == "style":
            self.capture_style = False
            return
        if self.skip_table_node_depth:
            self.skip_table_node_depth -= 1
            return
        if not self.capture_table or tag not in self.allowed_tags:
            return
        self.table_chunks.append(f"</{tag}>")
        if tag == "table":
            self.table_depth -= 1
            if self.table_depth <= 0:
                self.capture_table = False

    def handle_data(self, data: str) -> None:
        if self.capture_style:
            self.style_chunks.append(data)
        elif self.capture_table and not self.skip_table_node_depth:
            self.table_chunks.append(html_escape(data))

    def handle_entityref(self, name: str) -> None:
        if self.capture_table and not self.skip_table_node_depth:
            self.table_chunks.append(f"&{name};")

    def handle_charref(self, name: str) -> None:
        if self.capture_table and not self.skip_table_node_depth:
            self.table_chunks.append(f"&#{name};")

    def sanitized_html(self) -> str:
        table = "".join(self.table_chunks).strip()
        if not table:
            return ""
        style = "\n".join(self.style_chunks).strip()
        style_html = f"<style>{style}</style>" if style else ""
        return f'{style_html}<div class="google-sheet-table-source ritz grid-container" dir="ltr">{table}</div>'


class LoginPayload(BaseModel):
    username: str
    password: str


class CreateUserPayload(BaseModel):
    username: str
    full_name: str
    password: str
    role: Literal["admin", "viewer"]


class UpdateUserPayload(BaseModel):
    full_name: str
    role: Literal["admin", "viewer"]
    is_active: bool


class PasswordPayload(BaseModel):
    password: str


class ChangePasswordPayload(BaseModel):
    current_password: str
    new_password: str


class WebsitePayload(BaseModel):
    id: int | None = None
    name: str
    url: str
    requires_otp: bool = False
    is_active: bool = True


class CredentialPayload(BaseModel):
    id: int | None = None
    website_id: int
    login_username: str
    password: str
    notes: str = ""


class PermissionPayload(BaseModel):
    feature_codes: list[str]


class BulkPermissionPayload(BaseModel):
    user_ids: list[int]
    feature_codes: list[str]


class BulkDataPermissionPayload(BaseModel):
    user_ids: list[int]
    region_codes: list[str]


class DataRegionPayload(BaseModel):
    code: str
    name: str
    is_active: bool = True
    sort_order: int = 0


class SystemConnectionPayload(BaseModel):
    name: str
    connection_type: str
    description: str = ""
    config: dict[str, Any] = Field(default_factory=dict)
    is_active: bool = False


class SqlReportPayload(BaseModel):
    id: int | None = None
    ten_bao_cao: str
    ma_bao_cao: str
    cau_lenh_sql: str
    cac_tham_so: list[str] = Field(default_factory=list)


class OneBssReportPayload(BaseModel):
    id: int | None = None
    ma_bao_cao: str = ""
    ten_bao_cao: str
    danh_sach_bien: list[str] = Field(default_factory=list)
    parameters: dict[str, Any] = Field(default_factory=dict)
    otp_service_code: str = "onebss"
    report_url: str
    storage_link: str = ""


class RunReportPayload(BaseModel):
    ma_bao_cao: str
    filters: dict[str, Any] = Field(default_factory=dict)
    page: int = 1
    page_size: int = 20
    search: str = ""
    search_columns: list[str] = Field(default_factory=list)


class ExportLoadedReportPayload(BaseModel):
    ma_bao_cao: str
    columns: list[str] = Field(default_factory=list)
    rows: list[dict[str, Any]] = Field(default_factory=list)
    search: str = ""


class RunOneBssReportPayload(BaseModel):
    ma_bao_cao: str
    parameters: dict[str, Any] = Field(default_factory=dict)
    otp: str = ""
    session_id: str = ""
    otp_request_id: str = ""
    otp_source: str = ""
    job_id: str = ""


class OneBssTaskOtpPayload(BaseModel):
    otp: str = ""
    otp_request_id: str = ""
    otp_source: str = "manual"


class OneBssWorkerClaimPayload(BaseModel):
    worker_id: str = "onebss-workstation"


class OneBssWorkerStatusPayload(BaseModel):
    status: str
    message: str = ""
    worker_id: str = ""
    worker_session_id: str = ""
    details: dict[str, Any] = Field(default_factory=dict)


class OneBssWorkerResultPayload(BaseModel):
    ok: bool = True
    status: str = ""
    message: str = ""
    file_name: str = ""
    file_path: str = ""
    storage_link: str = ""
    storage_status: str = ""
    duration_ms: int = 0
    details: dict[str, Any] = Field(default_factory=dict)


class DashboardLayoutPayload(BaseModel):
    page_id: str
    page_name: str = ""
    parent_code: str | None = None
    layout: dict[str, Any] = Field(default_factory=dict)


class CreateMenuPayload(BaseModel):
    name: str


class SystemRolePayload(BaseModel):
    code: str
    name: str
    description: str = ""
    is_active: bool = True
    sort_order: int = 0


class FeatureLayoutItem(BaseModel):
    code: str
    name: str
    parent_code: str | None = None
    sort_order: int = 0


class FeatureLayoutPayload(BaseModel):
    features: list[FeatureLayoutItem]


class WorkTaskPayload(BaseModel):
    task_id: str = ""
    ten_cong_viec: str
    type: Literal["Daily", "Weekly", "Once"] = "Daily"
    time: str = "07:00"
    weekday: str = ""
    once_date: str = ""
    group: str = ""
    check: bool = False


class ZaloSendMessagePayload(BaseModel):
    chat_id: str = ""
    text: str = "Tin nhan test tu Bot VNPT Can Tho."


class ZaloAutoMessagePayload(BaseModel):
    schedule_id: str = ""
    name: str
    page_url: str = "/"
    page_label: str = ""
    schedule_type: Literal["TimeWindow", "Daily", "Weekly", "Monthly"] = "Daily"
    time_slots: list[str] = Field(default_factory=list)
    run_time: str = "07:00"
    weekday: str = ""
    month_day: int = 1
    target_type: Literal["person", "group"] = "group"
    chat_id: str = ""
    chat_name: str = ""
    caption: str = ""
    photo_url: str = ""
    is_active: bool = True


class ZaloCapturePayload(BaseModel):
    image_base64: str
    mime_type: str = "image/png"
    page_url: str = ""


class DataMiningSchedulePayload(BaseModel):
    schedule_id: str = ""
    name: str
    report_url: str
    schedule_type: Literal["Daily", "Weekly", "Monthly"] = "Daily"
    run_time: str = "07:00"
    weekday: str = ""
    month_day: int = 1
    storage_link: str = ""
    file_name_template: str = ""
    parameters: dict[str, Any] = Field(default_factory=dict)
    is_active: bool = True


class DataMiningRunPayload(BaseModel):
    otp: str = ""
    allow_device_registration: bool = True
    parameters: dict[str, Any] = Field(default_factory=dict)


class PageCapturePayload(BaseModel):
    page_url: str = "/"


def build_app_repository() -> AppRepository:
    return build_repository(get_settings())


def build_auth_service() -> AuthService:
    return AuthService(build_app_repository())


def build_database_service() -> DatabaseService:
    settings = get_settings()
    repository = build_app_repository()
    return DatabaseService(InternalApiClient.from_repository(settings, repository), repository)


def build_vault_service() -> VaultService:
    settings = get_settings()
    return VaultService(build_app_repository(), settings.session_secret.get_secret_value())


def build_connection_service() -> ConnectionService:
    settings = get_settings()
    return ConnectionService(build_app_repository(), settings)


def notify_if_failed(title: str, result: dict, extra: dict | None = None) -> None:
    if result.get("ok"):
        return
    settings = get_settings()
    details = extra or {}
    connection_name = result.get("connection_name")
    if connection_name:
        details = {**details, "ket_noi": connection_name}
    TelegramNotifier(settings).send_message(title, result.get("message", "Không rõ lỗi."), details or None)


def compact_log_text(value: Any, limit: int = 1000) -> str:
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return f"{text[:limit]}..."


def add_zalo_message_log(
    repository: AppRepository,
    *,
    direction: str,
    event_name: str = "",
    chat_id: str = "",
    chat_type: str = "",
    sender_id: str = "",
    sender_name: str = "",
    message_id: str = "",
    text: str = "",
    ok: bool = True,
    raw_preview: str = "",
    raw_keys: list[str] | None = None,
    result_keys: list[str] | None = None,
    message_keys: list[str] | None = None,
) -> None:
    action = "zalo_message_received" if direction == "in" else "zalo_message_sent"
    if direction == "out" and not ok:
        action = "zalo_message_send_failed"
    details = {
        "direction": direction,
        "event_name": event_name,
        "chat_id": chat_id,
        "chat_type": chat_type,
        "sender_id": sender_id,
        "sender_name": sender_name,
        "message_id": message_id,
        "text": compact_log_text(text),
        "ok": ok,
        "raw_preview": compact_log_text(raw_preview, 1800),
        "raw_keys": raw_keys or [],
        "result_keys": result_keys or [],
        "message_keys": message_keys or [],
    }
    repository.add_audit_log("zalo_bot", action, json.dumps(details, ensure_ascii=False))


def parse_zalo_message_log(row: dict[str, Any]) -> dict[str, Any] | None:
    if row.get("action") not in {"zalo_message_received", "zalo_message_sent", "zalo_message_send_failed"}:
        return None
    try:
        details = json.loads(row.get("details") or "{}")
    except json.JSONDecodeError:
        details = {"text": row.get("details") or ""}
    return {
        "id": row.get("id"),
        "created_at": row.get("created_at"),
        "action": row.get("action"),
        "direction": details.get("direction") or ("out" if row.get("action") != "zalo_message_received" else "in"),
        "event_name": details.get("event_name") or "",
        "chat_id": details.get("chat_id") or "",
        "chat_type": details.get("chat_type") or "",
        "sender_id": details.get("sender_id") or "",
        "sender_name": details.get("sender_name") or "",
        "message_id": details.get("message_id") or "",
        "text": details.get("text") or "",
        "raw_preview": details.get("raw_preview") or "",
        "raw_keys": details.get("raw_keys") or [],
        "result_keys": details.get("result_keys") or [],
        "message_keys": details.get("message_keys") or [],
        "ok": bool(details.get("ok", row.get("action") != "zalo_message_send_failed")),
    }


def latest_zalo_chat_id(repository: AppRepository) -> str:
    for row in repository.list_audit_logs(limit=500):
        parsed = parse_zalo_message_log(row)
        if parsed and parsed.get("direction") == "in" and parsed.get("chat_id"):
            return str(parsed["chat_id"])
    return ""


def raise_work_task_schema_error(error: RuntimeError) -> None:
    if "work_tasks" in str(error):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Supabase chưa có bảng work_tasks. Hãy chạy lại file sql/supabase_upgrade_admin_modules.sql.",
        ) from error
    raise error


def raise_zalo_auto_message_schema_error(error: RuntimeError) -> None:
    if "zalo_auto_messages" in str(error) or "zalo_message_captures" in str(error):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Supabase chua co bang zalo_auto_messages/zalo_message_captures. Hay chay lai file sql/supabase_upgrade_admin_modules.sql.",
        ) from error
    raise error


def raise_data_mining_schema_error(error: RuntimeError) -> None:
    if "data_mining_schedules" in str(error) or "data_mining_runs" in str(error):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Supabase chua co bang data_mining_schedules/data_mining_runs. Hay chay lai file sql/supabase_upgrade_admin_modules.sql.",
        ) from error
    raise error


def raise_sql_report_schema_error(error: RuntimeError) -> None:
    if "sql_reports" in str(error):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Supabase chưa có bảng sql_reports. Hãy chạy lại file sql/supabase_upgrade_admin_modules.sql.",
        ) from error
    raise error


def raise_onebss_report_schema_error(error: RuntimeError) -> None:
    error_text = str(error)
    is_missing_table_error = (
        "PGRST205" in error_text
        or "42P01" in error_text
        or "Could not find the table" in error_text
        or "relation" in error_text and "does not exist" in error_text
        or "Supabase REST loi 404" in error_text
    )
    is_missing_column_error = (
        "PGRST204" in error_text
        or "Could not find" in error_text and "column" in error_text
        or "no such column" in error_text
        or "has no column named" in error_text
    )
    if ("onebss_reports" in error_text or "onebss_report_runs" in error_text) and is_missing_table_error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Supabase chua co bang onebss_reports/onebss_report_runs. Hay chay lai file sql/supabase_upgrade_admin_modules.sql.",
        ) from error
    if ("onebss_report_runs" in error_text or "worker_id" in error_text or "otp_request_id" in error_text) and is_missing_column_error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Bang lich su OneBSS chua co cac cot hang doi may tram. Hay chay migration moi nhat hoac khoi dong lai ung dung de SQLite tu nang cap.",
        ) from error
    raise error


def raise_onebss_operation_error(error: Exception, action: str) -> None:
    try:
        raise_onebss_report_schema_error(error)  # type: ignore[arg-type]
    except HTTPException:
        raise
    except Exception:
        logger.exception("%s", action)
        detail = str(error).replace("\n", " ").strip()
        if len(detail) > 300:
            detail = f"{detail[:297]}..."
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"{action}: {type(error).__name__}: {detail or 'khong co chi tiet loi'}",
        ) from error


def raise_dashboard_layout_schema_error(error: RuntimeError) -> None:
    if "dashboard_layouts" in str(error):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Supabase chưa có bảng dashboard_layouts. Hãy chạy lại file sql/supabase_upgrade_admin_modules.sql.",
        ) from error
    raise error


def _dynamic_report_history_int(value: Any, default: int = 0) -> int:
    try:
        return max(0, int(value or default))
    except (TypeError, ValueError):
        return default


def _dynamic_report_history_report(ma_bao_cao: str, report: dict[str, Any] | None = None) -> dict[str, str]:
    source = report if isinstance(report, dict) else {}
    code = str(source.get("ma_bao_cao") or ma_bao_cao or "").strip().upper()
    name = str(source.get("ten_bao_cao") or "").strip()
    if code and name:
        return {"ma_bao_cao": code, "ten_bao_cao": name}
    try:
        stored = build_app_repository().get_sql_report_by_code(code or ma_bao_cao.strip().upper())
        if stored:
            code = str(stored.get("ma_bao_cao") or code or ma_bao_cao).strip().upper()
            name = str(stored.get("ten_bao_cao") or name or code).strip()
    except Exception:
        logger.exception("Cannot resolve dynamic report history metadata")
    return {"ma_bao_cao": code or str(ma_bao_cao or "").strip().upper(), "ten_bao_cao": name or code}


def _record_dynamic_report_history(
    *,
    actor: str,
    event_type: str,
    status_value: str,
    ma_bao_cao: str,
    report: dict[str, Any] | None = None,
    report_name: str = "",
    rows: Any = 0,
    total: Any = 0,
    message: str = "",
    filters: dict[str, Any] | None = None,
    search: str = "",
    search_columns: list[str] | None = None,
    history_id: str = "",
    job_id: str = "",
    drive_url: str = "",
    download_url: str = "",
    file_name: str = "",
    details: dict[str, Any] | None = None,
) -> None:
    try:
        report_info = _dynamic_report_history_report(ma_bao_cao, report)
        if report_name:
            report_info["ten_bao_cao"] = report_name
        safe_history_id = history_id or job_id or uuid.uuid4().hex
        safe_rows = _dynamic_report_history_int(rows)
        safe_total = _dynamic_report_history_int(total, safe_rows)
        payload = {
            "history_id": safe_history_id,
            "event_type": str(event_type or "load").strip().lower() or "load",
            "status": str(status_value or "").strip().lower() or "unknown",
            "ma_bao_cao": report_info["ma_bao_cao"],
            "report_code": report_info["ma_bao_cao"],
            "ten_bao_cao": report_info["ten_bao_cao"],
            "report_name": report_info["ten_bao_cao"],
            "rows": safe_rows,
            "total": safe_total,
            "message": compact_log_text(message, 800),
            "filters": filters if isinstance(filters, dict) else {},
            "search": compact_log_text(search, 300),
            "search_columns": search_columns or [],
            "job_id": job_id,
            "drive_url": drive_url,
            "download_url": download_url,
            "file_name": file_name,
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "details": details if isinstance(details, dict) else {},
        }
        build_app_repository().add_audit_log(
            actor or "system",
            DYNAMIC_REPORT_HISTORY_ACTION,
            json.dumps(payload, ensure_ascii=False, default=str),
        )
    except Exception:
        logger.exception("Cannot write dynamic report history")


def _parse_dynamic_report_history_log(row: dict[str, Any]) -> dict[str, Any] | None:
    if row.get("action") != DYNAMIC_REPORT_HISTORY_ACTION:
        return None
    raw_details = row.get("details")
    if isinstance(raw_details, dict):
        details = raw_details
    else:
        try:
            details = json.loads(raw_details or "{}")
        except (TypeError, json.JSONDecodeError):
            return None
    if not isinstance(details, dict):
        return None
    history_id = str(details.get("history_id") or details.get("job_id") or row.get("id") or "").strip()
    rows = _dynamic_report_history_int(details.get("rows"))
    total = _dynamic_report_history_int(details.get("total"), rows)
    return {
        "id": row.get("id"),
        "history_id": history_id,
        "job_id": details.get("job_id") or "",
        "event_type": details.get("event_type") or "load",
        "status": details.get("status") or "unknown",
        "ma_bao_cao": details.get("ma_bao_cao") or details.get("report_code") or "",
        "report_code": details.get("report_code") or details.get("ma_bao_cao") or "",
        "ten_bao_cao": details.get("ten_bao_cao") or details.get("report_name") or "",
        "report_name": details.get("report_name") or details.get("ten_bao_cao") or "",
        "rows": rows,
        "total": total,
        "message": details.get("message") or "",
        "drive_url": details.get("drive_url") or "",
        "download_url": details.get("download_url") or "",
        "file_name": details.get("file_name") or "",
        "created_by": row.get("actor") or "",
        "created_at": details.get("created_at") or row.get("created_at") or "",
    }


def _list_dynamic_report_history(limit: int = 30) -> list[dict[str, Any]]:
    safe_limit = min(max(int(limit or 30), 1), 100)
    active_jobs = [
        _dynamic_report_export_job_response(str(job.get("job_id") or ""), job)
        for job in _list_dynamic_report_export_job_snapshots(limit=50)
        if str(job.get("status") or "").lower() in DYNAMIC_REPORT_EXPORT_ACTIVE_STATUSES
    ]
    rows = build_app_repository().list_audit_logs(limit=min(max(safe_limit * 8, 100), 800))
    items: list[dict[str, Any]] = active_jobs[:safe_limit]
    seen: set[str] = {str(item.get("history_id") or item.get("job_id") or item.get("id") or "") for item in items}
    for row in rows:
        item = _parse_dynamic_report_history_log(row)
        if not item:
            continue
        key = item.get("history_id") or str(item.get("id") or "")
        if key in seen:
            continue
        seen.add(key)
        items.append(item)
        if len(items) >= safe_limit:
            break
    return items


TIME_PATTERN = re.compile(r"^\d{2}:\d{2}$")


def normalize_clock_time(value: str, field_name: str) -> str:
    text = str(value or "").strip()[:5]
    if not TIME_PATTERN.match(text):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field_name} chua dung dinh dang HH:MM.")
    hour, minute = [int(part) for part in text.split(":", 1)]
    if hour > 23 or minute > 59:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field_name} chua dung dinh dang HH:MM.")
    return f"{hour:02d}:{minute:02d}"


def normalize_zalo_time_slots(values: list[str]) -> list[str]:
    slots = []
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        for part in re.split(r"[\s,;]+", text):
            if part:
                slots.append(normalize_clock_time(part, "Khung gio"))
    return sorted(set(slots))


def normalize_zalo_page_url(value: str) -> str:
    raw_value = str(value or "").strip() or "/"
    parsed = urlparse(raw_value)
    if parsed.scheme and parsed.scheme not in {"http", "https"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Duong dan trang chuc nang khong hop le.")
    if not parsed.scheme and not raw_value.startswith("/"):
        raw_value = f"/{raw_value}"
    return raw_value


def validate_zalo_photo_url(value: str) -> str:
    photo_url = str(value or "").strip()
    if not photo_url:
        return ""
    parsed = urlparse(photo_url)
    if parsed.scheme != "https" or not parsed.netloc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="URL anh gui Zalo phai la link HTTPS.")
    return photo_url


def normalize_zalo_auto_message_payload(payload: ZaloAutoMessagePayload, schedule_id: str) -> dict[str, Any]:
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ten lich gui Zalo khong duoc de trong.")
    time_slots = normalize_zalo_time_slots(payload.time_slots)
    run_time = normalize_clock_time(payload.run_time, "Gio gui")
    if payload.schedule_type == "TimeWindow" and not time_slots:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Lich theo khung gio can nhap it nhat mot khung gio.")
    if payload.schedule_type == "Weekly" and not payload.weekday.strip():
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Lich hang tuan can nhap thu trong tuan.")
    return {
        "schedule_id": schedule_id,
        "name": name,
        "page_url": normalize_zalo_page_url(payload.page_url),
        "page_label": payload.page_label.strip(),
        "schedule_type": payload.schedule_type,
        "time_slots": time_slots,
        "run_time": run_time,
        "weekday": payload.weekday.strip(),
        "month_day": min(max(int(payload.month_day or 1), 1), 31),
        "target_type": payload.target_type,
        "chat_id": payload.chat_id.strip(),
        "chat_name": payload.chat_name.strip(),
        "caption": payload.caption.strip(),
        "photo_url": validate_zalo_photo_url(payload.photo_url),
        "is_active": payload.is_active,
    }


def normalize_data_mining_schedule_payload(payload: DataMiningSchedulePayload, schedule_id: str) -> dict[str, Any]:
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ten lich dao du lieu khong duoc de trong.")
    try:
        report_url = normalize_onebss_report_url(payload.report_url)
    except OneBssDownloadError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    run_time = str(payload.run_time or "07:00").strip()
    if not re.fullmatch(r"\d{2}:\d{2}", run_time):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Thoi gian lay phai dung dinh dang HH:MM.")
    month_day = min(max(int(payload.month_day or 1), 1), 31)
    return {
        "schedule_id": schedule_id,
        "name": name,
        "report_url": report_url,
        "schedule_type": payload.schedule_type,
        "run_time": run_time,
        "weekday": payload.weekday.strip(),
        "month_day": month_day,
        "storage_link": payload.storage_link.strip(),
        "file_name_template": payload.file_name_template.strip(),
        "parameters": payload.parameters if isinstance(payload.parameters, dict) else {},
        "is_active": bool(payload.is_active),
    }


def decode_capture_image(payload: ZaloCapturePayload) -> tuple[str, str]:
    mime_type = payload.mime_type.strip().lower() or "image/png"
    if mime_type not in {"image/png", "image/jpeg"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chi ho tro anh PNG hoac JPEG.")
    raw_image = payload.image_base64.strip()
    if raw_image.startswith("data:"):
        header, _, body = raw_image.partition(",")
        if "image/jpeg" in header:
            mime_type = "image/jpeg"
        elif "image/png" in header:
            mime_type = "image/png"
        raw_image = body
    try:
        image_bytes = base64.b64decode(raw_image, validate=True)
    except (binascii.Error, ValueError) as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Du lieu anh chup khong hop le.") from error
    if len(image_bytes) > 10 * 1024 * 1024:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Anh chup qua lon, hay chup vung gon hon.")
    if mime_type == "image/png" and not image_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File anh khong phai PNG hop le.")
    if mime_type == "image/jpeg" and not image_bytes.startswith(b"\xff\xd8"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File anh khong phai JPEG hop le.")
    return base64.b64encode(image_bytes).decode("ascii"), mime_type


def enrich_zalo_auto_message(repository: AppRepository, schedule: dict[str, Any]) -> dict[str, Any]:
    latest_capture = repository.get_latest_zalo_message_capture(str(schedule.get("schedule_id") or ""), include_image=False)
    schedule["latest_capture"] = latest_capture
    schedule["latest_capture_url"] = capture_public_url(get_settings(), latest_capture)
    return schedule


def normalize_dashboard_code(value: Any, label: str, *, uppercase: bool = True) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{label} không được để trống.")
    if not re.fullmatch(r"[A-Za-z0-9_-]+", normalized):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{label} chỉ được chứa chữ, số, dấu gạch dưới hoặc gạch ngang.")
    return normalized.upper() if uppercase else normalized


def normalize_dashboard_sql_code(value: Any) -> str:
    raw_value = str(value or "").strip()
    candidate = raw_value
    if "(" in candidate:
        candidate = candidate.split("(", 1)[0].strip()
    if not re.fullmatch(r"[A-Za-z0-9_-]+", candidate):
        match = re.search(r"[A-Za-z0-9_-]+", raw_value)
        candidate = match.group(0) if match else raw_value
    return normalize_dashboard_code(candidate, "Mã SQL")


def normalize_dashboard_layout(payload: DashboardLayoutPayload) -> tuple[str, str, dict[str, Any]]:
    page_id = normalize_dashboard_code(payload.page_id, "Mã trang")
    page_name = payload.page_name.strip() or page_id
    raw_layout = payload.layout if isinstance(payload.layout, dict) else {}
    tabs = raw_layout.get("tabs")
    if not isinstance(tabs, list) or not tabs:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Dashboard cần có ít nhất một Tab.")

    normalized_tabs = []
    seen_tabs: set[str] = set()
    for tab_index, tab in enumerate(tabs, start=1):
        if not isinstance(tab, dict):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cấu trúc Tab không hợp lệ.")
        tab_id = normalize_dashboard_code(tab.get("tab_id") or f"tab_{tab_index}", "Mã Tab", uppercase=False)
        if tab_id in seen_tabs:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Mã Tab {tab_id} bị trùng.")
        seen_tabs.add(tab_id)
        tab_name = str(tab.get("tab_name") or f"Tab {tab_index}").strip()
        if not tab_name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên Tab không được để trống.")

        normalized_rows = []
        grid_layout = tab.get("grid_layout") if isinstance(tab.get("grid_layout"), list) else []
        for row_index, row in enumerate(grid_layout, start=1):
            if not isinstance(row, dict):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cấu trúc dòng Layout không hợp lệ.")
            layout_type = str(row.get("layout_type") or "").strip()
            if layout_type not in DASHBOARD_LAYOUT_TYPES:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Loại Layout không hợp lệ.")
            max_position = DASHBOARD_LAYOUT_TYPES[layout_type]
            row_id = int(row.get("row_id") or row_index)
            normalized_widgets = []
            for widget in row.get("widgets") or []:
                if not isinstance(widget, dict):
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cấu trúc hiển thị không hợp lệ.")
                widget_type = str(widget.get("type") or "").strip()
                sql_code = str(widget.get("sql_code") or "").strip()
                title = str(widget.get("title") or "").strip()
                text_content = str(widget.get("text_content") or "").strip()
                raw_chart_config = widget.get("chart_config") if isinstance(widget.get("chart_config"), dict) else {}
                chart_config = {str(key).strip(): value for key, value in raw_chart_config.items() if str(key).strip()}
                if widget_type == "text_title":
                    if not title and not text_content:
                        continue
                elif widget_type == "google_sheet_embed":
                    if not str(chart_config.get("embed_url") or "").strip():
                        continue
                elif not sql_code:
                    continue
                if widget_type not in DASHBOARD_WIDGET_TYPES:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Loại hiển thị không hợp lệ.")
                position = int(widget.get("position") or 0)
                if position < 1 or position > max_position:
                    raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Vị trí hiển thị không khớp số cột Layout.")
                raw_filters = widget.get("filters") if isinstance(widget.get("filters"), dict) else {}
                filters = {str(key).strip(): value for key, value in raw_filters.items() if str(key).strip()}
                raw_report_id = widget.get("report_id")
                try:
                    report_id = int(raw_report_id) if raw_report_id not in (None, "") else None
                except (TypeError, ValueError):
                    report_id = None
                normalized_widgets.append({
                    "position": position,
                    "type": widget_type,
                    "title": title,
                    "sql_code": normalize_dashboard_sql_code(sql_code) if sql_code else "",
                    "report_id": report_id,
                    "filters": filters,
                    "chart_config": chart_config,
                    "text_content": text_content,
                    "icon_url": str(widget.get("icon_url") or "").strip(),
                })
            normalized_rows.append({
                "row_id": row_id,
                "layout_type": layout_type,
                "widgets": sorted(normalized_widgets, key=lambda item: item["position"]),
            })

        normalized_tabs.append({
            "tab_id": tab_id,
            "tab_name": tab_name,
            "order": tab_index,
            "grid_layout": normalized_rows,
        })

    return page_id, page_name, {"page_id": page_id, "tabs": normalized_tabs}


def dashboard_feature_has_ancestor(code: str, parent_by_code: dict[str, str | None], ancestor_code: str) -> bool:
    current_code: str | None = code
    visited: set[str] = set()
    while current_code and current_code not in visited:
        if current_code == ancestor_code:
            return True
        visited.add(current_code)
        current_code = parent_by_code.get(current_code)
    return False


def dashboard_page_id_from_feature_code(code: str) -> str:
    if code == "dashboard":
        return DEFAULT_DASHBOARD_PAGE_ID
    page_id = re.sub(r"[^A-Za-z0-9]+", "", code).upper()
    return page_id or DEFAULT_DASHBOARD_PAGE_ID


def is_dashboard_layout_feature(code: str, parent_by_code: dict[str, str | None]) -> bool:
    if code in DASHBOARD_LAYOUT_EXCLUDED_FEATURE_CODES:
        return False
    return (
        dashboard_feature_has_ancestor(code, parent_by_code, "dashboard")
        or dashboard_feature_has_ancestor(code, parent_by_code, "baocaomoi")
        or dashboard_feature_has_ancestor(code, parent_by_code, "truyvansql")
        or dashboard_feature_has_ancestor(code, parent_by_code, "new_reports")
        or dashboard_feature_has_ancestor(code, parent_by_code, "reports")
    )


def dashboard_layout_parent_candidates(features: list[dict]) -> list[dict]:
    candidates: list[dict] = []
    for feature in features:
        code = str(feature.get("code") or "")
        if not code or feature.get("parent_code"):
            continue
        if code in DASHBOARD_LAYOUT_PARENT_EXCLUDED_FEATURE_CODES:
            continue
        candidates.append(feature)
    return candidates


def validate_dashboard_layout_parent_code(features: list[dict], parent_code: str | None) -> str | None:
    raw_parent_code = str(parent_code or "").strip()
    if not raw_parent_code:
        return None
    normalized_parent_code = normalize_feature_code(raw_parent_code)
    selected_feature = next((feature for feature in features if str(feature.get("code") or "") == raw_parent_code), None)
    if not selected_feature and normalized_parent_code:
        selected_feature = next(
            (feature for feature in features if normalize_feature_code(feature.get("code")) == normalized_parent_code),
            None,
        )
    valid_parent_codes = {str(feature.get("code") or "") for feature in dashboard_layout_parent_candidates(features)}
    if not selected_feature or str(selected_feature.get("code") or "") not in valid_parent_codes:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Mục menu cha không hợp lệ.")
    return str(selected_feature.get("code") or "")


def feature_parent_code_for_page(features: list[dict], page_id: str) -> tuple[str, str | None]:
    feature_code = dashboard_feature_code_for_page(page_id)
    feature = next((item for item in features if str(item.get("code") or "") == feature_code), None)
    if not feature:
        feature = next(
            (item for item in features if normalize_feature_code(item.get("code")) == normalize_feature_code(feature_code)),
            None,
        )
    return (str(feature.get("code") or feature_code) if feature else feature_code, feature.get("parent_code") if feature else None)


def visible_features_for_user(features: list[dict], user: dict) -> list[dict]:
    if user.get("role") == "admin":
        return features
    allowed_codes = {str(code) for code in user.get("permissions", []) if str(code)}
    by_code = {str(feature.get("code") or ""): feature for feature in features if feature.get("code")}
    visible_codes: set[str] = set()
    for code in allowed_codes:
        current_code = code
        visited: set[str] = set()
        while current_code and current_code in by_code and current_code not in visited:
            visited.add(current_code)
            visible_codes.add(current_code)
            current_code = str(by_code[current_code].get("parent_code") or "")
    return [feature for feature in features if str(feature.get("code") or "") in visible_codes]


def visible_dashboard_layouts_for_user(layouts: list[dict], user: dict) -> list[dict]:
    if user.get("role") == "admin":
        return layouts
    allowed_codes = {str(code) for code in user.get("permissions", []) if str(code)}
    normalized_allowed_codes = {normalize_feature_code(code) for code in allowed_codes}
    visible_layouts: list[dict] = []
    for layout in layouts:
        page_id = str(layout.get("page_id") or "")
        feature_code = dashboard_feature_code_for_page(page_id)
        if feature_code in allowed_codes or normalize_feature_code(feature_code) in normalized_allowed_codes:
            visible_layouts.append(layout)
    return visible_layouts


def user_can_view_dashboard_page(user: dict, page_id: str, features: list[dict]) -> bool:
    if user.get("role") == "admin":
        return True
    feature_code, _ = feature_parent_code_for_page(features, page_id)
    allowed_codes = {str(code) for code in user.get("permissions", []) if str(code)}
    normalized_allowed_codes = {normalize_feature_code(code) for code in allowed_codes}
    return feature_code in allowed_codes or normalize_feature_code(feature_code) in normalized_allowed_codes


def require_dashboard_page_access(request: Request, page_id: str, features: list[dict]) -> dict:
    user = current_user(request)
    if not user_can_view_dashboard_page(user, page_id, features):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bạn chưa được cấp quyền sử dụng chức năng này.")
    return user


def dashboard_layout_response(repository: AppRepository, page_id: str, features: list[dict]) -> dict:
    try:
        layout = repository.get_dashboard_layout(page_id)
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)
    if not layout:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy trang dashboard.")
    feature_code, parent_code = feature_parent_code_for_page(features, page_id)
    layout["feature_code"] = feature_code
    layout["parent_code"] = parent_code
    if isinstance(layout.get("layout"), dict):
        layout["layout"]["parent_code"] = parent_code
    return layout


def build_dashboard_layout_pages(features: list[dict], layouts: list[dict]) -> list[dict]:
    parent_by_code = {str(feature.get("code") or ""): feature.get("parent_code") for feature in features}
    layout_by_id = {str(layout.get("page_id") or ""): layout for layout in layouts if layout.get("page_id")}
    layout_by_feature_code = {dashboard_feature_code_for_page(page_id): layout for page_id, layout in layout_by_id.items()}
    layout_feature_codes = set(layout_by_feature_code)
    designable_codes = {
        code for code in parent_by_code
        if code
        and code not in DASHBOARD_LAYOUT_EXCLUDED_FEATURE_CODES
        and (is_dashboard_layout_feature(code, parent_by_code) or code in layout_feature_codes)
    }
    non_designable_page_ids = {
        dashboard_page_id_from_feature_code(code)
        for code in parent_by_code
        if code and code not in designable_codes
    }

    pages: list[dict] = []
    included_page_ids: set[str] = set()
    ordered_features = sorted(
        features,
        key=lambda feature: (
            int(feature.get("sort_order") or 0),
            str(feature.get("name") or ""),
            str(feature.get("code") or ""),
        ),
    )
    for feature in ordered_features:
        code = str(feature.get("code") or "")
        if code not in designable_codes:
            continue
        normalized_code = normalize_feature_code(code)
        layout = layout_by_feature_code.get(code) or layout_by_feature_code.get(normalized_code)
        page_id = str(layout.get("page_id") or "") if layout else dashboard_page_id_from_feature_code(code)
        if page_id in included_page_ids:
            continue
        pages.append({
            "page_id": page_id,
            "page_name": str(feature.get("name") or (layout.get("page_name") if layout else page_id)),
            "layout_page_name": layout.get("page_name") if layout else "",
            "feature_code": code,
            "feature_name": str(feature.get("name") or ""),
            "parent_code": feature.get("parent_code"),
            "saved": bool(layout),
            "unsaved": not bool(layout),
            "created_at": layout.get("created_at") if layout else None,
            "updated_at": layout.get("updated_at") if layout else None,
        })
        included_page_ids.add(page_id)

    for layout in layouts:
        page_id = str(layout.get("page_id") or "")
        if not page_id or page_id in included_page_ids or page_id in non_designable_page_ids:
            continue
        pages.append({
            "page_id": page_id,
            "page_name": layout.get("page_name") or page_id,
            "layout_page_name": layout.get("page_name") or page_id,
            "feature_code": "",
            "feature_name": "",
            "parent_code": None,
            "saved": True,
            "unsaved": False,
            "created_at": layout.get("created_at"),
            "updated_at": layout.get("updated_at"),
        })
        included_page_ids.add(page_id)
    return pages


def validate_report_sql(sql: str) -> str:
    normalized = sql.strip()
    lowered = normalized.lower()
    if not (lowered.startswith("select") or lowered.startswith("define ")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Câu SQL phải bắt đầu bằng DEFINE hoặc SELECT.")
    if not normalized.endswith(";"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Câu SQL báo cáo phải kết thúc bằng dấu chấm phẩy (;).")
    executable_part = normalized[:-1].strip()
    statements = [part.strip() for part in executable_part.split(";") if part.strip()]
    if len(statements) != 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chỉ cho phép một câu SQL báo cáo và dấu chấm phẩy ở cuối.")
    non_define_lines = [line.strip() for line in executable_part.splitlines() if line.strip() and not line.strip().lower().startswith("define ")]
    if not non_define_lines or not non_define_lines[0].lower().startswith("select"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Sau các dòng DEFINE phải là câu lệnh SELECT.")
    blocked_words = (" insert ", " update ", " delete ", " drop ", " alter ", " truncate ", " merge ")
    padded = f" {executable_part.lower()} "
    if any(word in padded for word in blocked_words):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Câu SQL báo cáo không được chứa lệnh thay đổi dữ liệu.")
    return normalized


def validate_feature_layout_payload(existing_features: list[dict], items: list[FeatureLayoutItem]) -> None:
    valid_codes = {feature["code"] for feature in existing_features}
    parent_by_code = {feature["code"]: feature.get("parent_code") for feature in existing_features}
    seen_codes: set[str] = set()

    for item in items:
        if item.code in seen_codes:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Chức năng {item.code} bị lặp trong cấu trúc menu.")
        seen_codes.add(item.code)
        if item.code not in valid_codes:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Chức năng {item.code} không hợp lệ.")
        if not item.name.strip():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Tên hiển thị của {item.code} không được để trống.")
        if item.parent_code and item.parent_code not in valid_codes:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Chức năng cha {item.parent_code} không hợp lệ.")
        if item.parent_code == item.code:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chức năng cha không được trùng chính nó.")
        parent_by_code[item.code] = item.parent_code

    for code in valid_codes:
        visited: set[str] = set()
        parent_code = parent_by_code.get(code)
        while parent_code:
            if parent_code == code or parent_code in visited:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cấu trúc menu không được tạo vòng lặp cha/con.")
            visited.add(parent_code)
            parent_code = parent_by_code.get(parent_code)


def notify_login_failed_threshold(request: Request, username: str) -> None:
    display_username = username.strip() or "unknown"
    client_host = request.client.host if request.client else "unknown"
    repository = build_app_repository()
    try:
        failures = repository.record_login_failure(display_username, client_host)
    except RuntimeError as error:
        try:
            repository.add_audit_log("system", "login_attempt_counter_failed", str(error)[:500])
        except Exception:
            pass
        failures = 5
    if failures >= 5:
        sent = TelegramNotifier(get_settings()).send_message(
            "Canh bao dang nhap sai",
            f"Tai khoan {display_username} dang nhap sai {failures} lan lien tiep.",
            {"ip": client_host, "nguong_canh_bao": 5, "so_lan_sai": failures},
        )
        try:
            build_app_repository().add_audit_log(
                "system",
                "telegram_login_alert_sent" if sent else "telegram_login_alert_failed",
                f"Login failed alert for {display_username}: {failures} failures",
            )
        except Exception:
            pass


def reset_failed_login_counter(username: str) -> None:
    try:
        build_app_repository().reset_login_failures(username)
    except Exception:
        pass


def normalize_email_username(email: str) -> str:
    return (email or "").strip().split("@", 1)[0].lower()


def normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_employee_workbook(content: bytes) -> list[dict]:
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_cell(cell.value).upper() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    index = {header: position for position, header in enumerate(headers)}
    required = ["MÃ NHÂN VIÊN", "TÊN NHÂN VIÊN", "THƯ ĐIỆN TỬ"]
    if not all(column in index for column in required):
        raise ValueError("File Excel thiếu cột MÃ NHÂN VIÊN, TÊN NHÂN VIÊN hoặc THƯ ĐIỆN TỬ.")
    employees = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        employee_code = normalize_cell(row[index["MÃ NHÂN VIÊN"]])
        full_name = normalize_cell(row[index["TÊN NHÂN VIÊN"]])
        email = normalize_cell(row[index["THƯ ĐIỆN TỬ"]]).lower()
        if not employee_code and not full_name and not email:
            continue
        employees.append({
            "employee_code": employee_code,
            "full_name": full_name,
            "email": email,
            "phone": normalize_cell(row[index.get("SỐ ĐIỆN THOẠI", -1)]) if "SỐ ĐIỆN THOẠI" in index else "",
            "birth_date": normalize_cell(row[index.get("NGÀY SINH", -1)])[:10] if "NGÀY SINH" in index else "",
            "gender": normalize_cell(row[index.get("GIỚI TÍNH", -1)]) if "GIỚI TÍNH" in index else "",
            "department": normalize_cell(row[index.get("PHÒNG BAN", -1)]) if "PHÒNG BAN" in index else "",
            "job_title": normalize_cell(row[index.get("VTCV", -1)]) if "VTCV" in index else "",
        })
    return employees



def current_user(request: Request) -> dict:
    session_user = request.session.get("user")
    if not session_user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Bạn chưa đăng nhập.")
    user = build_app_repository().get_user_by_id(int(session_user["id"]))
    if not user or not user["is_active"]:
        request.session.clear()
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Phiên đăng nhập không còn hợp lệ.")
    public = AuthService.public_user(user)
    public["permissions"] = build_app_repository().get_user_permissions(public["id"])
    return public


def admin_user(request: Request) -> dict:
    user = current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail=ADMIN_ONLY_MESSAGE)
    return user


def onebss_worker_token(request: Request) -> str:
    settings = get_settings()
    expected = settings.internal_api_token.get_secret_value().strip() if getattr(settings, "internal_api_token", None) else ""
    if not expected:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Chua cau hinh INTERNAL_API_TOKEN cho may tram OneBSS.")
    provided = request.headers.get("x-worker-token", "").strip()
    authorization = request.headers.get("authorization", "").strip()
    if not provided and authorization.lower().startswith("bearer "):
        provided = authorization[7:].strip()
    if not provided or not hmac.compare_digest(provided, expected):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token may tram OneBSS khong hop le.")
    return provided


def require_feature(request: Request, feature_code: str) -> dict:
    user = current_user(request)
    if user["role"] == "admin":
        return user
    if feature_code not in build_app_repository().get_user_permissions(user["id"]):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Bạn chưa được cấp quyền sử dụng chức năng này.")
    return user


def normalize_google_sheet_public_url(raw_url: str) -> str:
    parsed = urlparse(str(raw_url or "").strip())
    if parsed.scheme != "https" or parsed.netloc != "docs.google.com" or not parsed.path.startswith("/spreadsheets/"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Link Google Sheet không hợp lệ.")
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["headers"] = "false"
    query["widget"] = "false"
    query.setdefault("single", "true")
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, parsed.params, urlencode(query), parsed.fragment))


def google_drive_oauth_serializer(settings: Any) -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(settings.session_secret.get_secret_value(), salt="google-drive-oauth")


def google_drive_oauth_redirect_uri(request: Request) -> str:
    settings = get_settings()
    configured = str(getattr(settings, "google_drive_oauth_redirect_uri", "") or "").strip()
    if configured:
        return configured
    public_base = str(getattr(settings, "app_public_url", "") or "").strip().rstrip("/")
    if not public_base:
        forwarded_proto = (request.headers.get("x-forwarded-proto") or "").split(",", 1)[0].strip()
        forwarded_host = (request.headers.get("x-forwarded-host") or request.headers.get("host") or "").split(",", 1)[0].strip()
        if forwarded_proto and forwarded_host:
            public_base = f"{forwarded_proto}://{forwarded_host}"
        else:
            public_base = str(request.base_url).rstrip("/")
    return f"{public_base}/api/google-drive/oauth/callback"


def google_drive_oauth_result_page(ok: bool, message: str) -> HTMLResponse:
    safe_message = html_escape(message)
    safe_status = "true" if ok else "false"
    return HTMLResponse(
        f"""<!doctype html>
<html lang="vi"><head><meta charset="utf-8"><title>Google Drive</title>
<style>body{{font-family:Arial,sans-serif;background:#082f49;color:#fff;margin:0;display:grid;min-height:100vh;place-items:center}}main{{max-width:520px;padding:28px;border:1px solid rgba(125,211,252,.35);border-radius:12px;background:rgba(2,6,23,.5)}}a{{color:#7dd3fc}}</style>
</head><body><main><h1>{'Da ket noi Google Drive' if ok else 'Ket noi Google Drive loi'}</h1><p>{safe_message}</p><p>Co the dong cua so nay va quay lai trang quan tri.</p><p><a href="/">Quay lai he thong</a></p></main>
<script>
try {{
  if (window.opener) {{
    window.opener.postMessage({{ type: "google-drive-oauth", ok: {safe_status}, message: {json.dumps(message)} }}, window.location.origin);
    setTimeout(() => window.close(), 1000);
  }}
}} catch (error) {{}}
</script></body></html>""",
        status_code=status.HTTP_200_OK if ok else status.HTTP_400_BAD_REQUEST,
    )


def is_sensitive_connection_key(key: str) -> bool:
    lowered = str(key or "").lower()
    return any(part in lowered for part in ("pass", "secret", "token", "credential", "private_key", "client_secret"))


def public_connection_config(config: dict[str, Any]) -> tuple[dict[str, Any], list[str]]:
    public: dict[str, Any] = {}
    protected: list[str] = []
    for key, value in (config or {}).items():
        if is_sensitive_connection_key(str(key)):
            protected.append(str(key))
        else:
            public[key] = value
    return public, protected


def merge_protected_connection_config(existing: dict[str, Any] | None, incoming: dict[str, Any] | None) -> dict[str, Any]:
    merged = dict(incoming or {})
    for key, value in (existing or {}).items():
        if is_sensitive_connection_key(str(key)) and key not in merged:
            merged[key] = value
    return merged


@router.get("/login", response_class=HTMLResponse)
def login_page(request: Request) -> Response:
    next_path = request.query_params.get("next") or "/"
    if request.session.get("user"):
        if not next_path.startswith("/") or next_path.startswith("//"):
            next_path = "/"
        return RedirectResponse(next_path, status_code=status.HTTP_303_SEE_OTHER)
    return templates.TemplateResponse(
        request=request,
        name="login.html",
        context={"app_name": get_settings().app_name},
    )


@router.get("/favicon.ico", include_in_schema=False)
def favicon() -> Response:
    return RedirectResponse("/static/images/system-logo.png", status_code=status.HTTP_307_TEMPORARY_REDIRECT)


@router.get("/", response_class=HTMLResponse)
def dashboard(request: Request) -> Response:
    return render_index_page(request, "")


def render_index_page(request: Request, feature_path: str) -> Response:
    try:
        user = current_user(request)
    except HTTPException:
        next_path = "/" + feature_path.strip("/") if feature_path else "/"
        next_query = f"?next={quote(next_path, safe='/')}" if next_path != "/" else ""
        return RedirectResponse(f"/login{next_query}", status_code=status.HTTP_303_SEE_OTHER)
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"app_name": get_settings().app_name, "user": user},
    )


@router.post("/api/auth/login")
def login(request: Request, payload: LoginPayload) -> dict:
    user = build_auth_service().authenticate(payload.username, payload.password)
    if not user:
        notify_login_failed_threshold(request, payload.username)
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Tên đăng nhập hoặc mật khẩu không đúng.")
    reset_failed_login_counter(payload.username)
    request.session.clear()
    request.session["user"] = user
    return {"ok": True, "user": user}


@router.post("/api/auth/logout")
def logout(request: Request) -> dict:
    user = request.session.get("user")
    if user:
        build_app_repository().add_audit_log(user["username"], "logout", "Đăng xuất")
    request.session.clear()
    return {"ok": True}


@router.get("/api/auth/me")
def me(request: Request) -> dict:
    return {"user": current_user(request)}


@router.post("/api/auth/change-password")
def change_password(request: Request, payload: ChangePasswordPayload) -> dict:
    user = current_user(request)
    try:
        build_auth_service().change_own_password(
            user["id"], user["username"], payload.current_password, payload.new_password
        )
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    user["must_change_password"] = 0
    request.session["user"] = user
    return {"ok": True, "message": "Đổi mật khẩu thành công."}


@router.get("/api/health/database")
def database_health(request: Request) -> dict:
    current_user(request)
    result = build_database_service().get_connection_status()
    notify_if_failed("Lỗi kết nối API dữ liệu nội bộ", result)
    return result


@router.get("/api/system/status")
def system_status(request: Request) -> dict:
    current_user(request)
    settings = get_settings()
    return {
        "internal_api": {
            "url": settings.internal_api_url,
            "mock_mode": settings.internal_api_mock_mode,
        },
        "query_policy": {
            "pagination_required": True,
            "page_size_min": 20,
            "page_size_max": 50,
            "data_source": "internal_fastapi",
        },
    }


@router.get("/api/dashboard/datcoc-test")
def dashboard_datcoc_test(request: Request) -> dict:
    current_user(request)
    return build_database_service().run_dashboard_datcoc_test()


@router.get("/api/dashboard/fiber")
def dashboard_fiber(request: Request) -> dict:
    current_user(request)
    return build_database_service().run_dashboard_fiber()


@router.get("/api/admin/users")
def list_users(request: Request) -> dict:
    admin_user(request)
    return {"users": build_app_repository().list_users()}


@router.post("/api/admin/users")
def create_user(request: Request, payload: CreateUserPayload) -> dict:
    actor = admin_user(request)
    try:
        user = build_auth_service().create_user(
            actor["username"], payload.username, payload.full_name, payload.password, payload.role
        )
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    return {"ok": True, "user": user}


@router.get("/api/admin/users/import-template")
def download_user_import_template(request: Request) -> Response:
    admin_user(request)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "NguoiDung"
    sheet.append(["MÃ NHÂN VIÊN", "TÊN NHÂN VIÊN", "SỐ ĐIỆN THOẠI", "THƯ ĐIỆN TỬ", "NGÀY SINH", "GIỚI TÍNH", "PHÒNG BAN", "VTCV"])
    sheet.append(["VNPT008888", "Nguyen Van A", "0912345678", "vana@vnpt.vn", "1990-01-01", "Nam", "Phong kinh doanh", "Nhan vien"])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="mau_import_nguoi_dung.xlsx"'},
    )


@router.post("/api/admin/users/import")
async def import_users(request: Request, file: UploadFile = File(...)) -> dict:
    actor = admin_user(request)
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chỉ hỗ trợ file Excel .xlsx.")
    content = await file.read(MAX_USER_IMPORT_BYTES + 1)
    if len(content) > MAX_USER_IMPORT_BYTES:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File import quá lớn. Giới hạn hiện tại là 2MB.")
    try:
        employees = parse_employee_workbook(content)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    repository = build_app_repository()
    auth_service = AuthService(repository)
    created = []
    skipped = []
    for employee in employees:
        username = normalize_email_username(employee["email"])
        if not employee["employee_code"] or not employee["full_name"] or not username:
            skipped.append({"employee_code": employee.get("employee_code"), "reason": "Thiếu mã nhân viên, họ tên hoặc email."})
            continue
        try:
            existing_user = repository.get_user_by_employee_or_email(employee["employee_code"], employee["email"])
        except RuntimeError as error:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
        if existing_user:
            skipped.append({"employee_code": employee["employee_code"], "reason": "Mã nhân viên hoặc email đã tồn tại."})
            continue
        try:
            user = auth_service.create_user(actor["username"], username, employee["full_name"], employee["employee_code"], "viewer", employee)
            created.append(user)
        except (ValueError, sqlite3.IntegrityError, RuntimeError) as error:
            skipped.append({"employee_code": employee["employee_code"], "reason": str(error)})
    repository.add_audit_log(actor["username"], "users_imported", f"Import users: created={len(created)}, skipped={len(skipped)}")
    return {"ok": True, "created_count": len(created), "skipped_count": len(skipped), "created": created, "skipped": skipped}


@router.put("/api/admin/users/{user_id}")
def update_user(request: Request, user_id: int, payload: UpdateUserPayload) -> dict:
    actor = admin_user(request)
    try:
        user = build_auth_service().update_user(
            actor["username"], actor["id"], user_id, payload.full_name, payload.role, payload.is_active
        )
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    return {"ok": True, "user": user}


@router.delete("/api/admin/users/{user_id}")
def delete_user(request: Request, user_id: int) -> dict:
    actor = admin_user(request)
    user = build_app_repository().get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy người dùng.")
    if user["id"] == actor["id"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Bạn không thể tự xóa tài khoản đang đăng nhập.")
    if user["role"] == "admin" and build_app_repository().count_active_admins() <= 1:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Hệ thống phải còn ít nhất một admin.")
    build_app_repository().delete_user(user_id)
    build_app_repository().add_audit_log(actor["username"], "user_deleted", f"Xoa user {user['username']}")
    return {"ok": True}


@router.post("/api/admin/users/{user_id}/reset-password")
def reset_password(request: Request, user_id: int, payload: PasswordPayload) -> dict:
    actor = admin_user(request)
    try:
        build_auth_service().reset_password(actor["username"], user_id, payload.password)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    return {"ok": True, "message": "Đặt lại mật khẩu thành công."}


@router.get("/api/admin/audit-logs")
def audit_logs(request: Request) -> dict:
    admin_user(request)
    return {"logs": build_app_repository().list_audit_logs()}


@router.get("/api/admin/system")
def system_info(request: Request) -> dict:
    admin_user(request)
    settings = get_settings()
    users = build_app_repository().list_users()
    return {
        "app_name": settings.app_name,
        "environment": settings.app_env,
        "internal_api_url": settings.internal_api_url,
        "internal_api_mock_mode": settings.internal_api_mock_mode,
        "storage_backend": settings.app_database_backend,
        "supabase_rest_url": settings.supabase_rest_url,
        "user_count": len(users),
        "active_user_count": sum(1 for user in users if user["is_active"]),
    }


@router.get("/api/admin/storage/health")
def storage_health(request: Request) -> dict:
    admin_user(request)
    repository = build_app_repository()
    if hasattr(repository, "health_check"):
        return repository.health_check()
    return {"ok": True, "backend": "sqlite"}


@router.get("/api/google-drive/oauth/status")
def google_drive_oauth_status_api(request: Request) -> dict:
    admin_user(request)
    return google_drive_oauth_status(get_settings(), build_app_repository())


@router.post("/api/google-drive/oauth/start")
def google_drive_oauth_start(request: Request) -> dict:
    actor = admin_user(request)
    settings = get_settings()
    if not google_drive_oauth_client_configured(settings):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Chua cau hinh GOOGLE_DRIVE_OAUTH_CLIENT_ID/GOOGLE_DRIVE_OAUTH_CLIENT_SECRET "
                "tren Render. Redirect URI can khai bao tren Google Cloud: "
                f"{google_drive_oauth_redirect_uri(request)}"
            ),
        )
    redirect_uri = google_drive_oauth_redirect_uri(request)
    state_value = google_drive_oauth_serializer(settings).dumps(
        {"user_id": actor["id"], "username": actor["username"], "redirect_uri": redirect_uri}
    )
    query = urlencode(
        {
            "client_id": settings.google_drive_oauth_client_id,
            "redirect_uri": redirect_uri,
            "response_type": "code",
            "scope": " ".join(GOOGLE_DRIVE_OAUTH_SCOPES),
            "access_type": "offline",
            "prompt": "consent",
            "include_granted_scopes": "true",
            "state": state_value,
        }
    )
    return {
        "authorization_url": f"https://accounts.google.com/o/oauth2/v2/auth?{query}",
        "redirect_uri": redirect_uri,
    }


@router.get("/api/google-drive/oauth/callback", response_class=HTMLResponse)
def google_drive_oauth_callback(request: Request) -> HTMLResponse:
    settings = get_settings()
    error_text = request.query_params.get("error") or ""
    if error_text:
        return google_drive_oauth_result_page(False, f"Google tu choi ket noi: {error_text}")
    code = request.query_params.get("code") or ""
    state_value = request.query_params.get("state") or ""
    if not code or not state_value:
        return google_drive_oauth_result_page(False, "Google callback thieu code/state.")
    try:
        state_payload = google_drive_oauth_serializer(settings).loads(state_value, max_age=15 * 60)
    except SignatureExpired:
        return google_drive_oauth_result_page(False, "Phien ket noi Google Drive da het han. Hay bam ket noi lai.")
    except BadSignature:
        return google_drive_oauth_result_page(False, "Ma xac thuc Google Drive khong hop le.")
    try:
        actor = admin_user(request)
    except HTTPException:
        return google_drive_oauth_result_page(False, "Hay dang nhap trang quan tri roi ket noi Google Drive lai.")
    if str(state_payload.get("username") or "") != str(actor.get("username") or ""):
        return google_drive_oauth_result_page(False, "Tai khoan quan tri khong khop phien ket noi Google Drive.")
    redirect_uri = str(state_payload.get("redirect_uri") or google_drive_oauth_redirect_uri(request))
    try:
        with httpx.Client(timeout=30) as client:
            token_response = client.post(
                "https://oauth2.googleapis.com/token",
                data={
                    "code": code,
                    "client_id": settings.google_drive_oauth_client_id,
                    "client_secret": settings.google_drive_oauth_client_secret.get_secret_value(),
                    "redirect_uri": redirect_uri,
                    "grant_type": "authorization_code",
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
            token_data = token_response.json() if token_response.content else {}
            if token_response.status_code >= 400:
                return google_drive_oauth_result_page(
                    False,
                    str(token_data.get("error_description") or token_data.get("error") or "Google khong tra token."),
                )
            email = ""
            access_token = str(token_data.get("access_token") or "")
            if access_token:
                user_response = client.get(
                    "https://www.googleapis.com/oauth2/v2/userinfo",
                    headers={"Authorization": f"Bearer {access_token}"},
                )
                if user_response.status_code < 400:
                    user_info = user_response.json()
                    email = str(user_info.get("email") or "")
    except Exception as error:
        logger.exception("Cannot finish Google Drive OAuth")
        return google_drive_oauth_result_page(False, f"Khong hoan tat ket noi Google Drive: {str(error)[:300]}")
    token_data["connected_at"] = datetime.now().isoformat(timespec="seconds")
    try:
        save_google_drive_oauth_tokens(settings, build_app_repository(), token_data, email=email)
    except GoogleDriveConfigurationError as error:
        return google_drive_oauth_result_page(False, str(error))
    build_app_repository().add_audit_log(actor["username"], "google_drive_oauth_connected", "Ket noi Google Drive OAuth")
    return google_drive_oauth_result_page(True, f"Da ket noi Google Drive{f' voi {email}' if email else ''}.")


@router.post("/api/google-drive/oauth/disconnect")
def google_drive_oauth_disconnect(request: Request) -> dict:
    actor = admin_user(request)
    clear_google_drive_oauth_tokens(build_app_repository())
    build_app_repository().add_audit_log(actor["username"], "google_drive_oauth_disconnected", "Ngat ket noi Google Drive OAuth")
    return {"ok": True, "message": "Da ngat ket noi Google Drive OAuth."}


@router.get("/api/admin/connections")
def system_connections(request: Request) -> dict:
    admin_user(request)
    connections = build_app_repository().list_system_connections()
    for connection in connections:
        config = connection.get("config", {})
        public_config, protected_keys = public_connection_config(config)
        connection["config"] = public_config
        connection["protected_config_keys"] = protected_keys
        if config.get("secret_ref"):
            connection["secret_ref"] = config["secret_ref"]
    return {"connections": connections}


@router.post("/api/admin/connections/{code}/test")
def test_system_connection(request: Request, code: str) -> dict:
    actor = admin_user(request)
    try:
        result = build_connection_service().test_connection(code)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(error)) from error
    build_app_repository().add_audit_log(actor["username"], "system_connection_tested", f"Kiểm tra kết nối {code}: {result['ok']}")
    notify_if_failed("Lỗi kiểm tra kết nối", result, {"ma_ket_noi": code})
    return result


@router.put("/api/admin/connections/{code}")
def save_system_connection(request: Request, code: str, payload: SystemConnectionPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    existing = repository.get_system_connection_by_code(code.strip()) or {}
    existing_config = existing.get("config") if isinstance(existing.get("config"), dict) else {}
    repository.upsert_system_connection(
        code.strip(),
        payload.name.strip(),
        payload.connection_type.strip(),
        payload.description.strip(),
        merge_protected_connection_config(existing_config, payload.config),
        payload.is_active,
    )
    repository.add_audit_log(actor["username"], "system_connection_saved", f"Luu ket noi {code}")
    return {"ok": True}


@router.get("/api/admin/sql-reports")
def list_sql_reports(request: Request) -> dict:
    admin_user(request)
    try:
        reports = build_app_repository().list_sql_reports()
    except RuntimeError as error:
        raise_sql_report_schema_error(error)
    return {"reports": reports}


@router.post("/api/admin/sql-reports")
def save_sql_report(request: Request, payload: SqlReportPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    existing_report = repository.get_sql_report_by_id(payload.id) if payload.id else None
    ten_bao_cao = payload.ten_bao_cao.strip()
    ma_bao_cao = payload.ma_bao_cao.strip().upper()
    cau_lenh_sql = validate_report_sql(payload.cau_lenh_sql)
    cac_tham_so = [item.strip() for item in payload.cac_tham_so if item.strip()]
    if not ten_bao_cao or not ma_bao_cao:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên báo cáo và mã báo cáo là bắt buộc.")
    try:
        report_id = repository.save_sql_report(payload.id, ten_bao_cao, ma_bao_cao, cau_lenh_sql, cac_tham_so)
    except sqlite3.IntegrityError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Mã báo cáo đã tồn tại.") from error
    except RuntimeError as error:
        raise_sql_report_schema_error(error)
    repository.add_audit_log(actor["username"], "sql_report_saved", f"Lưu cấu hình SQL {ma_bao_cao}")
    if hasattr(repository, "delete_dashboard_chart_cache_for_sql_report"):
        repository.delete_dashboard_chart_cache_for_sql_report(
            report_id=report_id,
            report_codes=[ma_bao_cao, (existing_report or {}).get("ma_bao_cao")],
        )
    return {"ok": True, "id": report_id}


@router.delete("/api/admin/sql-reports/{report_id}")
def delete_sql_report(request: Request, report_id: int) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    existing_report = repository.get_sql_report_by_id(report_id)
    if hasattr(repository, "delete_dashboard_chart_cache_for_sql_report"):
        repository.delete_dashboard_chart_cache_for_sql_report(
            report_id=report_id,
            report_codes=[(existing_report or {}).get("ma_bao_cao")],
        )
    try:
        repository.delete_sql_report(report_id)
    except RuntimeError as error:
        raise_sql_report_schema_error(error)
    repository.add_audit_log(actor["username"], "sql_report_deleted", f"Xóa cấu hình SQL {report_id}")
    return {"ok": True}


@router.get("/api/admin/onebss-reports")
def list_admin_onebss_reports(request: Request) -> dict:
    admin_user(request)
    try:
        return {"reports": build_app_repository().list_onebss_reports()}
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)


@router.post("/api/admin/onebss-reports")
def save_admin_onebss_report(request: Request, payload: OneBssReportPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    ten_bao_cao = payload.ten_bao_cao.strip()
    if not ten_bao_cao:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ten bao cao OneBSS khong duoc de trong.")
    try:
        report_url = normalize_onebss_report_url(payload.report_url)
    except OneBssDownloadError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    ma_bao_cao = payload.ma_bao_cao.strip().upper() or repository.generate_onebss_report_code()
    variables = [item.strip() for item in payload.danh_sach_bien if item.strip()]
    try:
        report_id = repository.save_onebss_report(
            payload.id,
            ma_bao_cao,
            ten_bao_cao,
            variables,
            payload.parameters if isinstance(payload.parameters, dict) else {},
            report_url,
            payload.storage_link.strip(),
            payload.otp_service_code.strip().lower() or "onebss",
        )
    except sqlite3.IntegrityError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Ma bao cao OneBSS da ton tai.") from error
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)
    repository.add_audit_log(actor["username"], "onebss_report_saved", f"Luu cau hinh OneBSS {ma_bao_cao}")
    return {"ok": True, "id": report_id, "ma_bao_cao": ma_bao_cao}


@router.delete("/api/admin/onebss-reports/{report_id}")
def delete_admin_onebss_report(request: Request, report_id: int) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        repository.delete_onebss_report(report_id)
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)
    repository.add_audit_log(actor["username"], "onebss_report_deleted", f"Xoa cau hinh OneBSS {report_id}")
    return {"ok": True}


@router.get("/api/admin/dashboard-layouts")
def list_dashboard_layouts(request: Request) -> dict:
    admin_user(request)
    try:
        layouts = build_app_repository().list_dashboard_layouts()
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)
    return {"layouts": layouts}


@router.get("/api/dashboard-layouts")
def list_visible_dashboard_layouts(request: Request) -> dict:
    user = current_user(request)
    try:
        layouts = build_app_repository().list_dashboard_layouts()
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)
    return {"layouts": visible_dashboard_layouts_for_user(layouts, user)}


@router.get("/api/admin/dashboard-layout-pages")
def list_dashboard_layout_pages(request: Request) -> dict:
    admin_user(request)
    repository = build_app_repository()
    try:
        layouts = repository.list_dashboard_layouts()
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)
    features = repository.list_features()
    return {"pages": build_dashboard_layout_pages(features, layouts)}


@router.get("/api/navigation")
def navigation(request: Request) -> dict:
    user = current_user(request)
    repository = build_app_repository()
    features = repository.list_features()
    try:
        layouts = repository.list_dashboard_layouts()
    except RuntimeError:
        layouts = []
    return {
        "features": visible_features_for_user(features, user),
        "dashboard_layouts": visible_dashboard_layouts_for_user(layouts, user),
    }


@router.get("/api/admin/dashboard-layouts/{page_id}")
def get_dashboard_layout(request: Request, page_id: str) -> dict:
    admin_user(request)
    repository = build_app_repository()
    safe_page_id = normalize_dashboard_code(page_id, "Mã trang")
    features = repository.list_features()
    return dashboard_layout_response(repository, safe_page_id, features)


@router.get("/api/dashboard-layouts/{page_id}")
def get_visible_dashboard_layout(request: Request, page_id: str) -> dict:
    repository = build_app_repository()
    safe_page_id = normalize_dashboard_code(page_id, "Mã trang")
    features = repository.list_features()
    require_dashboard_page_access(request, safe_page_id, features)
    return dashboard_layout_response(repository, safe_page_id, features)


@router.post("/api/admin/dashboard-layouts")
def save_dashboard_layout(request: Request, payload: DashboardLayoutPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    page_id, page_name, layout = normalize_dashboard_layout(payload)
    features = repository.list_features()
    parent_code = validate_dashboard_layout_parent_code(features, payload.parent_code)
    try:
        feature_code = repository.save_dashboard_layout(page_id, page_name, layout, parent_code)
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    repository.add_audit_log(actor["username"], "dashboard_layout_saved", f"Lưu layout dashboard {page_id}")
    refreshed_features = repository.list_features()
    feature_code, saved_parent_code = feature_parent_code_for_page(refreshed_features, page_id)
    layout["parent_code"] = saved_parent_code
    return {"ok": True, "page_id": page_id, "feature_code": feature_code, "parent_code": saved_parent_code, "layout": layout}


@router.delete("/api/admin/dashboard-layouts/{page_id}")
def delete_dashboard_layout(request: Request, page_id: str) -> dict:
    actor = admin_user(request)
    safe_page_id = normalize_dashboard_code(page_id, "Mã trang")
    try:
        build_app_repository().delete_dashboard_layout(safe_page_id)
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)
    build_app_repository().add_audit_log(actor["username"], "dashboard_layout_deleted", f"Xóa layout dashboard {safe_page_id}")
    return {"ok": True}


@router.delete("/api/admin/dashboard-layout-pages/{feature_code}")
def purge_unsaved_dashboard_layout_page(request: Request, feature_code: str) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    raw_code = str(feature_code or "").strip()
    normalized_code = normalize_feature_code(raw_code)
    features = repository.list_features()
    feature = next((item for item in features if str(item.get("code") or "") == raw_code), None)
    if not feature and normalized_code:
        feature = next((item for item in features if normalize_feature_code(item.get("code")) == normalized_code), None)
    if not feature:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy mục Dashboard chưa lưu.")
    code = str(feature.get("code") or "")
    if code in DASHBOARD_LAYOUT_EXCLUDED_FEATURE_CODES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Không được xóa mục hệ thống.")
    try:
        layouts = repository.list_dashboard_layouts()
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)
    page = next((item for item in build_dashboard_layout_pages(features, layouts) if item.get("feature_code") == code), None)
    if not page or page.get("saved"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chỉ xóa hẳn được mục Dashboard đang ở trạng thái Chưa lưu.")
    repository.delete_feature(code)
    repository.add_audit_log(actor["username"], "dashboard_layout_page_purged", f"Xóa hẳn mục dashboard chưa lưu {code}")
    return {"ok": True}


@router.get("/api/admin/dashboard-layouts/{page_id}/tabs/{tab_id}/data")
def load_dashboard_layout_tab_data(request: Request, page_id: str, tab_id: str) -> dict:
    admin_user(request)
    safe_page_id = normalize_dashboard_code(page_id, "Mã trang")
    safe_tab_id = normalize_dashboard_code(tab_id, "Mã Tab", uppercase=False)
    try:
        return build_database_service().run_dashboard_layout_tab(page_id=safe_page_id, tab_id=safe_tab_id)
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)


@router.get("/api/dashboard-layouts/{page_id}/tabs/{tab_id}/data")
def load_visible_dashboard_layout_tab_data(request: Request, page_id: str, tab_id: str) -> dict:
    repository = build_app_repository()
    features = repository.list_features()
    safe_page_id = normalize_dashboard_code(page_id, "Mã trang")
    require_dashboard_page_access(request, safe_page_id, features)
    safe_tab_id = normalize_dashboard_code(tab_id, "Mã Tab", uppercase=False)
    try:
        return build_database_service().run_dashboard_layout_tab(page_id=safe_page_id, tab_id=safe_tab_id)
    except RuntimeError as error:
        raise_dashboard_layout_schema_error(error)


@router.get("/api/google-sheet-table")
def load_google_sheet_table(request: Request, url: str) -> dict:
    current_user(request)
    safe_url = normalize_google_sheet_public_url(url)
    try:
        with httpx.Client(timeout=20, follow_redirects=True) as client:
            response = client.get(safe_url)
            response.raise_for_status()
    except httpx.TimeoutException:
        raise HTTPException(status_code=status.HTTP_504_GATEWAY_TIMEOUT, detail="Google Sheet phản hồi quá lâu.")
    except httpx.HTTPError as error:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=f"Không tải được Google Sheet: {error}")

    extractor = GoogleSheetTableExtractor()
    extractor.feed(response.text)
    table_html = extractor.sanitized_html()
    if not table_html:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Không tìm thấy bảng trong link Google Sheet.")
    return {"ok": True, "html": table_html}


@router.get("/api/reports/configs")
def list_report_configs(request: Request) -> dict:
    admin_user(request)
    try:
        reports = build_app_repository().list_sql_reports()
    except RuntimeError as error:
        raise_sql_report_schema_error(error)
    return {
        "reports": [
            {
                "id": report["id"],
                "ten_bao_cao": report["ten_bao_cao"],
                "ma_bao_cao": report["ma_bao_cao"],
                "cac_tham_so": report.get("cac_tham_so") or [],
            }
            for report in reports
        ]
    }


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return str(value)


def _dynamic_report_excel_workbook(result: dict[str, Any]) -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "TruyVanSQL"
    columns = [str(column) for column in (result.get("columns") or [])]
    rows = result.get("rows") or []
    if not columns and rows and isinstance(rows[0], dict):
        columns = list(rows[0].keys())
    if not columns:
        columns = ["Ket qua"]

    sheet.append(columns)
    for row in rows:
        if isinstance(row, dict):
            sheet.append([_excel_cell_value(row.get(column)) for column in columns])
        else:
            sheet.append([_excel_cell_value(row)])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="075FB0")
        cell.alignment = openpyxl.styles.Alignment(vertical="center")

    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column_cells[:200])
        sheet.column_dimensions[letter].width = min(max(width + 2, 10), 42)
    return workbook


class DynamicReportStreamingWorkbook:
    def __init__(self) -> None:
        self.workbook = openpyxl.Workbook(write_only=True)
        self.columns: list[str] = []
        self.sheet = None
        self.sheet_index = 0
        self.sheet_rows = 0
        self.total_rows = 0
        self.header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        self.header_fill = openpyxl.styles.PatternFill("solid", fgColor="075FB0")
        self.header_alignment = openpyxl.styles.Alignment(vertical="center")

    def append_rows(self, rows: list[dict[str, Any]], columns: list[str]) -> None:
        if not rows:
            return
        if not self.columns:
            self.columns = [str(column) for column in columns]
            if not self.columns and isinstance(rows[0], dict):
                self.columns = list(rows[0].keys())
            if not self.columns:
                self.columns = ["Ket qua"]
        for row in rows:
            if self.sheet is None or self.sheet_rows >= EXCEL_MAX_ROWS_PER_SHEET:
                self._start_sheet()
            if isinstance(row, dict):
                self.sheet.append([_excel_cell_value(row.get(column)) for column in self.columns])
            else:
                self.sheet.append([_excel_cell_value(row)])
            self.sheet_rows += 1
            self.total_rows += 1

    def _start_sheet(self) -> None:
        self.sheet_index += 1
        title = "TruyVanSQL" if self.sheet_index == 1 else f"TruyVanSQL_{self.sheet_index}"
        self.sheet = self.workbook.create_sheet(title[:31])
        self.sheet.freeze_panes = "A2"
        self.sheet_rows = 0
        header = []
        for index, column in enumerate(self.columns, start=1):
            cell = WriteOnlyCell(self.sheet, value=column)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            header.append(cell)
            self.sheet.column_dimensions[get_column_letter(index)].width = min(max(len(str(column)) + 2, 10), 42)
        self.sheet.append(header)
        self.sheet_rows = 1

    def save(self, target_path: Path) -> None:
        if self.sheet is None:
            self.columns = self.columns or ["Ket qua"]
            self._start_sheet()
        self.workbook.save(target_path)


def _dynamic_report_export_filename(result: dict[str, Any]) -> str:
    report = result.get("report") if isinstance(result.get("report"), dict) else {}
    code = str(report.get("ma_bao_cao") or "truy_van_sql").strip().lower()
    code = re.sub(r"[^a-z0-9_-]+", "_", code)
    code = code.strip("_") or "truy_van_sql"
    return f"{code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def _dynamic_report_export_max_workers() -> int:
    configured = int(getattr(get_settings(), "dynamic_report_export_max_workers", 1) or 1)
    return max(1, min(configured, 4))


def _dynamic_report_export_semaphore() -> threading.Semaphore:
    global DYNAMIC_REPORT_EXPORT_SEMAPHORE, DYNAMIC_REPORT_EXPORT_SEMAPHORE_WORKERS
    workers = _dynamic_report_export_max_workers()
    with DYNAMIC_REPORT_EXPORT_SEMAPHORE_LOCK:
        if DYNAMIC_REPORT_EXPORT_SEMAPHORE is None or DYNAMIC_REPORT_EXPORT_SEMAPHORE_WORKERS != workers:
            DYNAMIC_REPORT_EXPORT_SEMAPHORE = threading.Semaphore(workers)
            DYNAMIC_REPORT_EXPORT_SEMAPHORE_WORKERS = workers
        return DYNAMIC_REPORT_EXPORT_SEMAPHORE


def _dynamic_report_export_job_path(job_id: str) -> Path:
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "", str(job_id or ""))
    return DYNAMIC_REPORT_EXPORT_JOB_DIR / f"{safe_id or 'unknown'}.json"


def _dynamic_report_export_json_value(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(key): _dynamic_report_export_json_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_dynamic_report_export_json_value(item) for item in value]
    if isinstance(value, tuple):
        return [_dynamic_report_export_json_value(item) for item in value]
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _persist_dynamic_report_export_job(job_id: str, job: dict[str, Any]) -> None:
    try:
        with DYNAMIC_REPORT_EXPORT_JOB_FILE_LOCK:
            DYNAMIC_REPORT_EXPORT_JOB_DIR.mkdir(parents=True, exist_ok=True)
            target_path = _dynamic_report_export_job_path(job_id)
            target_path.write_text(
                json.dumps(_dynamic_report_export_json_value(job), ensure_ascii=False, default=str),
                encoding="utf-8",
            )
    except Exception as error:
        logger.warning("Cannot persist dynamic report export job %s: %s", job_id, error)


def _load_dynamic_report_export_job(job_id: str) -> dict[str, Any] | None:
    path = _dynamic_report_export_job_path(job_id)
    if not path.exists():
        return None
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except Exception as error:
        logger.warning("Cannot load dynamic report export job %s: %s", job_id, error)
        return None
    return loaded if isinstance(loaded, dict) else None


def _cleanup_dynamic_report_export_jobs() -> None:
    now = time.time()
    expired_paths: list[str] = []
    expired_job_paths: list[Path] = []
    with DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        expired_ids = [
            job_id
            for job_id, job in DYNAMIC_REPORT_EXPORT_JOBS.items()
            if job.get("status") not in DYNAMIC_REPORT_EXPORT_ACTIVE_STATUSES
            and now - float(job.get("updated_at") or job.get("created_at") or now) > DYNAMIC_REPORT_EXPORT_JOB_TTL_SECONDS
        ]
        for job_id in expired_ids:
            job = DYNAMIC_REPORT_EXPORT_JOBS.pop(job_id, {})
            if job.get("path"):
                expired_paths.append(str(job["path"]))
            expired_job_paths.append(_dynamic_report_export_job_path(job_id))
    for path_text in expired_paths:
        try:
            Path(path_text).unlink(missing_ok=True)
        except OSError:
            logger.warning("Cannot remove old dynamic report export file: %s", path_text)
    if DYNAMIC_REPORT_EXPORT_JOB_DIR.exists():
        for temp_job_path in DYNAMIC_REPORT_EXPORT_JOB_DIR.glob("*.tmp"):
            try:
                temp_job_path.unlink(missing_ok=True)
            except OSError:
                logger.warning("Cannot remove old dynamic report export temp metadata: %s", temp_job_path)
        for job_path in DYNAMIC_REPORT_EXPORT_JOB_DIR.glob("*.json"):
            if job_path in expired_job_paths:
                continue
            try:
                loaded = json.loads(job_path.read_text(encoding="utf-8"))
                if (
                    isinstance(loaded, dict)
                    and loaded.get("status") not in DYNAMIC_REPORT_EXPORT_ACTIVE_STATUSES
                    and now - float(loaded.get("updated_at") or loaded.get("created_at") or now) > DYNAMIC_REPORT_EXPORT_JOB_TTL_SECONDS
                ):
                    expired_job_paths.append(job_path)
            except Exception:
                expired_job_paths.append(job_path)
    for job_path in expired_job_paths:
        try:
            job_path.unlink(missing_ok=True)
        except OSError:
            logger.warning("Cannot remove old dynamic report export job metadata: %s", job_path)


def _set_dynamic_report_export_job(job_id: str, **updates: Any) -> None:
    snapshot: dict[str, Any] | None = None
    with DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        job = DYNAMIC_REPORT_EXPORT_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)
        job["updated_at"] = time.time()
        snapshot = dict(job)
    if snapshot:
        _persist_dynamic_report_export_job(job_id, snapshot)


def _get_dynamic_report_export_job(job_id: str) -> dict[str, Any] | None:
    with DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        job = DYNAMIC_REPORT_EXPORT_JOBS.get(job_id)
        if job:
            return dict(job)
    loaded = _load_dynamic_report_export_job(job_id)
    if not loaded:
        return None
    with DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        DYNAMIC_REPORT_EXPORT_JOBS[job_id] = dict(loaded)
    return dict(loaded)


def _dynamic_report_job_sort_value(job: dict[str, Any]) -> float:
    value = job.get("created_at") or job.get("updated_at") or 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).timestamp()
    except (TypeError, ValueError):
        return 0.0


def _is_dynamic_report_export_cancelled(job_id: str) -> bool:
    job = _get_dynamic_report_export_job(job_id) or {}
    return str(job.get("status") or "").lower() in {"cancel_requested", "cancelled"}


def _dynamic_report_export_cancelled_result(job_id: str, payload: RunReportPayload, actor: str) -> None:
    message = "Lenh xuat file da bi ngung."
    _set_dynamic_report_export_job(job_id, status="cancelled", message=message, progress={})
    _record_dynamic_report_history(
        actor=actor,
        event_type="export",
        status_value="cancelled",
        ma_bao_cao=payload.ma_bao_cao,
        message=message,
        filters=payload.filters,
        search=payload.search,
        search_columns=payload.search_columns,
        history_id=job_id,
        job_id=job_id,
    )


def _list_dynamic_report_export_job_snapshots(limit: int = 100) -> list[dict[str, Any]]:
    jobs_by_id: dict[str, dict[str, Any]] = {}
    with DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        for job_id, job in DYNAMIC_REPORT_EXPORT_JOBS.items():
            jobs_by_id[job_id] = dict(job)
    if DYNAMIC_REPORT_EXPORT_JOB_DIR.exists():
        for job_path in DYNAMIC_REPORT_EXPORT_JOB_DIR.glob("*.json"):
            try:
                loaded = json.loads(job_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            if not isinstance(loaded, dict):
                continue
            job_id = str(loaded.get("job_id") or job_path.stem)
            current = jobs_by_id.get(job_id)
            if not current or float(loaded.get("updated_at") or 0) >= float(current.get("updated_at") or 0):
                jobs_by_id[job_id] = loaded
    jobs = list(jobs_by_id.values())
    active_jobs = sorted(
        [job for job in jobs if str(job.get("status") or "").lower() in DYNAMIC_REPORT_EXPORT_ACTIVE_STATUSES],
        key=_dynamic_report_job_sort_value,
    )
    for position, job in enumerate(active_jobs, start=1):
        job["queue_position"] = position
        if str(job.get("status") or "").lower() == "queued":
            job["message"] = job.get("message") or f"Dang cho xu ly o vi tri {position}."
    jobs.sort(key=_dynamic_report_job_sort_value, reverse=True)
    return jobs[: min(max(int(limit or 100), 1), 200)]


def _dynamic_report_export_job_response(job_id: str, job: dict[str, Any]) -> dict[str, Any]:
    status_value = str(job.get("status") or "queued")
    response = {
        "ok": status_value not in {"failed", "cancelled"},
        "job_id": job_id,
        "history_id": job.get("history_id") or job_id,
        "event_type": "export",
        "status": status_value,
        "message": job.get("message") or "",
        "progress": job.get("progress") or {},
        "rows": int(job.get("rows") or 0),
        "total": int(job.get("total") or 0),
        "report_code": job.get("report_code") or "",
        "report_name": job.get("report_name") or "",
        "ma_bao_cao": job.get("report_code") or "",
        "ten_bao_cao": job.get("report_name") or "",
        "created_at": job.get("created_at") or "",
        "queue_position": int(job.get("queue_position") or 0),
        "can_cancel": status_value in DYNAMIC_REPORT_EXPORT_ACTIVE_STATUSES,
    }
    if status_value == "complete":
        drive_url = str(job.get("drive_url") or "")
        response["download_url"] = drive_url or f"/api/reports/export-jobs/{job_id}/download"
        response["drive_url"] = drive_url
        response["filename"] = job.get("filename") or ""
    if status_value in {"failed", "cancelled"}:
        response["details"] = job.get("details") or {}
    return response


def _run_dynamic_report_export_job(job_id: str, payload: RunReportPayload, created_by: str = "system") -> None:
    writer = DynamicReportStreamingWorkbook()
    target_path: Path | None = None
    semaphore = _dynamic_report_export_semaphore()
    acquired = False

    def progress_callback(progress: dict[str, Any]) -> None:
        if _is_dynamic_report_export_cancelled(job_id):
            raise DynamicReportExportCancelled()
        rows = int(progress.get("rows") or 0)
        total = int(progress.get("total") or 0)
        message = f"Đang xuất Excel: đã lấy {rows:,}" + (f"/{total:,} dòng." if total else " dòng.")
        _set_dynamic_report_export_job(job_id, status="running", message=message, progress=progress)

    if _is_dynamic_report_export_cancelled(job_id):
        _dynamic_report_export_cancelled_result(job_id, payload, created_by)
        return

    _set_dynamic_report_export_job(job_id, status="queued", message="Đang chờ lượt xuất Excel để tránh quá tải API dữ liệu nội bộ.")
    try:
        semaphore.acquire()
        acquired = True
        if _is_dynamic_report_export_cancelled(job_id):
            _dynamic_report_export_cancelled_result(job_id, payload, created_by)
            return
        _set_dynamic_report_export_job(job_id, status="running", message="Đang chuẩn bị dữ liệu xuất Excel.")
        DYNAMIC_REPORT_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        target_path = DYNAMIC_REPORT_EXPORT_DIR / f"{job_id}.xlsx"
        drive_folder_id = google_drive_folder_id(get_settings(), "")

        if drive_folder_id:
            filename = _dynamic_report_export_filename({"report": {"ma_bao_cao": payload.ma_bao_cao.strip().upper()}})
            _set_dynamic_report_export_job(job_id, status="running", message="Đang yêu cầu máy trạm xuất Excel và upload Google Drive.")
            remote_result = build_database_service().export_dynamic_report_to_drive(
                ma_bao_cao=payload.ma_bao_cao.strip().upper(),
                filters=payload.filters,
                search=payload.search,
                search_columns=payload.search_columns,
                drive_folder_id=drive_folder_id,
                file_name=filename,
            )
            if _is_dynamic_report_export_cancelled(job_id):
                _dynamic_report_export_cancelled_result(job_id, payload, created_by)
                return
            if remote_result.get("ok"):
                drive_url = str(
                    remote_result.get("drive_url")
                    or remote_result.get("storage_link")
                    or remote_result.get("web_view_link")
                    or remote_result.get("web_content_link")
                    or ""
                ).strip()
                rows = int(remote_result.get("rows") or remote_result.get("export_rows") or remote_result.get("row_count") or 0)
                total = int(remote_result.get("total") or rows)
                _set_dynamic_report_export_job(
                    job_id,
                    status="complete",
                    message=f"Đã xuất Excel trên máy trạm và upload Google Drive{f' đủ {rows:,} dòng' if rows else ''}.",
                    filename=str(remote_result.get("file_name") or filename),
                    drive_url=drive_url,
                    rows=rows,
                    total=total,
                    details=remote_result,
                )
                _record_dynamic_report_history(
                    actor=created_by,
                    event_type="export",
                    status_value="complete",
                    ma_bao_cao=payload.ma_bao_cao,
                    rows=rows,
                    total=total,
                    message=f"Da xuat Excel tren may tram va upload Google Drive{f' du {rows:,} dong' if rows else ''}.",
                    filters=payload.filters,
                    search=payload.search,
                    search_columns=payload.search_columns,
                    history_id=job_id,
                    job_id=job_id,
                    drive_url=drive_url,
                    file_name=str(remote_result.get("file_name") or filename),
                    details=remote_result,
                )
                return
            _set_dynamic_report_export_job(
                job_id,
                status="failed",
                message=remote_result.get("message") or "Máy trạm chưa xuất được file Excel lên Google Drive.",
                details=remote_result,
            )
            _record_dynamic_report_history(
                actor=created_by,
                event_type="export",
                status_value="failed",
                ma_bao_cao=payload.ma_bao_cao,
                message=str(remote_result.get("message") or "May tram chua xuat duoc file Excel len Google Drive."),
                filters=payload.filters,
                search=payload.search,
                search_columns=payload.search_columns,
                history_id=job_id,
                job_id=job_id,
                details=remote_result,
            )
            return

        result = build_database_service().export_dynamic_report(
            ma_bao_cao=payload.ma_bao_cao.strip().upper(),
            filters=payload.filters,
            search=payload.search,
            search_columns=payload.search_columns,
            progress_callback=progress_callback,
            page_callback=writer.append_rows,
            collect_rows=False,
        )
        if _is_dynamic_report_export_cancelled(job_id):
            _dynamic_report_export_cancelled_result(job_id, payload, created_by)
            if target_path:
                target_path.unlink(missing_ok=True)
            return
        if not result.get("ok"):
            target_path.unlink(missing_ok=True)
            _set_dynamic_report_export_job(
                job_id,
                status="failed",
                message=result.get("message") or "Không xuất được file Excel.",
                details=result.get("details") or {},
            )
            _record_dynamic_report_history(
                actor=created_by,
                event_type="export",
                status_value="failed",
                ma_bao_cao=payload.ma_bao_cao,
                report=result.get("report") if isinstance(result.get("report"), dict) else None,
                message=str(result.get("message") or "Khong xuat duoc file Excel."),
                filters=payload.filters,
                search=payload.search,
                search_columns=payload.search_columns,
                history_id=job_id,
                job_id=job_id,
                details=result.get("details") if isinstance(result.get("details"), dict) else {},
            )
            return

        filename = _dynamic_report_export_filename(result)
        writer.save(target_path)
        details = result.get("details") if isinstance(result.get("details"), dict) else {}
        rows = int(details.get("export_rows") or writer.total_rows)
        _set_dynamic_report_export_job(
            job_id,
            status="complete",
            message=f"Đã tạo file Excel đủ {rows:,} dòng.",
            filename=filename,
            path=str(target_path),
            rows=rows,
            total=int(details.get("fetched_total") or rows),
            details=details,
        )
        _record_dynamic_report_history(
            actor=created_by,
            event_type="export",
            status_value="complete",
            ma_bao_cao=payload.ma_bao_cao,
            report=result.get("report") if isinstance(result.get("report"), dict) else None,
            rows=rows,
            total=int(details.get("fetched_total") or rows),
            message=f"Da tao file Excel du {rows:,} dong.",
            filters=payload.filters,
            search=payload.search,
            search_columns=payload.search_columns,
            history_id=job_id,
            job_id=job_id,
            download_url=f"/api/reports/export-jobs/{job_id}/download",
            file_name=filename,
            details=details,
        )
    except DynamicReportExportCancelled:
        if target_path:
            try:
                target_path.unlink(missing_ok=True)
            except Exception:
                pass
        _dynamic_report_export_cancelled_result(job_id, payload, created_by)
    except Exception as error:
        logger.exception("Dynamic report export job failed: %s", error)
        if target_path:
            try:
                target_path.unlink(missing_ok=True)
            except Exception:
                pass
        _set_dynamic_report_export_job(job_id, status="failed", message=f"Không xuất được file Excel: {error}")
        _record_dynamic_report_history(
            actor=created_by,
            event_type="export",
            status_value="failed",
            ma_bao_cao=payload.ma_bao_cao,
            message=f"Khong xuat duoc file Excel: {error}",
            filters=payload.filters,
            search=payload.search,
            search_columns=payload.search_columns,
            history_id=job_id,
            job_id=job_id,
            details={"error": str(error)},
        )
    finally:
        if acquired:
            semaphore.release()


@router.post("/api/reports/run")
def run_dynamic_report(request: Request, payload: RunReportPayload) -> dict:
    actor = admin_user(request)
    try:
        result = build_database_service().run_dynamic_report(
            ma_bao_cao=payload.ma_bao_cao.strip().upper(),
            filters=payload.filters,
            page=payload.page,
            page_size=payload.page_size,
            search=payload.search,
            search_columns=payload.search_columns,
        )
    except RuntimeError as error:
        raise_sql_report_schema_error(error)
    pagination = result.get("pagination") if isinstance(result.get("pagination"), dict) else {}
    rows = result.get("rows") if isinstance(result.get("rows"), list) else []
    _record_dynamic_report_history(
        actor=actor["username"],
        event_type="load",
        status_value="success" if result.get("ok", True) else "failed",
        ma_bao_cao=payload.ma_bao_cao,
        report=result.get("report") if isinstance(result.get("report"), dict) else None,
        rows=len(rows),
        total=pagination.get("total") or len(rows),
        message=str(result.get("message") or ""),
        filters=payload.filters,
        search=payload.search,
        search_columns=payload.search_columns,
    )
    return result


@router.get("/api/reports/history")
def list_dynamic_report_history(request: Request, limit: int = 30) -> dict:
    admin_user(request)
    _cleanup_dynamic_report_export_jobs()
    return {"items": _list_dynamic_report_history(limit)}


@router.post("/api/reports/export")
def export_dynamic_report(request: Request, payload: RunReportPayload) -> Response:
    admin_user(request)
    try:
        result = build_database_service().export_dynamic_report(
            ma_bao_cao=payload.ma_bao_cao.strip().upper(),
            filters=payload.filters,
            search=payload.search,
            search_columns=payload.search_columns,
        )
    except RuntimeError as error:
        raise_sql_report_schema_error(error)

    if not result.get("ok"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.get("message") or "Không xuất được file Excel.")

    workbook = _dynamic_report_excel_workbook(result)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = _dynamic_report_export_filename(result)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}"},
    )


@router.post("/api/reports/export-jobs")
def start_dynamic_report_export_job(request: Request, payload: RunReportPayload) -> dict:
    actor = admin_user(request)
    _cleanup_dynamic_report_export_jobs()
    job_id = uuid.uuid4().hex
    now = time.time()
    report_info = _dynamic_report_history_report(payload.ma_bao_cao)
    with DYNAMIC_REPORT_EXPORT_JOBS_LOCK:
        DYNAMIC_REPORT_EXPORT_JOBS[job_id] = {
            "job_id": job_id,
            "status": "queued",
            "message": "Đã đưa yêu cầu xuất Excel vào hàng đợi.",
            "created_at": now,
            "updated_at": now,
            "progress": {},
            "created_by": actor["username"],
            "report_code": report_info["ma_bao_cao"],
            "report_name": report_info["ten_bao_cao"],
        }
        job_snapshot = dict(DYNAMIC_REPORT_EXPORT_JOBS[job_id])
    _persist_dynamic_report_export_job(job_id, job_snapshot)
    _record_dynamic_report_history(
        actor=actor["username"],
        event_type="export",
        status_value="queued",
        ma_bao_cao=report_info["ma_bao_cao"],
        report_name=report_info["ten_bao_cao"],
        message="Da dua yeu cau xuat Excel vao hang doi.",
        filters=payload.filters,
        search=payload.search,
        search_columns=payload.search_columns,
        history_id=job_id,
        job_id=job_id,
    )
    thread = threading.Thread(target=_run_dynamic_report_export_job, args=(job_id, payload, actor["username"]), daemon=True)
    thread.start()
    return {
        "ok": True,
        "job_id": job_id,
        "history_id": job_id,
        "event_type": "export",
        "status": "queued",
        "message": "Dang xuat file Excel o che do nen.",
        "report_code": report_info["ma_bao_cao"],
        "report_name": report_info["ten_bao_cao"],
        "queue_position": 0,
        "can_cancel": True,
    }
    return {"ok": True, "job_id": job_id, "status": "queued", "message": "Đang xuất file Excel ở chế độ nền."}


@router.get("/api/reports/export-jobs")
def list_dynamic_report_export_jobs(request: Request, limit: int = 100) -> dict:
    admin_user(request)
    _cleanup_dynamic_report_export_jobs()
    jobs = [
        _dynamic_report_export_job_response(str(job.get("job_id") or ""), job)
        for job in _list_dynamic_report_export_job_snapshots(limit)
    ]
    return {"jobs": jobs}


@router.delete("/api/reports/export-jobs/{job_id}")
def cancel_dynamic_report_export_job(request: Request, job_id: str) -> dict:
    actor = admin_user(request)
    _cleanup_dynamic_report_export_jobs()
    job = _get_dynamic_report_export_job(job_id)
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay job xuat Excel.")
    status_value = str(job.get("status") or "queued").lower()
    if status_value in DYNAMIC_REPORT_EXPORT_FINAL_STATUSES:
        return _dynamic_report_export_job_response(job_id, job)
    now = time.time()
    if status_value == "queued":
        _set_dynamic_report_export_job(
            job_id,
            status="cancelled",
            message="Lenh xuat file da bi xoa khoi hang doi.",
            cancelled_by=actor["username"],
            cancelled_at=now,
            progress={},
        )
        job = _get_dynamic_report_export_job(job_id) or job
        _record_dynamic_report_history(
            actor=actor["username"],
            event_type="export",
            status_value="cancelled",
            ma_bao_cao=str(job.get("report_code") or ""),
            report_name=str(job.get("report_name") or ""),
            message="Lenh xuat file da bi xoa khoi hang doi.",
            history_id=job_id,
            job_id=job_id,
        )
        return _dynamic_report_export_job_response(job_id, job)
    _set_dynamic_report_export_job(
        job_id,
        status="cancel_requested",
        message="Dang yeu cau ngung lenh xuat file nay, he thong se chuyen sang lenh tiep theo.",
        cancelled_by=actor["username"],
        cancelled_at=now,
    )
    job = _get_dynamic_report_export_job(job_id) or job
    return _dynamic_report_export_job_response(job_id, job)


@router.get("/api/reports/export-jobs/{job_id}")
def get_dynamic_report_export_job(request: Request, job_id: str) -> dict:
    admin_user(request)
    _cleanup_dynamic_report_export_jobs()
    job = _get_dynamic_report_export_job(job_id)
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy job xuất Excel.")
    return _dynamic_report_export_job_response(job_id, job)


@router.get("/api/reports/export-jobs/{job_id}/download")
def download_dynamic_report_export_job(request: Request, job_id: str) -> Response:
    admin_user(request)
    job = _get_dynamic_report_export_job(job_id)
    if not job:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy job xuất Excel.")
    if job.get("status") != "complete":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="File Excel chưa tạo xong.")
    file_path = Path(str(job.get("path") or ""))
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File Excel không còn tồn tại. Hãy xuất lại.")
    return FileResponse(
        file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=str(job.get("filename") or file_path.name),
    )


@router.post("/api/reports/export-loaded")
def export_loaded_dynamic_report(request: Request, payload: ExportLoadedReportPayload) -> Response:
    admin_user(request)
    max_rows = get_settings().dynamic_report_export_max_rows
    rows = payload.rows[:max_rows]
    result = {
        "ok": True,
        "report": {"ma_bao_cao": payload.ma_bao_cao.strip().upper() or "TRUY_VAN_SQL"},
        "columns": payload.columns,
        "rows": rows,
    }
    workbook = _dynamic_report_excel_workbook(result)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = _dynamic_report_export_filename(result)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}"},
    )


def _onebss_report_job_path(job_id: str) -> Path:
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "", str(job_id or ""))
    return ONEBSS_REPORT_JOB_DIR / f"{safe_id or 'unknown'}.json"


def _persist_onebss_report_job(job_id: str, job: dict[str, Any]) -> None:
    try:
        with ONEBSS_REPORT_JOB_FILE_LOCK:
            ONEBSS_REPORT_JOB_DIR.mkdir(parents=True, exist_ok=True)
            _onebss_report_job_path(job_id).write_text(
                json.dumps(_dynamic_report_export_json_value(job), ensure_ascii=False, default=str),
                encoding="utf-8",
            )
    except Exception as error:
        logger.warning("Cannot persist OneBSS report job %s: %s", job_id, error)


def _load_onebss_report_job(job_id: str) -> dict[str, Any] | None:
    path = _onebss_report_job_path(job_id)
    if not path.exists():
        return None
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
    except Exception as error:
        logger.warning("Cannot load OneBSS report job %s: %s", job_id, error)
        return None
    return loaded if isinstance(loaded, dict) else None


def _set_onebss_report_job(job_id: str, **updates: Any) -> None:
    snapshot: dict[str, Any] | None = None
    with ONEBSS_REPORT_JOBS_LOCK:
        job = ONEBSS_REPORT_JOBS.get(job_id)
        if not job:
            return
        job.update(updates)
        job["updated_at"] = time.time()
        snapshot = dict(job)
    if snapshot:
        _persist_onebss_report_job(job_id, snapshot)


def _get_onebss_report_job(job_id: str) -> dict[str, Any] | None:
    with ONEBSS_REPORT_JOBS_LOCK:
        job = ONEBSS_REPORT_JOBS.get(job_id)
        if job:
            return dict(job)
    loaded = _load_onebss_report_job(job_id)
    if not loaded:
        return None
    with ONEBSS_REPORT_JOBS_LOCK:
        ONEBSS_REPORT_JOBS[job_id] = dict(loaded)
    return dict(loaded)


def _cleanup_onebss_report_jobs() -> None:
    now = time.time()
    expired_paths: list[Path] = []
    with ONEBSS_REPORT_JOBS_LOCK:
        expired_ids = [
            job_id
            for job_id, job in ONEBSS_REPORT_JOBS.items()
            if str(job.get("status") or "").lower() not in ONEBSS_REPORT_ACTIVE_STATUSES
            and now - float(job.get("updated_at") or job.get("created_at") or now) > ONEBSS_REPORT_JOB_TTL_SECONDS
        ]
        for job_id in expired_ids:
            ONEBSS_REPORT_JOBS.pop(job_id, None)
            expired_paths.append(_onebss_report_job_path(job_id))
    if ONEBSS_REPORT_JOB_DIR.exists():
        for job_path in ONEBSS_REPORT_JOB_DIR.glob("*.json"):
            try:
                loaded = json.loads(job_path.read_text(encoding="utf-8"))
            except Exception:
                expired_paths.append(job_path)
                continue
            if (
                isinstance(loaded, dict)
                and str(loaded.get("status") or "").lower() not in ONEBSS_REPORT_ACTIVE_STATUSES
                and now - float(loaded.get("updated_at") or loaded.get("created_at") or now) > ONEBSS_REPORT_JOB_TTL_SECONDS
            ):
                expired_paths.append(job_path)
    for path in expired_paths:
        try:
            path.unlink(missing_ok=True)
        except OSError:
            logger.warning("Cannot remove old OneBSS report job metadata: %s", path)


def _list_onebss_report_job_snapshots(limit: int = 100) -> list[dict[str, Any]]:
    jobs_by_id: dict[str, dict[str, Any]] = {}
    with ONEBSS_REPORT_JOBS_LOCK:
        for job_id, job in ONEBSS_REPORT_JOBS.items():
            jobs_by_id[job_id] = dict(job)
    if ONEBSS_REPORT_JOB_DIR.exists():
        for job_path in ONEBSS_REPORT_JOB_DIR.glob("*.json"):
            try:
                loaded = json.loads(job_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            if not isinstance(loaded, dict):
                continue
            job_id = str(loaded.get("job_id") or job_path.stem)
            current = jobs_by_id.get(job_id)
            if not current or float(loaded.get("updated_at") or 0) >= float(current.get("updated_at") or 0):
                jobs_by_id[job_id] = loaded
    jobs = list(jobs_by_id.values())
    jobs.sort(key=_dynamic_report_job_sort_value, reverse=True)
    return jobs[: min(max(int(limit or 100), 1), 200)]


def _onebss_report_download_url(run: dict[str, Any]) -> str:
    file_path = str(run.get("file_path") or "").strip()
    run_id = str(run.get("run_id") or run.get("job_id") or "").strip()
    if not file_path or not run_id:
        return ""
    storage_link = str(run.get("storage_link") or "")
    storage_status = str(run.get("storage_status") or "")
    if storage_link.startswith(("http://", "https://")) and (
        storage_status.lower().startswith("uploaded_google_drive:")
        or "/file/d/" in storage_link
        or "/spreadsheets/d/" in storage_link
        or "?id=" in storage_link
        or "&id=" in storage_link
    ):
        return ""
    return f"/api/onebss-reports/runs/{quote(run_id)}/download"


def _decorate_onebss_report_run(run: dict[str, Any]) -> dict[str, Any]:
    decorated = dict(run)
    download_url = _onebss_report_download_url(decorated)
    if download_url:
        decorated["download_url"] = download_url
    return decorated


def _onebss_report_job_response(job_id: str, job: dict[str, Any]) -> dict[str, Any]:
    status_value = str(job.get("status") or "queued")
    response = {
        "ok": status_value not in {"failed", "cancelled", "storage_failed", "google_drive_not_configured", "google_drive_upload_failed"},
        "job_id": job_id,
        "run_id": job.get("run_id") or job_id,
        "status": status_value,
        "message": job.get("message") or "",
        "ma_bao_cao": job.get("ma_bao_cao") or "",
        "ten_bao_cao": job.get("ten_bao_cao") or "",
        "parameters": job.get("parameters") if isinstance(job.get("parameters"), dict) else {},
        "started_at": job.get("started_at") or "",
        "finished_at": job.get("finished_at") or "",
        "created_by": job.get("created_by") or "",
        "session_id": job.get("session_id") or job.get("worker_session_id") or "",
        "worker_session_id": job.get("worker_session_id") or "",
        "otp_request_id": job.get("otp_request_id") or "",
        "can_resume_otp": status_value in {"otp_required", "otp_invalid", "manual_otp_required"},
        "can_poll": status_value in ONEBSS_REPORT_ACTIVE_STATUSES,
    }
    result = job.get("result") if isinstance(job.get("result"), dict) else {}
    run = job.get("run") if isinstance(job.get("run"), dict) else {}
    for key in ("file_name", "file_path", "storage_link", "storage_status", "duration_ms"):
        response[key] = run.get(key) or result.get(key) or job.get(key) or ""
    download_url = _onebss_report_download_url(response)
    if download_url:
        response["download_url"] = download_url
    if result:
        response["result"] = result
    if run:
        response["run"] = _decorate_onebss_report_run(run)
    return response


def _onebss_job_matches(job: dict[str, Any], ma_bao_cao: str) -> bool:
    return not ma_bao_cao or str(job.get("ma_bao_cao") or "").strip().upper() == ma_bao_cao


def _active_onebss_job_runs(ma_bao_cao: str, limit: int = 50) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for job in _list_onebss_report_job_snapshots(limit=limit):
        status_value = str(job.get("status") or "").lower()
        if status_value not in ONEBSS_REPORT_ACTIVE_STATUSES:
            continue
        if not _onebss_job_matches(job, ma_bao_cao):
            continue
        rows.append(_decorate_onebss_report_run(_onebss_report_job_response(str(job.get("job_id") or ""), job)))
    return rows


def _save_onebss_job_result(
    repository: AppRepository,
    *,
    job_id: str,
    report: dict[str, Any],
    result: dict[str, Any],
    parameters: dict[str, Any],
    started_at: str,
    created_by: str,
) -> dict[str, Any]:
    finished_at = result.get("finished_at") or datetime.now().isoformat(timespec="seconds")
    return repository.save_onebss_report_run({
        "run_id": job_id,
        "ma_bao_cao": report.get("ma_bao_cao"),
        "ten_bao_cao": report.get("ten_bao_cao"),
        "status": result.get("status") or ("success" if result.get("ok") else "failed"),
        "message": result.get("message") or "",
        "file_name": result.get("file_name") or "",
        "file_path": result.get("file_path") or "",
        "storage_link": result.get("storage_link") or "",
        "storage_status": result.get("storage_status") or "",
        "parameters": result.get("parameters") if isinstance(result.get("parameters"), dict) else parameters,
        "started_at": started_at,
        "finished_at": finished_at,
        "duration_ms": int(result.get("duration_ms") or 0),
        "created_by": created_by,
    })


def _run_onebss_report_job(
    job_id: str,
    *,
    report: dict[str, Any],
    parameters: dict[str, Any],
    created_by: str,
    otp: str = "",
    session_id: str = "",
    otp_request_id: str = "",
    otp_source: str = "",
) -> None:
    settings = get_settings()
    started_at = str((_get_onebss_report_job(job_id) or {}).get("started_at") or datetime.now().isoformat(timespec="seconds"))
    _set_onebss_report_job(job_id, status="queued", message="Dang cho luot lay du lieu OneBSS.", session_id=session_id, otp_request_id=otp_request_id)
    with ONEBSS_REPORT_SEMAPHORE:
        _set_onebss_report_job(job_id, status="running", message="Dang dang nhap va lay du lieu OneBSS.", session_id=session_id, otp_request_id=otp_request_id)
        try:
            result = run_onebss_report_request(
                settings,
                report,
                parameters,
                otp=otp,
                session_id=session_id,
                created_by=created_by,
            )
        except Exception as error:
            logger.exception("Unhandled OneBSS report background job error")
            result = {
                "ok": False,
                "status": "failed",
                "message": f"Loi khi khoi chay OneBSS: {error}",
                "parameters": parameters,
            }

        status_value = str(result.get("status") or ("success" if result.get("ok") else "failed"))
        if status_value in {"otp_required", "otp_invalid", "manual_otp_required"} and result.get("session_id"):
            _set_onebss_report_job(
                job_id,
                status=status_value,
                message=result.get("message") or "OneBSS yeu cau OTP.",
                session_id=result.get("session_id") or session_id,
                otp_request_id=result.get("otp_request_id") or otp_request_id,
                parameters=result.get("parameters") if isinstance(result.get("parameters"), dict) else parameters,
                result=result,
            )
            return

        if result.get("ok") and otp_source == "manual" and otp_request_id:
            cancel_onebss_mobile_gateway_otp(settings, otp_request_id, "manual_otp_used")

        repository = build_app_repository()
        try:
            run = _save_onebss_job_result(
                repository,
                job_id=job_id,
                report=report,
                result=result,
                parameters=parameters,
                started_at=started_at,
                created_by=created_by,
            )
        except Exception as error:
            logger.exception("Cannot save OneBSS background run")
            _set_onebss_report_job(
                job_id,
                status="failed",
                message=f"Da chay OneBSS nhung khong luu duoc lich su: {error}",
                result=result,
                finished_at=datetime.now().isoformat(timespec="seconds"),
            )
            return

        try:
            repository.add_audit_log(created_by, "onebss_report_run", f"Chay bao cao OneBSS {report.get('ma_bao_cao')}: {result.get('ok')}")
        except Exception:
            logger.exception("Cannot write OneBSS report audit log")
        _set_onebss_report_job(
            job_id,
            status=run.get("status") or status_value,
            message=run.get("message") or result.get("message") or "",
            result=result,
            run=run,
            run_id=run.get("run_id") or job_id,
            file_name=run.get("file_name") or result.get("file_name") or "",
            file_path=run.get("file_path") or result.get("file_path") or "",
            storage_link=run.get("storage_link") or result.get("storage_link") or "",
            storage_status=run.get("storage_status") or result.get("storage_status") or "",
            finished_at=run.get("finished_at") or datetime.now().isoformat(timespec="seconds"),
            session_id="",
            otp_request_id="",
        )


def _start_onebss_report_job_thread(job_id: str, **kwargs: Any) -> None:
    thread = threading.Thread(target=_run_onebss_report_job, args=(job_id,), kwargs=kwargs, daemon=True)
    thread.start()


@router.get("/api/onebss-reports/configs")
def list_onebss_report_configs(request: Request) -> dict:
    admin_user(request)
    try:
        reports = build_app_repository().list_onebss_reports()
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)
    return {"reports": reports}


@router.get("/api/onebss-reports/runs")
def list_onebss_report_runs(request: Request, ma_bao_cao: str = "", limit: int = 50) -> dict:
    admin_user(request)
    report_code = ma_bao_cao.strip().upper()
    try:
        runs = [
            _decorate_onebss_report_run(run)
            for run in build_app_repository().list_onebss_report_runs(ma_bao_cao=report_code, limit=limit)
        ]
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)
    return {"runs": runs}


@router.delete("/api/onebss-reports/runs")
@router.post("/api/onebss-reports/runs/clear")
def clear_onebss_report_runs(request: Request, ma_bao_cao: str = "") -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    report_code = ma_bao_cao.strip().upper()
    try:
        deleted = repository.clear_onebss_report_runs(ma_bao_cao=report_code)
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)
    try:
        scope = report_code or "all"
        repository.add_audit_log(actor["username"], "onebss_report_runs_cleared", f"Xoa lich su chay OneBSS: {scope} ({deleted})")
    except Exception:
        logger.exception("Cannot write OneBSS clear history audit log")
    return {"ok": True, "deleted": deleted}


@router.get("/api/onebss-reports/otp-requests/{otp_request_id}")
def get_onebss_otp_request(request: Request, otp_request_id: str) -> dict:
    admin_user(request)
    result = inspect_onebss_mobile_gateway_otp(get_settings(), otp_request_id)
    if result.get("status") == "missing":
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=result.get("message") or "Khong tim thay OTP request.")
    return result


@router.get("/api/onebss-reports/jobs/{job_id}")
def get_onebss_report_job(request: Request, job_id: str) -> dict:
    admin_user(request)
    try:
        run = build_app_repository().get_onebss_report_run(job_id.strip())
    except (RuntimeError, sqlite3.Error, AttributeError) as error:
        raise_onebss_operation_error(error, "Khong kiem tra duoc task OneBSS")
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay job OneBSS.")
    return _onebss_report_job_response(job_id, {**run, "job_id": job_id})


def ensure_onebss_task_otp_request(repository: AppRepository, run: dict[str, Any], report: dict[str, Any], worker_session_id: str = "") -> dict[str, Any]:
    existing_request_id = str(run.get("otp_request_id") or "").strip()
    if existing_request_id:
        return {"otp_request_id": existing_request_id}
    settings = get_settings()
    session_label = worker_session_id or str(run.get("worker_session_id") or "") or str(run.get("run_id") or "")
    result = start_onebss_otp_mobile_gateway_request(
        settings,
        session_label,
        run.get("parameters") if isinstance(run.get("parameters"), dict) else {},
        otp_service_code=str(report.get("otp_service_code") or "onebss"),
        report_url=str(report.get("report_url") or ""),
        fallback_message="May tram bao OneBSS yeu cau OTP. Hay nhap OTP neu Mobile Gateway chua tu lay duoc.",
    )
    request_id = str(result.get("otp_request_id") or "").strip()
    updates = {
        "status": result.get("status") or "otp_required",
        "message": result.get("message") or "May tram dang doi OTP OneBSS.",
        "worker_session_id": session_label,
    }
    if request_id:
        updates["otp_request_id"] = request_id
    updated = repository.update_onebss_report_run(str(run.get("run_id") or ""), updates) or run
    return {"otp_request_id": request_id, "run": updated, "status": updated.get("status"), "message": updated.get("message")}


@router.post("/api/onebss-reports/jobs/{job_id}/otp")
def submit_onebss_report_job_otp(request: Request, job_id: str, payload: OneBssTaskOtpPayload) -> dict:
    admin_user(request)
    repository = build_app_repository()
    run = repository.get_onebss_report_run(job_id.strip())
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay job OneBSS.")
    otp_request_id = payload.otp_request_id.strip() or str(run.get("otp_request_id") or "")
    source = payload.otp_source.strip().lower()
    result: dict[str, Any]
    if payload.otp.strip():
        result = match_onebss_mobile_gateway_manual_otp(get_settings(), otp_request_id, payload.otp, source_id=job_id)
    else:
        result = inspect_onebss_mobile_gateway_otp(get_settings(), otp_request_id)
    if result.get("ok") or result.get("status") == "matched":
        run = repository.update_onebss_report_run(
            job_id.strip(),
            {
                "status": "running",
                "message": "Da nhan OTP cho task OneBSS, may tram se tiep tuc dang nhap.",
                "otp_request_id": otp_request_id,
            },
        ) or run
    return {"ok": bool(result.get("ok") or result.get("status") == "matched"), "otp": result, "run": _decorate_onebss_report_run(run)}


@router.get("/api/onebss-reports/runs/{run_id}/download")
def download_onebss_report_run(request: Request, run_id: str) -> Response:
    admin_user(request)
    run_id = run_id.strip()
    try:
        run = build_app_repository().get_onebss_report_run(run_id)
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay lan lay du lieu OneBSS.")
    file_path = Path(str(run.get("file_path") or ""))
    if not file_path.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="File ket qua OneBSS khong con ton tai tren may chu.")
    suffix = file_path.suffix.lower()
    media_type = "application/zip" if suffix == ".zip" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(file_path, media_type=media_type, filename=str(run.get("file_name") or file_path.name))


@router.post("/api/onebss-reports/run")
def run_onebss_report(request: Request, payload: RunOneBssReportPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    ma_bao_cao = payload.ma_bao_cao.strip().upper()
    try:
        report = repository.get_onebss_report_by_code(ma_bao_cao)
    except (RuntimeError, sqlite3.Error, AttributeError) as error:
        raise_onebss_operation_error(error, "Khong doc duoc cau hinh bao cao OneBSS")
    if not report:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay cau hinh bao cao OneBSS.")
    started_at = datetime.now().isoformat(timespec="seconds")
    run_parameters = payload.parameters if isinstance(payload.parameters, dict) and payload.parameters else report.get("parameters") or {}
    job_id = re.sub(r"[^A-Za-z0-9_-]+", "", payload.job_id.strip()) or uuid.uuid4().hex
    if payload.job_id.strip() and (payload.otp.strip() or payload.otp_request_id.strip()):
        return submit_onebss_report_job_otp(
            request,
            job_id,
            OneBssTaskOtpPayload(otp=payload.otp, otp_request_id=payload.otp_request_id, otp_source=payload.otp_source or "manual"),
        )
    try:
        existing_run = repository.get_onebss_report_run(job_id)
    except (RuntimeError, sqlite3.Error, AttributeError) as error:
        raise_onebss_operation_error(error, "Khong kiem tra duoc task OneBSS dang ton tai")
    if existing_run and str(existing_run.get("status") or "").lower() in ONEBSS_REPORT_ACTIVE_STATUSES:
        return _onebss_report_job_response(job_id, {**existing_run, "job_id": job_id})
    try:
        run = repository.save_onebss_report_run({
            "run_id": job_id,
            "ma_bao_cao": report.get("ma_bao_cao") or ma_bao_cao,
            "ten_bao_cao": report.get("ten_bao_cao") or ma_bao_cao,
            "status": "queued",
            "message": "Da dua yeu cau lay du lieu OneBSS vao hang doi may tram.",
            "parameters": run_parameters,
            "started_at": started_at,
            "finished_at": "",
            "created_by": actor["username"],
        })
    except (RuntimeError, sqlite3.Error, AttributeError) as error:
        raise_onebss_operation_error(error, "Khong tao duoc task lay bao cao OneBSS")
    try:
        repository.add_audit_log(actor["username"], "onebss_report_job_queued", f"Dua bao cao OneBSS {ma_bao_cao} vao hang doi: {job_id}")
    except Exception:
        logger.exception("Cannot write OneBSS report queued audit log")
    return _onebss_report_job_response(job_id, {**run, "job_id": job_id})


@router.post("/api/onebss-worker/tasks/claim")
def claim_onebss_worker_task(request: Request, payload: OneBssWorkerClaimPayload) -> dict:
    onebss_worker_token(request)
    repository = build_app_repository()
    try:
        run = repository.claim_next_onebss_report_run(payload.worker_id)
    except RuntimeError as error:
        raise_onebss_report_schema_error(error)
    if not run:
        return {"ok": True, "task": None, "message": "Khong co task OneBSS dang cho."}
    report = repository.get_onebss_report_by_code(str(run.get("ma_bao_cao") or ""))
    if not report:
        repository.update_onebss_report_run(
            str(run.get("run_id") or ""),
            {
                "status": "failed",
                "message": "Khong tim thay cau hinh bao cao OneBSS cho task.",
                "finished_at": datetime.now().isoformat(timespec="seconds"),
            },
        )
        return {"ok": False, "task": None, "message": "Khong tim thay cau hinh bao cao OneBSS."}
    folder_id = google_drive_folder_id(get_settings(), str(report.get("storage_link") or ""))
    return {
        "ok": True,
        "task": {
            "run_id": run.get("run_id"),
            "job_id": run.get("run_id"),
            "ma_bao_cao": run.get("ma_bao_cao"),
            "ten_bao_cao": run.get("ten_bao_cao"),
            "parameters": run.get("parameters") if isinstance(run.get("parameters"), dict) else {},
            "report": report,
            "report_url": report.get("report_url") or "",
            "storage_link": report.get("storage_link") or "",
            "drive_folder_id": folder_id,
            "otp_service_code": report.get("otp_service_code") or "onebss",
            "created_by": run.get("created_by") or "",
            "started_at": run.get("started_at") or "",
        },
    }


@router.post("/api/onebss-worker/tasks/{run_id}/status")
def update_onebss_worker_task_status(request: Request, run_id: str, payload: OneBssWorkerStatusPayload) -> dict:
    onebss_worker_token(request)
    repository = build_app_repository()
    run = repository.get_onebss_report_run(run_id.strip())
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay task OneBSS.")
    report = repository.get_onebss_report_by_code(str(run.get("ma_bao_cao") or "")) or {}
    status_value = payload.status.strip().lower()
    worker_session_id = payload.worker_session_id.strip()
    if status_value in {"waiting_otp", "otp_required", "manual_otp_required"}:
        otp_info = ensure_onebss_task_otp_request(repository, run, report, worker_session_id)
        updated_run = otp_info.get("run") if isinstance(otp_info.get("run"), dict) else repository.get_onebss_report_run(run_id.strip())
        return {"ok": True, "run": _decorate_onebss_report_run(updated_run or run), "otp_request_id": otp_info.get("otp_request_id") or ""}
    updated = repository.update_onebss_report_run(
        run_id.strip(),
        {
            "status": status_value or "running",
            "message": payload.message or "May tram dang xu ly task OneBSS.",
            "worker_id": payload.worker_id or run.get("worker_id") or "",
            "worker_session_id": worker_session_id or run.get("worker_session_id") or "",
        },
    )
    return {"ok": True, "run": _decorate_onebss_report_run(updated or run)}


@router.get("/api/onebss-worker/tasks/{run_id}/otp")
def get_onebss_worker_task_otp(request: Request, run_id: str) -> dict:
    onebss_worker_token(request)
    repository = build_app_repository()
    run = repository.get_onebss_report_run(run_id.strip())
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay task OneBSS.")
    request_id = str(run.get("otp_request_id") or "").strip()
    if not request_id:
        return {"ok": False, "status": "waiting", "message": "Task chua co OTP request.", "otp": ""}
    result = consume_onebss_mobile_gateway_otp(get_settings(), request_id)
    if result.get("ok") and result.get("otp"):
        updated = repository.update_onebss_report_run(
            run_id.strip(),
            {
                "status": "running",
                "message": "May tram da nhan OTP va dang tiep tuc dang nhap OneBSS.",
            },
        )
        return {"ok": True, "status": "matched", "otp": result.get("otp"), "run": _decorate_onebss_report_run(updated or run)}
    return {"ok": False, "status": result.get("status") or "waiting", "message": result.get("message") or "Chua co OTP.", "otp": ""}


@router.post("/api/onebss-worker/tasks/{run_id}/result")
def finish_onebss_worker_task(request: Request, run_id: str, payload: OneBssWorkerResultPayload) -> dict:
    onebss_worker_token(request)
    repository = build_app_repository()
    run = repository.get_onebss_report_run(run_id.strip())
    if not run:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay task OneBSS.")
    finished_at = datetime.now().isoformat(timespec="seconds")
    status_value = payload.status.strip().lower() or ("success" if payload.ok else "failed")
    updated = repository.update_onebss_report_run(
        run_id.strip(),
        {
            "status": status_value,
            "message": payload.message or ("Da lay bao cao OneBSS tren may tram." if payload.ok else "May tram khong lay duoc bao cao OneBSS."),
            "file_name": payload.file_name,
            "file_path": payload.file_path,
            "storage_link": payload.storage_link,
            "storage_status": payload.storage_status,
            "duration_ms": payload.duration_ms,
            "finished_at": finished_at,
        },
    )
    try:
        repository.add_audit_log(
            run.get("created_by") or "onebss-worker",
            "onebss_worker_task_finished",
            f"May tram tra ket qua task OneBSS {run_id}: {status_value}",
        )
    except Exception:
        logger.exception("Cannot write OneBSS worker result audit log")
    return {"ok": payload.ok, "run": _decorate_onebss_report_run(updated or run)}


@router.post("/api/admin/telegram/test-message")
def send_telegram_test_message(request: Request) -> dict:
    actor = admin_user(request)
    result = TelegramNotifier(get_settings()).test()
    build_app_repository().add_audit_log(
        actor["username"],
        "telegram_test_message_sent" if result.get("ok") else "telegram_test_message_failed",
        f"Kiem tra gui tin nhan Telegram: {result.get('message')}",
    )
    if not result.get("ok"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=result.get("message") or "Không gửi được Telegram. Kiểm tra TELEGRAM_TOKEN, MY_TELEGRAM_ID và hãy bấm Start trong bot.",
        )
    return result


@router.get("/api/test/telegram-alert")
def test_telegram_alert(request: Request) -> dict:
    actor = admin_user(request)
    sent = TelegramNotifier(get_settings()).send_message(
        "TEST Telegram",
        "\u26a0\ufe0f [TEST] Hệ thống kiểm tra kết nối Bot Telegram hoạt động bình thường!",
        {"route": "/api/test/telegram-alert", "actor": actor["username"]},
    )
    try:
        build_app_repository().add_audit_log(
            actor["username"],
            "telegram_admin_test_sent" if sent else "telegram_admin_test_failed",
            "GET /api/test/telegram-alert",
        )
    except Exception:
        pass
    if not sent:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Không gửi được Telegram. Kiểm tra token, chat ID và log Render.",
        )
    return {"ok": True, "message": "Đã gửi tin nhắn test Telegram."}


@router.post("/api/admin/zalo/webhook/setup")
def setup_zalo_webhook(request: Request) -> dict:
    actor = admin_user(request)
    result = ZaloBotClient(get_settings()).configure_webhook()
    build_app_repository().add_audit_log(
        actor["username"],
        "zalo_webhook_setup" if result.get("ok") else "zalo_webhook_setup_failed",
        f"Cai webhook Zalo: {result.get('message')}",
    )
    if not result.get("ok"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=result.get("message") or "Khong cai dat duoc webhook Zalo.",
        )
    return result


@router.post("/api/zalo/webhook")
async def zalo_webhook(request: Request) -> dict:
    client = ZaloBotClient(get_settings())
    if not client.webhook_secret:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="Zalo webhook chua duoc cau hinh.")
    if not client.has_valid_webhook_secret(request.headers.get("X-Bot-Api-Secret-Token")):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Zalo webhook secret khong hop le.")
    try:
        payload = await request.json()
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Payload Zalo khong hop le.") from error
    result = client.handle_webhook(payload)
    try:
        repository = build_app_repository()
        add_zalo_message_log(
            repository,
            direction="in",
            event_name=result.get("event_name") or "",
            chat_id=result.get("chat_id") or "",
            chat_type=result.get("chat_type") or "",
            sender_id=result.get("from_id") or "",
            sender_name=result.get("from_name") or "",
            message_id=result.get("message_id") or "",
            text=result.get("text") or "",
            raw_preview=result.get("raw_preview") or "",
            raw_keys=result.get("raw_keys") or [],
            result_keys=result.get("result_keys") or [],
            message_keys=result.get("message_keys") or [],
        )
        if result.get("reply_text"):
            add_zalo_message_log(
                repository,
                direction="out",
                event_name=result.get("event_name") or "",
                chat_id=result.get("chat_id") or "",
                chat_type=result.get("chat_type") or "",
                text=result.get("reply_text") or "",
                ok=bool(result.get("auto_replied")),
            )
    except Exception:
        pass
    return result


@router.get("/api/admin/zalo/message-logs")
def list_zalo_message_logs(request: Request, limit: int = 100) -> dict:
    admin_user(request)
    safe_limit = min(max(int(limit or 100), 1), 500)
    rows = build_app_repository().list_audit_logs(limit=min(max(safe_limit * 4, 100), 500))
    logs = []
    for row in rows:
        parsed = parse_zalo_message_log(row)
        if parsed:
            logs.append(parsed)
        if len(logs) >= safe_limit:
            break
    return {"logs": logs}


@router.post("/api/admin/zalo/send-test-message")
def send_zalo_test_message(request: Request, payload: ZaloSendMessagePayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    chat_id = payload.chat_id.strip() or latest_zalo_chat_id(repository)
    text = payload.text.strip() or "Tin nhan test tu Bot VNPT Can Tho."
    if not chat_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Chua co chat_id Zalo. Hay nhan tin rieng hoac mention bot trong nhom truoc, roi bam lam moi log.",
        )
    sent = ZaloBotClient(get_settings()).send_message(chat_id, text)
    add_zalo_message_log(
        repository,
        direction="out",
        event_name="manual_test",
        chat_id=chat_id,
        text=text,
        ok=sent,
    )
    repository.add_audit_log(actor["username"], "zalo_manual_test_sent" if sent else "zalo_manual_test_failed", f"Gui test Zalo toi {chat_id}: {sent}")
    if not sent:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Khong gui duoc tin nhan Zalo test.")
    return {"ok": True, "message": "Da gui tin nhan test Zalo.", "chat_id": chat_id}


@router.get("/api/admin/zalo/auto-messages")
def list_zalo_auto_messages(request: Request) -> dict:
    admin_user(request)
    repository = build_app_repository()
    try:
        schedules = repository.list_zalo_auto_messages()
        return {"schedules": [enrich_zalo_auto_message(repository, schedule) for schedule in schedules]}
    except RuntimeError as error:
        raise_zalo_auto_message_schema_error(error)


@router.post("/api/admin/zalo/auto-messages")
def save_zalo_auto_message(request: Request, payload: ZaloAutoMessagePayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        schedule_id = payload.schedule_id.strip() or repository.generate_zalo_auto_message_id()
        schedule_payload = normalize_zalo_auto_message_payload(payload, schedule_id)
        repository.save_zalo_auto_message(schedule_payload)
        schedule = repository.get_zalo_auto_message(schedule_id)
    except RuntimeError as error:
        raise_zalo_auto_message_schema_error(error)
    repository.add_audit_log(actor["username"], "zalo_auto_message_saved", f"Luu lich gui Zalo {schedule_id}")
    return {"ok": True, "schedule": enrich_zalo_auto_message(repository, schedule)}


@router.delete("/api/admin/zalo/auto-messages/{schedule_id}")
def delete_zalo_auto_message(request: Request, schedule_id: str) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        if not repository.get_zalo_auto_message(schedule_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay lich gui Zalo.")
        repository.delete_zalo_auto_message(schedule_id)
    except RuntimeError as error:
        raise_zalo_auto_message_schema_error(error)
    repository.add_audit_log(actor["username"], "zalo_auto_message_deleted", f"Xoa lich gui Zalo {schedule_id}")
    return {"ok": True}


@router.post("/api/admin/zalo/auto-messages/{schedule_id}/captures")
def upload_zalo_auto_message_capture(request: Request, schedule_id: str, payload: ZaloCapturePayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        schedule = repository.get_zalo_auto_message(schedule_id)
        if not schedule:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay lich gui Zalo.")
        image_base64, mime_type = decode_capture_image(payload)
        capture = repository.save_zalo_message_capture(
            schedule_id,
            image_base64,
            mime_type,
            normalize_zalo_page_url(payload.page_url or schedule.get("page_url") or "/"),
            actor["username"],
        )
    except RuntimeError as error:
        raise_zalo_auto_message_schema_error(error)
    repository.add_audit_log(actor["username"], "zalo_auto_message_capture_saved", f"Luu anh chup cho lich {schedule_id}")
    return {"ok": True, "capture": capture, "capture_url": capture_public_url(get_settings(), capture)}


@router.post("/api/admin/dashboard/capture")
def capture_dashboard_page(request: Request, payload: PageCapturePayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    page_url = normalize_zalo_page_url(payload.page_url or "/")
    try:
        image_bytes = capture_page_screenshot_bytes(repository, get_settings(), page_url)
    except Exception as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)[:500]) from error
    repository.add_audit_log(actor["username"], "dashboard_capture_generated", f"Chup anh Dashboard {page_url}")
    return {
        "ok": True,
        "mime_type": "image/png",
        "image_base64": base64.b64encode(image_bytes).decode("ascii"),
    }


@router.post("/api/admin/zalo/auto-messages/{schedule_id}/send-now")
def send_zalo_auto_message_now(request: Request, schedule_id: str) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        schedule = repository.get_zalo_auto_message(schedule_id)
        if not schedule:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay lich gui Zalo.")
        result = send_zalo_auto_message(repository, get_settings(), schedule)
    except RuntimeError as error:
        raise_zalo_auto_message_schema_error(error)
    repository.add_audit_log(actor["username"], "zalo_auto_message_manual_send", f"Gui thu lich Zalo {schedule_id}: {result.get('ok')}")
    if not result.get("ok"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=result.get("message") or "Khong gui duoc lich Zalo.")
    return result


@router.get("/api/zalo/auto-message-captures/{capture_id}")
def get_zalo_auto_message_capture(capture_id: str, token: str = "") -> Response:
    repository = build_app_repository()
    try:
        capture = repository.get_zalo_message_capture(capture_id)
    except RuntimeError as error:
        raise_zalo_auto_message_schema_error(error)
    if not capture or not hmac.compare_digest(str(token or ""), str(capture.get("public_token") or "")):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay anh.")
    try:
        image_bytes = base64.b64decode(str(capture.get("image_base64") or ""), validate=True)
    except (binascii.Error, ValueError) as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Anh khong hop le.") from error
    return Response(
        content=image_bytes,
        media_type=str(capture.get("mime_type") or "image/png"),
        headers={"Cache-Control": "public, max-age=86400"},
    )


@router.get("/api/admin/data-mining/schedules")
def list_data_mining_schedules(request: Request) -> dict:
    admin_user(request)
    repository = build_app_repository()
    try:
        return {"schedules": repository.list_data_mining_schedules()}
    except RuntimeError as error:
        raise_data_mining_schema_error(error)


@router.post("/api/admin/data-mining/schedules")
def save_data_mining_schedule(request: Request, payload: DataMiningSchedulePayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        schedule_id = payload.schedule_id.strip() or repository.generate_data_mining_schedule_id()
        schedule_payload = normalize_data_mining_schedule_payload(payload, schedule_id)
        repository.save_data_mining_schedule(schedule_payload)
        schedule = repository.get_data_mining_schedule(schedule_id)
    except RuntimeError as error:
        raise_data_mining_schema_error(error)
    repository.add_audit_log(actor["username"], "data_mining_schedule_saved", f"Luu lich dao du lieu {schedule_id}")
    return {"ok": True, "schedule": schedule}


@router.delete("/api/admin/data-mining/schedules/{schedule_id}")
def delete_data_mining_schedule(request: Request, schedule_id: str) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        if not repository.get_data_mining_schedule(schedule_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay lich dao du lieu.")
        repository.delete_data_mining_schedule(schedule_id)
    except RuntimeError as error:
        raise_data_mining_schema_error(error)
    repository.add_audit_log(actor["username"], "data_mining_schedule_deleted", f"Xoa lich dao du lieu {schedule_id}")
    return {"ok": True}


@router.get("/api/admin/data-mining/runs")
def list_data_mining_runs(request: Request, schedule_id: str = "", limit: int = 50) -> dict:
    admin_user(request)
    repository = build_app_repository()
    try:
        return {"runs": repository.list_data_mining_runs(schedule_id=schedule_id.strip(), limit=limit)}
    except RuntimeError as error:
        raise_data_mining_schema_error(error)


@router.post("/api/admin/data-mining/schedules/{schedule_id}/run-now")
def run_data_mining_schedule_now(request: Request, schedule_id: str, payload: DataMiningRunPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        schedule = repository.get_data_mining_schedule(schedule_id)
        if not schedule:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Khong tim thay lich dao du lieu.")
        result = run_data_mining_schedule(
            repository,
            get_settings(),
            schedule,
            otp=payload.otp.strip(),
            created_by=actor["username"],
            allow_device_registration=payload.allow_device_registration,
            interactive=True,
            parameter_overrides=payload.parameters if isinstance(payload.parameters, dict) else {},
        )
        run_key = f"manual:{result.get('run_id') or datetime.now().isoformat(timespec='seconds')}"
        repository.mark_data_mining_schedule_run(schedule_id, run_key, bool(result.get("ok")), result)
        refreshed_schedule = repository.get_data_mining_schedule(schedule_id)
    except RuntimeError as error:
        raise_data_mining_schema_error(error)
    repository.add_audit_log(actor["username"], "data_mining_manual_run", f"Chay thu lich dao du lieu {schedule_id}: {result.get('ok')}")
    return {"ok": bool(result.get("ok")), "result": result, "schedule": refreshed_schedule}


@router.get("/api/admin/work-tasks")
def list_work_tasks(request: Request, include_completed: bool = False) -> dict:
    admin_user(request)
    try:
        return {"tasks": build_app_repository().list_work_tasks(include_completed=include_completed)}
    except RuntimeError as error:
        raise_work_task_schema_error(error)


@router.post("/api/admin/work-tasks")
def save_work_task(request: Request, payload: WorkTaskPayload) -> dict:
    actor = admin_user(request)
    task_name = payload.ten_cong_viec.strip()
    if not task_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tên công việc không được để trống.")
    if not payload.time or len(payload.time.split(":")) < 2:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Thời gian chạy chưa đúng định dạng HH:MM.")
    if payload.type == "Once" and not payload.once_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Lịch chạy một lần cần nhập ngày chạy.")
    if payload.type == "Weekly" and not payload.weekday:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Lịch hàng tuần cần nhập ngày trong tuần.")
    repository = build_app_repository()
    try:
        task_id = payload.task_id.strip() or repository.generate_work_task_id()
    except RuntimeError as error:
        raise_work_task_schema_error(error)
    try:
        repository.save_work_task({
            "task_id": task_id,
            "ten_cong_viec": task_name,
            "type": payload.type,
            "time": payload.time[:5],
            "weekday": payload.weekday.strip(),
            "once_date": payload.once_date.strip(),
            "group": payload.group.strip(),
            "check": payload.check,
        })
    except RuntimeError as error:
        raise_work_task_schema_error(error)
    repository.add_audit_log(actor["username"], "work_task_saved", f"Luu lich cong viec {task_id}")
    return {"ok": True, "task": repository.get_work_task(task_id)}


@router.delete("/api/admin/work-tasks/{task_id}")
def delete_work_task(request: Request, task_id: str) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        if not repository.get_work_task(task_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy tác vụ.")
        repository.delete_work_task(task_id)
    except RuntimeError as error:
        raise_work_task_schema_error(error)
    repository.add_audit_log(actor["username"], "work_task_deleted", f"Xoa lich cong viec {task_id}")
    return {"ok": True}


@router.post("/api/admin/work-tasks/{task_id}/complete")
def complete_work_task(request: Request, task_id: str) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        if not repository.get_work_task(task_id):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy tác vụ.")
        repository.complete_work_task(task_id)
    except RuntimeError as error:
        raise_work_task_schema_error(error)
    repository.add_audit_log(actor["username"], "work_task_completed", f"Hoan thanh va an lich cong viec {task_id}")
    return {"ok": True, "message": "Đã đánh dấu hoàn thành và ẩn lịch công việc."}


@router.get("/api/notifications")
def notifications(request: Request) -> dict:
    current_user(request)
    return {
        "notifications": [
            {
                "title": "Thông báo hệ thống",
                "message": "Kênh thông báo nội bộ đã sẵn sàng. Thông báo của quản trị viên sẽ hiển thị tại đây.",
                "created_at": "2026-06-07T00:00:00+07:00",
            }
        ]
    }


@router.get("/api/websites")
def websites(request: Request) -> dict:
    require_feature(request, "xemdanhsachtaikhoan")
    return {"websites": build_app_repository().list_websites(active_only=True)}


@router.get("/api/credentials")
def credentials(request: Request) -> dict:
    user = require_feature(request, "xemdanhsachtaikhoan")
    return {"credentials": build_app_repository().list_credentials(user["id"])}


@router.post("/api/credentials")
def save_credential(request: Request, payload: CredentialPayload) -> dict:
    user = require_feature(request, "themvasuataikhoan")
    try:
        build_vault_service().save_credential(
            user, payload.id, payload.website_id, payload.login_username, payload.password, payload.notes
        )
    except (ValueError, sqlite3.IntegrityError) as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    return {"ok": True}


@router.post("/api/credentials/{credential_id}/reveal")
def reveal_credential(request: Request, credential_id: int) -> dict:
    user = require_feature(request, "xemmatkhaudaluu")
    try:
        password = build_vault_service().reveal_password(user, credential_id)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(error)) from error
    return {"password": password}


@router.delete("/api/credentials/{credential_id}")
def delete_credential(request: Request, credential_id: int) -> dict:
    user = require_feature(request, "themvasuataikhoan")
    try:
        build_vault_service().delete_credential(user, credential_id)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(error)) from error
    return {"ok": True}


@router.get("/api/admin/websites")
def admin_websites(request: Request) -> dict:
    admin_user(request)
    return {"websites": build_app_repository().list_websites()}


@router.post("/api/admin/websites")
def save_admin_website(request: Request, payload: WebsitePayload) -> dict:
    actor = admin_user(request)
    try:
        website = build_vault_service().save_website(
            actor["username"], payload.id, payload.name, payload.url, payload.requires_otp, payload.is_active
        )
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    return {"ok": True, "website": website}



@router.get("/api/admin/features")
def features(request: Request) -> dict:
    admin_user(request)
    return {"features": build_app_repository().list_features()}


@router.post("/api/admin/features/menu")
def create_menu_feature(request: Request, payload: CreateMenuPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    try:
        feature = repository.create_menu_feature(payload.name)
    except ValueError as error:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    repository.add_audit_log(actor["username"], "feature_menu_created", f"Create menu {feature.get('code')}")
    return {"ok": True, "feature": feature}


@router.put("/api/admin/features/layout")
def save_feature_layout(request: Request, payload: FeatureLayoutPayload) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    existing_features = repository.list_features()
    validate_feature_layout_payload(existing_features, payload.features)
    for item in payload.features:
        try:
            repository.update_feature_layout(item.code, item.name.strip(), item.parent_code, item.sort_order)
        except ValueError as error:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(error)) from error
    repository.add_audit_log(actor["username"], "feature_layout_saved", f"Cap nhat cau truc menu {len(payload.features)} chuc nang")
    return {"ok": True}


@router.get("/api/admin/roles")
def list_roles(request: Request) -> dict:
    admin_user(request)
    return {"roles": build_app_repository().list_system_roles()}


@router.post("/api/admin/roles")
def save_role(request: Request, payload: SystemRolePayload) -> dict:
    actor = admin_user(request)
    code = payload.code.strip().lower()
    if not code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Mã vai trò không được để trống.")
    build_app_repository().save_system_role(
        code,
        payload.name.strip(),
        payload.description.strip(),
        payload.is_active,
        payload.sort_order,
    )
    build_app_repository().add_audit_log(actor["username"], "role_saved", f"Luu vai tro {code}")
    return {"ok": True}


@router.delete("/api/admin/roles/{code}")
def delete_role(request: Request, code: str) -> dict:
    actor = admin_user(request)
    build_app_repository().delete_system_role(code.strip().lower())
    build_app_repository().add_audit_log(actor["username"], "role_deleted", f"Xoa vai tro {code}")
    return {"ok": True}


@router.get("/api/admin/users/{user_id}/permissions")
def user_permissions(request: Request, user_id: int) -> dict:
    admin_user(request)
    return {"feature_codes": build_app_repository().get_user_permissions(user_id)}


@router.put("/api/admin/users/{user_id}/permissions")
def update_permissions(request: Request, user_id: int, payload: PermissionPayload) -> dict:
    actor = admin_user(request)
    valid_codes = {feature["code"] for feature in build_app_repository().list_features()}
    if not set(payload.feature_codes).issubset(valid_codes):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh sách quyền không hợp lệ.")
    build_app_repository().set_user_permissions(user_id, payload.feature_codes)
    build_app_repository().add_audit_log(actor["username"], "permissions_updated", f"Cập nhật quyền người dùng #{user_id}")
    return {"ok": True}


@router.put("/api/admin/permissions/bulk")
def update_bulk_permissions(request: Request, payload: BulkPermissionPayload) -> dict:
    actor = admin_user(request)
    valid_codes = {feature["code"] for feature in build_app_repository().list_features()}
    if not payload.user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chưa chọn người dùng.")
    if not set(payload.feature_codes).issubset(valid_codes):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh sách quyền không hợp lệ.")
    build_app_repository().set_bulk_user_permissions(payload.user_ids, payload.feature_codes)
    build_app_repository().add_audit_log(actor["username"], "bulk_permissions_updated", f"Cap quyen cho {len(payload.user_ids)} user")
    return {"ok": True}


@router.get("/api/admin/regions")
def list_regions(request: Request) -> dict:
    admin_user(request)
    return {"regions": build_app_repository().list_data_regions()}


@router.post("/api/admin/regions")
def save_region(request: Request, payload: DataRegionPayload) -> dict:
    actor = admin_user(request)
    build_app_repository().save_data_region(payload.code.strip(), payload.name.strip(), payload.is_active, payload.sort_order)
    build_app_repository().add_audit_log(actor["username"], "region_saved", f"Luu phan vung {payload.code}")
    return {"ok": True}


@router.delete("/api/admin/regions/{code}")
def delete_region(request: Request, code: str) -> dict:
    actor = admin_user(request)
    repository = build_app_repository()
    repository.delete_data_region(code.strip())
    repository.add_audit_log(actor["username"], "region_deleted", f"Xoa phan vung {code}")
    return {"ok": True}


@router.get("/api/admin/users/{user_id}/data-permissions")
def user_data_permissions(request: Request, user_id: int) -> dict:
    admin_user(request)
    return {"region_codes": build_app_repository().get_user_data_permissions(user_id)}


@router.put("/api/admin/data-permissions/bulk")
def update_bulk_data_permissions(request: Request, payload: BulkDataPermissionPayload) -> dict:
    actor = admin_user(request)
    valid_codes = {region["code"] for region in build_app_repository().list_data_regions()}
    if not payload.user_ids:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chưa chọn người dùng.")
    if not set(payload.region_codes).issubset(valid_codes):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Danh sách phân vùng không hợp lệ.")
    build_app_repository().set_bulk_user_data_permissions(payload.user_ids, payload.region_codes)
    build_app_repository().add_audit_log(actor["username"], "bulk_data_permissions_updated", f"Cap vung du lieu cho {len(payload.user_ids)} user")
    return {"ok": True}


@router.get("/{feature_path:path}", response_class=HTMLResponse)
def feature_page(request: Request, feature_path: str) -> Response:
    if feature_path.startswith("api/"):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy API.")
    normalize_feature_code(feature_path)
    return render_index_page(request, feature_path)
