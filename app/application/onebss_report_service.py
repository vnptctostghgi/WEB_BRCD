from __future__ import annotations

import base64
import hashlib
import json
import logging
import re
import threading
import time
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime
from itertools import product
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, unquote, urlparse

import httpx

from app.application.onebss_data_mining_service import (
    EXPORT_DIRECT_SCRIPT,
    LOCAL_TIMEZONE,
    ONEBSS_STATE_PATH,
    REPORT_COMPONENT_READY_SCRIPT,
    OneBssDownloadError,
    OneBssReportDownloader,
    build_target_file_path,
    normalize_onebss_report_url,
    save_downloaded_file,
    with_resolved_schedule_parameters,
)
from app.application.zalo_auto_message_service import install_playwright_chromium, playwright_needs_browser_install
from app.data_access.repository_factory import build_repository
from app.modules.mobile_gateway.exceptions import OtpServiceError
from app.modules.mobile_gateway.otp_service import OtpService
from app.modules.mobile_gateway.repository import MobileGatewayRepository
from app.settings import Settings, get_settings


logger = logging.getLogger(__name__)
OTP_PATTERN = re.compile(r"^\d{4,8}$")
PENDING_SESSION_TTL_SECONDS = 10 * 60
ONEBSS_API_BASE_URL = "https://api-onebss.vnpt.vn"
ONEBSS_API_TOKEN_ID = "97388db0-6ce9-11ea-bc55-0242ac130003"
ONEBSS_API_CLIENT_ID = "clientapp"
ONEBSS_API_CLIENT_SECRET = "password"
ONEBSS_API_META_KEYS = {"baocao_id"}
ONEBSS_EXPORT_TIMEOUT_SECONDS = 900
KNOWN_ONEBSS_REPORT_IDS = {
    "PHATTRIENTHUEBAO/BIENDONGPHATTRIENTHUEBAO/RP_BSS_28429": 41668,
}
OTP_TEXT_NEEDLES = ["OTP", "mã OTP", "ma OTP", "mã xác thực", "ma xac thuc", "mã xác nhận", "ma xac nhan"]
OTP_INVALID_TEXT_NEEDLES = ["Số OTP không hợp lệ", "OTP không hợp lệ", "OTP khong hop le"]
OTP_REQUEST_TEXT_NEEDLES = [
    "Gửi yêu cầu",
    "Gui yeu cau",
    "Gửi mã",
    "Gui ma",
    "Gửi OTP",
    "Gui OTP",
    "Xác nhận",
    "Xac nhan",
    "xác nhận đăng nhập",
    "xac nhan dang nhap",
]
OTP_REQUEST_BUTTON_TEXTS = [
    "Gửi yêu cầu",
    "Gui yeu cau",
    "Gửi mã",
    "Gui ma",
    "Gửi OTP",
    "Gui OTP",
    "Xác nhận",
    "Xac nhan",
    "Tiếp tục",
    "Tiep tuc",
    "Đăng nhập",
    "Dang nhap",
]
LOGIN_ERROR_TEXT_NEEDLES = [
    "sai mật khẩu",
    "sai mat khau",
    "không đúng",
    "khong dung",
    "không hợp lệ",
    "khong hop le",
    "tài khoản hoặc mật khẩu",
    "tai khoan hoac mat khau",
]


@dataclass
class PendingOneBssSession:
    session_id: str
    playwright: Any
    browser: Any
    context: Any
    page: Any
    report: dict[str, Any]
    parameters: dict[str, Any]
    created_by: str
    created_at: float


@dataclass
class PendingOneBssApiSession:
    session_id: str
    secret_code: str
    report: dict[str, Any]
    parameters: dict[str, Any]
    username: str
    mobile_id: str
    device_id: str
    created_by: str
    created_at: float


@dataclass
class OneBssApiToken:
    access_token: str
    token_type: str
    username: str
    mobile_id: str
    device_id: str
    expires_at: float


@dataclass
class OneBssParameterRun:
    parameters: dict[str, Any]
    source_values: dict[str, Any]


@dataclass
class OneBssDownloadedFile:
    file_path: Path
    suggested_filename: str
    export_info: dict[str, Any]
    parameters: dict[str, Any]
    source_values: dict[str, Any]


PENDING_ONEBSS_SESSIONS: dict[str, PendingOneBssSession] = {}
PENDING_ONEBSS_API_SESSIONS: dict[str, PendingOneBssApiSession] = {}
ONEBSS_API_TOKENS: dict[str, OneBssApiToken] = {}
PENDING_ONEBSS_LOCK = threading.Lock()


def run_onebss_report_request(
    settings: Settings,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    otp: str = "",
    session_id: str = "",
    created_by: str = "",
) -> dict[str, Any]:
    cleanup_expired_onebss_sessions()
    resolved_parameters = with_resolved_schedule_parameters({"parameters": parameters if isinstance(parameters, dict) else {}})["parameters"]
    if session_id.startswith("api-"):
        return continue_onebss_api_session(settings, session_id, otp, resolved_parameters)
    if session_id:
        return continue_onebss_report_session(settings, session_id, otp, resolved_parameters)
    return start_onebss_api_session(settings, report, resolved_parameters, created_by=created_by)


def onebss_report_otp_service_code(report: dict[str, Any] | None) -> str:
    value = str((report or {}).get("otp_service_code") or "onebss").strip().lower()
    return value or "onebss"


def onebss_manual_otp_response(
    session_id: str,
    parameters: dict[str, Any],
    message: str,
    *,
    status: str = "otp_required",
    report_url: str = "",
    otp_request_id: str = "",
) -> dict[str, Any]:
    response: dict[str, Any] = {
        "ok": False,
        "status": status,
        "message": message,
        "session_id": session_id,
        "parameters": parameters,
    }
    if report_url:
        response["report_url"] = report_url
    if otp_request_id:
        response["otp_request_id"] = otp_request_id
    return response


def resolve_onebss_otp_with_mobile_gateway(
    settings: Settings,
    session_id: str,
    parameters: dict[str, Any],
    *,
    otp_service_code: str = "onebss",
    report_url: str = "",
    fallback_message: str = "OneBSS da gui OTP ve dien thoai. Hay nhap OTP neu Mobile Gateway chua tu lay duoc.",
) -> dict[str, Any]:
    if not bool(getattr(settings, "mobile_gateway_enabled", True)):
        return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)

    request_id = ""
    manual_fallback_enabled = True
    service_code = str(otp_service_code or "onebss").strip().lower() or "onebss"
    try:
        repository = MobileGatewayRepository(build_repository(settings), settings)
        config = repository.get_otp_configuration(service_code) or {}
        manual_fallback_enabled = bool(config.get("manual_fallback_enabled", True))
        if config and not bool(config.get("auto_fill_enabled", True)):
            return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)
        service = OtpService(repository)
        otp_request = service.create_request(service_code, job_id=session_id)
        request_id = str(otp_request.get("request_id") or "")
        timeout_seconds = int((config or {}).get("wait_timeout_seconds") or otp_request.get("wait_timeout_seconds") or 120)
        code = service.wait_for_code(request_id, timeout_seconds)
    except OtpServiceError as error:
        logger.info("OneBSS Mobile Gateway OTP is not available: %s", str(error)[:160])
        return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)
    except Exception:
        logger.exception("Cannot wait OneBSS OTP from Mobile Gateway")
        return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)

    if code:
        if session_id.startswith("api-"):
            return continue_onebss_api_session(settings, session_id, code, parameters)
        return continue_onebss_report_session(settings, session_id, code, parameters)

    status_value = "manual_otp_required" if manual_fallback_enabled else "otp_timeout"
    message = (
        "Chua nhan duoc OTP tu Mobile Gateway trong thoi gian cho. Hay nhap OTP thu cong."
        if manual_fallback_enabled
        else "Chua nhan duoc OTP tu Mobile Gateway trong thoi gian cho."
    )
    return onebss_manual_otp_response(
        session_id,
        parameters,
        message,
        status=status_value,
        report_url=report_url,
        otp_request_id=request_id,
    )


def start_onebss_otp_mobile_gateway_request(
    settings: Settings,
    session_id: str,
    parameters: dict[str, Any],
    *,
    otp_service_code: str = "onebss",
    report_url: str = "",
    fallback_message: str = "OneBSS da gui OTP ve dien thoai. Hay nhap OTP neu Mobile Gateway chua tu lay duoc.",
) -> dict[str, Any]:
    if not bool(getattr(settings, "mobile_gateway_enabled", True)):
        return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)

    service_code = str(otp_service_code or "onebss").strip().lower() or "onebss"
    try:
        repository = MobileGatewayRepository(build_repository(settings), settings)
        config = repository.get_otp_configuration(service_code) or {}
        if config and not bool(config.get("auto_fill_enabled", True)):
            return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)
        service = OtpService(repository)
        otp_request = service.create_request(service_code, job_id=session_id)
        request_id = str(otp_request.get("request_id") or "")
    except OtpServiceError as error:
        logger.info("OneBSS Mobile Gateway OTP is not available: %s", str(error)[:160])
        return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)
    except Exception:
        logger.exception("Cannot create OneBSS OTP request from Mobile Gateway")
        return onebss_manual_otp_response(session_id, parameters, fallback_message, report_url=report_url)

    message = "OneBSS da gui OTP ve dien thoai. He thong dang kiem tra tin nhan; co the nhap OTP thu cong neu nhan duoc truoc."
    return onebss_manual_otp_response(
        session_id,
        parameters,
        message,
        status="otp_required",
        report_url=report_url,
        otp_request_id=request_id,
    )


def inspect_onebss_mobile_gateway_otp(settings: Settings, request_id: str) -> dict[str, Any]:
    request_id = str(request_id or "").strip()
    if not request_id:
        return {"ok": False, "status": "missing", "message": "Thieu OTP request."}
    if not bool(getattr(settings, "mobile_gateway_enabled", True)):
        return {"ok": False, "status": "disabled", "message": "Mobile Gateway chua duoc bat."}

    try:
        repository = MobileGatewayRepository(build_repository(settings), settings)
        repository.expire_otp_requests()
        otp_request = repository.get_otp_request(request_id)
        if otp_request and otp_request.get("status") == "waiting":
            OtpService(repository).match_latest_for_request(otp_request)
            otp_request = repository.get_otp_request(request_id)
    except Exception:
        logger.exception("Cannot inspect OneBSS OTP request")
        return {"ok": False, "status": "failed", "message": "Khong kiem tra duoc OTP tu Mobile Gateway."}

    if not otp_request:
        return {"ok": False, "status": "missing", "message": "Khong tim thay OTP request."}

    status_value = str(otp_request.get("status") or "waiting")
    messages = {
        "waiting": "Dang doi tin nhan OTP.",
        "matched": "Da boc tach OTP tu Mobile Gateway.",
        "expired": "OTP request da het thoi gian cho.",
        "cancelled": "OTP request da bi huy.",
        "consumed": "OTP request da duoc su dung.",
    }
    return {
        "ok": status_value == "matched",
        "status": status_value,
        "message": messages.get(status_value, "Chua co OTP moi."),
        "code_masked": otp_request.get("code_masked") or "",
        "source_type": otp_request.get("matched_source_type") or "",
        "source_id": otp_request.get("matched_source_id") or "",
        "matched_at": otp_request.get("matched_at") or "",
    }


def consume_onebss_mobile_gateway_otp(settings: Settings, request_id: str) -> dict[str, Any]:
    request_id = str(request_id or "").strip()
    if not request_id:
        return {"ok": False, "status": "missing", "message": "Thieu OTP request."}
    if not bool(getattr(settings, "mobile_gateway_enabled", True)):
        return {"ok": False, "status": "disabled", "message": "Mobile Gateway chua duoc bat."}

    try:
        repository = MobileGatewayRepository(build_repository(settings), settings)
        repository.expire_otp_requests()
        otp_request = repository.get_otp_request(request_id)
        if otp_request and otp_request.get("status") == "waiting":
            OtpService(repository).match_latest_for_request(otp_request)
            otp_request = repository.get_otp_request(request_id)
    except Exception:
        logger.exception("Cannot inspect OneBSS OTP request")
        return {"ok": False, "status": "failed", "message": "Khong kiem tra duoc OTP tu Mobile Gateway."}

    if not otp_request:
        return {"ok": False, "status": "missing", "message": "Khong tim thay OTP request."}

    status_value = str(otp_request.get("status") or "waiting")
    if status_value == "matched":
        code = OtpService(repository).consume_code(request_id)
        if code:
            return {
                "ok": True,
                "status": "matched",
                "otp": code,
                "source_type": otp_request.get("matched_source_type") or "",
                "source_id": otp_request.get("matched_source_id") or "",
                "matched_at": otp_request.get("matched_at") or "",
            }
        return {"ok": False, "status": "consume_failed", "message": "Da nhan OTP nhung khong doc duoc ma."}

    messages = {
        "waiting": "Dang doi tin nhan OTP.",
        "expired": "OTP request da het thoi gian cho.",
        "cancelled": "OTP request da bi huy.",
        "consumed": "OTP request da duoc su dung.",
    }
    return {
        "ok": False,
        "status": status_value,
        "message": messages.get(status_value, "Chua co OTP moi."),
        "code_masked": otp_request.get("code_masked") or "",
    }


def cancel_onebss_mobile_gateway_otp(settings: Settings, request_id: str, reason: str = "manual_otp_used") -> None:
    request_id = str(request_id or "").strip()
    if not request_id or not bool(getattr(settings, "mobile_gateway_enabled", True)):
        return
    try:
        repository = MobileGatewayRepository(build_repository(settings), settings)
        OtpService(repository).cancel_request(request_id, reason)
    except Exception:
        logger.exception("Cannot cancel OneBSS OTP request %s", request_id)


def match_onebss_mobile_gateway_manual_otp(settings: Settings, request_id: str, code: str, source_id: str = "") -> dict[str, Any]:
    request_id = str(request_id or "").strip()
    code = re.sub(r"\D+", "", str(code or ""))[:8]
    if not request_id:
        return {"ok": False, "status": "missing_request", "message": "Chua co OTP request cho task OneBSS."}
    if not OTP_PATTERN.fullmatch(code):
        return {"ok": False, "status": "invalid_otp", "message": "OTP phai tu 4 den 8 chu so."}
    if not bool(getattr(settings, "mobile_gateway_enabled", True)):
        return {"ok": False, "status": "mobile_gateway_disabled", "message": "Mobile Gateway chua bat."}
    try:
        repository = MobileGatewayRepository(build_repository(settings), settings)
        matched = repository.match_otp_request(request_id, "manual", source_id or request_id, code)
        if matched:
            return {"ok": True, "status": "matched", "message": "Da ghi OTP thu cong cho task OneBSS."}
        request = repository.get_otp_request(request_id)
        status_value = str((request or {}).get("status") or "missing")
        return {
            "ok": status_value in {"matched", "consumed"},
            "status": status_value,
            "message": "OTP request da co ma khac hoac khong con cho OTP.",
        }
    except Exception as error:
        logger.exception("Cannot match manual OneBSS OTP request %s", request_id)
        return {"ok": False, "status": "failed", "message": str(error)[:300]}


def start_onebss_api_session(
    settings: Settings,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    created_by: str = "",
) -> dict[str, Any]:
    try:
        report_url = normalize_onebss_report_url(report.get("report_url"))
    except OneBssDownloadError as error:
        return {"ok": False, "status": "invalid_report_url", "message": str(error), "parameters": parameters}
    credentials = onebss_api_credentials(settings)
    if not credentials:
        return {"ok": False, "status": "missing_credentials", "message": "Chua cau hinh ONEBSS_USERNAME/ONEBSS_PASSWORD.", "parameters": parameters}
    username, password = credentials
    mobile_id, device_id = onebss_api_device_ids(username)
    cached_token = get_valid_onebss_api_token(username)
    if cached_token:
        try:
            return finish_onebss_report_download_api(settings, cached_token, report, parameters)
        except OneBssDownloadError as error:
            if onebss_error_looks_auth_related(error):
                forget_onebss_api_token(username)
            return {"ok": False, "status": "failed", "message": str(error)[:1000], "parameters": parameters}

    try:
        with httpx.Client(timeout=onebss_api_timeout(settings, minimum_seconds=30)) as client:
            response = client.post(
                f"{ONEBSS_API_BASE_URL}/quantri/user/xacthuc_tapdoan_v2",
                headers=onebss_api_base_headers(),
                json={
                    "username": username,
                    "password": password,
                    "mobile_id": mobile_id,
                    "device_id": device_id,
                },
            )
    except httpx.HTTPError as error:
        logger.exception("Cannot start OneBSS API login")
        return {"ok": False, "status": "login_failed", "message": f"Khong ket noi duoc API dang nhap OneBSS: {error}", "parameters": parameters}

    data = parse_onebss_json_response(response)
    secret_code = str(((data.get("data") if isinstance(data, dict) else {}) or {}).get("secretCode") or "").strip()
    if response.status_code == 200 and secret_code:
        pending = keep_onebss_api_session(secret_code, report, parameters, username, mobile_id, device_id, created_by)
        return start_onebss_otp_mobile_gateway_request(
            settings,
            pending.session_id,
            parameters,
            otp_service_code=onebss_report_otp_service_code(report),
            report_url=report_url,
            fallback_message="OneBSS da gui OTP ve dien thoai. Hay nhap OTP.",
        )
    return {
        "ok": False,
        "status": "login_failed",
        "message": onebss_api_error_message(data, response, fallback="Dang nhap OneBSS chua thanh cong."),
        "parameters": parameters,
        "report_url": report_url,
    }


def continue_onebss_api_session(
    settings: Settings,
    session_id: str,
    otp: str,
    parameters: dict[str, Any],
) -> dict[str, Any]:
    pending = get_onebss_api_session(session_id)
    if not pending:
        return {"ok": False, "status": "otp_session_expired", "message": "Phien OTP OneBSS da het han, hay bam lay bao cao lai."}
    if not OTP_PATTERN.fullmatch(str(otp or "").strip()):
        return {"ok": False, "status": "otp_required", "message": "OTP phai tu 4 den 8 chu so.", "session_id": session_id, "parameters": pending.parameters}

    effective_parameters = parameters if parameters else pending.parameters
    try:
        with httpx.Client(timeout=onebss_api_timeout(settings, minimum_seconds=30)) as client:
            response = client.post(
                f"{ONEBSS_API_BASE_URL}/quantri/oauth/token",
                headers=onebss_api_base_headers(),
                json={
                    "grant_type": "password",
                    "client_id": ONEBSS_API_CLIENT_ID,
                    "client_secret": ONEBSS_API_CLIENT_SECRET,
                    "otp": str(otp).strip(),
                    "secretCode": pending.secret_code,
                },
            )
    except httpx.HTTPError as error:
        logger.exception("Cannot verify OneBSS API OTP")
        return {
            "ok": False,
            "status": "otp_invalid",
            "message": f"Khong kiem tra duoc OTP OneBSS: {error}",
            "session_id": session_id,
            "parameters": effective_parameters,
        }

    data = parse_onebss_json_response(response)
    token_payload = onebss_token_payload(data)
    access_token = str(token_payload.get("access_token") or "").strip()
    if response.status_code != 200 or not access_token:
        status = "otp_invalid" if response.status_code in {400, 401, 403} else "login_failed"
        return {
            "ok": False,
            "status": status,
            "message": onebss_api_error_message(data, response, fallback="OTP OneBSS khong hop le hoac da het han."),
            "session_id": session_id,
            "parameters": effective_parameters,
        }

    token = remember_onebss_api_token(
        pending.username,
        token_payload,
        mobile_id=pending.mobile_id,
        device_id=pending.device_id,
    )
    pop_onebss_api_session(session_id)
    try:
        return finish_onebss_report_download_api(settings, token, pending.report, effective_parameters)
    except Exception as error:
        logger.exception("Cannot finish OneBSS API report request")
        return {"ok": False, "status": "failed", "message": str(error)[:1000], "parameters": effective_parameters}


def start_onebss_report_session(
    settings: Settings,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    created_by: str = "",
) -> dict[str, Any]:
    report_url = normalize_onebss_report_url(report.get("report_url"))
    username = str(getattr(settings, "onebss_username", "") or "").strip()
    password = settings.onebss_password.get_secret_value() if getattr(settings, "onebss_password", None) else ""
    if not username or not password:
        return {"ok": False, "status": "missing_credentials", "message": "Chua cau hinh ONEBSS_USERNAME/ONEBSS_PASSWORD."}

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return {"ok": False, "status": "missing_playwright", "message": "May chu chua cai Playwright."}

    playwright: Any | None = None
    browser: Any | None = None
    context: Any | None = None
    try:
        playwright = sync_playwright().start()
        try:
            browser = playwright.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"], timeout=30000)
        except Exception as launch_error:
            if playwright_needs_browser_install(launch_error):
                install_playwright_chromium()
                browser = playwright.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"], timeout=30000)
            else:
                raise
        context_options: dict[str, Any] = {
            "accept_downloads": True,
            "locale": "vi-VN",
            "viewport": {"width": 1440, "height": 920},
        }
        if ONEBSS_STATE_PATH.exists():
            context_options["storage_state"] = str(ONEBSS_STATE_PATH)
        context = browser.new_context(**context_options)
        page = context.new_page()
        page.set_default_timeout(30000)
        page.set_default_navigation_timeout(30000)
        helper = OneBssReportDownloader(settings)
        goto_onebss_page(page, report_url, step="open_report")
        wait_for_onebss_network_quiet(page, timeout_ms=8000, pause_ms=1000)
        if helper._is_login_page(page):
            if not wait_for_onebss_login_form(page, timeout_ms=30000):
                close_browser_stack(browser, context, playwright)
                return {
                    "ok": False,
                    "status": "login_form_not_ready",
                    "message": f"Trang dang nhap OneBSS chua hien form tren may chu. {onebss_login_diagnostic_text(page)}",
                    "parameters": parameters,
                }
            username_filled = helper._fill_first(page, ["input[name='username']", "input[placeholder*='Tài khoản']", "input[placeholder*='Tên']", "input[type='text']"], username)
            password_filled = helper._fill_first(page, ["input[name='password']", "input[placeholder*='Mật khẩu']", "input[type='password']"], password)
            if not username_filled or not password_filled:
                close_browser_stack(browser, context, playwright)
                return {
                    "ok": False,
                    "status": "login_form_not_ready",
                    "message": f"Khong dien duoc tai khoan/mat khau OneBSS tren may chu. {onebss_login_diagnostic_text(page)}",
                    "parameters": parameters,
                }
            click_onebss_button(page, helper, ["Đăng nhập", "Dang nhap", "Login"])
            wait_for_onebss_auth_transition(page, helper, timeout_ms=30000)
            if page_contains(page, OTP_TEXT_NEEDLES):
                pending = keep_onebss_session(playwright, browser, context, page, report, parameters, created_by)
                return start_onebss_otp_mobile_gateway_request(
                    settings,
                    pending.session_id,
                    parameters,
                    otp_service_code=onebss_report_otp_service_code(report),
                    fallback_message="OneBSS yeu cau OTP. Hay nhap OTP neu Mobile Gateway chua tu lay duoc.",
                )
            device_result = handle_onebss_device_registration(page, helper, parameters)
            if device_result:
                close_browser_stack(browser, context, playwright)
                return device_result
            otp_request_result = handle_onebss_otp_request(settings, page, helper, playwright, browser, context, report, parameters, created_by)
            if otp_request_result:
                return otp_request_result
            if helper._is_login_page(page):
                close_browser_stack(browser, context, playwright)
                return {"ok": False, "status": "login_failed", "message": onebss_login_failed_message(page), "parameters": parameters}
        result = finish_onebss_report_download(settings, helper, context, page, report, parameters)
        close_browser_stack(browser, context, playwright)
        return result
    except Exception as error:
        logger.exception("Cannot start OneBSS report request")
        close_browser_stack(browser, context, playwright)
        return {"ok": False, "status": "failed", "message": str(error)[:1000], "parameters": parameters}


def continue_onebss_report_session(
    settings: Settings,
    session_id: str,
    otp: str,
    parameters: dict[str, Any],
) -> dict[str, Any]:
    pending = get_onebss_session(session_id)
    if not pending:
        return {"ok": False, "status": "otp_session_expired", "message": "Phien OTP OneBSS da het han, hay bam lay bao cao lai."}
    if not OTP_PATTERN.fullmatch(str(otp or "").strip()):
        return {"ok": False, "status": "otp_required", "message": "OTP phai tu 4 den 8 chu so.", "session_id": session_id, "parameters": pending.parameters}

    page = pending.page
    context = pending.context
    helper = OneBssReportDownloader(settings)
    try:
        pending.parameters = parameters if parameters else pending.parameters
        helper._fill_otp(page, str(otp).strip())
        click_onebss_button(page, helper, ["Xác nhận", "Xac nhan", "Gửi yêu cầu", "Gui yeu cau", "Đăng nhập", "Dang nhap"])
        wait_for_onebss_auth_transition(page, helper, timeout_ms=30000)
        if page_contains(page, OTP_INVALID_TEXT_NEEDLES):
            return {
                "ok": False,
                "status": "otp_invalid",
                "message": "OTP OneBSS khong hop le, hay nhap lai OTP moi.",
                "session_id": session_id,
                "parameters": pending.parameters,
            }
        if page_contains(page, OTP_TEXT_NEEDLES) and helper._is_login_page(page):
            return {"ok": False, "status": "otp_required", "message": "OneBSS van dang yeu cau OTP.", "session_id": session_id, "parameters": pending.parameters}
        device_result = handle_onebss_device_registration(page, helper, pending.parameters)
        if device_result:
            pop_onebss_session(session_id)
            close_browser_stack(pending.browser, pending.context, pending.playwright)
            return device_result
        if helper._is_login_page(page):
            pop_onebss_session(session_id)
            close_browser_stack(pending.browser, pending.context, pending.playwright)
            return {"ok": False, "status": "login_failed", "message": onebss_login_failed_message(page), "parameters": pending.parameters}
        report_url = normalize_onebss_report_url(pending.report.get("report_url"))
        if page.url != report_url:
            goto_onebss_page(page, report_url, step="open_report_after_otp")
            wait_for_onebss_network_quiet(page, timeout_ms=8000, pause_ms=1000)
        result = finish_onebss_report_download(settings, helper, context, page, pending.report, pending.parameters)
        pop_onebss_session(session_id)
        close_browser_stack(pending.browser, pending.context, pending.playwright)
        return result
    except Exception as error:
        logger.exception("Cannot continue OneBSS report request")
        pop_onebss_session(session_id)
        close_browser_stack(pending.browser, pending.context, pending.playwright)
        return {"ok": False, "status": "failed", "message": str(error)[:1000], "parameters": pending.parameters}


def finish_onebss_report_download_api(
    settings: Settings,
    token: OneBssApiToken,
    report: dict[str, Any],
    parameters: dict[str, Any],
) -> dict[str, Any]:
    started = time.monotonic()
    report_url = normalize_onebss_report_url(report.get("report_url"))
    schedule_like = {
        "report_url": report_url,
        "storage_link": report.get("storage_link") or "",
        "file_name_template": report.get("ten_bao_cao") or report.get("ma_bao_cao") or "",
    }
    parameter_runs, merge_config, each_keys = build_onebss_parameter_runs(parameters)
    downloaded_files: list[OneBssDownloadedFile] = []
    temporary_files: list[Path] = []
    try:
        if len(parameter_runs) == 1:
            downloaded = download_onebss_report_file_api(
                settings,
                token,
                report,
                parameter_runs[0].parameters,
                source_values=parameter_runs[0].source_values,
            )
            downloaded_files.append(downloaded)
            target_file = downloaded.file_path
            storage_result = save_downloaded_file(settings, target_file, str(report.get("storage_link") or ""))
            ok = bool(storage_result.get("ok", True))
            export_info = downloaded.export_info
            return {
                "ok": ok,
                "status": "success" if ok else str(storage_result.get("status") or "storage_failed"),
                "message": storage_result.get("message") or "Da tai bao cao OneBSS.",
                "file_name": target_file.name,
                "file_path": str(target_file),
                "storage_link": storage_result.get("storage_link") or str(report.get("storage_link") or ""),
                "storage_status": storage_result.get("storage_status") or "",
                "report_id": export_info.get("report_id") or "",
                "report_title": export_info.get("title") or report.get("ten_bao_cao") or "",
                "parameters": export_info.get("params") or parameter_runs[0].parameters,
                "duration_ms": int((time.monotonic() - started) * 1000),
                "finished_at": datetime.now(LOCAL_TIMEZONE).isoformat(timespec="seconds"),
            }

        for index, parameter_run in enumerate(parameter_runs, start=1):
            target_file = build_onebss_temp_file_path(settings, index)
            temporary_files.append(target_file)
            downloaded_files.append(
                download_onebss_report_file_api(
                    settings,
                    token,
                    report,
                    parameter_run.parameters,
                    target_file=target_file,
                    source_values=parameter_run.source_values,
                )
            )

        return finalize_onebss_multiple_downloads(
            settings,
            report,
            parameters,
            schedule_like,
            downloaded_files,
            merge_config,
            each_keys,
            started,
        )
    finally:
        for temporary_file in temporary_files:
            try:
                temporary_file.unlink(missing_ok=True)
            except Exception:
                pass


def download_onebss_report_file_api(
    settings: Settings,
    token: OneBssApiToken,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    target_file: Path | None = None,
    source_values: dict[str, Any] | None = None,
) -> OneBssDownloadedFile:
    download_source = str(
        parameters.get("$download_source")
        or parameters.get("$onebss_download_source")
        or "grid"
    ).strip().lower()
    if download_source in {"excel", "export", "run_v5", "xlsx"}:
        try:
            return download_onebss_export_file_api(
                settings,
                token,
                report,
                parameters,
                target_file=target_file,
                source_values=source_values,
            )
        except OneBssDownloadError as error:
            if "405" not in str(error) and "method not allowed" not in str(error).lower():
                raise
            logger.info("OneBSS direct Excel export is not usable, fallback to grid data API: %s", error)
        return download_onebss_grid_file_api(
            settings,
            token,
            report,
            parameters,
            target_file=target_file,
            source_values=source_values,
        )
    if download_source in {"grid", "run_v7", "json", ""}:
        try:
            return download_onebss_grid_file_api(
                settings,
                token,
                report,
                parameters,
                target_file=target_file,
                source_values=source_values,
            )
        except OneBssDownloadError as error:
            if not should_fallback_to_onebss_excel_export(error):
                raise
            logger.info("OneBSS grid data API is not usable, fallback to direct Excel export: %s", error)
    return download_onebss_export_file_api(
        settings,
        token,
        report,
        parameters,
        target_file=target_file,
        source_values=source_values,
    )


def download_onebss_report_file_api_grid_first(
    settings: Settings,
    token: OneBssApiToken,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    target_file: Path | None = None,
    source_values: dict[str, Any] | None = None,
) -> OneBssDownloadedFile:
    try:
        return download_onebss_grid_file_api(
            settings,
            token,
            report,
            parameters,
            target_file=target_file,
            source_values=source_values,
        )
    except OneBssDownloadError as error:
        if not should_fallback_to_onebss_excel_export(error):
            raise
        logger.info("OneBSS grid data API is not usable, fallback to direct Excel export: %s", error)
    return download_onebss_export_file_api(
        settings,
        token,
        report,
        parameters,
        target_file=target_file,
        source_values=source_values,
    )


def download_onebss_grid_file_api(
    settings: Settings,
    token: OneBssApiToken,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    target_file: Path | None = None,
    source_values: dict[str, Any] | None = None,
) -> OneBssDownloadedFile:
    report_url = normalize_onebss_report_url(report.get("report_url"))
    report_id = onebss_report_id(report, parameters, token)
    export_params = onebss_export_parameters(parameters)
    request_id = datetime.now(LOCAL_TIMEZONE).strftime("%d%m%y%H%M%S") + str(int(time.time() * 1000) % 1000).zfill(3)
    headers = {**onebss_api_auth_headers(token), "apiKey": "x"}
    payload = {"baocao_id": report_id, "params": export_params}
    timeout = onebss_api_timeout(settings, minimum_seconds=ONEBSS_EXPORT_TIMEOUT_SECONDS)
    try:
        with httpx.Client(timeout=timeout) as client:
            response = client.post(
                f"{ONEBSS_API_BASE_URL}/web-report/report/bi/run_v7",
                params={"requestId": request_id},
                headers=headers,
                json=payload,
            )
            response = poll_onebss_json_if_needed(client, response, headers, payload, settings)
    except httpx.TimeoutException as error:
        raise OneBssDownloadError(
            f"OneBSS tra du lieu luoi qua lau va bi het thoi gian cho sau {ONEBSS_EXPORT_TIMEOUT_SECONDS} giay."
        ) from error
    except httpx.HTTPError as error:
        raise OneBssDownloadError(f"Khong goi duoc API du lieu luoi OneBSS: {error}") from error

    data = parse_onebss_json_response(response)
    rows = onebss_grid_rows(data, response)
    if not rows:
        raise OneBssDownloadError("OneBSS run_v7 grid khong co du lieu; fallback sang export Excel truc tiep.")
    suggested_filename = f"onebss_grid_{report_id}.xlsx"
    if target_file is None:
        schedule_like = {
            "report_url": report_url,
            "storage_link": report.get("storage_link") or "",
            "file_name_template": report.get("ten_bao_cao") or report.get("ma_bao_cao") or "",
        }
        target_file = build_target_file_path(
            settings,
            schedule_like,
            suggested_filename=suggested_filename,
            report_title=str(report.get("ten_bao_cao") or ""),
        )
    write_onebss_grid_excel(rows, target_file, sheet_name="DATA")
    return OneBssDownloadedFile(
        file_path=target_file,
        suggested_filename=suggested_filename,
        export_info={
            "ok": True,
            "source": "run_v7_grid",
            "report_id": report_id,
            "title": report.get("ten_bao_cao") or "",
            "params": export_params,
            "row_count": len(rows),
        },
        parameters=export_params,
        source_values=source_values or {},
    )


def download_onebss_export_file_api(
    settings: Settings,
    token: OneBssApiToken,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    target_file: Path | None = None,
    source_values: dict[str, Any] | None = None,
) -> OneBssDownloadedFile:
    report_url = normalize_onebss_report_url(report.get("report_url"))
    report_id = onebss_report_id(report, parameters, token)
    export_params = onebss_export_parameters(parameters)
    request_id = datetime.now(LOCAL_TIMEZONE).strftime("%d%m%y%H%M%S") + str(int(time.time() * 1000) % 1000).zfill(3)
    headers = {**onebss_api_auth_headers(token), "apiKey": "x"}
    timeout = onebss_api_timeout(settings, minimum_seconds=ONEBSS_EXPORT_TIMEOUT_SECONDS)
    try:
        with httpx.Client(timeout=timeout) as client:
            response = client.post(
                f"{ONEBSS_API_BASE_URL}/web-report/report/bi/run_v5",
                params={"requestId": request_id},
                headers=headers,
                json={"baocao_id": report_id, "params": export_params},
            )
            response = poll_onebss_export_if_needed(client, response, headers, settings)
    except httpx.TimeoutException as error:
        raise OneBssDownloadError(
            f"OneBSS tao file qua lau va bi het thoi gian cho sau {ONEBSS_EXPORT_TIMEOUT_SECONDS} giay. "
            "Hay thu rut ngan khoang ngay hoac chay tung phan vung neu bao cao qua lon."
        ) from error
    except httpx.HTTPError as error:
        raise OneBssDownloadError(f"Khong goi duoc API xuat OneBSS: {error}") from error

    ensure_onebss_file_response(response)
    suggested_filename = onebss_response_filename(response, fallback=f"onebss_{report_id}.xlsx")
    if target_file is None:
        schedule_like = {
            "report_url": report_url,
            "storage_link": report.get("storage_link") or "",
            "file_name_template": report.get("ten_bao_cao") or report.get("ma_bao_cao") or "",
        }
        target_file = build_target_file_path(
            settings,
            schedule_like,
            suggested_filename=suggested_filename,
            report_title=str(report.get("ten_bao_cao") or ""),
        )
    target_file.parent.mkdir(parents=True, exist_ok=True)
    target_file.write_bytes(response.content)
    return OneBssDownloadedFile(
        file_path=target_file,
        suggested_filename=suggested_filename,
        export_info={
            "ok": True,
            "report_id": report_id,
            "title": report.get("ten_bao_cao") or "",
            "params": export_params,
        },
        parameters=export_params,
        source_values=source_values or {},
    )


def finish_onebss_report_download(
    settings: Settings,
    helper: OneBssReportDownloader,
    context: Any,
    page: Any,
    report: dict[str, Any],
    parameters: dict[str, Any],
) -> dict[str, Any]:
    started = time.monotonic()
    report_url = normalize_onebss_report_url(report.get("report_url"))
    schedule_like = {
        "report_url": report_url,
        "storage_link": report.get("storage_link") or "",
        "file_name_template": report.get("ten_bao_cao") or report.get("ma_bao_cao") or "",
    }
    parameter_runs, merge_config, each_keys = build_onebss_parameter_runs(parameters)
    downloaded_files: list[OneBssDownloadedFile] = []
    temporary_files: list[Path] = []
    try:
        if len(parameter_runs) == 1:
            downloaded = download_onebss_report_file(
                settings,
                helper,
                page,
                report,
                parameter_runs[0].parameters,
                source_values=parameter_runs[0].source_values,
            )
            downloaded_files.append(downloaded)
            target_file = downloaded.file_path
            storage_result = save_downloaded_file(settings, target_file, str(report.get("storage_link") or ""))
            context.storage_state(path=str(ONEBSS_STATE_PATH))
            ok = bool(storage_result.get("ok", True))
            export_info = downloaded.export_info
            return {
                "ok": ok,
                "status": "success" if ok else str(storage_result.get("status") or "storage_failed"),
                "message": storage_result.get("message") or "Da tai bao cao OneBSS.",
                "file_name": target_file.name,
                "file_path": str(target_file),
                "storage_link": storage_result.get("storage_link") or str(report.get("storage_link") or ""),
                "storage_status": storage_result.get("storage_status") or "",
                "report_id": export_info.get("report_id") or "",
                "report_title": export_info.get("title") or report.get("ten_bao_cao") or "",
                "parameters": export_info.get("params") or parameter_runs[0].parameters,
                "duration_ms": int((time.monotonic() - started) * 1000),
                "finished_at": datetime.now(LOCAL_TIMEZONE).isoformat(timespec="seconds"),
            }

        for index, parameter_run in enumerate(parameter_runs, start=1):
            target_file = build_onebss_temp_file_path(settings, index)
            temporary_files.append(target_file)
            downloaded_files.append(
                download_onebss_report_file(
                    settings,
                    helper,
                    page,
                    report,
                    parameter_run.parameters,
                    target_file=target_file,
                    source_values=parameter_run.source_values,
                )
            )

        context.storage_state(path=str(ONEBSS_STATE_PATH))
        return finalize_onebss_multiple_downloads(
            settings,
            report,
            parameters,
            schedule_like,
            downloaded_files,
            merge_config,
            each_keys,
            started,
        )
    finally:
        for temporary_file in temporary_files:
            try:
                temporary_file.unlink(missing_ok=True)
            except Exception:
                pass


def download_onebss_report_file(
    settings: Settings,
    helper: OneBssReportDownloader,
    page: Any,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    target_file: Path | None = None,
    source_values: dict[str, Any] | None = None,
) -> OneBssDownloadedFile:
    report_url = normalize_onebss_report_url(report.get("report_url"))
    page.wait_for_function(REPORT_COMPONENT_READY_SCRIPT, timeout=90000)
    export_info: dict[str, Any] = {}
    with page.expect_download(timeout=helper.timeout_ms) as download_info:
        export_info = page.evaluate(EXPORT_DIRECT_SCRIPT, parameters)
    download = download_info.value
    if export_info and not export_info.get("ok", True):
        raise OneBssDownloadError(str(export_info.get("message") or "OneBSS khong tao duoc file xuat."))
    if target_file is None:
        schedule_like = {
            "report_url": report_url,
            "storage_link": report.get("storage_link") or "",
            "file_name_template": report.get("ten_bao_cao") or report.get("ma_bao_cao") or "",
        }
        target_file = build_target_file_path(
            settings,
            schedule_like,
            suggested_filename=download.suggested_filename,
            report_title=str(export_info.get("title") or report.get("ten_bao_cao") or ""),
        )
    target_file.parent.mkdir(parents=True, exist_ok=True)
    download.save_as(str(target_file))
    return OneBssDownloadedFile(
        file_path=target_file,
        suggested_filename=str(download.suggested_filename or target_file.name),
        export_info=export_info,
        parameters=export_info.get("params") if isinstance(export_info.get("params"), dict) else parameters,
        source_values=source_values or {},
    )


def onebss_api_credentials(settings: Settings) -> tuple[str, str] | None:
    username = str(getattr(settings, "onebss_username", "") or "").strip()
    password = settings.onebss_password.get_secret_value() if getattr(settings, "onebss_password", None) else ""
    if not username or not password:
        return None
    if "@" not in username:
        username = f"{username}@vnpt.vn"
    return username, password


def onebss_api_device_ids(username: str) -> tuple[str, str]:
    normalized = username.strip().lower()
    if normalized.startswith("quyennt.cto"):
        return "9568FAAF6355", "DEV-1c4cef88"
    digest = hashlib.sha1(normalized.encode("utf-8")).hexdigest()
    return digest[:12].upper(), f"DEV-{digest[12:20]}"


def onebss_api_timeout(settings: Settings, *, minimum_seconds: int = 30) -> httpx.Timeout:
    configured = int(getattr(settings, "onebss_download_timeout_seconds", 180) or 180)
    total = max(minimum_seconds, configured, 30)
    return httpx.Timeout(float(total), connect=20.0)


def onebss_api_base_headers() -> dict[str, str]:
    return {
        "Content-Type": "application/json",
        "Token-id": ONEBSS_API_TOKEN_ID,
        "Mac-address": "WEB",
    }


def onebss_api_auth_headers(token: OneBssApiToken) -> dict[str, str]:
    token_type = token.token_type.capitalize() if token.token_type else "Bearer"
    return {
        **onebss_api_base_headers(),
        "Authorization": f"{token_type} {token.access_token}",
        "SelectedMenuId": "",
        "SelectedPath": "",
        "App-secret": onebss_app_secret(token.username, token.device_id),
    }


def onebss_app_secret(username: str, device_id: str) -> str:
    device_info = {
        "device_id": device_id,
        "device_ip": "Unknown",
        "device_name": "Chrome 124",
        "mac_address": "Unknown",
        "mobile_id": "Unknown",
        "app_id": 3,
        "app_version": "Unknown",
        "os_version": "Windows 10",
    }
    text = json.dumps(device_info, ensure_ascii=False, separators=(",", ":"))
    return base64.b64encode(text.encode("utf-8")).decode("ascii")


def parse_onebss_json_response(response: httpx.Response) -> dict[str, Any]:
    try:
        data = response.json()
    except ValueError:
        return {"message": response.text[:1000], "status_code": response.status_code}
    return data if isinstance(data, dict) else {"data": data}


def onebss_token_payload(data: dict[str, Any]) -> dict[str, Any]:
    payload: Any = data.get("data") if isinstance(data, dict) and "data" in data else data
    if isinstance(payload, str):
        try:
            payload = json.loads(payload)
        except ValueError:
            return {}
    return payload if isinstance(payload, dict) else {}


def onebss_api_error_message(data: dict[str, Any], response: httpx.Response, *, fallback: str) -> str:
    for key in ("message", "error_description", "error"):
        value = data.get(key) if isinstance(data, dict) else ""
        if value:
            return str(value).strip()[:1000]
    nested = data.get("data") if isinstance(data, dict) else None
    if isinstance(nested, dict):
        for key in ("message", "error_description", "error"):
            value = nested.get(key)
            if value:
                return str(value).strip()[:1000]
    if response.status_code >= 400:
        return f"{fallback} HTTP {response.status_code}."
    return fallback


def onebss_error_looks_auth_related(error: Exception) -> bool:
    text = str(error).lower()
    return any(needle in text for needle in ("401", "403", "token", "unauthorized", "forbidden", "het han", "dang nhap"))


def remember_onebss_api_token(
    username: str,
    token_payload: dict[str, Any],
    *,
    mobile_id: str,
    device_id: str,
) -> OneBssApiToken:
    expires_in = token_payload.get("expires_in") or token_payload.get("expires")
    try:
        expires_after = int(expires_in)
    except (TypeError, ValueError):
        expires_after = 20 * 60
    token = OneBssApiToken(
        access_token=str(token_payload.get("access_token") or ""),
        token_type=str(token_payload.get("token_type") or "Bearer"),
        username=username,
        mobile_id=mobile_id,
        device_id=device_id,
        expires_at=time.time() + max(60, expires_after - 60),
    )
    with PENDING_ONEBSS_LOCK:
        ONEBSS_API_TOKENS[username] = token
    return token


def get_valid_onebss_api_token(username: str) -> OneBssApiToken | None:
    with PENDING_ONEBSS_LOCK:
        token = ONEBSS_API_TOKENS.get(username)
        if token and token.expires_at > time.time() + 30:
            return token
        if token:
            ONEBSS_API_TOKENS.pop(username, None)
    return None


def forget_onebss_api_token(username: str) -> None:
    with PENDING_ONEBSS_LOCK:
        ONEBSS_API_TOKENS.pop(username, None)


def onebss_export_parameters(parameters: dict[str, Any]) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for key, value in parameters.items():
        key_text = str(key)
        if key_text.startswith("$") or key_text in ONEBSS_API_META_KEYS:
            continue
        output[key_text] = value
    return output


def onebss_report_id(report: dict[str, Any], parameters: dict[str, Any], token: OneBssApiToken) -> int:
    for candidate in (
        parameters.get("$baocao_id"),
        parameters.get("baocao_id"),
        report.get("baocao_id"),
        report.get("report_id"),
    ):
        if candidate not in {None, ""}:
            try:
                return int(candidate)
            except (TypeError, ValueError):
                break
    path = onebss_report_path(normalize_onebss_report_url(report.get("report_url")))
    if path in KNOWN_ONEBSS_REPORT_IDS:
        return KNOWN_ONEBSS_REPORT_IDS[path]
    return resolve_onebss_report_id_api(report, token)


def resolve_onebss_report_id_api(report: dict[str, Any], token: OneBssApiToken) -> int:
    report_url = normalize_onebss_report_url(report.get("report_url"))
    path = onebss_report_path(report_url)
    if not path:
        raise OneBssDownloadError("Chua co $baocao_id va khong doc duoc path bao cao OneBSS tu link.")
    headers = onebss_api_auth_headers(token)
    try:
        with httpx.Client(timeout=httpx.Timeout(60.0, connect=20.0)) as client:
            response = client.post(f"{ONEBSS_API_BASE_URL}/web-quantri/nguoidung/report_list", headers=headers, json={})
    except httpx.HTTPError as error:
        raise OneBssDownloadError(f"Khong lay duoc danh sach bao cao OneBSS de tim ma bao cao: {error}") from error
    data = parse_onebss_json_response(response)
    if response.status_code >= 400:
        raise OneBssDownloadError(onebss_api_error_message(data, response, fallback="Khong lay duoc danh sach bao cao OneBSS."))
    for item in iter_onebss_dicts(data):
        item_path = str(item.get("path") or item.get("report_path") or item.get("duong_dan") or "").strip()
        if item_path and item_path == path:
            for key in ("report_id", "baocao_id", "id"):
                value = item.get(key)
                if value not in {None, ""}:
                    return int(value)
    raise OneBssDownloadError("Khong tim thay ma bao cao OneBSS tu link. Hay them \"$baocao_id\": 41668 vao cau hinh tham so.")


def onebss_report_path(report_url: str) -> str:
    parsed = urlparse(report_url)
    query_text = parsed.fragment.split("?", 1)[1] if "?" in parsed.fragment else parsed.query
    query = parse_qs(query_text)
    return unquote(str(query.get("path", [""])[0] or "")).strip()


def iter_onebss_dicts(value: Any):
    if isinstance(value, dict):
        yield value
        for child in value.values():
            yield from iter_onebss_dicts(child)
    elif isinstance(value, list):
        for child in value:
            yield from iter_onebss_dicts(child)


def poll_onebss_export_if_needed(
    client: httpx.Client,
    response: httpx.Response,
    headers: dict[str, str],
    settings: Settings,
) -> httpx.Response:
    if response.status_code != 202:
        return response
    location = response.headers.get("location", "").strip() or response.text.strip().strip('"')
    if not location:
        raise OneBssDownloadError("OneBSS dang tao file nhung khong tra link kiem tra file.")
    poll_url = onebss_export_poll_url(location)
    deadline = time.monotonic() + max(ONEBSS_EXPORT_TIMEOUT_SECONDS, int(getattr(settings, "onebss_download_timeout_seconds", 180) or 180))
    last_response = response
    while time.monotonic() < deadline:
        time.sleep(10)
        last_response = client.get(poll_url, headers=headers)
        if last_response.status_code != 202:
            return last_response
    raise OneBssDownloadError(f"OneBSS chua tao xong file sau thoi gian cho. HTTP cuoi: {last_response.status_code}.")


def poll_onebss_json_if_needed(
    client: httpx.Client,
    response: httpx.Response,
    headers: dict[str, str],
    payload: dict[str, Any],
    settings: Settings,
) -> httpx.Response:
    if response.status_code != 202:
        return response
    location = response.headers.get("location", "").strip() or response.text.strip().strip('"')
    if not location:
        raise OneBssDownloadError("OneBSS dang tao du lieu luoi nhung khong tra link kiem tra.")
    poll_url = onebss_export_poll_url(location)
    deadline = time.monotonic() + max(ONEBSS_EXPORT_TIMEOUT_SECONDS, int(getattr(settings, "onebss_download_timeout_seconds", 180) or 180))
    last_response = response
    while time.monotonic() < deadline:
        time.sleep(10)
        last_response = client.post(poll_url, headers=headers, json=payload)
        if last_response.status_code != 202:
            return last_response
    raise OneBssDownloadError(f"OneBSS chua tra du lieu luoi sau thoi gian cho. HTTP cuoi: {last_response.status_code}.")


def onebss_export_poll_url(location: str) -> str:
    if location.startswith("http://") or location.startswith("https://"):
        return location
    if location.startswith("/web-report/"):
        return f"{ONEBSS_API_BASE_URL}{location}"
    if location.startswith("/"):
        return f"{ONEBSS_API_BASE_URL}/web-report{location}"
    return f"{ONEBSS_API_BASE_URL}/web-report/{location}"


def ensure_onebss_file_response(response: httpx.Response) -> None:
    if response.status_code >= 400:
        data = parse_onebss_json_response(response)
        raise OneBssDownloadError(onebss_api_error_message(data, response, fallback="OneBSS khong tra file bao cao."))
    content_type = response.headers.get("content-type", "").lower()
    body = response.content or b""
    if not body:
        raise OneBssDownloadError("OneBSS tra file rong.")
    looks_like_file = (
        b"PK\x03\x04" in body[:8]
        or body[:8].startswith(b"\xd0\xcf\x11\xe0")
        or "attachment" in response.headers.get("content-disposition", "").lower()
        or "spreadsheet" in content_type
        or "octet-stream" in content_type
    )
    if looks_like_file:
        return
    text = response.text[:1000] if response.encoding else body[:1000].decode("utf-8", errors="ignore")
    raise OneBssDownloadError(f"OneBSS chua tra file Excel: {text}")


def onebss_response_filename(response: httpx.Response, *, fallback: str) -> str:
    disposition = response.headers.get("content-disposition", "")
    match = re.search(r"filename\*=UTF-8''([^;]+)", disposition, flags=re.IGNORECASE)
    if match:
        return unquote(match.group(1).strip().strip('"')) or fallback
    match = re.search(r"filename=\"?([^\";]+)\"?", disposition, flags=re.IGNORECASE)
    if match:
        return unquote(match.group(1).strip()) or fallback
    return fallback


def onebss_grid_rows(data: dict[str, Any], response: httpx.Response | None = None) -> list[Any]:
    if response is not None and response.status_code >= 400:
        raise OneBssDownloadError(onebss_api_error_message(data, response, fallback="OneBSS khong tra du lieu luoi."))
    if isinstance(data, dict):
        error_code = str(data.get("error_code") or "")
        if error_code and error_code != "BSS-00000000":
            message = data.get("message") or data.get("message_detail") or data.get("error") or "OneBSS tra loi khi lay du lieu luoi."
            raise OneBssDownloadError(str(message))
        payload = data.get("data")
        if isinstance(payload, list):
            return payload
        if isinstance(payload, dict):
            for key in ("data", "rows", "items", "records"):
                rows = payload.get(key)
                if isinstance(rows, list):
                    return rows
    return []


def write_onebss_grid_excel(rows: list[Any], target_file: Path, *, sheet_name: str = "DATA") -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as error:
        raise OneBssDownloadError("May chu chua cai openpyxl de tao file Excel tu du lieu OneBSS.") from error

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = safe_excel_sheet_title(sheet_name)
    if rows and all(isinstance(row, dict) for row in rows):
        headers = onebss_grid_headers(rows)
        worksheet.append(headers)
        for row in rows:
            worksheet.append([excel_cell_value(row.get(header)) for header in headers])
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        apply_basic_excel_widths(worksheet, len(headers))
    elif rows:
        worksheet.append(["VALUE"])
        for row in rows:
            worksheet.append([excel_cell_value(row)])
        worksheet["A1"].font = Font(bold=True)
        worksheet.freeze_panes = "A2"
        apply_basic_excel_widths(worksheet, 1)
    else:
        worksheet.append(["NO_DATA"])
        worksheet["A1"].font = Font(bold=True)
        apply_basic_excel_widths(worksheet, 1)
    target_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target_file)


def onebss_grid_headers(rows: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        for key in row.keys():
            key_text = str(key)
            if key_text not in seen:
                seen.add(key_text)
                headers.append(key_text)
    return headers or ["VALUE"]


def excel_cell_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def should_fallback_to_onebss_excel_export(error: Exception) -> bool:
    text = str(error).lower()
    fallback_needles = ("chưa hỗ trợ", "chua ho tro", "header", "run_v7", "404", "not found", "khong co du lieu", "no data")
    return any(needle in text for needle in fallback_needles)


def build_onebss_temp_file_path(settings: Settings, index: int) -> Path:
    base_dir = Path(str(getattr(settings, "data_mining_download_dir", "data/data_mining_downloads") or "data/data_mining_downloads"))
    base_dir.mkdir(parents=True, exist_ok=True)
    return (base_dir / f".onebss_part_{uuid.uuid4().hex}_{index}.xlsx").resolve()


def merged_excel_suggested_filename(suggested_filename: str) -> str:
    suffix = Path(str(suggested_filename or "")).suffix.lower()
    if suffix == ".xlsx":
        return suggested_filename
    stem = Path(str(suggested_filename or "onebss_report")).stem or "onebss_report"
    return f"{stem}.xlsx"


def split_archive_suggested_filename(suggested_filename: str) -> str:
    stem = Path(str(suggested_filename or "onebss_report")).stem or "onebss_report"
    return f"{stem}_parts.zip"


def onebss_merge_mode(merge_config: dict[str, Any]) -> str:
    return str(merge_config.get("mode") or merge_config.get("$mode") or "").strip().lower()


def should_merge_onebss_excel_parts(merge_config: dict[str, Any]) -> bool:
    return onebss_merge_mode(merge_config) in {"append", "merge", "merged", "single", "sheet", "sheets"}


def archive_onebss_downloaded_files(
    downloaded_files: list[OneBssDownloadedFile],
    target_file: Path,
    each_keys: list[str],
) -> None:
    if not downloaded_files:
        raise OneBssDownloadError("Khong co file OneBSS nao de dong goi.")
    target_file.parent.mkdir(parents=True, exist_ok=True)
    used_names: set[str] = set()
    with zipfile.ZipFile(target_file, "w", compression=zipfile.ZIP_DEFLATED, allowZip64=True) as archive:
        for index, downloaded in enumerate(downloaded_files, start=1):
            arcname = onebss_archive_member_name(downloaded, each_keys, index)
            while arcname.lower() in used_names:
                arcname = f"{index:02d}_{arcname}"
            used_names.add(arcname.lower())
            archive.write(downloaded.file_path, arcname)


def onebss_archive_member_name(downloaded: OneBssDownloadedFile, each_keys: list[str], index: int) -> str:
    suffix = downloaded.file_path.suffix or Path(str(downloaded.suggested_filename or "")).suffix or ".xlsx"
    source_parts: list[str] = []
    for key in each_keys:
        value = downloaded.source_values.get(key)
        if value not in {None, ""}:
            source_parts.append(f"{key}_{value}")
    source_label = "_".join(source_parts) or str(index).zfill(2)
    stem = Path(str(downloaded.suggested_filename or downloaded.file_path.name or f"onebss_{index}")).stem
    return f"{safe_zip_name(stem)}_{safe_zip_name(source_label)}{suffix}"


def safe_zip_name(value: Any) -> str:
    text = str(value or "").strip() or "onebss"
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", text)
    text = re.sub(r"\s+", "_", text).strip("._")
    return text[:120] or "onebss"


def finalize_onebss_multiple_downloads(
    settings: Settings,
    report: dict[str, Any],
    parameters: dict[str, Any],
    schedule_like: dict[str, Any],
    downloaded_files: list[OneBssDownloadedFile],
    merge_config: dict[str, Any],
    each_keys: list[str],
    started: float,
) -> dict[str, Any]:
    first_export = downloaded_files[0].export_info if downloaded_files else {}
    if should_merge_onebss_excel_parts(merge_config):
        target_file = build_target_file_path(
            settings,
            schedule_like,
            suggested_filename=merged_excel_suggested_filename(downloaded_files[0].suggested_filename if downloaded_files else ".xlsx"),
            report_title=str(first_export.get("title") or report.get("ten_bao_cao") or ""),
        )
        merge_onebss_excel_files(downloaded_files, target_file, merge_config, each_keys)
        result_kind = "merged"
        result_count_key = "merged_file_count"
        result_message = f"Da tai va gop {len(downloaded_files)} file OneBSS thanh 1 file."
    else:
        target_file = build_target_file_path(
            settings,
            schedule_like,
            suggested_filename=split_archive_suggested_filename(downloaded_files[0].suggested_filename if downloaded_files else ".xlsx"),
            report_title=str(first_export.get("title") or report.get("ten_bao_cao") or ""),
        )
        archive_onebss_downloaded_files(downloaded_files, target_file, each_keys)
        result_kind = "split_archive"
        result_count_key = "split_file_count"
        result_message = f"Da tai {len(downloaded_files)} file OneBSS rieng va dong goi thanh 1 file zip."

    storage_result = save_downloaded_file(settings, target_file, str(report.get("storage_link") or ""))
    ok = bool(storage_result.get("ok", True))
    status = "success" if ok else str(storage_result.get("status") or "storage_failed")
    storage_message = storage_result.get("message") or "Da tai bao cao OneBSS."
    message = f"{result_message} {storage_message}" if ok else storage_message
    return {
        "ok": ok,
        "status": status,
        "message": message,
        "file_name": target_file.name,
        "file_path": str(target_file),
        "storage_link": storage_result.get("storage_link") or str(report.get("storage_link") or ""),
        "storage_status": storage_result.get("storage_status") or "",
        "report_id": first_export.get("report_id") or "",
        "report_title": first_export.get("title") or report.get("ten_bao_cao") or "",
        "parameters": parameters,
        "run_parameters": [downloaded.parameters for downloaded in downloaded_files],
        result_count_key: len(downloaded_files),
        "output_mode": result_kind,
        "duration_ms": int((time.monotonic() - started) * 1000),
        "finished_at": datetime.now(LOCAL_TIMEZONE).isoformat(timespec="seconds"),
    }


def build_onebss_parameter_runs(parameters: dict[str, Any]) -> tuple[list[OneBssParameterRun], dict[str, Any], list[str]]:
    merge_config = parameters.get("$merge_excel") if isinstance(parameters.get("$merge_excel"), dict) else {}
    base_parameters: dict[str, Any] = {}
    each_items: list[tuple[str, list[Any]]] = []
    for key, value in parameters.items():
        key_text = str(key)
        if key_text.startswith("$") or key_text in ONEBSS_API_META_KEYS:
            continue
        if isinstance(value, dict) and "$each" in value:
            each_values = value.get("$each")
            if not isinstance(each_values, list) or not each_values:
                raise OneBssDownloadError(f"Bien {key} dung $each nhung danh sach gia tri dang rong hoac khong hop le.")
            each_items.append((key, each_values))
            continue
        base_parameters[key] = value

    if not each_items:
        return [OneBssParameterRun(parameters=base_parameters, source_values={})], merge_config, []

    runs: list[OneBssParameterRun] = []
    for values in product(*(item[1] for item in each_items)):
        run_parameters = {**base_parameters}
        source_values: dict[str, Any] = {}
        for (key, _), value in zip(each_items, values):
            run_parameters[key] = value
            source_values[key] = value
        runs.append(OneBssParameterRun(parameters=run_parameters, source_values=source_values))
    return runs, merge_config, [item[0] for item in each_items]


def merge_onebss_excel_files(
    downloaded_files: list[OneBssDownloadedFile],
    target_file: Path,
    merge_config: dict[str, Any],
    each_keys: list[str],
) -> None:
    if not downloaded_files:
        raise OneBssDownloadError("Khong co file OneBSS nao de gop.")
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
    except ImportError as error:
        raise OneBssDownloadError("May chu chua cai openpyxl de gop file Excel OneBSS.") from error

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = safe_excel_sheet_title(str(merge_config.get("sheet") or "DATA"))
    source_column_name = merge_source_column_name(merge_config, each_keys)
    header_signature: tuple[str, ...] | None = None
    header_output_row: int | None = None
    max_columns = 1

    for file_index, downloaded in enumerate(downloaded_files):
        workbook = load_workbook(downloaded.file_path, data_only=False, read_only=True)
        try:
            sheet_name = str(merge_config.get("source_sheet") or "").strip()
            worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
            rows = worksheet_rows(worksheet)
        finally:
            workbook.close()
        if not rows:
            continue
        header_index = detect_header_row_index(rows, merge_config)
        if file_index == 0:
            selected_rows = list(enumerate(rows))
            header_signature = row_signature(rows[header_index]) if 0 <= header_index < len(rows) else None
        else:
            start_index = header_index + 1 if header_signature and row_signature(rows[header_index]) == header_signature else int(merge_config.get("skip_rows", 1) or 0)
            selected_rows = list(enumerate(rows[start_index:], start=start_index))

        source_label = merge_source_label(downloaded.source_values, source_column_name)
        for source_row_index, row in selected_rows:
            if file_index > 0 and not row_has_content(row):
                continue
            output_row = trim_trailing_empty_cells(row)
            if source_column_name:
                if source_row_index == header_index:
                    output_row.append(source_column_name)
                elif source_row_index > header_index and row_has_content(row):
                    output_row.append(source_label)
                else:
                    output_row.append("")
            output_sheet.append(output_row)
            max_columns = max(max_columns, len(output_row))
            if file_index == 0 and source_row_index == header_index:
                header_output_row = output_sheet.max_row

    if output_sheet.max_row == 1 and not row_has_content([cell.value for cell in output_sheet[1]]):
        raise OneBssDownloadError("Cac file OneBSS tai ve khong co du lieu de gop.")
    if header_output_row:
        for cell in output_sheet[header_output_row]:
            cell.font = Font(bold=True)
        output_sheet.freeze_panes = f"A{header_output_row + 1}"
    apply_basic_excel_widths(output_sheet, max_columns)
    target_file.parent.mkdir(parents=True, exist_ok=True)
    output_workbook.save(target_file)


def worksheet_rows(worksheet: Any) -> list[list[Any]]:
    rows = [[cell.value for cell in row] for row in worksheet.iter_rows()]
    while rows and not row_has_content(rows[-1]):
        rows.pop()
    return rows


def detect_header_row_index(rows: list[list[Any]], merge_config: dict[str, Any]) -> int:
    configured = merge_config.get("header_row")
    if configured not in {None, ""}:
        try:
            return max(0, min(len(rows) - 1, int(configured) - 1))
        except (TypeError, ValueError):
            pass
    min_cells = int(merge_config.get("min_header_cells") or 2)
    for index, row in enumerate(rows):
        if sum(1 for value in row if value not in {None, ""}) >= min_cells:
            return index
    return 0


def row_has_content(row: list[Any]) -> bool:
    return any(value not in {None, ""} for value in row)


def trim_trailing_empty_cells(row: list[Any]) -> list[Any]:
    trimmed = list(row)
    while trimmed and trimmed[-1] in {None, ""}:
        trimmed.pop()
    return trimmed


def row_signature(row: list[Any]) -> tuple[str, ...]:
    return tuple(str(value or "").strip().lower() for value in trim_trailing_empty_cells(row))


def merge_source_column_name(merge_config: dict[str, Any], each_keys: list[str]) -> str:
    value = merge_config.get("source_column")
    if value is False:
        return ""
    if isinstance(value, str) and value.strip():
        return value.strip()
    if value is None and len(each_keys) == 1:
        return each_keys[0]
    if value is True and each_keys:
        return "_".join(each_keys)
    return ""


def merge_source_label(source_values: dict[str, Any], source_column_name: str) -> Any:
    if not source_values or not source_column_name:
        return ""
    if source_column_name in source_values:
        return source_values[source_column_name]
    if len(source_values) == 1:
        return next(iter(source_values.values()))
    return ", ".join(f"{key}={value}" for key, value in source_values.items())


def safe_excel_sheet_title(value: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "_", value).strip() or "DATA"
    return title[:31]


def apply_basic_excel_widths(worksheet: Any, max_columns: int) -> None:
    for column_index in range(1, min(max_columns, 80) + 1):
        letter = worksheet.cell(row=1, column=column_index).column_letter
        width = 10
        for row_index in range(1, min(worksheet.max_row, 200) + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            if value not in {None, ""}:
                width = max(width, min(len(str(value)) + 2, 45))
        worksheet.column_dimensions[letter].width = width


def keep_onebss_session(
    playwright: Any,
    browser: Any,
    context: Any,
    page: Any,
    report: dict[str, Any],
    parameters: dict[str, Any],
    created_by: str,
) -> PendingOneBssSession:
    session = PendingOneBssSession(
        session_id=uuid.uuid4().hex,
        playwright=playwright,
        browser=browser,
        context=context,
        page=page,
        report=report,
        parameters=parameters,
        created_by=created_by,
        created_at=time.time(),
    )
    with PENDING_ONEBSS_LOCK:
        PENDING_ONEBSS_SESSIONS[session.session_id] = session
    return session


def keep_onebss_api_session(
    secret_code: str,
    report: dict[str, Any],
    parameters: dict[str, Any],
    username: str,
    mobile_id: str,
    device_id: str,
    created_by: str,
) -> PendingOneBssApiSession:
    session = PendingOneBssApiSession(
        session_id=f"api-{uuid.uuid4().hex}",
        secret_code=secret_code,
        report=report,
        parameters=parameters,
        username=username,
        mobile_id=mobile_id,
        device_id=device_id,
        created_by=created_by,
        created_at=time.time(),
    )
    with PENDING_ONEBSS_LOCK:
        PENDING_ONEBSS_API_SESSIONS[session.session_id] = session
    return session


def get_onebss_session(session_id: str) -> PendingOneBssSession | None:
    with PENDING_ONEBSS_LOCK:
        return PENDING_ONEBSS_SESSIONS.get(session_id)


def get_onebss_api_session(session_id: str) -> PendingOneBssApiSession | None:
    with PENDING_ONEBSS_LOCK:
        return PENDING_ONEBSS_API_SESSIONS.get(session_id)


def pop_onebss_session(session_id: str) -> PendingOneBssSession | None:
    with PENDING_ONEBSS_LOCK:
        return PENDING_ONEBSS_SESSIONS.pop(session_id, None)


def pop_onebss_api_session(session_id: str) -> PendingOneBssApiSession | None:
    with PENDING_ONEBSS_LOCK:
        return PENDING_ONEBSS_API_SESSIONS.pop(session_id, None)


def cleanup_expired_onebss_sessions() -> None:
    now = time.time()
    expired: list[PendingOneBssSession] = []
    with PENDING_ONEBSS_LOCK:
        for key, session in list(PENDING_ONEBSS_SESSIONS.items()):
            if now - session.created_at > PENDING_SESSION_TTL_SECONDS:
                expired.append(PENDING_ONEBSS_SESSIONS.pop(key))
        for key, session in list(PENDING_ONEBSS_API_SESSIONS.items()):
            if now - session.created_at > PENDING_SESSION_TTL_SECONDS:
                PENDING_ONEBSS_API_SESSIONS.pop(key, None)
        for key, token in list(ONEBSS_API_TOKENS.items()):
            if token.expires_at <= now:
                ONEBSS_API_TOKENS.pop(key, None)
    for session in expired:
        close_browser_stack(session.browser, session.context, session.playwright)


def close_browser_stack(browser: Any, context: Any, playwright: Any | None = None) -> None:
    try:
        context.close()
    except Exception:
        pass
    try:
        browser.close()
    except Exception:
        pass
    if playwright is not None:
        try:
            playwright.stop()
        except Exception:
            pass


def wait_for_onebss_network_quiet(page: Any, *, timeout_ms: int = 8000, pause_ms: int = 500) -> bool:
    try:
        page.wait_for_load_state("networkidle", timeout=timeout_ms)
        quiet = True
    except Exception:
        quiet = False
        try:
            logger.info("OneBSS network stayed active after %sms at %s", timeout_ms, page.url)
        except Exception:
            pass
    if pause_ms > 0:
        try:
            page.wait_for_timeout(pause_ms)
        except Exception:
            time.sleep(pause_ms / 1000)
    return quiet


def goto_onebss_page(page: Any, url: str, *, step: str, timeout_ms: int = 30000) -> None:
    logger.info("OneBSS step=%s goto=%s", step, url)
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
    except Exception as error:
        logger.exception("OneBSS step=%s cannot open page", step)
        raise OneBssDownloadError(
            f"May chu khong mo duoc trang OneBSS trong {int(timeout_ms / 1000)} giay o buoc {step}: {str(error)[:500]}"
        ) from error


def wait_for_onebss_login_form(page: Any, timeout_ms: int = 30000) -> bool:
    deadline = time.monotonic() + (timeout_ms / 1000)
    while time.monotonic() < deadline:
        state = onebss_login_page_state(page)
        if int(state.get("passwordInputCount") or 0) > 0 and int(state.get("inputCount") or 0) >= 2:
            return True
        try:
            page.wait_for_timeout(500)
        except Exception:
            time.sleep(0.5)
    return False


def onebss_login_page_state(page: Any) -> dict[str, Any]:
    try:
        state = page.evaluate(
            """
            () => {
              const inputs = Array.from(document.querySelectorAll('input'));
              const buttons = Array.from(document.querySelectorAll('button'));
              const inputInfo = inputs.slice(0, 8).map((input) => ({
                type: input.getAttribute('type') || '',
                name: input.getAttribute('name') || '',
                placeholder: input.getAttribute('placeholder') || '',
                autocomplete: input.getAttribute('autocomplete') || '',
                visible: !!(input.offsetWidth || input.offsetHeight || input.getClientRects().length),
              }));
              const buttonTexts = buttons.slice(0, 8).map((button) => (button.innerText || button.textContent || '').replace(/\\s+/g, ' ').trim()).filter(Boolean);
              return {
                url: location.href,
                title: document.title || '',
                readyState: document.readyState || '',
                htmlLength: document.documentElement ? document.documentElement.outerHTML.length : 0,
                bodyLength: document.body ? document.body.innerText.length : 0,
                bodyText: document.body ? document.body.innerText.replace(/\\s+/g, ' ').trim().slice(0, 300) : '',
                inputCount: inputs.length,
                passwordInputCount: inputs.filter((input) => String(input.getAttribute('type') || '').toLowerCase() === 'password').length,
                buttonCount: buttons.length,
                buttonTexts,
                inputInfo,
              };
            }
            """
        )
        return state if isinstance(state, dict) else {}
    except Exception as error:
        return {"diagnosticError": str(error)[:200], "url": str(getattr(page, "url", "") or "")}


def onebss_login_diagnostic_text(page: Any) -> str:
    state = onebss_login_page_state(page)
    parts = [
        f"Trang hien tai: {state.get('url') or getattr(page, 'url', '')}",
        f"title={state.get('title') or ''}",
        f"ready={state.get('readyState') or ''}",
        f"html={state.get('htmlLength') or 0}",
        f"body={state.get('bodyLength') or 0}",
        f"inputs={state.get('inputCount') or 0}",
        f"password_inputs={state.get('passwordInputCount') or 0}",
        f"buttons={state.get('buttonCount') or 0}",
    ]
    button_texts = state.get("buttonTexts")
    if isinstance(button_texts, list) and button_texts:
        parts.append(f"button_texts={', '.join(str(item)[:40] for item in button_texts[:4])}")
    body_text = str(state.get("bodyText") or "")
    if body_text:
        parts.append(f"Noi dung: {body_text[:300]}")
    if state.get("diagnosticError"):
        parts.append(f"diagnostic_error={state.get('diagnosticError')}")
    return ". ".join(parts)


def wait_for_onebss_auth_transition(page: Any, helper: OneBssReportDownloader, timeout_ms: int = 12000) -> None:
    deadline = time.monotonic() + (timeout_ms / 1000)
    device_needles = ["ĐĂNG KÝ THIẾT BỊ", "DANG KY THIET BI", "đăng ký thiết bị", "dang ky thiet bi"]
    while time.monotonic() < deadline:
        if (
            page_contains(page, OTP_TEXT_NEEDLES)
            or page_contains(page, OTP_INVALID_TEXT_NEEDLES)
            or page_contains(page, LOGIN_ERROR_TEXT_NEEDLES)
            or page_contains(page, device_needles)
            or page_contains(page, OTP_REQUEST_TEXT_NEEDLES)
        ):
            return
        try:
            if not helper._is_login_page(page):
                return
        except Exception:
            return
        try:
            page.wait_for_timeout(500)
        except Exception:
            time.sleep(0.5)


def click_onebss_button(page: Any, helper: OneBssReportDownloader, texts: list[str]) -> bool:
    if helper._click_button_text(page, texts):
        return True
    try:
        clicked = page.evaluate(
            """
            (texts) => {
              const normalize = (value) => String(value || '')
                .normalize('NFD')
                .replace(/[\\u0300-\\u036f]/g, '')
                .replace(/\\s+/g, ' ')
                .trim()
                .toLowerCase();
              const wanted = texts.map(normalize).filter(Boolean);
              const buttons = Array.from(document.querySelectorAll('button'));
              const visible = buttons.filter((button) => {
                const rect = button.getBoundingClientRect();
                const style = window.getComputedStyle(button);
                return rect.width > 0 && rect.height > 0 && style.visibility !== 'hidden' && style.display !== 'none' && !button.disabled;
              });
              const matched = visible.find((button) => {
                const text = normalize(button.innerText || button.textContent || button.value);
                return wanted.some((item) => text === item || text.includes(item));
              });
              const target = matched || (visible.length === 1 ? visible[0] : null);
              if (!target) return false;
              target.click();
              return true;
            }
            """,
            texts,
        )
        if clicked:
            return True
    except Exception:
        pass
    try:
        page.keyboard.press("Enter")
        return True
    except Exception:
        return False


def handle_onebss_otp_request(
    settings: Settings | Any,
    page: Any,
    helper: OneBssReportDownloader,
    playwright: Any,
    browser: Any,
    context: Any,
    report: dict[str, Any],
    parameters: dict[str, Any] | str,
    created_by: str | None = None,
) -> dict[str, Any] | None:
    legacy_manual_otp = False
    if not isinstance(settings, Settings):
        old_page = settings
        old_helper = page
        old_playwright = helper
        old_browser = playwright
        old_context = browser
        old_report = context
        old_parameters = report
        old_created_by = parameters
        settings = get_settings()
        page = old_page
        helper = old_helper
        playwright = old_playwright
        browser = old_browser
        context = old_context
        report = old_report if isinstance(old_report, dict) else {}
        parameters = old_parameters if isinstance(old_parameters, dict) else {}
        created_by = str(old_created_by or "system")
        legacy_manual_otp = True
    else:
        parameters = parameters if isinstance(parameters, dict) else {}
        created_by = str(created_by or "system")

    if page_contains(page, LOGIN_ERROR_TEXT_NEEDLES):
        return None
    url = str(getattr(page, "url", "") or "").lower()
    url_indicates_otp_flow = "auth/login" in url and ("username=" in url or "deviceid=" in url)
    if not url_indicates_otp_flow and not page_contains(page, OTP_REQUEST_TEXT_NEEDLES):
        return None

    clicked_any = False
    for _ in range(3):
        if page_contains(page, OTP_TEXT_NEEDLES):
            break
        clicked = click_onebss_button(page, helper, OTP_REQUEST_BUTTON_TEXTS)
        if not clicked:
            break
        clicked_any = True
        wait_for_onebss_network_quiet(page, timeout_ms=5000, pause_ms=1000)

    if not clicked_any and not page_contains(page, OTP_TEXT_NEEDLES) and not url_indicates_otp_flow:
        return None

    pending = keep_onebss_session(playwright, browser, context, page, report, parameters, created_by)
    if legacy_manual_otp:
        return onebss_manual_otp_response(
            pending.session_id,
            parameters,
            "OneBSS da gui yeu cau OTP. Hay nhap ma OTP khi dien thoai nhan duoc.",
            status="otp_required",
        )
    return start_onebss_otp_mobile_gateway_request(
        settings,
        pending.session_id,
        parameters,
        otp_service_code=onebss_report_otp_service_code(report),
        fallback_message="OneBSS da gui yeu cau OTP. Hay nhap ma OTP khi dien thoai nhan duoc.",
    )


def handle_onebss_device_registration(page: Any, helper: OneBssReportDownloader, parameters: dict[str, Any]) -> dict[str, Any] | None:
    if not page_contains(page, ["ĐĂNG KÝ THIẾT BỊ", "DANG KY THIET BI", "đăng ký thiết bị", "dang ky thiet bi"]):
        return None
    try:
        checkbox = page.locator("input[type='checkbox']").first
        if checkbox.count():
            checkbox.check(force=True, timeout=10000)
        clicked = helper._click_button_text(
            page,
            ["Gửi yêu cầu đăng ký", "Gui yeu cau dang ky", "Gửi yêu cầu", "Gui yeu cau", "Xác nhận", "Xac nhan"],
        )
        if not clicked:
            return {
                "ok": False,
                "status": "device_registration_required",
                "message": "OneBSS yeu cau dang ky thiet bi moi nhung khong tim thay nut gui yeu cau.",
                "parameters": parameters,
            }
        wait_for_onebss_network_quiet(page, timeout_ms=8000, pause_ms=1000)
    except Exception as error:
        logger.exception("Cannot send OneBSS device registration request")
        return {
            "ok": False,
            "status": "device_registration_failed",
            "message": f"Khong gui duoc yeu cau dang ky thiet bi OneBSS: {error}",
            "parameters": parameters,
        }
    if page_contains(page, ["ĐĂNG KÝ THIẾT BỊ", "DANG KY THIET BI", "chờ phê duyệt", "cho phe duyet"]):
        return {
            "ok": False,
            "status": "device_registration_pending",
            "message": "Da gui yeu cau dang ky thiet bi OneBSS va dang cho phe duyet.",
            "parameters": parameters,
        }
    return None


def onebss_login_failed_message(page: Any) -> str:
    return f"Dang nhap OneBSS chua thanh cong. {onebss_login_diagnostic_text(page)}"


def page_contains(page: Any, needles: list[str]) -> bool:
    try:
        text = page.locator("body").inner_text(timeout=5000).lower()
    except Exception:
        return False
    return any(needle.lower() in text for needle in needles)
