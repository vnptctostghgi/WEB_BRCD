from __future__ import annotations

import base64
import csv
import logging
import re
import time
import uuid
from datetime import datetime, timedelta, timezone
from io import StringIO
from pathlib import Path
from typing import Any, Callable
from urllib.parse import parse_qs, quote, urlparse
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import openpyxl

from app.application.database_service import DatabaseService
from app.application.google_drive_service import (
    GoogleDriveConfigurationError,
    google_drive_oauth_credentials,
    load_service_account_info,
)
from app.application.onebss_data_mining_service import resolve_dynamic_parameters, safe_filename_part
from app.application.zalo_auto_message_service import (
    install_playwright_chromium,
    playwright_needs_browser_install,
    public_base_url,
)
from app.application.zalo_bot import ZaloBotClient
from app.data_access.internal_api_client import InternalApiClient
from app.settings import Settings


logger = logging.getLogger(__name__)
try:
    LOCAL_TIMEZONE = ZoneInfo("Asia/Ho_Chi_Minh")
except ZoneInfoNotFoundError:
    LOCAL_TIMEZONE = timezone(timedelta(hours=7), name="Asia/Ho_Chi_Minh")

TASK_REPORT_AUTO_SOURCE_TYPES = {"onebss", "sql", "ftp"}
TASK_REPORT_AUTO_CREATED_BY = "task_report_auto"
TASK_REPORT_AUTO_CAPTURE_ROUTE = "/api/task-report-auto/captures"
DEFAULT_STEP_WAIT_SECONDS = 5
DEFAULT_SOURCE_TIMEOUT_SECONDS = 60 * 60
SPREADSHEETS_SCOPE = "https://www.googleapis.com/auth/spreadsheets"
DRIVE_SCOPE = "https://www.googleapis.com/auth/drive"


class TaskReportAutoError(RuntimeError):
    pass


class TaskReportAutoStepError(TaskReportAutoError):
    def __init__(self, step: str, message: str, details: dict[str, Any] | None = None) -> None:
        super().__init__(message)
        self.step = step
        self.details = details or {}


def normalize_task_report_auto_source_type(value: Any) -> str:
    source_type = str(value or "onebss").strip().lower()
    if source_type not in TASK_REPORT_AUTO_SOURCE_TYPES:
        raise ValueError("Nguon du lieu phai la onebss, sql hoac ftp.")
    return source_type


def normalize_task_report_auto_payload(payload: Any, task_id: str) -> dict[str, Any]:
    data = payload.model_dump() if hasattr(payload, "model_dump") else dict(payload or {})
    source_type = normalize_task_report_auto_source_type(data.get("source_type"))
    source_code = str(data.get("source_code") or "").strip().upper()
    if not source_code:
        raise ValueError("Chua chon ma bao cao nguon.")
    spreadsheet_id = extract_spreadsheet_id(data.get("spreadsheet_id") or data.get("spreadsheet_url"))
    if not spreadsheet_id:
        raise ValueError("Chua nhap link/ID Google Sheet dich.")
    if not str(data.get("sheet_name") or "").strip():
        raise ValueError("Chua nhap ten sheet/tab dich.")
    public_url = str(data.get("public_url") or "").strip()
    chat_id = str(data.get("chat_id") or "").strip()
    if chat_id and not public_url:
        raise ValueError("Can nhap link public web de chup anh truoc khi gui Zalo.")
    retry_limit = int(data.get("retry_limit") or 0)
    return {
        "task_id": task_id,
        "name": str(data.get("name") or "").strip() or source_code,
        "source_type": source_type,
        "source_code": source_code,
        "source_config": data.get("source_config") if isinstance(data.get("source_config"), dict) else {},
        "schedule_type": str(data.get("schedule_type") or "Daily").strip() or "Daily",
        "time_slots": normalize_time_slots(data.get("time_slots")),
        "run_time": normalize_time_text(data.get("run_time") or "07:00"),
        "weekday": str(data.get("weekday") or "").strip(),
        "month_day": min(max(int(data.get("month_day") or 1), 1), 31),
        "spreadsheet_id": spreadsheet_id,
        "spreadsheet_url": str(data.get("spreadsheet_url") or "").strip(),
        "sheet_name": str(data.get("sheet_name") or "DATA").strip() or "DATA",
        "public_url": public_url,
        "public_wait_selector": str(data.get("public_wait_selector") or "").strip(),
        "target_type": str(data.get("target_type") or "group").strip() or "group",
        "chat_id": chat_id,
        "chat_name": str(data.get("chat_name") or "").strip(),
        "caption": str(data.get("caption") or "").strip(),
        "retry_limit": min(max(retry_limit, 0), 5),
        "is_active": bool(data.get("is_active", True)),
    }


def normalize_time_slots(value: Any) -> list[str]:
    raw_items: list[Any]
    if isinstance(value, str):
        raw_items = re.split(r"[,;\s]+", value)
    elif isinstance(value, list):
        raw_items = value
    else:
        raw_items = []
    slots = []
    for item in raw_items:
        text = normalize_time_text(item)
        if text and text not in slots:
            slots.append(text)
    return slots


def normalize_time_text(value: Any) -> str:
    text = str(value or "").strip()
    match = re.match(r"^(\d{1,2}):(\d{2})", text)
    if not match:
        return ""
    hour = min(max(int(match.group(1)), 0), 23)
    minute = min(max(int(match.group(2)), 0), 59)
    return f"{hour:02d}:{minute:02d}"


def extract_spreadsheet_id(value: Any) -> str:
    raw_value = str(value or "").strip()
    if not raw_value:
        return ""
    parsed = urlparse(raw_value)
    if parsed.scheme in {"http", "https"}:
        parts = [part for part in parsed.path.split("/") if part]
        if "d" in parts:
            index = parts.index("d")
            if len(parts) > index + 1:
                return parts[index + 1]
        return parse_qs(parsed.query).get("id", [""])[0].strip()
    if "/" not in raw_value and len(raw_value) >= 20:
        return raw_value
    return ""


def quote_sheet_title(title: str) -> str:
    return "'" + str(title or "DATA").replace("'", "''") + "'"


def task_report_auto_capture_public_url(settings: Settings, capture: dict[str, Any] | None) -> str:
    if not capture:
        return ""
    base_url = public_base_url(settings)
    capture_id = str(capture.get("capture_id") or "").strip()
    token = str(capture.get("public_token") or "").strip()
    if not base_url or not capture_id or not token:
        return ""
    return f"{base_url}{TASK_REPORT_AUTO_CAPTURE_ROUTE}/{quote(capture_id)}?token={quote(token)}"


def task_report_auto_caption(task: dict[str, Any], run: dict[str, Any]) -> str:
    caption = str(task.get("caption") or "").strip()
    if caption:
        return caption
    title = str(task.get("name") or "Task report auto").strip()
    source = str(task.get("source_code") or "").strip()
    return f"{title}\n{source}" if source else title


def task_report_auto_created_by(run_id: str) -> str:
    return f"{TASK_REPORT_AUTO_CREATED_BY}:{run_id}"


def is_task_report_auto_created_by(value: Any) -> bool:
    return str(value or "").startswith(TASK_REPORT_AUTO_CREATED_BY)


class TaskReportAutoRunner:
    def __init__(
        self,
        repository: Any,
        settings: Settings,
        *,
        step_wait_seconds: int = DEFAULT_STEP_WAIT_SECONDS,
        poll_seconds: int = 5,
    ) -> None:
        self.repository = repository
        self.settings = settings
        self.step_wait_seconds = max(0, int(step_wait_seconds))
        self.poll_seconds = max(1, int(poll_seconds))

    def run_next_once(self) -> dict[str, Any] | None:
        run = self.repository.claim_next_task_report_auto_run()
        if not run:
            return None
        return self.run_claimed(run)

    def run_claimed(self, run: dict[str, Any]) -> dict[str, Any]:
        run_id = str(run.get("run_id") or "")
        task = self.repository.get_task_report_auto_task(str(run.get("task_id") or ""))
        if not task:
            result = {"ok": False, "status": "failed", "message": "Khong tim thay cau hinh Task report auto."}
            self._finish_run(run, result)
            return result

        started = time.monotonic()
        result: dict[str, Any] = {"ok": False, "status": "failed", "message": ""}
        try:
            source = self._run_step(run_id, task, "mine", lambda: self._run_source(task, run))
            self._wait_between_steps()
            sheet = self._run_step(run_id, task, "sheet", lambda: self._upload_source_to_sheet(task, source))
            self._wait_between_steps()
            capture = {"ok": True, "status": "separate_schedule", "message": "Buoc chup anh duoc cau hinh tai Lich gui Zalo."}
            zalo = {"ok": True, "status": "separate_schedule", "message": "Buoc gui tin duoc cau hinh tai Lich gui Zalo."}
            result = {
                "ok": True,
                "status": "success",
                "message": "Task report auto da chay thanh cong.",
                "source": source,
                "sheet": sheet,
                "capture": capture,
                "zalo": zalo,
            }
            self.repository.update_task_report_auto_run(
                run_id,
                {
                    "status": "success",
                    "current_step": "done",
                    "message": result["message"],
                    "result": result,
                    "source_run_id": str(source.get("run_id") or ""),
                    "capture_id": str(capture.get("capture", {}).get("capture_id") or ""),
                    "capture_url": str(capture.get("capture_url") or ""),
                    "finished_at": self.repository._now() if hasattr(self.repository, "_now") else datetime.now(timezone.utc).isoformat(timespec="seconds"),
                    "duration_ms": int((time.monotonic() - started) * 1000),
                },
            )
            self.repository.mark_task_report_auto_task_run(str(task.get("task_id") or ""), str(run.get("run_key") or ""), True, result)
            return result
        except TaskReportAutoStepError as error:
            result = {"ok": False, "status": "failed", "message": str(error), "failed_step": error.step, "details": error.details}
            self._finish_run(run, result, started=started, current_step=error.step)
            self.repository.mark_task_report_auto_task_run(str(task.get("task_id") or ""), str(run.get("run_key") or ""), False, result)
            return result
        except Exception as error:
            logger.exception("Task report auto run failed: %s", run_id)
            result = {"ok": False, "status": "failed", "message": str(error)[:1000] or "Task report auto phat sinh loi."}
            self._finish_run(run, result, started=started)
            self.repository.mark_task_report_auto_task_run(str(task.get("task_id") or ""), str(run.get("run_key") or ""), False, result)
            return result

    def _finish_run(
        self,
        run: dict[str, Any],
        result: dict[str, Any],
        *,
        started: float | None = None,
        current_step: str = "failed",
    ) -> None:
        run_id = str(run.get("run_id") or "")
        self.repository.update_task_report_auto_run(
            run_id,
            {
                "status": "failed",
                "current_step": current_step,
                "message": str(result.get("message") or "")[:1000],
                "result": result,
                "finished_at": self.repository._now() if hasattr(self.repository, "_now") else datetime.now(timezone.utc).isoformat(timespec="seconds"),
                "duration_ms": int((time.monotonic() - started) * 1000) if started else 0,
            },
        )

    def _run_step(self, run_id: str, task: dict[str, Any], step: str, callback: Callable[[], dict[str, Any]]) -> dict[str, Any]:
        retry_limit = int(task.get("retry_limit") or 0)
        last_error: Exception | None = None
        for attempt in range(1, retry_limit + 2):
            self.repository.update_task_report_auto_run(
                run_id,
                {
                    "status": "running",
                    "current_step": step,
                    "message": f"Dang chay buoc {step}, lan {attempt}/{retry_limit + 1}.",
                },
            )
            try:
                result = callback()
                if result.get("ok") is False:
                    raise TaskReportAutoStepError(step, str(result.get("message") or f"Buoc {step} chua thanh cong."), result)
                self.repository.append_task_report_auto_run_step(
                    run_id,
                    step,
                    {"ok": True, "attempt": attempt, "message": result.get("message") or "", "result": result},
                )
                return result
            except TaskReportAutoStepError as error:
                last_error = error
                details = error.details
                message = str(error)
            except Exception as error:
                last_error = error
                details = {"error_type": error.__class__.__name__}
                message = str(error)[:1000]
            self.repository.append_task_report_auto_run_step(
                run_id,
                step,
                {"ok": False, "attempt": attempt, "message": message, "details": details},
            )
            if attempt <= retry_limit:
                self._sleep(self.step_wait_seconds)
        if isinstance(last_error, TaskReportAutoStepError):
            raise last_error
        raise TaskReportAutoStepError(step, str(last_error or f"Buoc {step} that bai."))

    def _wait_between_steps(self) -> None:
        self._sleep(self.step_wait_seconds)

    @staticmethod
    def _sleep(seconds: int | float) -> None:
        if seconds > 0:
            time.sleep(seconds)

    def _run_source(self, task: dict[str, Any], run: dict[str, Any]) -> dict[str, Any]:
        source_type = str(task.get("source_type") or "").lower()
        if source_type == "onebss":
            return self._run_onebss_source(task, run)
        if source_type == "ftp":
            return self._run_ftp_source(task, run)
        if source_type == "sql":
            return self._run_sql_source(task, run)
        raise TaskReportAutoStepError("mine", "Nguon du lieu khong hop le.")

    def _run_onebss_source(self, task: dict[str, Any], run: dict[str, Any]) -> dict[str, Any]:
        report_code = str(task.get("source_code") or "").strip().upper()
        report = self.repository.get_onebss_report_by_code(report_code)
        if not report:
            raise TaskReportAutoStepError("mine", "Khong tim thay cau hinh bao cao OneBSS.")
        config = task.get("source_config") if isinstance(task.get("source_config"), dict) else {}
        parameters = report.get("parameters") if isinstance(report.get("parameters"), dict) else {}
        if isinstance(config.get("parameters"), dict):
            parameters = {**parameters, **config["parameters"]}
        run_id = safe_filename_part(f"TRA_{run.get('run_id') or ''}_{uuid.uuid4().hex[:8]}", "task_report_auto_onebss")
        self.repository.save_onebss_report_run(
            {
                "run_id": run_id,
                "ma_bao_cao": report.get("ma_bao_cao") or report_code,
                "ten_bao_cao": report.get("ten_bao_cao") or report_code,
                "status": "queued",
                "message": "Task report auto da dua OneBSS vao hang doi may tram.",
                "parameters": parameters,
                "started_at": datetime.now(LOCAL_TIMEZONE).isoformat(timespec="seconds"),
                "finished_at": "",
                "created_by": task_report_auto_created_by(str(run.get("run_id") or "")),
            }
        )
        finished = self._wait_source_run(
            lambda: self.repository.get_onebss_report_run(run_id),
            success_statuses={"success"},
            failure_statuses={"failed", "cancelled", "storage_failed", "google_drive_not_configured", "google_drive_upload_failed"},
            timeout_seconds=int(config.get("timeout_seconds") or DEFAULT_SOURCE_TIMEOUT_SECONDS),
        )
        return self._source_file_result(finished, "Da lay xong du lieu OneBSS.")

    def _run_ftp_source(self, task: dict[str, Any], run: dict[str, Any]) -> dict[str, Any]:
        report_code = str(task.get("source_code") or "").strip().upper()
        report = self.repository.get_ftp_report_by_code(report_code)
        if not report:
            raise TaskReportAutoStepError("mine", "Khong tim thay cau hinh bao cao FTP.")
        if not report.get("is_active"):
            raise TaskReportAutoStepError("mine", "Bao cao FTP dang tam tat.")
        config = task.get("source_config") if isinstance(task.get("source_config"), dict) else {}
        folder_path = str(config.get("folder_path") or report.get("folder_path") or "").strip()
        file_name_template = str(config.get("file_name_template") or report.get("file_name_template") or "").strip()
        variables = config.get("variables") if isinstance(config.get("variables"), dict) else {}
        file_name_template = apply_template_variables(file_name_template, variables)
        run_id = safe_filename_part(f"TRA_{run.get('run_id') or ''}_{uuid.uuid4().hex[:8]}", "task_report_auto_ftp")
        self.repository.save_ftp_report_run(
            {
                "run_id": run_id,
                "ma_bao_cao": report.get("ma_bao_cao") or report_code,
                "ten_bao_cao": report.get("ten_bao_cao") or report_code,
                "status": "queued",
                "message": "Task report auto da dua FTP vao hang doi may tram.",
                "folder_path": folder_path,
                "file_name_template": file_name_template,
                "started_at": datetime.now(LOCAL_TIMEZONE).isoformat(timespec="seconds"),
                "finished_at": "",
                "created_by": task_report_auto_created_by(str(run.get("run_id") or "")),
            }
        )
        finished = self._wait_source_run(
            lambda: self.repository.get_ftp_report_run(run_id),
            success_statuses={"success"},
            failure_statuses={"failed", "cancelled"},
            timeout_seconds=int(config.get("timeout_seconds") or DEFAULT_SOURCE_TIMEOUT_SECONDS),
        )
        return self._source_file_result(finished, "Da lay xong du lieu FTP.")

    def _run_sql_source(self, task: dict[str, Any], run: dict[str, Any]) -> dict[str, Any]:
        config = task.get("source_config") if isinstance(task.get("source_config"), dict) else {}
        filters = config.get("filters") if isinstance(config.get("filters"), dict) else {}
        if not filters and isinstance(config.get("parameters"), dict):
            filters = config["parameters"]
        filters = resolve_dynamic_parameters(filters)
        service = DatabaseService(InternalApiClient.from_repository(self.settings, self.repository), self.repository)
        if not bool(config.get("force_worker")):
            result = service.export_dynamic_report(
                ma_bao_cao=str(task.get("source_code") or "").strip().upper(),
                filters=filters,
                search=str(config.get("search") or ""),
                search_columns=config.get("search_columns") if isinstance(config.get("search_columns"), list) else [],
                report_id=config.get("report_id"),
                report_name=str(config.get("report_name") or ""),
                collect_rows=True,
            )
            if result.get("ok") is not False:
                rows = result.get("rows") if isinstance(result.get("rows"), list) else []
                columns = result.get("columns") if isinstance(result.get("columns"), list) else []
                return {
                    "ok": True,
                    "message": result.get("message") or "Da lay xong du lieu SQL.",
                    "source_type": "sql",
                    "rows": rows,
                    "columns": columns,
                    "run_id": "",
                    "result": result,
                }
            if not bool(config.get("allow_worker", True)):
                raise TaskReportAutoStepError("mine", result.get("message") or "Khong lay duoc du lieu SQL.", result)

        queued = self._queue_sql_worker_source(task, run, config, filters)
        return queued

    def _queue_sql_worker_source(self, task: dict[str, Any], run: dict[str, Any], config: dict[str, Any], filters: dict[str, Any]) -> dict[str, Any]:
        service = DatabaseService(InternalApiClient.from_repository(self.settings, self.repository), self.repository)
        page_size = int(config.get("page_size") or getattr(self.settings, "dynamic_report_export_page_size", 20000) or 20000)
        max_rows = int(config.get("max_rows") or getattr(self.settings, "dynamic_report_export_max_rows", 1000000) or 1000000)
        prepared = service.prepare_dynamic_report_query(
            ma_bao_cao=str(task.get("source_code") or "").strip().upper(),
            filters=filters,
            page=1,
            page_size=page_size,
            search=str(config.get("search") or ""),
            search_columns=config.get("search_columns") if isinstance(config.get("search_columns"), list) else [],
            report_id=config.get("report_id"),
            report_name=str(config.get("report_name") or ""),
        )
        if prepared.get("ok") is False:
            raise TaskReportAutoStepError("mine", prepared.get("message") or "Khong chuan bi duoc SQL worker.", prepared)
        run_id = safe_filename_part(f"TRA_{run.get('run_id') or ''}_{uuid.uuid4().hex[:8]}", "task_report_auto_sql")
        payload = {
            "ma_bao_cao": prepared.get("ma_bao_cao") or task.get("source_code") or "",
            "filters": filters,
            "page": 1,
            "page_size": page_size,
            "collect_all_pages": True,
            "max_rows": max_rows,
            "search": str(config.get("search") or ""),
            "search_columns": config.get("search_columns") if isinstance(config.get("search_columns"), list) else [],
            "report_id": config.get("report_id") if config.get("report_id") not in (None, "") else None,
            "report_name": str(config.get("report_name") or prepared.get("ten_bao_cao") or ""),
        }
        self.repository.save_report_run(
            {
                "run_id": run_id,
                "job_id": run_id,
                "run_type": "load",
                "status": "queued_worker",
                "message": "Task report auto da dua SQL vao hang doi may tram.",
                "created_by": task_report_auto_created_by(str(run.get("run_id") or "")),
                "report_code": prepared.get("ma_bao_cao") or task.get("source_code") or "",
                "report_name": prepared.get("ten_bao_cao") or "",
                "payload": payload,
                "details": {"task_report_auto": True},
            }
        )
        finished = self._wait_source_run(
            lambda: self.repository.get_report_run(run_id),
            success_statuses={"complete", "success"},
            failure_statuses={"failed", "cancelled"},
            timeout_seconds=int(config.get("timeout_seconds") or DEFAULT_SOURCE_TIMEOUT_SECONDS),
        )
        result = finished.get("result") if isinstance(finished.get("result"), dict) else {}
        rows = result.get("rows") if isinstance(result.get("rows"), list) else []
        columns = result.get("columns") if isinstance(result.get("columns"), list) else []
        return {
            "ok": True,
            "message": result.get("message") or finished.get("message") or "Da lay xong du lieu SQL qua may tram.",
            "source_type": "sql",
            "rows": rows,
            "columns": columns,
            "run_id": run_id,
            "result": result,
        }

    def _wait_source_run(
        self,
        getter: Callable[[], dict[str, Any] | None],
        *,
        success_statuses: set[str],
        failure_statuses: set[str],
        timeout_seconds: int,
    ) -> dict[str, Any]:
        deadline = time.monotonic() + max(1, int(timeout_seconds))
        last_run: dict[str, Any] | None = None
        while time.monotonic() < deadline:
            run = getter()
            if run:
                last_run = run
                status = str(run.get("status") or "").strip().lower()
                if status in success_statuses:
                    return run
                if status in failure_statuses:
                    raise TaskReportAutoStepError("mine", str(run.get("message") or "Buoc dao du lieu that bai."), run)
            self._sleep(self.poll_seconds)
        message = "Qua thoi gian cho ket qua dao du lieu."
        if last_run and last_run.get("message"):
            message = f"{message} Trang thai cuoi: {last_run.get('status')} - {last_run.get('message')}"
        raise TaskReportAutoStepError("mine", message, last_run or {})

    @staticmethod
    def _source_file_result(run: dict[str, Any], message: str) -> dict[str, Any]:
        file_path = str(run.get("file_path") or "").strip()
        if not file_path:
            raise TaskReportAutoStepError("mine", "Da chay nguon nhung chua co file ket qua.", run)
        if not Path(file_path).exists():
            raise TaskReportAutoStepError("mine", f"File ket qua khong ton tai tren web: {file_path}", run)
        return {
            "ok": True,
            "message": message,
            "source_type": "file",
            "run_id": run.get("run_id") or "",
            "file_path": file_path,
            "file_name": run.get("file_name") or Path(file_path).name,
            "run": run,
        }

    def _upload_source_to_sheet(self, task: dict[str, Any], source: dict[str, Any]) -> dict[str, Any]:
        if source.get("source_type") == "sql":
            values = values_from_sql_result(source.get("columns"), source.get("rows"))
        else:
            values = values_from_file(Path(str(source.get("file_path") or "")), task.get("source_config") if isinstance(task.get("source_config"), dict) else {})
        if not values:
            raise TaskReportAutoStepError("sheet", "Nguon du lieu khong co dong nao de nap Google Sheet.", source)
        spreadsheet_id = str(task.get("spreadsheet_id") or "").strip()
        sheet_name = str(task.get("sheet_name") or "DATA").strip() or "DATA"
        result = write_values_to_google_sheet(self.settings, self.repository, spreadsheet_id, sheet_name, values)
        return {
            "ok": True,
            "message": f"Da nap {len(values) - 1 if len(values) > 1 else len(values)} dong vao Google Sheet.",
            "spreadsheet_id": spreadsheet_id,
            "sheet_name": sheet_name,
            "updated_rows": len(values),
            **result,
        }

    def _capture_public_web(self, task: dict[str, Any], run_id: str) -> dict[str, Any]:
        public_url = str(task.get("public_url") or "").strip()
        selector = str(task.get("public_wait_selector") or "").strip()
        image_bytes = capture_public_web_screenshot_bytes(public_url, selector=selector)
        image_base64 = base64.b64encode(image_bytes).decode("ascii")
        capture = self.repository.save_task_report_auto_capture(
            str(task.get("task_id") or ""),
            run_id,
            image_base64,
            "image/png",
            public_url,
            TASK_REPORT_AUTO_CREATED_BY,
        )
        capture_url = task_report_auto_capture_public_url(self.settings, capture)
        if not capture_url:
            raise TaskReportAutoStepError("capture", "Da chup anh nhung khong tao duoc URL cong khai.")
        return {"ok": True, "message": "Da chup anh public web.", "capture": capture, "capture_url": capture_url, "page_url": public_url}

    def _send_zalo(self, task: dict[str, Any], run: dict[str, Any], capture: dict[str, Any]) -> dict[str, Any]:
        chat_id = str(task.get("chat_id") or "").strip()
        photo_url = str(capture.get("capture_url") or "").strip()
        if not chat_id:
            raise TaskReportAutoStepError("zalo", "Chua cau hinh chat_id Zalo.")
        if urlparse(photo_url).scheme != "https":
            raise TaskReportAutoStepError("zalo", "URL anh gui Zalo phai dung HTTPS.")
        caption = task_report_auto_caption(task, run)
        sent = ZaloBotClient(self.settings).send_photo(chat_id, photo_url, caption)
        try:
            self.repository.add_audit_log(
                TASK_REPORT_AUTO_CREATED_BY,
                "task_report_auto_zalo_sent" if sent else "task_report_auto_zalo_failed",
                f"{task.get('task_id')}: {task.get('name')} -> {chat_id}",
            )
        except Exception:
            logger.exception("Cannot write Task report auto Zalo audit log")
        if not sent:
            raise TaskReportAutoStepError("zalo", "Khong gui duoc anh qua Zalo.", {"chat_id": chat_id, "photo_url": photo_url})
        return {"ok": True, "message": "Da gui anh qua Zalo.", "chat_id": chat_id, "photo_url": photo_url, "caption": caption}


def apply_template_variables(template: str, variables: dict[str, Any]) -> str:
    result = str(template or "")
    for key, value in variables.items():
        safe_key = re.escape(str(key))
        result = re.sub(r"\{\{\s*" + safe_key + r"\s*\}\}", str(value), result)
        result = re.sub(r"\{\s*" + safe_key + r"\s*\}", str(value), result)
    return result


def values_from_sql_result(columns_value: Any, rows_value: Any) -> list[list[Any]]:
    rows = rows_value if isinstance(rows_value, list) else []
    columns = [str(column) for column in (columns_value if isinstance(columns_value, list) else [])]
    if not columns and rows and isinstance(rows[0], dict):
        columns = list(rows[0].keys())
    if not columns:
        columns = ["Ket qua"]
    values: list[list[Any]] = [columns]
    for row in rows:
        if isinstance(row, dict):
            values.append([sheet_cell_value(row.get(column)) for column in columns])
        else:
            values.append([sheet_cell_value(row)])
    return values


def values_from_file(path: Path, source_config: dict[str, Any] | None = None) -> list[list[Any]]:
    if not path.exists():
        raise TaskReportAutoStepError("sheet", f"Khong tim thay file nguon: {path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return values_from_csv(path)
    if suffix not in {".xlsx", ".xlsm"}:
        raise TaskReportAutoStepError("sheet", f"Chua ho tro nap dinh dang file {suffix or 'khong ro'} len Google Sheet.")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_name = str((source_config or {}).get("workbook_sheet_name") or "").strip()
        worksheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.worksheets[0]
        return [[sheet_cell_value(value) for value in row] for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def values_from_csv(path: Path) -> list[list[Any]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [[sheet_cell_value(cell) for cell in row] for row in csv.reader(StringIO(text), dialect)]


def sheet_cell_value(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return str(value)


def google_sheets_credentials(settings: Settings, repository: Any) -> Any:
    try:
        oauth_credentials = google_drive_oauth_credentials(settings, repository)
    except GoogleDriveConfigurationError as error:
        logger.warning("Google Drive OAuth is unavailable; falling back to the configured service account: %s", error)
        oauth_credentials = None
    if oauth_credentials:
        credentials, _ = oauth_credentials
        try:
            from google.auth.transport.requests import Request
        except ImportError as error:
            raise GoogleDriveConfigurationError("May chu chua cai thu vien Google OAuth.") from error
        credentials.refresh(Request())
        return credentials
    info = load_service_account_info(settings)
    try:
        from google.oauth2 import service_account
    except ImportError as error:
        raise GoogleDriveConfigurationError("May chu chua cai thu vien Google OAuth.") from error
    credentials = service_account.Credentials.from_service_account_info(info, scopes=[DRIVE_SCOPE, SPREADSHEETS_SCOPE])
    impersonated_user = str(getattr(settings, "google_drive_impersonated_user", "") or "").strip()
    return credentials.with_subject(impersonated_user) if impersonated_user else credentials


def write_values_to_google_sheet(
    settings: Settings,
    repository: Any,
    spreadsheet_id: str,
    sheet_name: str,
    values: list[list[Any]],
) -> dict[str, Any]:
    try:
        from googleapiclient.discovery import build
    except ImportError as error:
        raise GoogleDriveConfigurationError("May chu chua cai thu vien Google Sheets API.") from error
    credentials = google_sheets_credentials(settings, repository)
    sheets = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    ensure_google_sheet_tab(sheets, spreadsheet_id, sheet_name)
    target = quote_sheet_title(sheet_name)
    sheets.spreadsheets().values().clear(spreadsheetId=spreadsheet_id, range=target, body={}).execute()
    updated_cells = 0
    chunk_size = 5000
    for index in range(0, len(values), chunk_size):
        chunk = values[index:index + chunk_size]
        response = sheets.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"{target}!A{index + 1}",
            valueInputOption="RAW",
            body={"values": chunk},
        ).execute()
        updated_cells += int(response.get("updatedCells") or 0)
    return {"updated_cells": updated_cells}


def ensure_google_sheet_tab(sheets: Any, spreadsheet_id: str, sheet_name: str) -> None:
    metadata = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties(title)").execute()
    titles = {
        str((sheet.get("properties") or {}).get("title") or "")
        for sheet in metadata.get("sheets", [])
    }
    if sheet_name in titles:
        return
    sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": sheet_name}}}]},
    ).execute()


def capture_public_web_screenshot_bytes(public_url: str, *, selector: str = "") -> bytes:
    parsed = urlparse(str(public_url or "").strip())
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise RuntimeError("Link public web chua hop le.")
    try:
        from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
        from playwright.sync_api import sync_playwright
    except ImportError as error:
        raise RuntimeError("May chu chua cai Playwright de chup anh.") from error

    def run_capture(install_retry: bool = False) -> bytes:
        try:
            with sync_playwright() as playwright:
                browser = playwright.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
                try:
                    context = browser.new_context(viewport={"width": 1920, "height": 1080}, device_scale_factor=2, locale="vi-VN")
                    page = context.new_page()
                    page.goto(public_url, wait_until="networkidle", timeout=90000)
                    try:
                        if selector:
                            page.wait_for_selector(selector, state="visible", timeout=60000)
                            page.wait_for_timeout(1200)
                            image = page.locator(selector).first.screenshot(type="png")
                        else:
                            page.wait_for_selector("body", state="visible", timeout=30000)
                            page.wait_for_function("() => document.body && document.body.innerText.trim().length > 0", timeout=30000)
                            page.wait_for_timeout(1200)
                            clip_script = """
                                () => {
                                  const cells = [...document.querySelectorAll('table.waffle td')]
                                    .filter((cell) => (cell.innerText || cell.textContent || '').trim());
                                  if (!cells.length) return null;
                                  const boxes = cells.map((cell) => cell.getBoundingClientRect())
                                    .filter((box) => box.width > 0 && box.height > 0);
                                  if (!boxes.length) return null;
                                  const padding = 0;
                                  const left = Math.max(0, Math.min(...boxes.map((box) => box.left + window.scrollX)) - padding);
                                  const top = Math.max(0, Math.min(...boxes.map((box) => box.top + window.scrollY)) - padding);
                                  const right = Math.max(...boxes.map((box) => box.right + window.scrollX)) + padding;
                                  const bottom = Math.max(...boxes.map((box) => box.bottom + window.scrollY)) + padding;
                                  return {x: left, y: top, width: Math.max(1, right - left), height: Math.max(1, bottom - top)};
                                }
                                """
                            content_clip = page.evaluate(clip_script)
                            sheet_frame = next((frame for frame in page.frames if "/pubhtml/sheet" in frame.url), None)
                            if sheet_frame:
                                frame_clip = sheet_frame.evaluate(clip_script)
                                frame_box = sheet_frame.frame_element().bounding_box()
                                if frame_clip and frame_box:
                                    content_clip = {
                                        "x": frame_box["x"] + frame_clip["x"],
                                        "y": frame_box["y"] + frame_clip["y"],
                                        "width": frame_clip["width"],
                                        "height": frame_clip["height"],
                                    }
                            image = page.screenshot(type="png", clip=content_clip) if content_clip else page.screenshot(type="png", full_page=True)
                    except PlaywrightTimeoutError:
                        raise RuntimeError("Public web chua load xong de chup anh.") from None
                    if len(image) < 100:
                        raise RuntimeError("Anh chup public web qua nho, co the trang bi loi hoac trang trong.")
                    return image
                finally:
                    browser.close()
        except Exception as error:
            if not install_retry and playwright_needs_browser_install(error):
                install_playwright_chromium()
                return run_capture(install_retry=True)
            raise

    return run_capture()
