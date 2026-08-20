from __future__ import annotations

import argparse
import calendar
import csv
import base64
import fnmatch
import ftplib
import json
import mimetypes
import multiprocessing as mp
import os
import queue
import re
import sys
import threading
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlparse

import httpx

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.application.onebss_report_service import OneBssProgressCancelled, run_onebss_report_request
from app.settings import get_settings


class OneBssTaskCancelled(OneBssProgressCancelled):
    pass


class FtpTaskCancelled(Exception):
    pass


class SqlTaskCancelled(Exception):
    pass


TRANSIENT_HTTP_STATUS_CODES = {408, 425, 429, 500, 502, 503, 504}
WORKER_VERSION = "2026.08.20-sql-cancel-v42"
LOCAL_INTERNAL_API_URL = "http://127.0.0.1:8000/api/du-lieu-web"
LOCAL_DRIVE_UPLOAD_API_URL = "http://127.0.0.1:8000/api/du-lieu-web"
PUBLIC_DRIVE_UPLOAD_API_URL = "https://api.vnptcto.com/api/du-lieu-web"
TASK_KIND_ONEBSS = "onebss"
TASK_KIND_SQL = "sql"
TASK_KIND_FTP = "ftp"
TASK_KIND_LABELS = {
    TASK_KIND_ONEBSS: "OneBSS",
    TASK_KIND_SQL: "SQL",
    TASK_KIND_FTP: "FTP",
}


def env_int(names: str | list[str], default: int, *, minimum: int = 1, maximum: int = 32) -> int:
    env_names = [names] if isinstance(names, str) else names
    raw_value = ""
    for name in env_names:
        raw_value = str(os.getenv(name, "") or "").strip()
        if raw_value:
            break
    try:
        value = int(float(raw_value)) if raw_value else int(default)
    except (TypeError, ValueError):
        value = int(default)
    return max(int(minimum), min(int(value), int(maximum)))


def worker_concurrency_limits() -> dict[str, int]:
    total = env_int(
        ["VNPTCTO_WORKER_MAX_CONCURRENT_TASKS", "ONEBSS_WORKER_MAX_CONCURRENT_TASKS"],
        4,
        minimum=1,
        maximum=8,
    )
    return {
        "total": total,
        TASK_KIND_ONEBSS: min(env_int("ONEBSS_WORKER_MAX_ONEBSS_TASKS", 2, minimum=1, maximum=4), total),
        TASK_KIND_SQL: min(env_int(["SQL_WORKER_MAX_CONCURRENT_TASKS", "SQL_WORKER_MAX_TASKS"], 2, minimum=1, maximum=4), total),
        TASK_KIND_FTP: min(env_int(["FTP_WORKER_MAX_CONCURRENT_TASKS", "FTP_WORKER_MAX_TASKS"], 2, minimum=1, maximum=6), total),
    }


def worker_task_run_id(kind: str, task: dict[str, Any]) -> str:
    if kind == TASK_KIND_SQL:
        return str(task.get("run_id") or task.get("job_id") or "").strip()
    return str(task.get("run_id") or "").strip()


def worker_task_report_code(kind: str, task: dict[str, Any]) -> str:
    if kind == TASK_KIND_ONEBSS and isinstance(task.get("report"), dict):
        return str((task.get("report") or {}).get("ma_bao_cao") or "").strip()
    if kind == TASK_KIND_SQL:
        query = task.get("query") if isinstance(task.get("query"), dict) else {}
        return str(task.get("report_code") or query.get("ma_bao_cao") or "").strip()
    return str(task.get("ma_bao_cao") or "").strip()


def safe_worker_file_part(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(value or "").strip())
    return text.strip(".-")[:80] or "worker"


def onebss_worker_slot_state_path(worker_id: str, slot: int) -> Path:
    root = str(os.getenv("VNPTCTO_WORKSTATION_ROOT", "") or "").strip()
    base_dir = Path(root) / "data" / "onebss-sessions" if root else ROOT / "data" / "onebss-worker-sessions"
    base_dir.mkdir(parents=True, exist_ok=True)
    slot_number = max(1, int(slot or 1))
    return base_dir / f"{safe_worker_file_part(worker_id)}-slot-{slot_number}.json"


def task_run_workspace(base_dir: Path, kind: str, run_id: str) -> Path:
    task_id = safe_worker_file_part(run_id) if str(run_id or "").strip() else f"{kind}-{time.time_ns()}"
    target = base_dir / "runs" / task_id
    target.mkdir(parents=True, exist_ok=True)
    return target


class WorkerConcurrencyTracker:
    def __init__(self) -> None:
        self._lock = threading.Lock()
        self._active: dict[str, dict[str, Any]] = {}

    def counts(self) -> dict[str, int]:
        with self._lock:
            return self._counts_locked()

    def active_details(self) -> dict[str, Any]:
        with self._lock:
            counts = self._counts_locked()
            active_tasks = [
                {
                    "run_id": run_id,
                    "task_type": info.get("kind") or "",
                    "report": info.get("report") or "",
                    "slot": int(info.get("slot") or 0),
                    "elapsed_seconds": int(time.monotonic() - float(info.get("started_monotonic") or time.monotonic())),
                }
                for run_id, info in self._active.items()
            ]
        return {
            "concurrency_limits": worker_concurrency_limits(),
            "active_counts": counts,
            "active_tasks": active_tasks,
        }

    def can_start(self, kind: str) -> bool:
        limits = worker_concurrency_limits()
        with self._lock:
            counts = self._counts_locked()
            return counts["total"] < limits["total"] and counts.get(kind, 0) < limits.get(kind, 1)

    def try_start(self, kind: str, run_id: str, report_code: str = "") -> bool:
        return self.start_slot(kind, run_id, report_code) > 0

    def start_slot(self, kind: str, run_id: str, report_code: str = "") -> int:
        task_id = str(run_id or "").strip()
        if not task_id:
            task_id = f"{kind}-{time.time_ns()}"
        limits = worker_concurrency_limits()
        with self._lock:
            counts = self._counts_locked()
            if task_id in self._active:
                return 0
            if counts["total"] >= limits["total"] or counts.get(kind, 0) >= limits.get(kind, 1):
                return 0
            used_slots = {
                int(info.get("slot") or 0)
                for info in self._active.values()
                if str(info.get("kind") or "") == kind
            }
            slot = next((candidate for candidate in range(1, limits.get(kind, 1) + 1) if candidate not in used_slots), 1)
            self._active[task_id] = {
                "kind": kind,
                "report": report_code,
                "slot": slot,
                "started_monotonic": time.monotonic(),
            }
            return slot

    def finish(self, run_id: str) -> None:
        task_id = str(run_id or "").strip()
        with self._lock:
            self._active.pop(task_id, None)

    def _counts_locked(self) -> dict[str, int]:
        counts = {
            "total": len(self._active),
            TASK_KIND_ONEBSS: 0,
            TASK_KIND_SQL: 0,
            TASK_KIND_FTP: 0,
        }
        for info in self._active.values():
            kind = str(info.get("kind") or "")
            if kind in counts:
                counts[kind] += 1
        return counts


class WorkerTaskDispatcher:
    def __init__(self, base_url: str, headers: dict[str, str], worker_id: str, poll_seconds: float) -> None:
        self.base_url = base_url.rstrip("/")
        self.headers = dict(headers)
        self.worker_id = worker_id
        self.poll_seconds = poll_seconds
        self.tracker = WorkerConcurrencyTracker()
        self._threads_lock = threading.Lock()
        self._threads: list[threading.Thread] = []

    def can_start(self, kind: str) -> bool:
        return self.tracker.can_start(kind)

    def active_counts(self) -> dict[str, int]:
        return self.tracker.counts()

    def active_details(self) -> dict[str, Any]:
        return self.tracker.active_details()

    def has_active_tasks(self) -> bool:
        return self.active_counts()["total"] > 0

    def has_available_slot(self) -> bool:
        return any(self.can_start(kind) for kind in (TASK_KIND_ONEBSS, TASK_KIND_SQL, TASK_KIND_FTP))

    def start_task(self, kind: str, task: dict[str, Any]) -> bool:
        run_id = worker_task_run_id(kind, task) or f"{kind}-{time.time_ns()}"
        report_code = worker_task_report_code(kind, task)
        slot = self.tracker.start_slot(kind, run_id, report_code)
        if not slot:
            return False
        task_payload = dict(task)
        if kind == TASK_KIND_ONEBSS:
            task_payload["_worker_slot"] = slot
            task_payload["_worker_state_path"] = str(onebss_worker_slot_state_path(self.worker_id, slot))
        thread = threading.Thread(
            target=self._run_task,
            args=(kind, task_payload, run_id),
            name=f"vnptcto-{kind}-{run_id[:24]}",
            daemon=True,
        )
        with self._threads_lock:
            self._threads.append(thread)
        thread.start()
        return True

    def wait_until_idle(self) -> None:
        while self.has_active_tasks():
            with self._threads_lock:
                threads = list(self._threads)
            for thread in threads:
                thread.join(timeout=0.2)
            self.prune_threads()

    def prune_threads(self) -> None:
        with self._threads_lock:
            self._threads = [thread for thread in self._threads if thread.is_alive()]

    def _run_task(self, kind: str, task: dict[str, Any], run_id: str) -> None:
        label = TASK_KIND_LABELS.get(kind, kind.upper())
        try:
            with httpx.Client(base_url=self.base_url, headers=self.headers, timeout=httpx.Timeout(60.0, connect=20.0)) as task_client:
                if kind == TASK_KIND_ONEBSS:
                    process_task(task_client, task, self.worker_id, self.poll_seconds)
                elif kind == TASK_KIND_SQL:
                    process_sql_task(task_client, task, self.worker_id)
                elif kind == TASK_KIND_FTP:
                    process_ftp_task(task_client, task, self.worker_id)
                else:
                    raise RuntimeError(f"Loai task khong ho tro: {kind}")
        except Exception as error:
            print(
                f"Luong {label} {run_id} loi ngoai y muon: {describe_request_error(error)}",
                file=sys.stderr,
                flush=True,
            )
            traceback.print_exc(file=sys.stderr)
        finally:
            self.tracker.finish(run_id)


def worker_process_details() -> dict[str, Any]:
    return {
        "pid": os.getpid(),
        "python": sys.version.split()[0],
        "worker_version": WORKER_VERSION,
        "worker_process": f"Worker PID {os.getpid()} dang chay nen.",
    }


def response_is_cancelled(data: dict[str, Any]) -> bool:
    if bool(data.get("cancelled")) or str(data.get("status") or "").lower() in {"cancel_requested", "cancelled"}:
        return True
    run = data.get("run")
    return isinstance(run, dict) and str(run.get("status") or "").lower() in {"cancel_requested", "cancelled"}


def describe_request_error(error: Exception) -> str:
    if isinstance(error, httpx.HTTPStatusError):
        status_code = error.response.status_code if error.response is not None else "?"
        return f"HTTP {status_code}"
    return str(error)[:300] or error.__class__.__name__


def is_transient_request_error(error: Exception) -> bool:
    if isinstance(error, httpx.HTTPStatusError):
        return bool(error.response is not None and error.response.status_code in TRANSIENT_HTTP_STATUS_CODES)
    return isinstance(error, httpx.RequestError)


def transient_retry_delay_seconds(attempt: int) -> float:
    return min(60.0, max(5.0, float(attempt) * 5.0))


def request_json(client: httpx.Client, method: str, path: str, **kwargs: Any) -> dict[str, Any]:
    retry_forever = bool(kwargs.pop("_retry_forever", True))
    max_attempts = max(1, int(kwargs.pop("_max_attempts", 3) or 3))
    attempt = 0
    while True:
        try:
            response = client.request(method, path, **kwargs)
            response.raise_for_status()
            data = response.json()
            return data if isinstance(data, dict) else {"ok": True, "data": data}
        except (httpx.HTTPStatusError, httpx.RequestError) as error:
            if not is_transient_request_error(error):
                raise
            attempt += 1
            if not retry_forever and attempt >= max_attempts:
                return {"ok": False, "task": None, "transient_error": describe_request_error(error)}
            delay_seconds = transient_retry_delay_seconds(attempt)
            print(
                f"Ket noi web loi tam thoi ({describe_request_error(error)}). "
                f"May tram se thu lai sau {int(delay_seconds)} giay.",
                file=sys.stderr,
            )
            time.sleep(delay_seconds)


def send_heartbeat(client: httpx.Client, worker_id: str, status: str = "idle", message: str = "", details: dict[str, Any] | None = None) -> None:
    payload = {
        "worker_id": worker_id,
        "status": status,
        "roles": ["onebss_worker", "sql_report_worker", "sql_export_worker", "ftp_report_worker", "excel_export", "drive_upload"],
        "version": WORKER_VERSION,
        "local_time": datetime.now().isoformat(timespec="seconds"),
        "message": message,
        "details": {
            **worker_process_details(),
            **(details or {}),
        },
    }
    try:
        response = client.post("/api/workstation/heartbeat", json=payload)
        if response.status_code == 404:
            return
        response.raise_for_status()
    except Exception as error:
        print(f"Khong gui duoc heartbeat may tram: {describe_request_error(error)}", file=sys.stderr)


def onebss_worker_otp_wait_seconds() -> float:
    raw_value = str(os.getenv("ONEBSS_WORKER_OTP_WAIT_SECONDS", "180") or "180").strip()
    try:
        value = float(raw_value)
    except (TypeError, ValueError):
        value = 180.0
    return max(30.0, min(value, 900.0))


def onebss_task_guard_enabled() -> bool:
    disabled = str(os.getenv("ONEBSS_WORKER_DISABLE_TASK_GUARD", "") or "").strip().lower() in {"1", "true", "yes", "on"}
    if disabled:
        return False
    enabled = str(os.getenv("ONEBSS_WORKER_ENABLE_TASK_GUARD", "") or "").strip().lower() in {"1", "true", "yes", "on"}
    return enabled


def wait_for_otp(
    client: httpx.Client,
    run_id: str,
    poll_seconds: float,
    progress_callback=None,
    *,
    timeout_seconds: float | None = None,
) -> str:
    if progress_callback:
        progress_callback("Dang doi OTP tu tin nhan/Mobile Gateway.")
    started = time.monotonic()
    last_notice = time.monotonic()
    max_wait = onebss_worker_otp_wait_seconds() if timeout_seconds is None else max(30.0, float(timeout_seconds))
    sleep_seconds = min(max(float(poll_seconds or 1.0), 1.0), 2.0)
    while True:
        data = request_json(
            client,
            "GET",
            f"/api/onebss-worker/tasks/{run_id}/otp",
            timeout=6.0,
            _retry_forever=False,
            _max_attempts=1,
        )
        if response_is_cancelled(data):
            raise OneBssTaskCancelled(str(data.get("message") or "Task OneBSS da bi huy."))
        if data.get("ok") and data.get("otp"):
            if progress_callback:
                progress_callback("Da nhan duoc OTP tu Mobile Gateway.")
            return str(data["otp"])
        status_value = str(data.get("status") or "").strip().lower()
        if status_value in {"expired", "cancelled", "consumed", "consume_failed"}:
            raise TimeoutError(str(data.get("message") or "Khong lay duoc OTP OneBSS tu Mobile Gateway."))
        now = time.monotonic()
        elapsed = now - started
        if elapsed >= max_wait:
            raise TimeoutError(
                "Khong nhan duoc OTP OneBSS trong "
                f"{int(max_wait)} giay. Hay kiem tra Mobile Gateway/SMS OTP roi bam lay bao cao lai."
            )
        if progress_callback and now - last_notice >= 15:
            base_message = str(data.get("transient_error") or data.get("message") or "Dang doi OTP tu tin nhan/Mobile Gateway.")
            progress_callback(f"{base_message} Da cho {int(elapsed)} giay.")
            last_notice = now
        time.sleep(sleep_seconds)


def _unique_urls(urls: list[str]) -> list[str]:
    seen: set[str] = set()
    unique: list[str] = []
    for url in urls:
        value = str(url or "").strip()
        if not value:
            continue
        key = value.rstrip("/").lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(value)
    return unique


def _is_public_tunnel_api_url(url: str) -> bool:
    host = (urlparse(str(url or "")).hostname or "").lower()
    return host == "api.vnptcto.com"


def internal_drive_upload_api_urls() -> list[str]:
    configured = os.getenv("ONEBSS_DRIVE_UPLOAD_API_URL", "").strip()
    internal = os.getenv("INTERNAL_API_URL", "").strip()
    urls: list[str] = []
    if not configured or _is_public_tunnel_api_url(configured):
        urls.append(LOCAL_DRIVE_UPLOAD_API_URL)
    urls.extend([configured, internal, PUBLIC_DRIVE_UPLOAD_API_URL])
    return _unique_urls(urls)


def internal_sql_api_urls() -> list[str]:
    configured = os.getenv("SQL_WORKER_API_URL", "").strip() or os.getenv("INTERNAL_API_URL", "").strip()
    urls: list[str] = []
    if not configured or _is_public_tunnel_api_url(configured):
        urls.append(LOCAL_INTERNAL_API_URL)
    urls.append(configured)
    return _unique_urls(urls)


def upload_result_file_to_internal_drive(
    file_path: str,
    drive_folder_id: str,
    *,
    request_source: str = "onebss-worker",
    default_message: str = "Da upload file OneBSS len Google Drive qua API trung gian.",
    job_id: str = "",
) -> dict[str, Any]:
    folder_id = str(drive_folder_id or "").strip()
    if not folder_id:
        return {}
    source = Path(str(file_path or ""))
    if not source.exists() or not source.is_file() or source.stat().st_size <= 0:
        return {}
    token = os.getenv("INTERNAL_API_TOKEN", "").strip()
    api_urls = internal_drive_upload_api_urls()
    if not api_urls or not token:
        return {}
    mime_type = mimetypes.guess_type(source.name)[0] or "application/octet-stream"
    payload = {
        "action": "upload_file_to_drive",
        "source": request_source,
        "file_name": source.name,
        "job_id": str(job_id or "").strip(),
        "run_id": str(job_id or "").strip(),
        "file_base64": base64.b64encode(source.read_bytes()).decode("ascii"),
        "content_type": mime_type,
        "drive_folder_id": folder_id,
    }
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    timeout_seconds = float(os.getenv("ONEBSS_DRIVE_UPLOAD_TIMEOUT_SECONDS", "300") or "300")
    last_error: Exception | None = None
    for api_url in api_urls:
        try:
            with httpx.Client(timeout=httpx.Timeout(timeout_seconds, connect=10.0)) as internal_client:
                response = internal_client.post(api_url, json=payload, headers=headers)
            response.raise_for_status()
            break
        except (httpx.HTTPStatusError, httpx.RequestError) as error:
            last_error = error
            if api_url != api_urls[-1]:
                print(f"Khong upload Drive qua {api_url}: {describe_request_error(error)}. Thu endpoint tiep theo.", file=sys.stderr)
                continue
            raise
    else:
        if last_error:
            raise last_error
        return {}
    data = response.json()
    if not isinstance(data, dict) or not data.get("ok"):
        return {}
    drive_url = str(data.get("drive_url") or data.get("storage_link") or data.get("web_view_link") or data.get("web_content_link") or "").strip()
    if not drive_url:
        return {}
    file_id = str(data.get("file_id") or "").strip()
    return {
        "file_name": str(data.get("file_name") or source.name),
        "storage_link": drive_url,
        "storage_status": f"uploaded_google_drive:{file_id}" if file_id else "uploaded_google_drive",
        "message": str(data.get("message") or default_message),
    }


def upload_task_file(client: httpx.Client, upload_path: str, file_path: str, cancelled_error=OneBssTaskCancelled) -> dict[str, Any]:
    source = Path(str(file_path or ""))
    if not source.exists() or not source.is_file() or source.stat().st_size <= 0:
        return {}
    mime_type = mimetypes.guess_type(source.name)[0] or "application/octet-stream"
    attempt = 0
    while True:
        try:
            with source.open("rb") as handle:
                response = client.post(
                    upload_path,
                    files={"file": (source.name, handle, mime_type)},
                )
            response.raise_for_status()
            data = response.json()
            if isinstance(data, dict) and response_is_cancelled(data):
                raise cancelled_error(str(data.get("message") or "Task da bi huy."))
            uploaded = data.get("file") if isinstance(data.get("file"), dict) else {}
            return uploaded if isinstance(uploaded, dict) else {}
        except (httpx.HTTPStatusError, httpx.RequestError) as error:
            if not is_transient_request_error(error):
                raise
            attempt += 1
            delay_seconds = transient_retry_delay_seconds(attempt)
            print(
                f"Ket noi web loi tam thoi khi gui file ({describe_request_error(error)}). "
                f"May tram se thu lai sau {int(delay_seconds)} giay.",
                file=sys.stderr,
            )
            time.sleep(delay_seconds)


def upload_result_file(client: httpx.Client, run_id: str, file_path: str) -> dict[str, Any]:
    return upload_task_file(client, f"/api/onebss-worker/tasks/{run_id}/file", file_path)


def attach_worker_file_if_needed(client: httpx.Client, run_id: str, result: dict[str, Any], drive_folder_id: str = "", progress_callback=None) -> dict[str, Any]:
    storage_status = str(result.get("storage_status") or "").lower()
    if storage_status.startswith("uploaded_google_drive:"):
        return result
    try:
        if drive_folder_id and progress_callback:
            progress_callback("Dang upload file len Google Drive qua API trung gian.")
        drive_uploaded = upload_result_file_to_internal_drive(
            str(result.get("file_path") or ""),
            drive_folder_id,
            job_id=run_id,
        )
    except Exception as error:
        print(f"Cannot upload OneBSS result to Drive through internal API: {error}", file=sys.stderr)
        if progress_callback:
            progress_callback("Upload Google Drive qua API trung gian loi, dang gui file ve web.")
        drive_uploaded = {}
    if drive_uploaded:
        if progress_callback:
            progress_callback("Da upload file len Google Drive.")
        merged = {**result}
        for key in ("file_name", "storage_link", "storage_status"):
            if drive_uploaded.get(key):
                merged[key] = drive_uploaded.get(key)
        merged["ok"] = True
        merged["status"] = "success"
        merged["message"] = drive_uploaded.get("message") or "Da upload file OneBSS len Google Drive qua API trung gian."
        return merged
    if progress_callback:
        progress_callback("Dang gui file ket qua ve web de co link tai xuong.")
    uploaded = upload_result_file(client, run_id, str(result.get("file_path") or ""))
    if not uploaded:
        return result
    if progress_callback:
        progress_callback("Da gui file ket qua ve web.")
    merged = {**result}
    for key in ("file_name", "file_path", "storage_link", "storage_status"):
        if uploaded.get(key):
            merged[key] = uploaded.get(key)
    failed_only_at_storage = str(result.get("status") or "").lower() in {
        "google_drive_upload_failed",
        "google_drive_not_configured",
        "storage_failed",
    }
    if failed_only_at_storage:
        merged["ok"] = True
        merged["status"] = "success"
        merged["message"] = uploaded.get("message") or "Da tai bao cao OneBSS va gui file ve web."
    return merged


def onebss_task_timeout_seconds() -> float:
    raw_value = os.getenv("ONEBSS_TASK_TIMEOUT_SECONDS", "").strip() or os.getenv("ONEBSS_WORKER_TASK_TIMEOUT_SECONDS", "").strip() or "1200"
    try:
        value = float(raw_value)
    except ValueError:
        value = 1200.0
    return min(max(value, 120.0), 21600.0)


def onebss_worker_max_otp_attempts() -> int:
    """So lan thu OTP toi da truoc khi worker bao loi thay vi loop vo tan.

    Mac dinh 3. Co the cau hinh qua ONEBSS_WORKER_MAX_OTP_ATTEMPTS. Gia tri toi thieu 1.
    """
    raw_value = os.getenv("ONEBSS_WORKER_MAX_OTP_ATTEMPTS", "").strip()
    try:
        value = int(raw_value)
    except (TypeError, ValueError):
        return 3
    return max(1, value)


def terminate_process(process: mp.Process) -> None:
    if not process.is_alive():
        return
    process.terminate()
    process.join(timeout=10)
    if process.is_alive():
        try:
            process.kill()
        except AttributeError:
            process.terminate()
        process.join(timeout=5)


def _run_onebss_report_request_child(
    result_queue: mp.Queue,
    base_url: str,
    token: str,
    worker_id: str,
    run_id: str,
    report: dict[str, Any],
    parameters: dict[str, Any],
    otp: str,
    session_id: str,
    state_path: str,
) -> None:
    headers = {"Authorization": f"Bearer {token}"}
    with httpx.Client(base_url=base_url.rstrip("/"), headers=headers, timeout=httpx.Timeout(60.0, connect=20.0)) as child_client:
        def child_progress(message: str, status: str = "running") -> None:
            data = request_json(
                child_client,
                "POST",
                f"/api/onebss-worker/tasks/{run_id}/status",
                json={
                    "status": status,
                    "message": message,
                    "worker_id": worker_id,
                    "worker_session_id": session_id,
                    "details": {
                        **worker_process_details(),
                        "task_type": "onebss",
                        "process": "child",
                    },
                },
                timeout=10.0,
                _retry_forever=False,
            )
            if data.get("transient_error"):
                return
            if response_is_cancelled(data):
                raise OneBssTaskCancelled(str(data.get("message") or "Task OneBSS da bi huy."))

        try:
            settings = get_settings().model_copy(update={"mobile_gateway_enabled": False, "google_drive_folder_id": ""})
            result = run_onebss_report_request(
                settings,
                report,
                parameters,
                otp=otp,
                session_id=session_id,
                created_by=worker_id,
                progress_callback=child_progress,
                state_path=state_path or None,
            )
            result_queue.put({"ok": True, "result": result})
        except OneBssTaskCancelled as error:
            result_queue.put({"ok": False, "cancelled": True, "message": str(error)})
        except Exception as error:
            result_queue.put({"ok": False, "error": str(error)[:1000], "error_type": error.__class__.__name__})


def read_onebss_child_result(result_queue: mp.Queue) -> dict[str, Any]:
    for _ in range(20):
        try:
            data = result_queue.get_nowait()
            return data if isinstance(data, dict) else {}
        except queue.Empty:
            time.sleep(0.1)
    return {}


def run_onebss_report_request_guarded(
    client: httpx.Client,
    worker_id: str,
    run_id: str,
    report: dict[str, Any],
    parameters: dict[str, Any],
    *,
    otp: str,
    session_id: str,
    progress_callback=None,
    state_path: str | Path | None = None,
) -> dict[str, Any]:
    if not onebss_task_guard_enabled() or not hasattr(client, "base_url"):
        settings = get_settings().model_copy(update={"mobile_gateway_enabled": False, "google_drive_folder_id": ""})
        return run_onebss_report_request(
            settings,
            report,
            parameters,
            otp=otp,
            session_id=session_id,
            created_by=worker_id,
            progress_callback=progress_callback,
            state_path=state_path,
        )

    timeout_seconds = onebss_task_timeout_seconds()
    token = os.getenv("INTERNAL_API_TOKEN", "").strip()
    if not token:
        raise RuntimeError("Thieu INTERNAL_API_TOKEN de chay worker OneBSS.")
    result_queue: mp.Queue = mp.Queue()
    process = mp.Process(
        target=_run_onebss_report_request_child,
        args=(result_queue, str(client.base_url), token, worker_id, run_id, report, parameters, otp, session_id, str(state_path or "")),
        daemon=True,
    )
    started = time.monotonic()
    last_notice = started
    try:
        process.start()
        while process.is_alive():
            process.join(timeout=1.0)
            now = time.monotonic()
            elapsed = now - started
            if elapsed >= timeout_seconds:
                terminate_process(process)
                raise TimeoutError(
                    f"Task OneBSS vuot qua {int(timeout_seconds / 60)} phut nen worker da tu dung de tranh treo may tram."
                )
            if progress_callback and now - last_notice >= 30:
                last_notice = now
                progress_callback(f"May tram van dang xu ly OneBSS ({int(elapsed / 60)} phut).")
        payload = read_onebss_child_result(result_queue)
        if payload.get("cancelled"):
            raise OneBssTaskCancelled(str(payload.get("message") or "Task OneBSS da bi huy."))
        if payload.get("ok"):
            result = payload.get("result")
            if isinstance(result, dict):
                return result
            return {"ok": False, "status": "failed", "message": "Worker OneBSS tra ket qua khong hop le."}
        message = str(payload.get("error") or "").strip()
        if not message:
            message = f"Worker OneBSS dung bat thuong voi ma {process.exitcode}."
        raise RuntimeError(message)
    except Exception:
        terminate_process(process)
        raise
    finally:
        try:
            result_queue.close()
            result_queue.join_thread()
        except Exception:
            pass


def process_task(client: httpx.Client, task: dict[str, Any], worker_id: str, poll_seconds: float) -> None:
    run_id = str(task.get("run_id") or "")
    report = task.get("report") if isinstance(task.get("report"), dict) else {}
    parameters = task.get("parameters") if isinstance(task.get("parameters"), dict) else {}
    drive_folder_id = str(task.get("drive_folder_id") or "").strip()
    worker_slot = int(task.get("_worker_slot") or 0)
    worker_state_path = str(task.get("_worker_state_path") or "").strip()
    report_for_worker = {**report, "storage_link": ""}
    session_id = ""
    otp = ""
    started = time.monotonic()
    last_progress = {"message": "", "at": 0.0}
    last_heartbeat = {"at": 0.0}
    max_otp_attempts = onebss_worker_max_otp_attempts()
    otp_attempts = 0

    def send_progress(message: str, status: str = "running") -> None:
        text = str(message or "").strip()
        if not text:
            return
        now = time.monotonic()
        if text == last_progress["message"] and now - float(last_progress["at"] or 0) < 3:
            return
        last_progress["message"] = text
        last_progress["at"] = now
        data = request_json(
            client,
            "POST",
            f"/api/onebss-worker/tasks/{run_id}/status",
            json={
                "status": status,
                "message": text,
                "worker_id": worker_id,
                "worker_session_id": session_id,
                "details": {
                    **worker_process_details(),
                    "task_type": "onebss",
                    "worker_slot": worker_slot,
                    "worker_state_path": worker_state_path,
                },
            },
            timeout=8.0,
            _retry_forever=False,
            _max_attempts=1,
        )
        if data.get("transient_error"):
            print(
                f"Khong cap nhat duoc tien trinh OneBSS ({data.get('transient_error')}); worker tiep tuc xu ly.",
                file=sys.stderr,
            )
            return
        if response_is_cancelled(data):
            raise OneBssTaskCancelled(str(data.get("message") or "Task OneBSS da bi huy."))
        if now - float(last_heartbeat["at"] or 0) >= 25:
            last_heartbeat["at"] = now
            send_heartbeat(
                client,
                worker_id,
                "busy",
                text,
                {"run_id": run_id, "report": report.get("ma_bao_cao") or "", "task_type": "onebss"},
            )

    try:
        send_progress("May tram da nhan task OneBSS. Dang khoi tao phien chay.")
        while True:
            result = run_onebss_report_request_guarded(
                client,
                worker_id,
                run_id,
                report_for_worker,
                parameters,
                otp=otp,
                session_id=session_id,
                progress_callback=send_progress,
                state_path=worker_state_path or None,
            )
            status = str(result.get("status") or ("success" if result.get("ok") else "failed")).lower()
            if status == "otp_session_expired":
                if otp_attempts >= max_otp_attempts:
                    duration_ms = int((time.monotonic() - started) * 1000)
                    failure_message = (
                        f"Phien OTP OneBSS da het han sau {max_otp_attempts} lan thu. "
                        "Hay kiem tra OTP/Mobile Gateway roi bam lay bao cao lai."
                    )
                    try:
                        request_json(
                            client,
                            "POST",
                            f"/api/onebss-worker/tasks/{run_id}/result",
                            json={
                                "ok": False,
                                "status": "failed",
                                "message": failure_message,
                                "duration_ms": duration_ms,
                                "details": {
                                    "error_type": "OtpSessionExpired",
                                    "otp_attempts": otp_attempts,
                                    "max_otp_attempts": max_otp_attempts,
                                },
                            },
                        )
                    except Exception as update_error:
                        print(
                            f"Khong cap nhat duoc ket qua OneBSS: {describe_request_error(update_error)}",
                            file=sys.stderr,
                        )
                    return
                send_progress("Phien OTP OneBSS da het han. May tram dang dang nhap lai de lay OTP moi.")
                session_id = ""
                otp = ""
                continue
            if status in {"otp_required", "otp_invalid", "manual_otp_required"} and result.get("session_id"):
                session_id = str(result.get("session_id") or "")
                otp_attempts += 1
                if otp_attempts > max_otp_attempts:
                    duration_ms = int((time.monotonic() - started) * 1000)
                    failure_message = (
                        f"May tram khong nhan duoc OTP OneBSS sau {max_otp_attempts} lan thu. "
                        "Hay kiem tra Mobile Gateway / app Android dong bo SMS hoac nhap OTP thu cong."
                    )
                    print(failure_message, file=sys.stderr)
                    try:
                        request_json(
                            client,
                            "POST",
                            f"/api/onebss-worker/tasks/{run_id}/result",
                            json={
                                "ok": False,
                                "status": "failed",
                                "message": failure_message,
                                "duration_ms": duration_ms,
                                "details": {
                                    "error_type": "OtpExhausted",
                                    "otp_attempts": otp_attempts,
                                    "max_otp_attempts": max_otp_attempts,
                                },
                            },
                        )
                    except Exception as update_error:
                        print(
                            f"Khong cap nhat duoc ket qua OneBSS: {describe_request_error(update_error)}",
                            file=sys.stderr,
                        )
                    return
                status_response = request_json(
                    client,
                    "POST",
                    f"/api/onebss-worker/tasks/{run_id}/status",
                    json={
                        "status": status,
                        "message": result.get("message") or "May tram dang doi OTP OneBSS.",
                        "worker_id": worker_id,
                        "worker_session_id": session_id,
                    },
                )
                if response_is_cancelled(status_response):
                    return
                send_progress(
                    f"Dang thu lay OTP OneBSS (lan thu {otp_attempts}/{max_otp_attempts}).",
                    status,
                )
                otp = wait_for_otp(client, run_id, poll_seconds, lambda message: send_progress(message, status))
                continue

            duration_ms = int((time.monotonic() - started) * 1000)
            send_progress("Da hoan thanh buoc lay du lieu OneBSS. Dang xu ly file ket qua.")
            result = attach_worker_file_if_needed(client, run_id, result, drive_folder_id, send_progress)
            status = str(result.get("status") or ("success" if result.get("ok") else "failed")).lower()
            finish_response = request_json(
                client,
                "POST",
                f"/api/onebss-worker/tasks/{run_id}/result",
                json={
                    "ok": bool(result.get("ok")),
                    "status": status,
                    "message": result.get("message") or "",
                    "file_name": result.get("file_name") or "",
                    "file_path": result.get("file_path") or "",
                    "storage_link": result.get("storage_link") or "",
                    "storage_status": result.get("storage_status") or "",
                    "duration_ms": int(result.get("duration_ms") or duration_ms),
                    "details": result,
                },
            )
            if response_is_cancelled(finish_response):
                return
            return
    except OneBssTaskCancelled as error:
        print(str(error), file=sys.stderr)
        return
    except Exception as error:
        duration_ms = int((time.monotonic() - started) * 1000)
        error_message = str(error)[:500] or error.__class__.__name__
        print(f"Task OneBSS loi: {error_message}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        try:
            request_json(
                client,
                "POST",
                f"/api/onebss-worker/tasks/{run_id}/result",
                json={
                    "ok": False,
                    "status": "failed",
                    "message": f"May tram gap loi khi lay OneBSS: {error_message}",
                    "duration_ms": duration_ms,
                    "details": {"error_type": error.__class__.__name__},
                },
            )
        except Exception as update_error:
            print(
                f"Khong cap nhat duoc ket qua OneBSS: {describe_request_error(update_error)}",
                file=sys.stderr,
            )
        return


def run_sql_worker_query(task: dict[str, Any]) -> dict[str, Any]:
    query = dict(task.get("query") if isinstance(task.get("query"), dict) else {})
    if not query:
        raise RuntimeError("Task SQL khong co cau lenh truy van.")
    run_id = str(task.get("run_id") or task.get("job_id") or "").strip()
    if run_id:
        query.setdefault("run_id", run_id)
        query.setdefault("job_id", run_id)
        query.setdefault("worker_task_id", run_id)
    token = os.getenv("INTERNAL_API_TOKEN", "").strip()
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    timeout_seconds = float(os.getenv("SQL_WORKER_TIMEOUT_SECONDS", "1800") or "1800")
    api_urls = internal_sql_api_urls()
    if not api_urls:
        raise RuntimeError("Chua cau hinh URL API du lieu local cho SQL worker.")
    last_error: Exception | None = None
    for api_url in api_urls:
        try:
            with httpx.Client(timeout=httpx.Timeout(timeout_seconds, connect=10.0)) as internal_client:
                response = internal_client.post(api_url, json=query, headers=headers)
            response.raise_for_status()
            data = response.json()
            return data if isinstance(data, dict) else {"ok": True, "data": data}
        except (httpx.HTTPStatusError, httpx.RequestError) as error:
            last_error = error
            if api_url != api_urls[-1]:
                print(f"Khong chay SQL qua {api_url}: {describe_request_error(error)}. Thu endpoint tiep theo.", file=sys.stderr)
                continue
            raise
    if last_error:
        raise last_error
    return {"ok": False, "message": "API du lieu local khong tra ket qua."}


def int_from_values(*values: Any, default: int = 0) -> int:
    for value in values:
        if isinstance(value, list):
            return len(value)
        try:
            if value not in (None, ""):
                return int(value)
        except (TypeError, ValueError):
            continue
    return default


def run_sql_worker_query_all_pages(task: dict[str, Any], progress_callback=None) -> dict[str, Any]:
    query = dict(task.get("query") if isinstance(task.get("query"), dict) else {})
    pagination = dict(query.get("pagination") if isinstance(query.get("pagination"), dict) else {})
    page_size = max(1, min(int_from_values(pagination.get("page_size"), default=20000), 20000))
    start_page = max(1, int_from_values(pagination.get("page"), default=1))
    max_rows = max(1, int_from_values(query.get("max_rows"), os.getenv("EXPORT_MAX_ROWS", "1000000"), default=1000000))
    rows: list[dict[str, Any]] = []
    columns: list[str] = []
    total = 0
    page = start_page
    last_result: dict[str, Any] = {}

    while len(rows) < max_rows:
        page_query = {
            **query,
            "collect_all_pages": False,
            "pagination": {**pagination, "page": page, "page_size": page_size},
        }
        page_task = {**task, "query": page_query}
        result = run_sql_worker_query(page_task)
        last_result = result if isinstance(result, dict) else {}
        if last_result.get("ok") is False:
            return last_result
        page_rows = last_result.get("rows") or last_result.get("data") or []
        if not isinstance(page_rows, list):
            page_rows = []
        if not columns:
            columns = last_result.get("columns") if isinstance(last_result.get("columns"), list) else []
            if not columns and page_rows and isinstance(page_rows[0], dict):
                columns = list(page_rows[0].keys())
        if not total:
            total = int_from_values(last_result.get("total"), (last_result.get("pagination") or {}).get("total") if isinstance(last_result.get("pagination"), dict) else None)
        remaining = max_rows - len(rows)
        rows.extend(page_rows[:remaining])
        if progress_callback:
            progress_callback(
                f"May tram da lay {len(rows)}"
                + (f"/{total}" if total else "")
                + " dong SQL de nap Sheet."
            )
        if not page_rows:
            break
        if total and len(rows) >= total:
            break
        if len(page_rows) < page_size and not total:
            break
        page += 1

    details = last_result.get("details") if isinstance(last_result.get("details"), dict) else {}
    return {
        **last_result,
        "ok": bool(last_result.get("ok", True)),
        "message": last_result.get("message") or f"May tram da tai {len(rows)} dong SQL qua API local.",
        "columns": columns,
        "rows": rows,
        "total": total or len(rows),
        "page": start_page,
        "page_size": page_size,
        "pagination": {
            "page": start_page,
            "page_size": page_size,
            "total": total or len(rows),
            "fetched_rows": len(rows),
            "truncated": bool(total and len(rows) < total) or (not total and len(rows) >= max_rows),
        },
        "details": {
            **details,
            "collect_all_pages": True,
            "fetched_rows": len(rows),
            "max_rows": max_rows,
        },
    }


def _run_sql_worker_query_child(result_queue: mp.Queue, task: dict[str, Any]) -> None:
    try:
        result_queue.put({"ok": True, "result": run_sql_worker_query(task)})
    except Exception as error:
        result_queue.put({"ok": False, "error": str(error)[:1000], "error_type": error.__class__.__name__})


def run_sql_worker_query_cancellable(task: dict[str, Any], progress_callback) -> dict[str, Any]:
    """Run the blocking local SQL export in a process that can be stopped by the web job."""
    timeout_seconds = float(os.getenv("SQL_WORKER_TIMEOUT_SECONDS", "1800") or "1800")
    result_queue: mp.Queue = mp.Queue()
    process = mp.Process(target=_run_sql_worker_query_child, args=(result_queue, task), daemon=True)
    started = time.monotonic()
    last_status_check = 0.0
    payload: dict[str, Any] = {}
    try:
        process.start()
        while process.is_alive():
            process.join(timeout=1.0)
            now = time.monotonic()
            if now - started >= timeout_seconds:
                raise TimeoutError(f"Task SQL vuot qua {int(timeout_seconds / 60)} phut nen worker da tu dung.")
            if now - last_status_check >= 3.0:
                last_status_check = now
                response = progress_callback(
                    f"May tram van dang truy van SQL, da chay {int(now - started)} giay.",
                    details={"step": "oracle_export_drive", "elapsed_seconds": int(now - started)},
                )
                if response_is_cancelled(response):
                    raise SqlTaskCancelled(str(response.get("message") or "Lenh lay du lieu SQL da bi ngung."))
        for _ in range(20):
            try:
                payload = result_queue.get_nowait()
                break
            except queue.Empty:
                time.sleep(0.1)
        if payload.get("ok") and isinstance(payload.get("result"), dict):
            return payload["result"]
        message = str(payload.get("error") or "").strip()
        if not message:
            message = f"Worker SQL dung bat thuong voi ma {process.exitcode}."
        raise RuntimeError(message)
    except Exception:
        terminate_process(process)
        raise
    finally:
        try:
            result_queue.close()
            result_queue.join_thread()
        except Exception:
            pass


def process_sql_task(client: httpx.Client, task: dict[str, Any], worker_id: str) -> None:
    run_id = str(task.get("run_id") or task.get("job_id") or "")
    started = time.monotonic()

    def send_progress(message: str, status: str = "running_worker", details: dict[str, Any] | None = None) -> dict[str, Any]:
        response = request_json(
            client,
            "POST",
            f"/api/sql-worker/tasks/{run_id}/status",
            json={
                "status": status,
                "message": message,
                "worker_id": worker_id,
                "details": {
                    **worker_process_details(),
                    **(details or {}),
                },
            },
            timeout=8.0,
            _retry_forever=False,
            _max_attempts=1,
        )
        if response_is_cancelled(response):
            raise SqlTaskCancelled(str(response.get("message") or "Lenh lay du lieu SQL da bi ngung."))
        return response

    def start_export_heartbeat(report_code: str) -> threading.Event:
        stop_event = threading.Event()

        def heartbeat_loop() -> None:
            while not stop_event.wait(25):
                elapsed_seconds = int(time.monotonic() - started)
                try:
                    send_progress(
                        f"May tram van dang xuat Excel/upload Drive, da chay {elapsed_seconds} giay.",
                        details={
                            "step": "oracle_export_drive",
                            "report": report_code,
                            "elapsed_seconds": elapsed_seconds,
                            "worker_version": WORKER_VERSION,
                        },
                    )
                except Exception as error:
                    print(f"Khong cap nhat duoc tien trinh SQL: {describe_request_error(error)}", file=sys.stderr)

        threading.Thread(
            target=heartbeat_loop,
            name=f"vnptcto-sql-export-{run_id or 'task'}",
            daemon=True,
        ).start()
        return stop_event

    try:
        query = task.get("query") if isinstance(task.get("query"), dict) else {}
        action = str(query.get("action") or "").strip()
        report_code = str(task.get("report_code") or query.get("ma_bao_cao") or "")
        export_heartbeat: threading.Event | None = None
        if action == "export_sql_report_to_drive":
            send_progress(
                "May tram da nhan lenh lay du lieu SQL. Dang chuan bi ket noi Oracle noi bo.",
                details={"step": "received", "report": report_code, "api_urls": internal_sql_api_urls()},
            )
            send_progress(
                "Dang ket noi Oracle noi bo, xuat Excel va upload Google Drive.",
                details={"step": "oracle_export_drive", "report": report_code, "api_urls": internal_sql_api_urls()},
            )
            export_heartbeat = start_export_heartbeat(report_code)
        else:
            send_progress(
                "May tram da nhan task SQL. Dang goi API du lieu local.",
                details={"step": "received", "report": report_code, "api_urls": internal_sql_api_urls()},
            )
        try:
            if action == "run_sql_report" and bool(query.get("collect_all_pages")):
                result = run_sql_worker_query_all_pages(task, send_progress)
            else:
                result = (
                    run_sql_worker_query_cancellable(task, send_progress)
                    if action == "export_sql_report_to_drive"
                    else run_sql_worker_query(task)
                )
        finally:
            if export_heartbeat is not None:
                export_heartbeat.set()

        def result_int(*values: Any) -> int:
            return int_from_values(*values)

        if action == "export_sql_report_to_drive":
            columns = result.get("columns") if isinstance(result.get("columns"), list) else []
            pagination = result.get("pagination") if isinstance(result.get("pagination"), dict) else {}
            total = result_int(result.get("total"), result.get("rows"), pagination.get("total"))
            details = result.get("details") if isinstance(result.get("details"), dict) else {}
            drive_url = str(
                result.get("drive_url")
                or result.get("storage_link")
                or result.get("web_view_link")
                or result.get("web_content_link")
                or ""
            ).strip()
            send_progress(
                "Da nhan ket qua tu API local. Dang cap nhat file va link Drive len web.",
                details={"step": "returning_drive_link", "report": report_code, "drive_url": drive_url},
            )
            details = {
                **details,
                "duration_ms": int((time.monotonic() - started) * 1000),
                "file_id": str(result.get("file_id") or ""),
                "file_name": str(result.get("file_name") or query.get("file_name") or ""),
                "drive_url": drive_url,
                "storage_link": drive_url,
                "rows": total,
                "total": result_int(result.get("total"), total),
            }
            finish_response = request_json(
                client,
                "POST",
                f"/api/sql-worker/tasks/{run_id}/result",
                json={
                    "ok": bool(result.get("ok", True)),
                    "status": "success" if result.get("ok", True) else "failed",
                    "message": result.get("message") or "May tram da xuat Excel va upload Google Drive.",
                    "columns": columns,
                    "rows": [],
                    "pagination": {"page": 1, "page_size": result_int(pagination.get("page_size"), total), "total": total},
                    "report": task.get("report") if isinstance(task.get("report"), dict) else {},
                    "details": details,
                    "drive_url": drive_url,
                    "storage_link": drive_url,
                    "file_name": str(result.get("file_name") or query.get("file_name") or ""),
                    "file_id": str(result.get("file_id") or ""),
                    "total": result_int(result.get("total"), total),
                },
            )
            if response_is_cancelled(finish_response):
                return
            return

        rows = result.get("rows") or result.get("data") or []
        if not isinstance(rows, list):
            rows = []
        columns = result.get("columns") if isinstance(result.get("columns"), list) else []
        if not columns and rows and isinstance(rows[0], dict):
            columns = list(rows[0].keys())
        result_pagination = result.get("pagination") if isinstance(result.get("pagination"), dict) else {}
        pagination = {
            "page": int(result.get("page") or result_pagination.get("page") or ((task.get("query") or {}).get("pagination") or {}).get("page") or 1),
            "page_size": int(result.get("page_size") or result_pagination.get("page_size") or ((task.get("query") or {}).get("pagination") or {}).get("page_size") or len(rows) or 20),
            "total": int(result.get("total") or result_pagination.get("total") or len(rows)),
        }
        for key in ("fetched_rows", "truncated"):
            if key in result_pagination:
                pagination[key] = result_pagination[key]
        details = result.get("details") if isinstance(result.get("details"), dict) else {}
        details = {**details, "duration_ms": int((time.monotonic() - started) * 1000)}
        finish_response = request_json(
            client,
            "POST",
            f"/api/sql-worker/tasks/{run_id}/result",
            json={
                "ok": bool(result.get("ok", True)),
                "status": "success" if result.get("ok", True) else "failed",
                "message": result.get("message") or "May tram da tai du lieu SQL qua API local.",
                "columns": columns,
                "rows": rows,
                "pagination": pagination,
                "report": task.get("report") if isinstance(task.get("report"), dict) else {},
                "details": details,
            },
        )
        if response_is_cancelled(finish_response):
            return
    except SqlTaskCancelled as error:
        print(f"Task SQL da ngung: {error}", flush=True)
        return
    except Exception as error:
        message = str(error)[:500] or error.__class__.__name__
        print(f"Task SQL loi: {message}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        try:
            request_json(
                client,
                "POST",
                f"/api/sql-worker/tasks/{run_id}/result",
                json={
                    "ok": False,
                    "status": "failed",
                    "message": f"May tram khong chay duoc SQL qua API local: {message}",
                    "details": {"error_type": error.__class__.__name__},
                },
            )
        except Exception as update_error:
            print(f"Khong cap nhat duoc ket qua SQL: {describe_request_error(update_error)}", file=sys.stderr)


def safe_local_filename(value: str, fallback: str = "ftp_result") -> str:
    text = Path(str(value or fallback)).name.strip() or fallback
    text = "".join("_" if character in '<>:"/\\|?*' or ord(character) < 32 else character for character in text)
    text = text.strip(" .")
    return text or fallback


def normalize_ftp_variable_value(value: Any) -> str:
    text = str(value or "").strip()
    match = re.fullmatch(r"\{([0-9][0-9A-Za-z._/-]*)\}", text)
    return match.group(1) if match else text


def normalize_ftp_variables(value: Any) -> dict[str, str]:
    if isinstance(value, dict):
        return {
            str(key).strip(): normalize_ftp_variable_value(item)
            for key, item in value.items()
            if str(key).strip()
        }
    if not isinstance(value, str):
        return {}
    variables: dict[str, str] = {}
    for raw_line in value.replace("\r", "\n").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        separator = "=" if "=" in line else ":" if ":" in line else ""
        if not separator:
            continue
        key, item = line.split(separator, 1)
        key = key.strip()
        if key:
            variables[key] = normalize_ftp_variable_value(item)
    return variables


def _replace_ftp_template_tokens(template: str, values: dict[str, str]) -> str:
    output = str(template or "")
    for key in sorted(values, key=len, reverse=True):
        value = str(values.get(key) or "")
        output = output.replace(f"{{{{{key}}}}}", value).replace(f"{{{key}}}", value)
    return output


def _ftp_month_values(year: int, month: int, prefix: str = "") -> dict[str, str]:
    last_day = calendar.monthrange(year, month)[1]
    base = f"{year:04d}{month:02d}"
    return {
        f"{prefix}yyyy": f"{year:04d}",
        f"{prefix}yy": f"{year % 100:02d}",
        f"{prefix}mm": f"{month:02d}",
        f"{prefix}m": str(month),
        f"{prefix}yyyyMM": base,
        f"{prefix}yyyymm": base,
        f"{prefix}YYYYMM": base,
        f"{prefix}yyyymm01": f"{base}01",
        f"{prefix}yyyyMM01": f"{base}01",
        f"{prefix}firstday": f"{base}01",
        f"{prefix}first_day": f"{base}01",
        f"{prefix}lastday": f"{base}{last_day:02d}",
        f"{prefix}last_day": f"{base}{last_day:02d}",
        f"{prefix}last_dd": f"{last_day:02d}",
        f"{prefix}lastd": str(last_day),
    }


def ftp_template_values(now: datetime | None = None, variables: Any = None) -> dict[str, str]:
    current = now or datetime.now()
    yesterday = current - timedelta(days=1)
    values = {
        **_ftp_month_values(current.year, current.month),
        "yyyy": f"{current:%Y}",
        "YYYY": f"{current:%Y}",
        "yy": f"{current:%y}",
        "YY": f"{current:%y}",
        "mm": f"{current:%m}",
        "MM": f"{current:%m}",
        "m": str(current.month),
        "dd": f"{current:%d}",
        "DD": f"{current:%d}",
        "d": str(current.day),
        "yyyymmdd": f"{current:%Y%m%d}",
        "yyyyMMdd": f"{current:%Y%m%d}",
        "YYYYMMDD": f"{current:%Y%m%d}",
        "ddmmyyyy": f"{current:%d%m%Y}",
        "DDMMYYYY": f"{current:%d%m%Y}",
        "today": f"{current:%Y%m%d}",
        "today_ddmmyyyy": f"{current:%d%m%Y}",
        "yesterday": f"{yesterday:%Y%m%d}",
        "yesterday_ddmmyyyy": f"{yesterday:%d%m%Y}",
    }
    raw_variables = normalize_ftp_variables(variables)
    for _ in range(4):
        changed = False
        for key, raw_value in raw_variables.items():
            rendered = _replace_ftp_template_tokens(raw_value, values)
            if values.get(key) != rendered:
                values[key] = rendered
                changed = True
        if not changed:
            break
    thang = values.get("thang") or raw_variables.get("thang")
    if re.fullmatch(r"\d{6}", str(thang or "")):
        year = int(str(thang)[:4])
        month = int(str(thang)[4:6])
        if 1 <= month <= 12:
            values.update(_ftp_month_values(year, month, "thang_"))
            values["lastday"] = values["thang_lastday"]
            values["last_day"] = values["thang_last_day"]
            values["last_dd"] = values["thang_last_dd"]
    return values


def render_ftp_file_template(template: str, now: datetime | None = None, variables: Any = None) -> str:
    return _replace_ftp_template_tokens(str(template or "").strip(), ftp_template_values(now, variables)).strip()


def normalize_legacy_ftp_site_text(value: Any) -> str:
    text = str(value or "")
    text = re.sub(r"/DATA_BILLING/HGA(?=/|$)", "/DATA_BILLING/HAG", text, flags=re.IGNORECASE)
    return re.sub(r"\bHGA_", "HAG_", text, flags=re.IGNORECASE)


def decode_ftp_template_config(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    text = str(value or "").strip()
    if not text.startswith("{"):
        return {}
    try:
        config = json.loads(text)
    except json.JSONDecodeError:
        return {}
    if not isinstance(config, dict):
        return {}
    if not any(key in config for key in ("sources", "variables", "file_name_template", "output_file_name_template")):
        return {}
    return config


def _parse_ftp_location(base_config: dict[str, Any], folder_path: str, file_template: str) -> tuple[dict[str, Any], str, str]:
    ftp_config = dict(base_config)
    folder_value = normalize_legacy_ftp_site_text(folder_path).strip()
    template_value = normalize_legacy_ftp_site_text(file_template).strip()
    parsed = urlparse(folder_value)
    if parsed.scheme.lower() == "ftp":
        if parsed.hostname:
            ftp_config["host"] = parsed.hostname
        if parsed.port:
            ftp_config["port"] = parsed.port
        if parsed.username and not str(ftp_config.get("username") or "").strip():
            ftp_config["username"] = unquote(parsed.username)
        if parsed.password and not str(ftp_config.get("password") or "").strip():
            ftp_config["password"] = unquote(parsed.password)
        path = unquote(parsed.path or "/")
        if not template_value and path and not path.endswith("/"):
            template_value = Path(path).name
            folder_value = str(Path(path).parent).replace("\\", "/")
        else:
            folder_value = path or "/"
    return ftp_config, folder_value or "/", template_value


def ftp_modified_sort_key(ftp: ftplib.FTP, name: str) -> tuple[str, str]:
    try:
        response = ftp.sendcmd(f"MDTM {name}")
        stamp = response.split(maxsplit=1)[1].strip() if " " in response else response.strip()
        return (stamp, name)
    except Exception:
        return ("", name)


def parse_ftp_task(task: dict[str, Any]) -> tuple[dict[str, Any], str, str]:
    connection = task.get("connection") if isinstance(task.get("connection"), dict) else {}
    config = connection.get("config") if isinstance(connection.get("config"), dict) else {}
    template_config = decode_ftp_template_config(task.get("file_name_template"))
    folder_path = str(task.get("folder_path") or "").strip()
    file_template = str(
        template_config.get("file_name_template")
        or template_config.get("output_file_name_template")
        or task.get("file_name_template")
        or ""
    ).strip()
    return _parse_ftp_location(dict(config), folder_path, file_template)


def ftp_task_variables(task: dict[str, Any]) -> dict[str, str]:
    template_config = decode_ftp_template_config(task.get("file_name_template"))
    variables = normalize_ftp_variables(template_config.get("variables"))
    report = task.get("report") if isinstance(task.get("report"), dict) else {}
    report_config = decode_ftp_template_config(report.get("file_name_template"))
    for key, value in normalize_ftp_variables(report_config.get("variables")).items():
        variables.setdefault(key, value)
    variables.update(normalize_ftp_variables(task.get("variables")))
    return variables


def build_ftp_download_plan(task: dict[str, Any], now: datetime | None = None) -> dict[str, Any]:
    connection = task.get("connection") if isinstance(task.get("connection"), dict) else {}
    base_config = connection.get("config") if isinstance(connection.get("config"), dict) else {}
    template_config = decode_ftp_template_config(task.get("file_name_template"))
    report = task.get("report") if isinstance(task.get("report"), dict) else {}
    report_config = decode_ftp_template_config(report.get("file_name_template"))
    config = template_config or report_config
    variables = ftp_task_variables(task)
    sources_config = config.get("sources") if isinstance(config.get("sources"), list) else []
    source_items: list[dict[str, Any]] = []
    if sources_config:
        for index, source in enumerate(sources_config, start=1):
            if not isinstance(source, dict):
                continue
            label = str(source.get("name") or source.get("label") or source.get("source") or f"Nguon {index}").strip()
            if label.upper() == "HGA":
                label = "HAG"
            folder_path = normalize_legacy_ftp_site_text(render_ftp_file_template(str(source.get("folder_path") or ""), now, variables))
            file_template = normalize_legacy_ftp_site_text(render_ftp_file_template(str(source.get("file_name_template") or source.get("file") or ""), now, variables))
            source_config, folder_path, file_template = _parse_ftp_location(dict(base_config), folder_path, file_template)
            source_items.append({
                "name": label or f"Nguon {index}",
                "folder_path": folder_path,
                "file_name_template": file_template,
                "config": source_config,
            })
    else:
        source_config, folder_path, file_template = parse_ftp_task(task)
        folder_path = normalize_legacy_ftp_site_text(render_ftp_file_template(folder_path, now, variables))
        file_template = normalize_legacy_ftp_site_text(render_ftp_file_template(file_template, now, variables))
        source_items.append({
            "name": str(task.get("ma_bao_cao") or task.get("ten_bao_cao") or "FTP").strip() or "FTP",
            "folder_path": folder_path,
            "file_name_template": file_template,
            "config": source_config,
        })
    output_template = str(
        config.get("output_file_name_template")
        or config.get("output")
        or task.get("output_file_name_template")
        or f"{task.get('ma_bao_cao') or 'ftp'}_{{yyyyMMdd}}.xlsx"
    ).strip()
    output_name = normalize_legacy_ftp_site_text(render_ftp_file_template(output_template, now, variables)) or f"ftp_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    if sources_config and not Path(output_name).suffix:
        output_name = f"{output_name}.xlsx"
    return {
        "variables": variables,
        "sources": source_items,
        "output_file_name": safe_local_filename(output_name, "ftp_merged.xlsx"),
        "is_multi_source": len(source_items) > 1 or bool(sources_config),
    }


def _download_ftp_source_file(source: dict[str, Any], local_dir: Path, progress_callback=None) -> dict[str, Any]:
    config = source.get("config") if isinstance(source.get("config"), dict) else {}
    folder_path = str(source.get("folder_path") or "/").strip() or "/"
    resolved_name = str(source.get("file_name_template") or "").strip()
    host = str(config.get("host") or "").strip()
    username = str(config.get("username") or "").strip()
    password = str(config.get("password") or "")
    port = int(config.get("port") or 21)
    timeout = float(config.get("timeout_seconds") or 60)
    passive = config.get("passive", True) is not False
    if not host or not username or not password:
        raise RuntimeError("FTP thieu host, username hoac password.")
    if not resolved_name:
        raise RuntimeError("Ten file FTP sau khi render dang rong.")
    label = str(source.get("name") or "FTP")
    if progress_callback:
        progress_callback(f"Dang ket noi FTP {host}:{port} cho {label}.", "running", resolved_name)
    ftp = ftplib.FTP()
    try:
        ftp.connect(host=host, port=port, timeout=timeout)
        ftp.login(user=username, passwd=password)
        ftp.set_pasv(passive)
        if folder_path and folder_path not in {".", "/"}:
            _cwd_ftp_folder(ftp, folder_path, label)
        if progress_callback:
            progress_callback(f"Dang tim file {label}: {resolved_name}.", "running", resolved_name)
        wildcard = any(character in resolved_name for character in "*?[")
        names = ftp.nlst()
        if wildcard:
            matches = [name for name in names if fnmatch.fnmatch(Path(name).name, resolved_name)]
        else:
            matches = [name for name in names if Path(name).name.lower() == resolved_name.lower()]
            if not matches:
                matches = [resolved_name]
        if not matches:
            raise FileNotFoundError(f"Khong tim thay file FTP {label}: {resolved_name}")
        remote_name = sorted(matches, key=lambda item: ftp_modified_sort_key(ftp, item), reverse=True)[0]
        safe_label = safe_local_filename(label, "nguon")
        local_name = safe_local_filename(Path(remote_name).name or resolved_name, "ftp_result")
        local_path = (local_dir / f"{datetime.now():%Y%m%d_%H%M%S}_{safe_label}_{local_name}").resolve()
        if progress_callback:
            progress_callback(f"Dang tai file {label}: {Path(remote_name).name}.", "running", Path(remote_name).name)
        with local_path.open("wb") as handle:
            ftp.retrbinary(f"RETR {remote_name}", handle.write)
        if local_path.stat().st_size <= 0:
            local_path.unlink(missing_ok=True)
            raise RuntimeError(f"File FTP {label} tai ve rong.")
        return {
            "source": label,
            "folder_path": folder_path,
            "resolved_file_name": Path(remote_name).name,
            "file_name": local_path.name,
            "file_path": str(local_path),
        }
    finally:
        try:
            ftp.quit()
        except Exception:
            try:
                ftp.close()
            except Exception:
                pass


def _cwd_ftp_folder(ftp: ftplib.FTP, folder_path: str, label: str) -> None:
    try:
        ftp.cwd(folder_path)
        return
    except ftplib.all_errors as direct_error:
        if not folder_path.startswith("/"):
            raise RuntimeError(
                f"Khong vao duoc thu muc FTP {label}: {folder_path}. FTP tra ve: {direct_error}"
            ) from direct_error
        try:
            ftp.cwd("/")
            for segment in [part for part in folder_path.split("/") if part]:
                ftp.cwd(segment)
            return
        except ftplib.all_errors as segmented_error:
            raise RuntimeError(
                "Khong vao duoc thu muc FTP "
                f"{label}: {folder_path}. FTP tra ve: {direct_error}. "
                f"Thu vao tung cap cung loi: {segmented_error}"
            ) from segmented_error


def _read_ftp_csv_rows(path: Path) -> list[list[Any]]:
    for encoding in ("utf-8-sig", "cp1258", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = path.read_text(encoding="utf-8", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(text.splitlines(), dialect)]


def _read_ftp_xlsx_rows(path: Path) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("May tram chua cai openpyxl de gop file Excel FTP.") from error
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_ftp_tabular_rows(path: Path) -> list[list[Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_ftp_csv_rows(path)
    if suffix == ".xlsx":
        return _read_ftp_xlsx_rows(path)
    raise RuntimeError(f"Chuc nang gop nhieu file FTP chi ho tro CSV hoac XLSX, chua ho tro {suffix or 'file nay'}.")


def _merge_ftp_downloaded_files(downloads: list[dict[str, Any]], target_path: Path, progress_callback=None) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as error:
        raise RuntimeError("May tram chua cai openpyxl de xuat file tong hop FTP.") from error
    parsed: list[dict[str, Any]] = []
    headers: list[str] = []
    for item in downloads:
        source = str(item.get("source") or "FTP")
        rows = _read_ftp_tabular_rows(Path(str(item.get("file_path") or "")))
        rows = [row for row in rows if any(str(cell or "").strip() for cell in row)]
        if not rows:
            continue
        source_headers = [str(cell or "").strip() or f"Col{index}" for index, cell in enumerate(rows[0], start=1)]
        for header in source_headers:
            if header not in headers:
                headers.append(header)
        parsed.append({"source": source, "headers": source_headers, "rows": rows[1:], "file": item.get("resolved_file_name") or ""})
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TongHop"
    output_headers = ["Nguon", *headers]
    sheet.append(output_headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    row_count = 0
    for item in parsed:
        source_headers = item["headers"]
        for row in item["rows"]:
            mapped = {source_headers[index]: value for index, value in enumerate(row) if index < len(source_headers)}
            sheet.append([item["source"], *[mapped.get(header, "") for header in headers]])
            row_count += 1
    if not row_count:
        sheet.append(["", *["" for _ in headers]])
    meta = workbook.create_sheet("Nguon")
    meta.append(["Nguon", "File"])
    for item in parsed:
        meta.append([item["source"], item["file"]])
    if progress_callback:
        progress_callback(f"Da gop {len(downloads)} file FTP thanh {target_path.name}.", "running", target_path.name)
    workbook.save(target_path)


def download_ftp_report_file(task: dict[str, Any], progress_callback=None) -> dict[str, Any]:
    started = time.monotonic()
    plan = build_ftp_download_plan(task)
    sources = plan["sources"]
    if not sources:
        raise RuntimeError("Chua co nguon FTP nao de tai.")
    run_id = str(task.get("run_id") or task.get("job_id") or "").strip()
    base_dir = Path(str(get_settings().data_mining_download_dir or "data/data_mining_downloads")) / "ftp"
    local_dir = task_run_workspace(base_dir, "ftp", run_id)
    downloaded = [_download_ftp_source_file(source, local_dir, progress_callback) for source in sources]
    if not plan["is_multi_source"]:
        result = downloaded[0]
        duration_ms = int((time.monotonic() - started) * 1000)
        return {
            "ok": True,
            "status": "success",
            "message": "Da tai file FTP tren may tram.",
            "resolved_file_name": result["resolved_file_name"],
            "file_name": result["file_name"],
            "file_path": result["file_path"],
            "duration_ms": duration_ms,
        }
    output_path = (local_dir / f"{datetime.now():%Y%m%d_%H%M%S}_{plan['output_file_name']}").resolve()
    _merge_ftp_downloaded_files(downloaded, output_path, progress_callback)
    if output_path.stat().st_size <= 0:
        output_path.unlink(missing_ok=True)
        raise RuntimeError("File tong hop FTP rong.")
    duration_ms = int((time.monotonic() - started) * 1000)
    return {
        "ok": True,
        "status": "success",
        "message": f"Da tai va gop {len(downloaded)} file FTP tren may tram.",
        "resolved_file_name": plan["output_file_name"],
        "file_name": output_path.name,
        "file_path": str(output_path),
        "duration_ms": duration_ms,
        "sources": downloaded,
    }


def upload_ftp_result_file(client: httpx.Client, run_id: str, file_path: str) -> dict[str, Any]:
    return upload_task_file(client, f"/api/ftp-worker/tasks/{run_id}/file", file_path, FtpTaskCancelled)


def attach_ftp_file_if_needed(client: httpx.Client, run_id: str, result: dict[str, Any], drive_folder_id: str = "", progress_callback=None) -> dict[str, Any]:
    storage_status = str(result.get("storage_status") or "").lower()
    if storage_status.startswith("uploaded_google_drive"):
        return result
    try:
        if drive_folder_id and progress_callback:
            progress_callback("Dang upload file FTP len Google Drive qua API trung gian.")
        drive_uploaded = upload_result_file_to_internal_drive(
            str(result.get("file_path") or ""),
            drive_folder_id,
            request_source="ftp-worker",
            default_message="Da upload file FTP len Google Drive qua API trung gian.",
            job_id=run_id,
        )
    except Exception as error:
        print(f"Cannot upload FTP result to Drive through internal API: {error}", file=sys.stderr)
        if progress_callback:
            progress_callback("Upload Google Drive qua API trung gian loi, dang gui file FTP ve web.")
        drive_uploaded = {}
    if drive_uploaded:
        if progress_callback:
            progress_callback("Da upload file FTP len Google Drive.")
        merged = {**result}
        for key in ("file_name", "storage_link", "storage_status"):
            if drive_uploaded.get(key):
                merged[key] = drive_uploaded.get(key)
        merged["ok"] = True
        merged["status"] = "success"
        merged["message"] = drive_uploaded.get("message") or "Da upload file FTP len Google Drive qua API trung gian."
        return merged
    if progress_callback:
        progress_callback("Dang gui file FTP ket qua ve web de co link tai xuong.")
    uploaded = upload_ftp_result_file(client, run_id, str(result.get("file_path") or ""))
    if not uploaded:
        return result
    if progress_callback:
        progress_callback("Da gui file FTP ket qua ve web.")
    merged = {**result}
    for key in ("file_name", "file_path", "storage_link", "storage_status"):
        if uploaded.get(key):
            merged[key] = uploaded.get(key)
    return merged


def process_ftp_task(client: httpx.Client, task: dict[str, Any], worker_id: str) -> None:
    run_id = str(task.get("run_id") or "")
    drive_folder_id = str(task.get("drive_folder_id") or "").strip()
    started = time.monotonic()
    last_progress = {"message": "", "at": 0.0}

    def send_progress(message: str, status: str = "running", resolved_file_name: str = "") -> None:
        text = str(message or "").strip()
        if not text:
            return
        now = time.monotonic()
        if text == last_progress["message"] and now - float(last_progress["at"] or 0) < 3:
            return
        last_progress["message"] = text
        last_progress["at"] = now
        data = request_json(
            client,
            "POST",
            f"/api/ftp-worker/tasks/{run_id}/status",
            json={
                "status": status,
                "message": text,
                "worker_id": worker_id,
                "resolved_file_name": resolved_file_name,
                "details": {
                    **worker_process_details(),
                    "resolved_file_name": resolved_file_name,
                },
            },
        )
        if response_is_cancelled(data):
            raise FtpTaskCancelled(str(data.get("message") or "Task FTP da bi huy."))

    try:
        send_progress("May tram da nhan task FTP. Dang khoi tao ket noi.")
        result = download_ftp_report_file(task, send_progress)
        send_progress("Da tai file FTP. Dang xu ly link ket qua.", "running", str(result.get("resolved_file_name") or ""))
        result = attach_ftp_file_if_needed(client, run_id, result, drive_folder_id, send_progress)
        duration_ms = int(result.get("duration_ms") or int((time.monotonic() - started) * 1000))
        finish_response = request_json(
            client,
            "POST",
            f"/api/ftp-worker/tasks/{run_id}/result",
            json={
                "ok": bool(result.get("ok")),
                "status": result.get("status") or ("success" if result.get("ok") else "failed"),
                "message": result.get("message") or "",
                "resolved_file_name": result.get("resolved_file_name") or "",
                "file_name": result.get("file_name") or "",
                "file_path": result.get("file_path") or "",
                "storage_link": result.get("storage_link") or "",
                "storage_status": result.get("storage_status") or "",
                "duration_ms": duration_ms,
                "details": result,
            },
        )
        if response_is_cancelled(finish_response):
            return
    except FtpTaskCancelled as error:
        print(str(error), file=sys.stderr)
        return
    except Exception as error:
        message = str(error)[:500] or error.__class__.__name__
        print(f"Task FTP loi: {message}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        try:
            request_json(
                client,
                "POST",
                f"/api/ftp-worker/tasks/{run_id}/result",
                json={
                    "ok": False,
                    "status": "failed",
                    "message": message,
                    "duration_ms": int((time.monotonic() - started) * 1000),
                    "details": {"error_type": error.__class__.__name__},
                },
            )
        except Exception as update_error:
            print(f"Khong cap nhat duoc ket qua FTP: {describe_request_error(update_error)}", file=sys.stderr)


def poll_worker_once(
    client: httpx.Client,
    worker_id: str,
    poll_seconds: float,
    *,
    include_sql: bool = True,
    include_ftp: bool = True,
    dispatcher: WorkerTaskDispatcher | None = None,
) -> bool:
    if dispatcher is not None:
        dispatcher.prune_threads()

    if dispatcher is None or dispatcher.can_start(TASK_KIND_ONEBSS):
        claim = request_json(
            client,
            "POST",
            "/api/onebss-worker/tasks/claim",
            json={
                "worker_id": worker_id,
                "version": WORKER_VERSION,
                "details": {
                    **worker_process_details(),
                    **(dispatcher.active_details() if dispatcher is not None else {}),
                },
            },
            timeout=10.0,
            _retry_forever=False,
        )
        if claim.get("transient_error"):
            return False
        task = claim.get("task") if isinstance(claim.get("task"), dict) else None
        if task:
            run_id = worker_task_run_id(TASK_KIND_ONEBSS, task)
            report_code = worker_task_report_code(TASK_KIND_ONEBSS, task)
            print(f"Nhan task OneBSS {run_id} ({report_code}).", flush=True)
            try:
                status_response = request_json(
                    client,
                    "POST",
                    f"/api/onebss-worker/tasks/{run_id}/status",
                    json={
                        "status": "running",
                        "message": "May tram da nhan task OneBSS va dang chuan bi phien chay.",
                        "worker_id": worker_id,
                        "worker_session_id": "",
                        "details": {
                            **worker_process_details(),
                            "task_type": "onebss",
                            "process": "parent",
                            "stage": "task_claimed",
                            **(dispatcher.active_details() if dispatcher is not None else {}),
                        },
                    },
                    timeout=10.0,
                    _retry_forever=False,
                )
                if response_is_cancelled(status_response):
                    print(f"Task OneBSS {run_id} da bi huy truoc khi xu ly.", flush=True)
                    return True
                if status_response.get("transient_error"):
                    print(
                        f"Khong cap nhat duoc trang thai nhan task OneBSS {run_id}: {status_response.get('transient_error')}",
                        file=sys.stderr,
                        flush=True,
                    )
            except Exception as error:
                print(f"Khong cap nhat duoc trang thai nhan task OneBSS {run_id}: {describe_request_error(error)}", file=sys.stderr, flush=True)
            if dispatcher is not None and dispatcher.start_task(TASK_KIND_ONEBSS, task):
                send_heartbeat(
                    client,
                    worker_id,
                    "busy",
                    f"May tram da bat dau task OneBSS {run_id}.",
                    {
                        "run_id": run_id,
                        "report": report_code,
                        "task_type": "onebss",
                        **dispatcher.active_details(),
                    },
                )
                return True
            send_heartbeat(
                client,
                worker_id,
                "busy",
                f"Dang xu ly task {run_id}.",
                {"run_id": run_id, "report": report_code, "task_type": "onebss"},
            )
            process_task(client, task, worker_id, poll_seconds)
            send_heartbeat(client, worker_id, "idle", "May tram OneBSS da quay lai trang thai cho task.")
            return True

    if include_sql and (dispatcher is None or dispatcher.can_start(TASK_KIND_SQL)):
        sql_claim = request_json(
            client,
            "POST",
            "/api/sql-worker/tasks/claim",
            json={
                "worker_id": worker_id,
                "version": WORKER_VERSION,
                "details": {
                    **worker_process_details(),
                    **(dispatcher.active_details() if dispatcher is not None else {}),
                },
            },
            timeout=10.0,
            _retry_forever=False,
        )
        if sql_claim.get("transient_error"):
            return False
        sql_task = sql_claim.get("task") if isinstance(sql_claim.get("task"), dict) else None
        if sql_task:
            run_id = worker_task_run_id(TASK_KIND_SQL, sql_task)
            report_code = worker_task_report_code(TASK_KIND_SQL, sql_task)
            if dispatcher is not None and dispatcher.start_task(TASK_KIND_SQL, sql_task):
                send_heartbeat(
                    client,
                    worker_id,
                    "busy",
                    f"May tram da bat dau task SQL {run_id}.",
                    {"run_id": run_id, "report": report_code, "task_type": "sql", **dispatcher.active_details()},
                )
                return True
            send_heartbeat(
                client,
                worker_id,
                "busy",
                f"Dang xu ly task SQL {run_id}.",
                {"run_id": run_id, "report": report_code, "task_type": "sql"},
            )
            process_sql_task(client, sql_task, worker_id)
            send_heartbeat(client, worker_id, "idle", "May tram SQL da quay lai trang thai cho task.")
            return True

    if include_ftp and (dispatcher is None or dispatcher.can_start(TASK_KIND_FTP)):
        ftp_claim = request_json(
            client,
            "POST",
            "/api/ftp-worker/tasks/claim",
            json={
                "worker_id": worker_id,
                "version": WORKER_VERSION,
                "details": {
                    **worker_process_details(),
                    **(dispatcher.active_details() if dispatcher is not None else {}),
                },
            },
            timeout=10.0,
            _retry_forever=False,
        )
        if ftp_claim.get("transient_error"):
            return False
        ftp_task = ftp_claim.get("task") if isinstance(ftp_claim.get("task"), dict) else None
        if ftp_task:
            run_id = worker_task_run_id(TASK_KIND_FTP, ftp_task)
            report_code = worker_task_report_code(TASK_KIND_FTP, ftp_task)
            if dispatcher is not None and dispatcher.start_task(TASK_KIND_FTP, ftp_task):
                send_heartbeat(
                    client,
                    worker_id,
                    "busy",
                    f"May tram da bat dau task FTP {run_id}.",
                    {"run_id": run_id, "report": report_code, "task_type": "ftp", **dispatcher.active_details()},
                )
                return True
            send_heartbeat(
                client,
                worker_id,
                "busy",
                f"Dang xu ly task FTP {run_id}.",
                {"run_id": run_id, "report": report_code, "task_type": "ftp"},
            )
            process_ftp_task(client, ftp_task, worker_id)
            send_heartbeat(client, worker_id, "idle", "May tram FTP da quay lai trang thai cho task.")
            return True

    return False


def main() -> int:
    parser = argparse.ArgumentParser(description="Poll vnptcto.com for OneBSS/FTP export tasks and run them on this workstation.")
    parser.add_argument("--base-url", default=os.getenv("VNPTCTO_BASE_URL", "https://vnptcto.com"))
    parser.add_argument("--token", default=os.getenv("INTERNAL_API_TOKEN", ""))
    parser.add_argument("--worker-id", default=os.getenv("ONEBSS_WORKER_ID", "onebss-workstation"))
    parser.add_argument("--poll-seconds", type=float, default=float(os.getenv("ONEBSS_WORKER_POLL_SECONDS", "5")))
    parser.add_argument("--heartbeat-seconds", type=float, default=float(os.getenv("ONEBSS_WORKER_HEARTBEAT_SECONDS", "60")))
    parser.add_argument("--once", action="store_true")
    args = parser.parse_args()
    if not args.token:
        raise SystemExit("Missing INTERNAL_API_TOKEN or --token.")
    limits = worker_concurrency_limits()
    print(
        "Worker version "
        f"{WORKER_VERSION}; task_guard={'on' if onebss_task_guard_enabled() else 'off'}; "
        f"otp_wait={int(onebss_worker_otp_wait_seconds())}s; "
        f"parallel_total={limits['total']}; onebss={limits[TASK_KIND_ONEBSS]}; "
        f"sql={limits[TASK_KIND_SQL]}; ftp={limits[TASK_KIND_FTP]}.",
        flush=True,
    )

    headers = {"Authorization": f"Bearer {args.token}"}
    sql_poll_seconds = max(float(os.getenv("SQL_WORKER_POLL_SECONDS", "10") or "10"), args.poll_seconds)
    ftp_poll_seconds = max(float(os.getenv("FTP_WORKER_POLL_SECONDS", "30") or "30"), args.poll_seconds)
    next_sql_poll = 0.0
    next_ftp_poll = 0.0
    dispatcher = WorkerTaskDispatcher(args.base_url, headers, args.worker_id, args.poll_seconds)
    with httpx.Client(base_url=args.base_url.rstrip("/"), headers=headers, timeout=httpx.Timeout(60.0, connect=20.0)) as client:
        send_heartbeat(client, args.worker_id, "starting", "May tram OneBSS dang khoi dong.", dispatcher.active_details())
        last_heartbeat = time.monotonic()
        while True:
            now = time.monotonic()
            if now - last_heartbeat >= max(15.0, args.heartbeat_seconds):
                dispatcher.prune_threads()
                active_counts = dispatcher.active_counts()
                if active_counts["total"] > 0:
                    send_heartbeat(
                        client,
                        args.worker_id,
                        "busy",
                        f"May tram dang xu ly {active_counts['total']}/{worker_concurrency_limits()['total']} task.",
                        dispatcher.active_details(),
                    )
                else:
                    send_heartbeat(client, args.worker_id, "idle", "May tram OneBSS dang cho task.", dispatcher.active_details())
                last_heartbeat = now
            include_sql = now >= next_sql_poll
            include_ftp = now >= next_ftp_poll
            try:
                processed = poll_worker_once(
                    client,
                    args.worker_id,
                    args.poll_seconds,
                    include_sql=include_sql,
                    include_ftp=include_ftp,
                    dispatcher=dispatcher,
                )
            except Exception as error:
                processed = False
                print(
                    f"Vong lap worker loi: {describe_request_error(error)}. Worker se tiep tuc thu lai.",
                    file=sys.stderr,
                    flush=True,
                )
                traceback.print_exc(file=sys.stderr)
                time.sleep(max(5.0, args.poll_seconds))
            if include_sql:
                next_sql_poll = time.monotonic() + (0.0 if processed and dispatcher.has_available_slot() else sql_poll_seconds)
            if include_ftp:
                next_ftp_poll = time.monotonic() + (0.0 if processed and dispatcher.has_available_slot() else ftp_poll_seconds)
            if processed:
                last_heartbeat = time.monotonic()
            if args.once:
                dispatcher.wait_until_idle()
                return 0
            if processed and dispatcher.has_available_slot():
                time.sleep(min(1.0, max(0.2, args.poll_seconds / 5)))
            else:
                time.sleep(args.poll_seconds)


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as error:
        print(f"Worker dung do loi khong mong muon: {describe_request_error(error)}", file=sys.stderr, flush=True)
        traceback.print_exc(file=sys.stderr)
        raise SystemExit(1)
